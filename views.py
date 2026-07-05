# ============================================================
# backend/apps/accounts/permissions.py
# Custom permission classes for RBAC
# ============================================================
from rest_framework.permissions import BasePermission, SAFE_METHODS


class IsSuperAdmin(BasePermission):
    """Allow access only to super admins."""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.is_super_admin


class IsStaffOrAdmin(BasePermission):
    """Allow any authenticated staff or admin."""
    def has_permission(self, request, view):
        return request.user.is_authenticated


class IsSuperAdminOrReadOnly(BasePermission):
    """Read-only for staff; write access only for super admin."""
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.method in SAFE_METHODS:
            return True
        return request.user.is_super_admin


# ============================================================
# backend/apps/accounts/views.py
# ============================================================
from rest_framework import viewsets, status, generics, mixins
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.throttling import ScopedRateThrottle
from rest_framework.exceptions import ValidationError as DRFValidationError
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.contrib.auth.password_validation import validate_password
from django.conf import settings
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import FileResponse, HttpResponse
from django.shortcuts import render
from django.db import IntegrityError, connection, transaction
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import Sum, Count, Q, F, Exists, OuterRef, Subquery
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.html import escape
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
import base64
import io
import logging
import mimetypes
import re
import textwrap
import uuid
import zipfile
from whatsapp_service import WATIClient, log_whatsapp_message, send_candidate_document, send_candidate_message, send_whatsapp_message
from calendar import month_abbr, monthrange
from xml.etree import ElementTree

from crm.models import (
    Branch, UserTarget, UserMonthlyRating, BranchTarget, HistoricalAnalyticsEntry, Discount, BranchTransferRequest,
    DataImportHistory,
    RulesSigningRequest, UserSessionLog, WhatsAppMessage, WhatsAppTemplate, Notification, Lead, WalkIn, Payment,
    TeamNotice, TeamNoticeReply,
    PhoneNumberChangeHistory,
    PaymentInstallment, PaymentReasonRequest, PaymentReasonMessage, AdminReceipt, FollowUp, Enrollment, CourseChangeHistory,
    EnrollmentCounselorChangeHistory, EnrollmentRulesResetHistory,
    CounselorChangeRequest, CourseChangeRequest, LeadTransferHistory, WalkInAssignmentChangeRequest,
    CandidateStatusHistory,
    get_default_installment_schedule, get_enrollment_installment_schedule, get_saved_enrollment_installment_schedule,
    normalize_installment_schedule,
    enrollment_payable_fee, ENROLLMENT_PAYMENT_AMOUNT, LOW_FEE_SINGLE_PAYMENT_MAX_COURSE_FEE, MIN_INSTALLMENT_AMOUNT,
)
from crm.rating_service import calculate_kpi_rating
from serializers import (
    BranchSerializer, UserSerializer, UserTargetSerializer,
    BranchTargetSerializer, HistoricalAnalyticsEntrySerializer,
    CustomTokenObtainPairSerializer, UserPerformanceReportSerializer,
    UserMonitoringSerializer, UserMonthlyRatingSerializer,
    WhatsAppTemplateSerializer, NotificationSerializer, AdminReceiptSerializer,
    TeamNoticeSerializer, TeamNoticeReplySerializer, DataImportHistorySerializer,
    PaymentReasonRequestSerializer, CourseChangeHistorySerializer, CourseChangeRequestSerializer,
    CounselorChangeRequestSerializer, WalkInAssignmentChangeRequestSerializer, payment_installment_summary,
    user_identity_payload,
)


ENROLLMENT_COUNT_STATUSES = tuple(Enrollment.FINAL_STATUSES)
ENROLLMENT_VALUE_STATUSES = (Enrollment.Status.ENROLLED, Enrollment.Status.ACTIVE)
QUALIFICATION_KPI_FIELDS = (
    'expected_course_budget',
    'planned_joining_time',
    'primary_goal',
    'other_institutes_considering',
    'counselor_status',
    'competitor_status',
    'follow_up_priority',
    'conversion_probability',
)

COMMON_SOURCE_ALIASES = {
    'direct_walk_in': 'direct_walkin',
    'staff_referral': 'team_reference',
    'other': 'others',
}

LEAD_STATUS_FILTER_ALIASES = {
    'new_lead': [Lead.Status.NEW],
    'no_answer': [Lead.Status.NOT_ANSWERING, Lead.Status.CALL_NOT_ATTENDED],
    'continuous_no_answer': [Lead.Status.CONTINUOUSLY_NOT_ANSWERING_CALLS],
    'cna': [Lead.Status.CONTINUOUSLY_NOT_ANSWERING_CALLS],
    'na': [Lead.Status.NOT_ANSWERING, Lead.Status.CALL_NOT_ATTENDED],
    'will_walk_in': [Lead.Status.WILL_WALK_IN, Lead.Status.WALK_IN],
    'walk_in_completed': [Lead.Status.CONVERTED_TO_WALKIN],
    'ready_to_join': [Lead.Status.WILL_ENROLL],
    'joined': [Lead.Status.ENROLLED, Lead.Status.CONVERTED],
    'lost_to_competitor': [Lead.Status.JOINED_OTHER_INSTITUTE, Lead.Status.LOST],
}

WALKIN_STATUS_FILTER_ALIASES = {
    'new_lead': [WalkIn.Status.NEW],
    'new': [WalkIn.Status.NEW],
    'follow_up': [WalkIn.Status.FOLLOW_UP],
    'joined': [WalkIn.Status.CONVERTED],
    'walk_in_completed': [WalkIn.Status.NEW, WalkIn.Status.FOLLOW_UP],
    'not_interested': [WalkIn.Status.NOT_INTERESTED],
    'no_answer': [],
    'continuous_no_answer': [],
}


def normalize_source_filter(value, choices):
    value = str(value or '').strip().lower()
    if not value:
        return ''
    value = COMMON_SOURCE_ALIASES.get(value, value)
    valid_values = {choice[0] for choice in choices}
    if value == 'direct_walkin' and 'direct' in valid_values and value not in valid_values:
        value = 'direct'
    if value == 'team_reference' and 'staff_reference' in valid_values and value not in valid_values:
        value = 'staff_reference'
    return value if value in valid_values else ''


def filter_candidate_status(queryset, value, model, aliases=None):
    value = str(value or '').strip()
    if not value:
        return queryset
    aliases = aliases or {}
    if hasattr(model, 'CounselorStatus') and value in {choice[0] for choice in model.CounselorStatus.choices}:
        return queryset.filter(counselor_status=value)
    if value in aliases:
        query = Q(status__in=aliases[value])
        if hasattr(model, 'CounselorStatus'):
            query |= Q(counselor_status=value)
        return queryset.filter(query)
    return queryset.filter(status=value)


def create_status_history(record, record_type, old_status, new_status, user=None, remarks=''):
    old_status = str(old_status or '')
    new_status = str(new_status or '')
    if not new_status or old_status == new_status:
        return None
    return CandidateStatusHistory.objects.create(
        record_type=record_type,
        record_id=record.id,
        old_status=old_status,
        new_status=new_status,
        changed_by=user if getattr(user, 'is_authenticated', False) else None,
        remarks=str(remarks or '').strip(),
    )


def tracked_crm_status(record):
    if isinstance(record, Enrollment):
        return getattr(record, 'status', '')
    return getattr(record, 'counselor_status', '') or getattr(record, 'status', '')


def official_enrollment_queryset(queryset):
    return queryset.filter(status__in=ENROLLMENT_COUNT_STATUSES)


def enrollment_value_queryset(queryset):
    return queryset.filter(status__in=ENROLLMENT_VALUE_STATUSES)


def current_month_enrollment_queryset(queryset, year, month):
    return enrollment_value_queryset(queryset).filter(enrollment_date__year=year, enrollment_date__month=month)


def visible_candidate_queryset(queryset):
    model = queryset.model
    if 'is_deleted' in missing_model_columns(model, ['is_deleted']):
        return queryset
    return queryset.filter(is_deleted=False)


def visible_payment_queryset(queryset):
    if 'is_deleted' in missing_model_columns(Enrollment, ['is_deleted']):
        return queryset
    return queryset.filter(enrollment__is_deleted=False)


def candidate_identifier(candidate):
    if isinstance(candidate, Enrollment):
        return candidate.student_number or ''
    if isinstance(candidate, WalkIn):
        return candidate.candidate_number or ''
    return candidate.lead_number or ''


def candidate_record_type(candidate):
    if isinstance(candidate, Enrollment):
        return 'enrollment'
    if isinstance(candidate, WalkIn):
        return 'walkin'
    return 'lead'


def candidate_added_by(candidate):
    user = getattr(candidate, 'created_by', None) or getattr(candidate, 'enrolled_by', None)
    if not user:
        return ''
    return user.full_name or user.username or ''


def candidate_deleted_by(candidate):
    user = getattr(candidate, 'deleted_by', None)
    if not user:
        return ''
    return user.full_name or user.username or ''


def candidate_payload(candidate):
    return {
        'id': candidate.id,
        'record_type': candidate_record_type(candidate),
        'student_id': candidate_identifier(candidate),
        'name': candidate.name,
        'phone': candidate.phone,
        'branch': candidate.branch.name if candidate.branch else '',
        'branch_id': candidate.branch_id,
        'course': candidate.course.name if candidate.course else '',
        'status': candidate.status,
        'added_date': candidate.created_at,
        'added_by': candidate_added_by(candidate),
        'deleted_at': getattr(candidate, 'deleted_at', None),
        'deleted_by': candidate_deleted_by(candidate),
    }


def linked_candidate_records(candidate):
    records = [candidate]
    if isinstance(candidate, Enrollment):
        if candidate.walkin_id:
            records.append(candidate.walkin)
        if candidate.lead_id:
            records.append(candidate.lead)
        elif candidate.walkin_id and candidate.walkin.lead_id:
            records.append(candidate.walkin.lead)
    elif isinstance(candidate, WalkIn):
        if hasattr(candidate, 'enrollment') and candidate.enrollment:
            records.append(candidate.enrollment)
        if candidate.lead_id:
            records.append(candidate.lead)
    elif isinstance(candidate, Lead):
        if hasattr(candidate, 'walkin') and candidate.walkin:
            records.append(candidate.walkin)
            if hasattr(candidate.walkin, 'enrollment') and candidate.walkin.enrollment:
                records.append(candidate.walkin.enrollment)
        records.extend(list(candidate.enrollments.all()))
    unique = {}
    for record in records:
        unique[(candidate_record_type(record), record.id)] = record
    return list(unique.values())

User = get_user_model()
logger = logging.getLogger(__name__)
_TABLE_COLUMN_CACHE = {}
FORM_PROCESSING_ERROR_MESSAGE = 'Unable to process form at the moment. Please contact support.'
PUBLIC_LEAD_COURSE_NAMES = [
    'Artificial Intelligence',
    'Data Analytics',
    'Full Stack Python',
    'Full Stack Java',
    'MERN Stack',
    'Cyber Security',
    'Digital Marketing',
]


def missing_model_columns(model, field_names):
    cache_key = model._meta.db_table
    try:
        columns = _TABLE_COLUMN_CACHE.get(cache_key)
        if columns is None:
            with connection.cursor() as cursor:
                columns = {
                    column.name
                    for column in connection.introspection.get_table_description(cursor, model._meta.db_table)
                }
            _TABLE_COLUMN_CACHE[cache_key] = columns
        return [
            field_name
            for field_name in field_names
            if model._meta.get_field(field_name).column not in columns
        ]
    except Exception:
        logger.exception('Could not inspect columns for %s.', model._meta.db_table)
        return list(field_names)


def valid_walkin_enrollment(walkin):
    try:
        enrollment = getattr(walkin, 'enrollment', None)
    except Enrollment.DoesNotExist:
        enrollment = None
    if enrollment:
        return enrollment
    if walkin.converted_to_type == 'enrollment' and walkin.converted_record_id:
        return Enrollment.objects.filter(pk=walkin.converted_record_id, walkin=walkin).first()
    return None


def clear_stale_walkin_conversion(walkin):
    if not (
        walkin.converted_to_type
        or walkin.converted_record_id
        or walkin.converted_at
        or walkin.status == WalkIn.Status.CONVERTED
    ):
        return False
    if valid_walkin_enrollment(walkin):
        return False
    old_status = walkin.status
    walkin.converted_to_type = ''
    walkin.converted_record_id = None
    walkin.converted_at = None
    walkin.converted_by = None
    if walkin.status == WalkIn.Status.CONVERTED:
        walkin.status = WalkIn.Status.NEW
    walkin.save(update_fields=[
        'status', 'converted_to_type', 'converted_record_id',
        'converted_at', 'converted_by', 'updated_at',
    ])
    create_status_history(
        walkin,
        CandidateStatusHistory.RecordType.WALKIN,
        old_status,
        walkin.status,
        remarks='Stale enrollment conversion cleared.',
    )
    return True


def mark_walkin_enrollment_converted(walkin, enrollment, user, data=None):
    data = data or {}
    old_status = walkin.status
    walkin.name = data.get('name') or walkin.name
    walkin.phone = data.get('phone') or walkin.phone
    walkin.email = data.get('email') or walkin.email
    walkin.dob = data.get('dob') or walkin.dob
    walkin.location = data.get('location') or walkin.location
    walkin.pincode = data.get('pincode') or walkin.pincode
    if data.get('branch'):
        walkin.branch_id = data.get('branch')
    walkin.preferred_timing = data.get('preferred_timing') or walkin.preferred_timing
    walkin.qualification = data.get('qualification', walkin.qualification or '')
    walkin.degree = data.get('degree', walkin.degree or '')
    walkin.status = WalkIn.Status.CONVERTED
    walkin.follow_up_date = None
    walkin.converted_to_type = 'enrollment'
    walkin.converted_record_id = enrollment.id
    walkin.converted_at = timezone.now()
    walkin.converted_by = user
    walkin.save(update_fields=[
        'name', 'phone', 'email', 'dob', 'location', 'pincode',
        'branch', 'preferred_timing', 'qualification', 'degree',
        'status', 'follow_up_date', 'converted_to_type',
        'converted_record_id', 'converted_at', 'converted_by', 'updated_at',
    ])
    create_status_history(
        walkin,
        CandidateStatusHistory.RecordType.WALKIN,
        old_status,
        walkin.status,
        user,
        'Joined - No follow-up required',
    )
    system_remark = 'Joined - No follow-up required'
    if not FollowUp.objects.filter(
        record_type=FollowUp.RecordType.WALKIN,
        record_id=walkin.id,
        remarks=system_remark,
        next_follow_up_date__isnull=True,
    ).exists():
        FollowUp.objects.create(
            record_type=FollowUp.RecordType.WALKIN,
            record_id=walkin.id,
            follow_up_date=timezone.localdate(),
            next_follow_up_date=None,
            remarks=system_remark,
            updated_by=user,
        )


def mark_lead_enrollment_converted(lead, enrollment, user, data=None):
    data = data or {}
    old_status = lead.status
    lead.name = data.get('name') or lead.name
    lead.phone = data.get('phone') or lead.phone
    lead.dob = data.get('dob') or lead.dob
    lead.email = data.get('email') or lead.email
    lead.location = data.get('location') or lead.location
    lead.pincode = data.get('pincode') or lead.pincode
    lead.preferred_timing = data.get('preferred_timing') or lead.preferred_timing
    lead.qualification = data.get('qualification', lead.qualification or '')
    lead.degree = data.get('degree', lead.degree or '')
    if data.get('branch'):
        lead.branch_id = data.get('branch')
    lead.status = Lead.Status.CONVERTED
    lead.remarks = 'Joined'
    lead.next_follow_up_date = None
    lead.converted_to_type = 'enrollment'
    lead.converted_record_id = enrollment.id
    lead.converted_at = timezone.now()
    lead.converted_by = user
    lead.save(update_fields=[
        'name', 'phone', 'dob', 'email', 'location', 'pincode',
        'preferred_timing', 'qualification', 'degree', 'branch',
        'status', 'remarks', 'next_follow_up_date', 'converted_to_type',
        'converted_record_id', 'converted_at', 'converted_by', 'updated_at',
    ])
    create_status_history(
        lead,
        CandidateStatusHistory.RecordType.LEAD,
        old_status,
        lead.status,
        user,
        lead.remarks,
    )


def app_url(path):
    """Build a URL path inside the configured application mount."""
    base_path = getattr(settings, 'APP_BASE_PATH', '') or ''
    return f'{base_path}/{path.strip("/")}'


def build_login_url(request):
    return request.build_absolute_uri(app_url('login'))


def send_login_credentials_email(user, plain_password, login_url):
    body = (
        f'Hi {user.full_name},\n\n'
        'Your Indra Institute CRM account has been created.\n\n'
        f'Login URL: {login_url}\n'
        f'Username: {user.username}\n'
        f'Password: {plain_password}\n\n'
        'Please keep these credentials safe and do not share them with anyone.\n\n'
        '-Team IIE'
    )
    send_mail(
        subject='Indra Institute CRM - Login Credentials',
        message=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[user.email],
        fail_silently=False,
    )


WHATSAPP_PLACEHOLDERS = [
    'student_name',
    'candidate_name',
    'course_name',
    'branch_name',
    'phone_number',
    'total_fee',
    'paid_amount',
    'pending_amount',
    'installment_number',
    'installment_amount',
    'due_date',
    'next_payment_date',
    'follow_up_date',
    'rules_link',
    'institute_name',
]


WHATSAPP_MESSAGE_TYPE_MAP = {
    WhatsAppTemplate.TemplateType.LEAD_FOLLOW_UP: WhatsAppMessage.MsgType.FOLLOW_UP,
    WhatsAppTemplate.TemplateType.WALKIN_FOLLOW_UP: WhatsAppMessage.MsgType.WALKIN_REMIND,
    WhatsAppTemplate.TemplateType.PAYMENT_REMINDER: WhatsAppMessage.MsgType.PAYMENT_REMINDER,
    WhatsAppTemplate.TemplateType.BIRTHDAY_WISH: WhatsAppMessage.MsgType.BIRTHDAY,
    WhatsAppTemplate.TemplateType.RULES_FORM_LINK: WhatsAppMessage.MsgType.RULES_FORM_LINK,
    WhatsAppTemplate.TemplateType.OFFER_MESSAGE: WhatsAppMessage.MsgType.OFFER_MESSAGE,
}


def whatsapp_currency(value):
    if value in (None, ''):
        return ''
    return f'₹{float(value):,.0f}'


def whatsapp_date(value):
    if not value:
        return ''
    if isinstance(value, str):
        value = parse_date(value)
    if not value:
        return ''
    return value.strftime('%d %b %Y')


def render_whatsapp_template(body, values):
    message = body or ''
    normalized = {key: str(values.get(key) or '') for key in WHATSAPP_PLACEHOLDERS}
    for key, value in normalized.items():
        message = re.sub(r'{{\s*' + re.escape(key) + r'\s*}}', value, message)
    return re.sub(r'{{\s*[\w]+\s*}}', '', message).strip()


def active_whatsapp_template(template_type):
    return WhatsAppTemplate.objects.filter(
        template_type=template_type,
        is_active=True,
    ).order_by('name').first()


def whatsapp_send_payload(log):
    provider_response = log.provider_response or {}
    return {
        'whatsapp_sent': log.status == WhatsAppMessage.MsgStatus.SENT,
        'whatsapp_status': log.status,
        'whatsapp_error': log.error_message,
        'whatsapp_log_id': log.id,
        'whatsapp_url': provider_response.get('whatsapp_url') or provider_response.get('url') or '',
        'whatsapp_provider': log.provider,
    }


def candidate_values(record, extra=None):
    course = getattr(record, 'course', None)
    branch = getattr(record, 'branch', None)
    values = {
        'student_name': getattr(record, 'name', ''),
        'candidate_name': getattr(record, 'name', ''),
        'course_name': course.name if course else '',
        'branch_name': branch.name if branch else '',
        'phone_number': getattr(record, 'phone', ''),
        'follow_up_date': whatsapp_date(getattr(record, 'next_follow_up_date', None) or getattr(record, 'follow_up_date', None)),
        'institute_name': 'IIE',
    }
    values.update(extra or {})
    return values


def send_candidate_template(record, template_type, message_type, request, related_model):
    template = active_whatsapp_template(template_type)
    if not template:
        return None, Response({'detail': 'No active WhatsApp template configured for this message type.'}, status=400)
    values = candidate_values(record)
    log = send_candidate_message(
        candidate_name=getattr(record, 'name', ''),
        phone=getattr(record, 'phone', ''),
        message_type=message_type,
        message_body=template.message_body,
        template=template,
        values=values,
        sent_by=request.user,
        related_model=related_model,
        related_id=record.id,
        dedupe=False,
    )
    return log, Response({
        'detail': 'WhatsApp message processed.',
        'phone': getattr(record, 'phone', ''),
        'whatsapp_message': render_whatsapp_template(template.message_body, values),
        **whatsapp_send_payload(log),
    })


def rules_sign_view(request, token):
    """Serve the Rules signing SPA route with share-preview metadata."""
    title = 'IIE Rules & Regulations'
    description = 'Review and sign the IIE Rules & Regulations form.'
    html = render(request, 'index.html').content.decode('utf-8')

    html = re.sub(r'<title>.*?</title>', f'<title>{title}</title>', html, count=1, flags=re.IGNORECASE | re.DOTALL)
    meta = (
        f'<meta name="description" content="{description}" />\n'
        f'    <meta property="og:title" content="{title}" />\n'
        f'    <meta property="og:description" content="{description}" />\n'
        '    <meta property="og:type" content="website" />'
    )
    if 'property="og:title"' not in html:
        html = html.replace('<meta name="viewport" content="width=device-width, initial-scale=1.0" />', '<meta name="viewport" content="width=device-width, initial-scale=1.0" />\n    ' + meta)
    return HttpResponse(html)


def get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def touch_user_session(request):
    if not request.user.is_authenticated:
        return
    if request.user.is_super_admin:
        return
    session_log_id = request.data.get('session_log_id') if hasattr(request, 'data') else None
    if session_log_id:
        updated = UserSessionLog.objects.filter(
            id=session_log_id,
            user=request.user,
            is_active_session=True,
        ).update(last_seen_at=timezone.now())
        if updated:
            return
    latest = UserSessionLog.objects.filter(
        user=request.user,
        is_active_session=True,
    ).order_by('-login_at').first()
    if latest:
        latest.last_seen_at = timezone.now()
        latest.save(update_fields=['last_seen_at', 'updated_at'])


def apply_enrollment_discount(data, course, branch_id=None):
    discount_id = data.get('discount') or data.get('discount_id')
    if not discount_id:
        data['discount'] = None
        data['discount_amount'] = data.get('discount_amount') or 0
        return None

    try:
        discount = Discount.objects.prefetch_related('courses').get(pk=discount_id)
    except Discount.DoesNotExist:
        raise ValueError('Selected discount is not available.')

    if not discount.is_available_for_course(course.id, branch_id):
        raise ValueError('Selected discount is expired or not available for this course.')

    course_fee = Decimal(str(data.get('actual_fees') or course.actual_fees or 0))
    discount_amount = discount.calculate_amount(course_fee)
    data['discount'] = discount.id
    data['discount_amount'] = discount_amount
    data['discount_reason'] = discount.name
    return discount


def truthy_value(value):
    return value in (True, 'true', 'True', '1', 1, 'yes', 'on')


def normalized_date(value):
    if not value:
        return None
    if hasattr(value, 'date') and not hasattr(value, 'month'):
        return value
    if hasattr(value, 'isoformat') and hasattr(value, 'month'):
        return value
    if isinstance(value, str):
        parsed = parse_date(value)
        if parsed:
            return parsed
        for date_format in ('%d-%m-%Y', '%d/%m/%Y'):
            try:
                return timezone.datetime.strptime(value, date_format).date()
            except ValueError:
                continue
    return value


def apply_spot_conversion_discount(data, visit_date):
    is_applied = truthy_value(data.get('spot_conversion_discount_applied'))
    data['spot_conversion_discount_applied'] = is_applied
    if not is_applied:
        data['spot_conversion_discount_amount'] = 0
        return
    visit = normalized_date(visit_date)
    enrollment = normalized_date(data.get('enrollment_date'))
    if not visit or not enrollment or visit != enrollment:
        raise ValueError('Spot conversion discount is available only when visit date and enrollment date are the same.')
    data['spot_conversion_discount_amount'] = 2000


def make_json_safe_payload(data):
    payload = {}
    for key, value in data.items():
        if isinstance(value, Decimal):
            payload[key] = str(value)
        elif isinstance(value, (list, tuple)):
            payload[key] = value[0] if len(value) == 1 else list(value)
        elif hasattr(value, 'isoformat'):
            payload[key] = value.isoformat()
        else:
            payload[key] = value
    return payload


def normalize_phone_number(value):
    return re.sub(r'\D', '', str(value or ''))


def duplicate_phone_exists(new_phone, record_type, record_id):
    model_map = {
        'lead': Lead,
        'walkin': WalkIn,
        'enrollment': Enrollment,
    }
    for current_type, model in model_map.items():
        queryset = model.objects.exclude(phone__isnull=True).exclude(phone='')
        if current_type == record_type:
            queryset = queryset.exclude(pk=record_id)
        for phone in queryset.values_list('phone', flat=True):
            if normalize_phone_number(phone) == new_phone:
                return True
    return False


def matching_candidate_phone_records(phone):
    normalized = normalize_phone_number(phone)
    if not normalized:
        return []
    records = []
    for lead in Lead.objects.select_related('branch', 'course'):
        if normalize_phone_number(lead.phone) == normalized:
            records.append({
                'type': 'Lead',
                'id': lead.id,
                'name': lead.name,
                'branch_name': lead.branch.name if lead.branch else 'Unassigned',
                'course_name': lead.course.name if lead.course else '',
                'url': f'/leads/{lead.id}',
            })
    for walkin in WalkIn.objects.select_related('branch', 'course'):
        if normalize_phone_number(walkin.phone) == normalized:
            records.append({
                'type': 'Walk-in',
                'id': walkin.id,
                'name': walkin.name,
                'branch_name': walkin.branch.name if walkin.branch else '',
                'course_name': walkin.course.name if walkin.course else '',
                'url': f'/walkins/{walkin.id}',
            })
    for enrollment in Enrollment.objects.select_related('branch', 'course'):
        if normalize_phone_number(enrollment.phone) == normalized:
            records.append({
                'type': 'Student',
                'id': enrollment.id,
                'name': enrollment.name,
                'branch_name': enrollment.branch.name if enrollment.branch else '',
                'course_name': enrollment.course.name if enrollment.course else '',
                'url': f'/students/{enrollment.id}',
            })
    return records


class PhoneNumberUpdateView(APIView):
    """POST /api/phone-numbers/<record_type>/<record_id>/ - audited phone-only update."""
    permission_classes = [IsStaffOrAdmin]

    model_map = {
        'lead': Lead,
        'walkin': WalkIn,
        'enrollment': Enrollment,
        'student': Enrollment,
    }

    def get_record(self, record_type, record_id, user):
        model = self.model_map.get(record_type)
        if not model:
            return None
        queryset = model.objects.all()
        if not user.is_super_admin:
            queryset = queryset.filter(branch=user.branch)
        return queryset.filter(pk=record_id).first()

    def post(self, request, record_type, record_id):
        record = self.get_record(record_type, record_id, request.user)
        if not record:
            return Response({'detail': 'Record not found.'}, status=status.HTTP_404_NOT_FOUND)

        old_phone = normalize_phone_number(record.phone)
        new_phone = normalize_phone_number(request.data.get('phone'))
        if not new_phone:
            return Response({'phone': 'Phone number is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(new_phone) != 10:
            return Response({'phone': 'Phone number should be 10 digits.'}, status=status.HTTP_400_BAD_REQUEST)
        canonical_type = 'enrollment' if record_type == 'student' else record_type
        if new_phone == old_phone:
            return Response({'phone': record.phone, 'detail': 'Phone number unchanged.'})
        if duplicate_phone_exists(new_phone, canonical_type, int(record_id)):
            return Response(
                {'phone': 'This phone number already exists in another record.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            PhoneNumberChangeHistory.objects.create(
                record_type=canonical_type,
                record_id=record.id,
                old_phone_number=record.phone,
                new_phone_number=new_phone,
                changed_by=request.user,
            )
            record.phone = new_phone
            record.save(update_fields=['phone', 'updated_at'])

        return Response({
            'phone': record.phone,
            'detail': 'Phone number updated successfully.',
        })


ADMIN_APPROVAL_NOTIFICATION_TITLES = {
    'Counselor Change Request',
    'Counselor Change Approval Needed',
    'Counselor Change Approved by Counselor',
    'Counselor Change Rejected by Counselor',
    'Course Change Request',
    'Installment Added by User',
    'Installment Edited by User',
    'Installment Deleted by User',
    'Lead Assigned',
    'Lead Transferred',
    'Walk-in Transfer Request',
    'Walk-in Transferred',
    'Branch Transfer Request',
    'Branch Transfer Approval Required',
    'Student Status Change Request',
    'Enrollment Created',
    'New Enrollment Created',
    'Payment Awaiting Approval',
    'Payment Added',
    'Payment Reason Response Submitted',
    'Lead Transfer Approval Required',
    'Enrollment Approval Request',
}

ADMIN_ACTION_NOTIFICATION_KEYWORDS = (
    'counselor change',
    'course change request',
    'installment added',
    'installment edited',
    'installment deleted',
    'lead transfer',
    'lead transferred',
    'walk-in transfer',
    'walkin transfer',
    'walk-in transferred',
    'branch transfer',
    'student status change',
    'new enrollment created',
    'enrollment created',
    'enrollment approval',
    'payment added',
    'payment awaiting approval',
    'payment reason response submitted',
)

ADMIN_EXCLUDED_NOTIFICATION_KEYWORDS = (
    'rules & regulations',
    'rules form',
    'student handbook',
    'dashboard visit',
    'dashboard viewed',
    'performance hub',
    'page visit',
    'viewed form',
    'viewed page',
    'general activity',
)

SYSTEM_NOTIFICATION_KEYWORDS = (
    'error',
    'failed',
    'failure',
    'crashed',
)

ADMIN_VISIBLE_NOTIFICATION_CATEGORIES = (
    Notification.Category.APPROVAL,
    Notification.Category.SYSTEM,
)


def notification_category_for(title, notification_type=Notification.NType.INFO, category=None):
    if category:
        return category
    normalized_title = str(title or '').strip()
    if normalized_title in ADMIN_APPROVAL_NOTIFICATION_TITLES:
        return Notification.Category.APPROVAL
    lower_title = normalized_title.lower()
    if any(keyword in lower_title for keyword in ADMIN_ACTION_NOTIFICATION_KEYWORDS):
        return Notification.Category.APPROVAL
    if notification_type == Notification.NType.ERROR or any(keyword in lower_title for keyword in SYSTEM_NOTIFICATION_KEYWORDS):
        return Notification.Category.SYSTEM
    return Notification.Category.INFO


def admin_action_required_filter():
    title_filter = Q(title__in=ADMIN_APPROVAL_NOTIFICATION_TITLES)
    for keyword in ADMIN_ACTION_NOTIFICATION_KEYWORDS:
        title_filter |= Q(title__icontains=keyword)
    for keyword in SYSTEM_NOTIFICATION_KEYWORDS:
        title_filter |= Q(title__icontains=keyword)
    excluded_filter = Q()
    for keyword in ADMIN_EXCLUDED_NOTIFICATION_KEYWORDS:
        excluded_filter |= Q(title__icontains=keyword) | Q(message__icontains=keyword)
    return (Q(category__in=ADMIN_VISIBLE_NOTIFICATION_CATEGORIES) | title_filter) & ~excluded_filter


def create_user_notification(user, title, message, notification_type=Notification.NType.INFO, related_url='', category=None):
    if not user or not getattr(user, 'is_active', False):
        return None
    if is_admin_operational_noise(user, title):
        return None
    return Notification.objects.create(
        user=user,
        title=title,
        message=message,
        type=notification_type,
        category=notification_category_for(title, notification_type, category),
        status=Notification.Status.UNREAD,
        is_read=False,
        related_url=related_url,
    )


def notify_branch_users(branch, title, message, notification_type=Notification.NType.INFO, related_url=''):
    for user in User.objects.filter(branch=branch, is_active=True).exclude(role=User.Role.SUPER_ADMIN):
        create_user_notification(user, title, message, notification_type, related_url)


def create_notification_once(user, title, message, notification_type=Notification.NType.INFO, related_url='', category=None):
    if not user:
        return None
    if is_admin_operational_noise(user, title):
        return None
    notification = Notification.objects.filter(
        user=user,
        title=title,
        message=message,
        related_url=related_url,
    ).exclude(status=Notification.Status.RESOLVED).first()
    if notification:
        return notification
    notification = Notification.objects.create(
        user=user,
        title=title,
        message=message,
        type=notification_type,
        category=notification_category_for(title, notification_type, category),
        status=Notification.Status.UNREAD,
        is_read=False,
        related_url=related_url,
    )
    return notification


def notification_actor_name(user, fallback='Staff'):
    if not user:
        return fallback
    return (
        getattr(user, 'full_name', '')
        or getattr(user, 'username', '')
        or getattr(user, 'email', '')
        or fallback
    )


def counselor_request_url(request_id):
    return f'/counselor-change-requests?request={request_id}'


def candidate_url_for_counselor_request(change_request):
    if change_request.record_type == CounselorChangeRequest.RecordType.LEAD:
        return f'/leads/{change_request.lead_id}'
    return f'/enrollments/{change_request.enrollment_id}'


def active_admin_users():
    return User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True)


ADMIN_OPERATIONAL_NOTIFICATION_TITLES = {
    'Lead follow-up due today',
    'Missed lead follow-up',
    'Walk-in follow-up due today',
    'Missed walk-in follow-up',
    'Payment due today',
    'Birthday reminder',
    'Bill Generated',
    'Receipt Generated',
}


def is_admin_operational_noise(user, title):
    return bool(getattr(user, 'is_super_admin', False) and title in ADMIN_OPERATIONAL_NOTIFICATION_TITLES)


def notify_admin_users(title, message, notification_type=Notification.NType.INFO, related_url='', category=None):
    for admin_user in active_admin_users():
        create_user_notification(admin_user, title, message, notification_type, related_url, category=category)


def notify_admin_enrollment_created(enrollment, created_by=None):
    actor_name = (created_by.full_name or created_by.username) if created_by else 'System'
    course_name = enrollment.course.name if enrollment.course else 'Course not set'
    notify_admin_users(
        'New Enrollment Created',
        f'{enrollment.name} enrolled for {course_name} by {actor_name}.',
        Notification.NType.INFO,
        f'/enrollments/{enrollment.id}',
        category=Notification.Category.APPROVAL,
    )


def notify_counselor_request_submitted(change_request):
    url = counselor_request_url(change_request.id)
    message = (
        f'{change_request.candidate_name} counselor change requested: '
        f'{change_request.current_counselor.full_name if change_request.current_counselor else "Unassigned"} to '
        f'{change_request.requested_counselor.full_name if change_request.requested_counselor else "Unassigned"}.'
    )
    create_user_notification(change_request.current_counselor, 'Counselor Change Approval Needed', message, Notification.NType.WARNING, url)
    for admin_user in active_admin_users():
        create_user_notification(admin_user, 'Counselor Change Request', message, Notification.NType.INFO, url)


def notify_counselor_decision(change_request, approved):
    url = counselor_request_url(change_request.id)
    state = 'approved' if approved else 'rejected'
    title = f'Counselor Change {state.title()} by Counselor'
    message = f'{change_request.current_counselor.full_name if change_request.current_counselor else "Current counselor"} {state} transfer for {change_request.candidate_name}.'
    create_user_notification(change_request.requested_by, title, message, Notification.NType.SUCCESS if approved else Notification.NType.ERROR, url)
    for admin_user in active_admin_users():
        create_user_notification(admin_user, title, message, Notification.NType.SUCCESS if approved else Notification.NType.ERROR, url)


def notify_admin_decision(change_request, approved):
    url = candidate_url_for_counselor_request(change_request)
    state = 'approved' if approved else 'rejected'
    title = f'Counselor Change {state.title()}'
    message = f'Counselor change for {change_request.candidate_name} was {state}.'
    tone = Notification.NType.SUCCESS if approved else Notification.NType.ERROR
    for user in {
        change_request.requested_by,
        change_request.current_counselor,
        change_request.requested_counselor,
    }:
        create_user_notification(user, title, message, tone, url)


def current_counselor_for_record(record):
    if isinstance(record, Lead):
        return record.assigned_to or record.created_by
    return getattr(record, 'counselor', None) or record.created_by or record.enrolled_by


def create_counselor_change_request(record, record_type, requested_counselor, requested_by, reason):
    current_counselor = current_counselor_for_record(record)
    if current_counselor and current_counselor.id == requested_counselor.id:
        raise ValueError('This counselor is already assigned.')
    branch = getattr(record, 'branch', None)
    if branch and requested_counselor.branch_id != branch.id:
        raise ValueError('Select a counselor from the candidate branch.')
    if not str(reason or '').strip():
        raise ValueError('Reason for transfer is required.')
    filter_kwargs = {
        'record_type': record_type,
        'status__in': [
            CounselorChangeRequest.Status.PENDING_COUNSELOR,
            CounselorChangeRequest.Status.PENDING_ADMIN,
        ],
    }
    if record_type == CounselorChangeRequest.RecordType.LEAD:
        filter_kwargs['lead'] = record
    else:
        filter_kwargs['enrollment'] = record
    if CounselorChangeRequest.objects.filter(**filter_kwargs).exists():
        raise ValueError('This candidate already has a pending counselor change request.')

    change_request = CounselorChangeRequest.objects.create(
        record_type=record_type,
        lead=record if record_type == CounselorChangeRequest.RecordType.LEAD else None,
        enrollment=record if record_type == CounselorChangeRequest.RecordType.ENROLLMENT else None,
        branch=branch,
        candidate_name=record.name,
        candidate_phone=getattr(record, 'phone', '') or '',
        current_counselor=current_counselor,
        requested_counselor=requested_counselor,
        requested_by=requested_by,
        reason=str(reason or '').strip(),
    )
    notify_counselor_request_submitted(change_request)
    return change_request


def apply_counselor_change(change_request, admin_user, force=False):
    if change_request.record_type == CounselorChangeRequest.RecordType.LEAD:
        record = Lead.objects.select_for_update().get(pk=change_request.lead_id)
        record.assigned_to = change_request.requested_counselor
        record.created_by = change_request.requested_counselor
        record.save(update_fields=['assigned_to', 'created_by', 'updated_at'])
    else:
        record = Enrollment.objects.select_for_update().get(pk=change_request.enrollment_id)
        old_counselor = record.counselor or record.created_by or record.enrolled_by
        record.counselor = change_request.requested_counselor
        record.save(update_fields=['counselor', 'updated_at'])
        EnrollmentCounselorChangeHistory.objects.create(
            enrollment=record,
            old_counselor=old_counselor,
            new_counselor=change_request.requested_counselor,
            changed_by=admin_user,
            reason=change_request.reason,
        )
    change_request.status = CounselorChangeRequest.Status.APPROVED
    change_request.admin_decision_by = admin_user
    change_request.admin_decision_at = timezone.now()
    change_request.force_transfer = force
    return record


def walkin_assignment_user_field(field_type):
    if field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY:
        return 'assigned_to'
    if field_type == WalkInAssignmentChangeRequest.FieldType.COUNSELING_BY:
        return 'counseling_by'
    raise ValueError('Select a valid assignment field.')


def validate_walkin_assignment_user(walkin, field_type, user):
    if not user or not user.is_active or user.role != User.Role.STAFF or user.is_superuser:
        raise ValueError('Select a valid active staff user.')
    if field_type == WalkInAssignmentChangeRequest.FieldType.COUNSELING_BY:
        if not walkin.branch_id:
            raise ValueError('Candidate branch is required before assigning Counseling By.')
        if user.branch_id != walkin.branch_id:
            raise ValueError('Select an active user from the candidate branch for Counseling By.')
    return user


def create_walkin_assignment_change_request(walkin, field_type, requested_user, requested_by, reason, requested_walk_in_by=''):
    field_name = walkin_assignment_user_field(field_type)
    previous_user = getattr(walkin, field_name)
    fixed_walk_in_by_values = (WalkIn.WalkInBy.DIRECT, WalkIn.WalkInBy.FRIENDS_REFERENCE)
    requested_fixed_walk_in_by = (
        field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY
        and requested_walk_in_by in fixed_walk_in_by_values
    )
    current_fixed_walk_in_by = (
        field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY
        and walkin.walk_in_by in fixed_walk_in_by_values
    )
    if requested_fixed_walk_in_by:
        requested_user = None
    else:
        validate_walkin_assignment_user(walkin, field_type, requested_user)
    if previous_user and requested_user and previous_user.id == requested_user.id:
        raise ValueError('This user is already assigned.')
    if requested_fixed_walk_in_by and walkin.walk_in_by == requested_walk_in_by:
        raise ValueError(f'Walk-in By is already {requested_walk_in_by}.')
    if not previous_user and not current_fixed_walk_in_by:
        raise ValueError('This field is empty. Assign it directly before using change requests.')
    if not str(reason or '').strip():
        raise ValueError('Reason for change is required.')
    if WalkInAssignmentChangeRequest.objects.filter(
        walkin=walkin,
        field_type=field_type,
        status__in=[
            WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR,
            WalkInAssignmentChangeRequest.Status.PENDING_ADMIN,
        ],
    ).exists():
        raise ValueError('This walk-in already has a pending assignment change request for this field.')

    change_request = WalkInAssignmentChangeRequest.objects.create(
        walkin=walkin,
        field_type=field_type,
        branch=walkin.branch,
        candidate_name=walkin.name,
        candidate_phone=walkin.phone or '',
        previous_user=previous_user,
        previous_walk_in_by=walkin.walk_in_by if field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY else '',
        requested_user=requested_user,
        requested_walk_in_by=requested_walk_in_by if requested_fixed_walk_in_by else '',
        requested_by=requested_by,
        reason=str(reason or '').strip(),
        status=(
            WalkInAssignmentChangeRequest.Status.PENDING_ADMIN
            if requested_fixed_walk_in_by
            else WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR
        ),
    )
    title = 'Counselor Change Request'
    field_label = change_request.get_field_type_display()
    message = (
        f'{walkin.name} {field_label} change requested: '
        f'{walkin.walk_in_by if current_fixed_walk_in_by else (previous_user.full_name if previous_user else "Unassigned")} to '
        f'{requested_walk_in_by if requested_fixed_walk_in_by else (requested_user.full_name or requested_user.username)}.'
    )
    if requested_user:
        create_user_notification(
            requested_user,
            title,
            message,
            Notification.NType.WARNING,
            f'/admin/walkin-assignment-requests?request={change_request.id}',
            category=Notification.Category.APPROVAL,
        )
    else:
        notify_admin_users(
            title,
            message,
            Notification.NType.WARNING,
            f'/admin/walkin-assignment-requests?request={change_request.id}',
            category=Notification.Category.APPROVAL,
        )
    return change_request


def approve_walkin_assignment_by_counselor(change_request, counselor_user, remarks=''):
    if change_request.status != WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR:
        raise ValueError('Only requests awaiting counselor approval can be approved.')
    if change_request.requested_user_id != counselor_user.id:
        raise ValueError('Only the requested counselor can approve this request.')
    change_request.status = WalkInAssignmentChangeRequest.Status.PENDING_ADMIN
    change_request.counselor_reviewed_by = counselor_user
    change_request.counselor_reviewed_at = timezone.now()
    change_request.counselor_remarks = remarks or ''
    change_request.save(update_fields=[
        'status', 'counselor_reviewed_by', 'counselor_reviewed_at',
        'counselor_remarks', 'updated_at',
    ])
    notify_admin_users(
        'Approved Counselor Change Request Awaiting Final Approval',
        f'{change_request.requested_user.full_name or change_request.requested_user.username} approved the {change_request.get_field_type_display()} change request for {change_request.candidate_name}.',
        Notification.NType.WARNING,
        f'/admin/walkin-assignment-requests?request={change_request.id}',
        category=Notification.Category.APPROVAL,
    )
    return change_request


def apply_walkin_assignment_change(change_request, admin_user, remarks=''):
    field_name = walkin_assignment_user_field(change_request.field_type)
    walkin = WalkIn.objects.select_for_update().get(pk=change_request.walkin_id)
    requested_fixed_walk_in_by = (
        change_request.field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY
        and change_request.requested_walk_in_by in (
            WalkIn.WalkInBy.DIRECT,
            WalkIn.WalkInBy.FRIENDS_REFERENCE,
        )
    )
    if requested_fixed_walk_in_by:
        walkin.assigned_to = None
        walkin.walk_in_by = change_request.requested_walk_in_by
        walkin.save(update_fields=['assigned_to', 'walk_in_by', 'updated_at'])
    else:
        validate_walkin_assignment_user(walkin, change_request.field_type, change_request.requested_user)
        setattr(walkin, field_name, change_request.requested_user)
        update_fields = [field_name, 'updated_at']
        if change_request.field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY:
            walkin.walk_in_by = ''
            update_fields.append('walk_in_by')
        walkin.save(update_fields=update_fields)
    change_request.status = WalkInAssignmentChangeRequest.Status.APPROVED
    change_request.reviewed_by = admin_user
    change_request.reviewed_at = timezone.now()
    change_request.admin_remarks = remarks or ''
    change_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'admin_remarks', 'updated_at'])
    create_user_notification(
        change_request.requested_by,
        'Walk-in Assignment Change Approved',
        f'{change_request.get_field_type_display()} change for {change_request.candidate_name} was approved.',
        Notification.NType.SUCCESS,
        f'/walkins/{walkin.id}',
    )
    return walkin


def resolve_notifications(queryset):
    return queryset.exclude(status=Notification.Status.RESOLVED).update(
        status=Notification.Status.RESOLVED,
        is_read=True,
        resolved_at=timezone.now(),
    )


def mark_notifications_terminal(queryset, terminal_status):
    return queryset.update(
        status=terminal_status,
        is_read=True,
        resolved_at=timezone.now(),
    )


def mark_public_walkin_notifications_read(walkin_id):
    Notification.objects.filter(
        title='New public walk-in submitted',
        related_url=f'/walkins/{walkin_id}',
        status=Notification.Status.UNREAD,
    ).update(status=Notification.Status.READ, is_read=True)


def resolve_public_walkin_notifications(walkin_id):
    resolve_notifications(Notification.objects.filter(
        title='New public walk-in submitted',
        related_url=f'/walkins/{walkin_id}',
    ))


def resolve_rules_signed_notifications(enrollment_id):
    resolve_notifications(Notification.objects.filter(
        title='Rules & Regulations Signed',
        related_url=f'/enrollments/{enrollment_id}',
    ))


def payment_has_active_due(payment, due_date=None):
    due_date = due_date or timezone.localdate()
    return (
        payment.status in [Payment.Status.UNPAID, Payment.Status.PARTIAL]
        and payment.balance > 0
        and payment.next_payment_date == due_date
    )


def resolve_payment_due_notifications(payment_id):
    resolve_notifications(Notification.objects.filter(
        title='Payment due today',
        related_url=f'/payments/{payment_id}',
    ))


def resolve_payment_due_notifications_if_inactive(payment):
    if not payment_has_active_due(payment):
        resolve_payment_due_notifications(payment.id)


def notify_rules_signed(enrollment, submitted_at=None):
    title = 'Rules & Regulations Signed'
    submitted_value = submitted_at or timezone.now()
    course_name = enrollment.course.name if enrollment.course else ''
    branch_name = enrollment.branch.name if enrollment.branch else ''
    detail_lines = [
        f'{enrollment.name} has submitted the signed Rules & Regulations form.',
    ]
    if course_name:
        detail_lines.append(f'Course: {course_name}')
    if branch_name:
        detail_lines.append(f'Branch: {branch_name}')
    detail_lines.append(f'Submitted on: {submitted_value.strftime("%d %b %Y, %I:%M %p")}')
    message = '\n'.join(detail_lines)
    related_url = f'/enrollments/{enrollment.id}'
    recipients = list(User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True))
    if enrollment.branch_id:
        recipients.extend(
            User.objects.filter(branch=enrollment.branch, is_active=True).exclude(role=User.Role.SUPER_ADMIN)
        )
    seen = set()
    for user in recipients:
        if user.id in seen:
            continue
        seen.add(user.id)
        create_notification_once(user, title, message, Notification.NType.SUCCESS, related_url)


LEAD_CLOSED_FOLLOW_UP_STATUSES = [
    Lead.Status.COUNSELING_COMPLETED,
    Lead.Status.DEMO_ATTENDED,
    Lead.Status.WILL_WALK_IN,
    Lead.Status.WALK_IN,
    Lead.Status.ENROLLED,
    Lead.Status.CONVERTED,
    Lead.Status.CONVERTED_TO_WALKIN,
    Lead.Status.NOT_INTERESTED,
    Lead.Status.JOINED_OTHER_INSTITUTE,
    Lead.Status.DROPPED,
    Lead.Status.LOST,
]

WALKIN_CLOSED_FOLLOW_UP_STATUSES = [
    WalkIn.Status.CONVERTED,
    WalkIn.Status.NOT_INTERESTED,
    WalkIn.Status.TRANSFERRED,
]


def automated_lead_status_display(lead):
    status_value = effective_lead_status(lead)
    if status_value == Lead.Status.ENROLLED:
        return 'Enrolled'
    if status_value == Lead.Status.CONVERTED_TO_WALKIN:
        return 'Converted to Walk-in'
    if status_value == Lead.Status.NEW and lead.source == Lead.Source.MANUAL:
        return 'Follow-up'
    return dict(Lead.Status.choices).get(status_value, lead.get_status_display())


def automated_walkin_status_display(walkin):
    if walkin.converted_to_type == 'enrollment' or walkin.status == WalkIn.Status.CONVERTED:
        return 'Enrolled'
    return walkin.get_status_display()


def pending_follow_up_queryset(queryset, record_type, due_lookup, closed_statuses):
    completed_current_due = FollowUp.objects.filter(
        record_type=record_type,
        record_id=OuterRef('pk'),
        follow_up_date=OuterRef(due_lookup),
    )
    return queryset.exclude(
        status__in=closed_statuses,
    ).annotate(
        has_completed_current_due=Exists(completed_current_due),
    ).exclude(
        has_completed_current_due=True,
    )


def active_lead_follow_up_queryset(queryset, due_lookup, due_date):
    return pending_follow_up_queryset(
        queryset,
        FollowUp.RecordType.LEAD,
        due_lookup,
        LEAD_CLOSED_FOLLOW_UP_STATUSES,
    ).filter(**{due_lookup: due_date})


def active_walkin_follow_up_queryset(queryset, due_lookup, due_date):
    return pending_follow_up_queryset(
        queryset,
        FollowUp.RecordType.WALKIN,
        due_lookup,
        WALKIN_CLOSED_FOLLOW_UP_STATUSES,
    ).filter(**{due_lookup: due_date})


def missed_lead_follow_up_queryset(queryset, due_lookup, today):
    return pending_follow_up_queryset(
        queryset,
        FollowUp.RecordType.LEAD,
        due_lookup,
        LEAD_CLOSED_FOLLOW_UP_STATUSES,
    ).filter(**{f'{due_lookup}__lt': today})


def missed_walkin_follow_up_queryset(queryset, due_lookup, today):
    return pending_follow_up_queryset(
        queryset,
        FollowUp.RecordType.WALKIN,
        due_lookup,
        WALKIN_CLOSED_FOLLOW_UP_STATUSES,
    ).filter(**{f'{due_lookup}__lt': today})


def pending_duration_bounds(request):
    today = timezone.localdate()
    duration = request.query_params.get('duration') or ''
    if duration == 'today':
        return today, today
    if duration == 'tomorrow':
        tomorrow = today + timedelta(days=1)
        return tomorrow, tomorrow
    if duration == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    if duration == 'last7':
        return today - timedelta(days=6), today
    if duration == 'this_week':
        week_start = today - timedelta(days=(today.weekday() + 1) % 7)
        return week_start, week_start + timedelta(days=6)
    if duration == 'month':
        return today.replace(day=1), today
    if duration == 'this_month':
        start = today.replace(day=1)
        return start, start.replace(day=monthrange(start.year, start.month)[1])
    if duration == 'last_month':
        current_start = today.replace(day=1)
        end = current_start - timedelta(days=1)
        return end.replace(day=1), end
    if duration == 'last_3_months':
        return today - timedelta(days=90), today
    if duration == 'last_6_months':
        return today - timedelta(days=180), today
    if duration == 'custom':
        return (
            parse_date(request.query_params.get('date_from') or '') or None,
            parse_date(request.query_params.get('date_to') or '') or None,
        )
    return None, today


def apply_pending_date_filter(queryset, field_name, request):
    start, end = pending_duration_bounds(request)
    if start:
        queryset = queryset.filter(**{f'{field_name}__gte': start})
    if end:
        queryset = queryset.filter(**{f'{field_name}__lte': end})
    return queryset


def apply_overdue_follow_up_date_filter(queryset, field_name, request):
    today = timezone.localdate()
    overdue_end = today - timedelta(days=1)
    start, end = pending_duration_bounds(request)
    if end is None or end >= today:
        end = overdue_end
    if start and start >= today:
        return queryset.none()
    if start:
        queryset = queryset.filter(**{f'{field_name}__gte': start})
    return queryset.filter(**{f'{field_name}__lte': end})


class CRMSearchFilter(SearchFilter):
    def filter_queryset(self, request, queryset, view):
        search_terms = self.get_search_terms(request)
        if not search_terms:
            return queryset

        search_fields = self.get_search_fields(view, request)
        id_search_fields = getattr(view, 'id_search_fields', [])
        for term in search_terms:
            term_query = Q()
            for field in search_fields:
                term_query |= Q(**{f'{field}__icontains': term})
            if term.isdigit():
                for field in id_search_fields:
                    term_query |= Q(**{field: int(term)})
            queryset = queryset.filter(term_query)
        return queryset.distinct()


def apply_text_search(queryset, search, fields, id_fields=None):
    search = (search or '').strip()
    if not search:
        return queryset
    query = Q()
    for field in fields:
        query |= Q(**{f'{field}__icontains': search})
    if search.isdigit():
        for field in id_fields or []:
            query |= Q(**{field: int(search)})
    return queryset.filter(query).distinct()


def pending_staff_filter(request):
    raw = request.query_params.get('user') or request.query_params.get('counselor') or ''
    try:
        return int(raw) if raw else None
    except (TypeError, ValueError):
        return None


def truthy_query_param(value):
    return str(value or '').strip().lower() in {'1', 'true', 'yes', 'on'}


def prune_stale_follow_up_notifications(user, lead_due_ids, lead_missed_ids, walkin_due_ids, walkin_missed_ids):
    notification_rules = [
        ('Lead follow-up due today', {f'/leads/{record_id}' for record_id in lead_due_ids}),
        ('Missed lead follow-up', {f'/leads/{record_id}' for record_id in lead_missed_ids}),
        ('Walk-in follow-up due today', {f'/walkins/{record_id}' for record_id in walkin_due_ids}),
        ('Missed walk-in follow-up', {f'/walkins/{record_id}' for record_id in walkin_missed_ids}),
    ]
    for title, active_urls in notification_rules:
        stale_notifications = Notification.objects.filter(user=user, title=title)
        if active_urls:
            stale_notifications = stale_notifications.exclude(related_url__in=active_urls)
        resolve_notifications(stale_notifications)


def prune_stale_payment_due_notifications(user, active_payment_ids):
    active_urls = {f'/payments/{record_id}' for record_id in active_payment_ids}
    stale_notifications = Notification.objects.filter(user=user, title='Payment due today')
    if active_urls:
        stale_notifications = stale_notifications.exclude(related_url__in=active_urls)
    resolve_notifications(stale_notifications)


def clear_follow_up_notifications_for_record(record_type, record_id):
    if record_type == FollowUp.RecordType.LEAD:
        titles = ['Lead follow-up due today', 'Missed lead follow-up']
        related_url = f'/leads/{record_id}'
    else:
        titles = ['Walk-in follow-up due today', 'Missed walk-in follow-up']
        related_url = f'/walkins/{record_id}'
    resolve_notifications(Notification.objects.filter(title__in=titles, related_url=related_url))


def generate_smart_notifications(user):
    if not user.is_authenticated:
        return
    if user.is_super_admin:
        return
    today = timezone.localdate()
    lead_scope = visible_candidate_queryset(Lead.objects.filter(branch=user.branch))
    walkin_scope = visible_candidate_queryset(WalkIn.objects.filter(branch=user.branch))
    lead_due_qs = active_lead_follow_up_queryset(lead_scope, 'next_follow_up_date', today)
    lead_missed_qs = missed_lead_follow_up_queryset(lead_scope, 'next_follow_up_date', today)
    walkin_due_qs = active_walkin_follow_up_queryset(walkin_scope, 'follow_up_date', today)
    walkin_missed_qs = missed_walkin_follow_up_queryset(walkin_scope, 'follow_up_date', today)

    lead_due_ids = list(lead_due_qs.values_list('id', flat=True))
    lead_missed_ids = list(lead_missed_qs.values_list('id', flat=True))
    walkin_due_ids = list(walkin_due_qs.values_list('id', flat=True))
    walkin_missed_ids = list(walkin_missed_qs.values_list('id', flat=True))
    prune_stale_follow_up_notifications(user, lead_due_ids, lead_missed_ids, walkin_due_ids, walkin_missed_ids)

    for lead in lead_due_qs[:25]:
        create_notification_once(user, 'Lead follow-up due today', f'{lead.name} needs follow-up today.', Notification.NType.WARNING, f'/leads/{lead.id}')
    for lead in lead_missed_qs[:25]:
        create_notification_once(user, 'Missed lead follow-up', f'{lead.name} has a missed follow-up.', Notification.NType.ERROR, f'/leads/{lead.id}')
    for walkin in walkin_due_qs[:25]:
        create_notification_once(user, 'Walk-in follow-up due today', f'{walkin.name} needs follow-up today.', Notification.NType.WARNING, f'/walkins/{walkin.id}')
    for walkin in walkin_missed_qs[:25]:
        create_notification_once(user, 'Missed walk-in follow-up', f'{walkin.name} has a missed follow-up.', Notification.NType.ERROR, f'/walkins/{walkin.id}')
    payment_qs = visible_payment_queryset(Payment.objects.filter(
        next_payment_date=today,
        status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
        paid_amount__lt=F('total_fees'),
    ).select_related('enrollment'))
    payment_qs = payment_qs.filter(enrollment__branch=user.branch)
    payment_due_ids = list(payment_qs.values_list('id', flat=True))
    prune_stale_payment_due_notifications(user, payment_due_ids)
    for payment in payment_qs[:25]:
        create_notification_once(user, 'Payment due today', f'{payment.enrollment.name} has a payment due today.', Notification.NType.WARNING, f'/payments/{payment.id}')
    enrollment_qs = visible_candidate_queryset(Enrollment.objects.filter(
        status__in=Enrollment.FINAL_STATUSES,
        dob__month=today.month,
        dob__day=today.day,
    ))
    enrollment_qs = enrollment_qs.filter(branch=user.branch)
    for enrollment in enrollment_qs[:25]:
        create_notification_once(user, 'Birthday reminder', f'{enrollment.name} has a birthday today.', Notification.NType.WARNING, f'/students/{enrollment.id}')


def add_one_month(value):
    if not value:
        return None
    month = value.month + 1
    year = value.year + (month - 1) // 12
    month = ((month - 1) % 12) + 1
    day = min(value.day, monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def build_rules_installment_plan(enrollment):
    return [
        {
            **item,
            'date': item.get('due_date'),
        }
        for item in get_saved_enrollment_installment_schedule(enrollment)
    ]


def extract_rules_template_text():
    template_path = Path(settings.MEDIA_ROOT) / 'rules_templates' / 'rules_and_regulations.docx'
    if not template_path.exists():
        return ['Rules and Regulations template is not available.']
    try:
        with zipfile.ZipFile(template_path) as docx:
            xml = docx.read('word/document.xml')
        root = ElementTree.fromstring(xml)
        namespace = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        paragraphs = []
        for paragraph in root.findall('.//w:p', namespace):
            text = ''.join(node.text or '' for node in paragraph.findall('.//w:t', namespace)).strip()
            if text:
                paragraphs.append(text)
        return paragraphs or ['Rules and Regulations']
    except Exception:
        return ['Rules and Regulations template could not be rendered. Please contact IIE.']


def wrap_text(draw, text, font, max_width):
    words = str(text).split()
    lines = []
    current = ''
    for word in words:
        candidate = f'{current} {word}'.strip()
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines or ['']


def rules_duration_label(duration_months):
    months = int(duration_months or 0)
    if not months:
        return ''
    return '1 Month' if months == 1 else f'{months} Months'


def rules_pdf_date(value):
    if not value:
        return 'Not set'
    if isinstance(value, str):
        value = parse_date(value) or value
    if hasattr(value, 'strftime'):
        return value.strftime('%d %b %Y')
    return str(value)


def validate_data_image(image_data, label):
    if not image_data or ',' not in image_data:
        raise ValueError(f'{label} is required.')
    try:
        image_bytes = base64.b64decode(image_data.split(',', 1)[1])
        from PIL import Image
        Image.open(io.BytesIO(image_bytes)).verify()
        return image_bytes
    except ImportError as exc:
        raise RuntimeError('Form image processing dependency is unavailable.') from exc
    except ValueError:
        raise
    except Exception:
        raise ValueError(f'Invalid {label.lower()} image.')


def image_storage_extension(image_bytes):
    try:
        from PIL import Image
        with Image.open(io.BytesIO(image_bytes)) as image:
            image_format = (image.format or 'JPEG').lower()
    except Exception:
        return 'jpg'
    if image_format == 'jpeg':
        return 'jpg'
    return re.sub(r'[^a-z0-9]', '', image_format) or 'jpg'


def build_signed_rules_pdf(enrollment, signature_bytes=None, selfie_bytes=None, submitted_at=None):
    try:
        from PIL import Image, ImageDraw, ImageFont, ImageOps
    except ImportError as exc:
        raise RuntimeError('Signed Rules & Regulation PDF generation dependency is unavailable.') from exc

    width, height = 1240, 1754
    margin = 80
    content_width = width - margin * 2

    def load_font(weight, size):
        font_candidates = [
            Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'fonts' / f'libertinus-serif-{weight}.woff2',
            Path(settings.BASE_DIR) / 'staticfiles' / 'assets' / f'libertinus-serif-{weight}.woff2',
            Path('/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf'),
            Path('/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf'),
            Path('C:/Windows/Fonts/georgia.ttf'),
            Path('C:/Windows/Fonts/times.ttf'),
        ]
        for font_path in font_candidates:
            try:
                if font_path.exists():
                    return ImageFont.truetype(str(font_path), size=size)
            except Exception:
                continue
        return ImageFont.load_default()

    font = load_font(400, 21)
    value_font = load_font(600, 21)
    label_font = load_font(600, 17)
    section_font = load_font(700, 26)
    title_font = load_font(700, 36)
    small_font = load_font(400, 19)
    pages = []
    page = Image.new('RGB', (width, height), 'white')
    draw = ImageDraw.Draw(page)
    y = margin
    ink = '#111827'
    muted = '#64748b'
    border = '#cbd5e1'
    soft = '#f8fafc'

    def add_page():
        nonlocal page, draw, y
        pages.append(page)
        page = Image.new('RGB', (width, height), 'white')
        draw = ImageDraw.Draw(page)
        y = margin

    def ensure_space(required):
        if y + required > height - margin - 250:
            add_page()

    def write_wrapped(text, current_font=font, fill=ink, spacing=34, max_width=content_width, x=margin):
        nonlocal y
        for line in wrap_text(draw, text, current_font, max_width):
            if y > height - margin - 250:
                add_page()
            draw.text((x, y), line, fill=fill, font=current_font)
            y += spacing

    def write_section(title):
        nonlocal y
        ensure_space(70)
        if y > margin:
            y += 18
        draw.text((margin, y), title, fill=ink, font=section_font)
        y += 42

    def write_detail_grid(items):
        nonlocal y
        column_gap = 24
        row_gap = 18
        box_width = (content_width - column_gap) // 2
        box_height = 92
        for index in range(0, len(items), 2):
            ensure_space(box_height + row_gap)
            row_items = items[index:index + 2]
            for column, (label, value) in enumerate(row_items):
                x = margin + column * (box_width + column_gap)
                draw.rounded_rectangle((x, y, x + box_width, y + box_height), radius=10, fill=soft, outline=border, width=1)
                draw.text((x + 18, y + 16), str(label).upper(), fill=muted, font=label_font)
                value_lines = wrap_text(draw, value or 'Not set', value_font, box_width - 36)
                draw.text((x + 18, y + 46), value_lines[0], fill=ink, font=value_font)
                if len(value_lines) > 1:
                    draw.text((x + 18, y + 70), value_lines[1], fill=ink, font=small_font)
            y += box_height + row_gap

    def write_installment(item):
        nonlocal y
        ensure_space(112)
        draw.rounded_rectangle((margin, y, margin + content_width, y + 96), radius=10, fill=soft, outline=border, width=1)
        draw.text((margin + 18, y + 16), str(item.get('label') or 'Installment'), fill=ink, font=value_font)
        draw.text((margin + 18, y + 52), f'Due Date: {rules_pdf_date(item.get("date") or item.get("due_date"))}', fill=muted, font=small_font)
        amount_text = f'Rs {Decimal(str(item.get("amount") or 0)):,.0f}'
        amount_width = draw.textbbox((0, 0), amount_text, font=value_font)[2]
        draw.text((margin + content_width - amount_width - 18, y + 30), amount_text, fill=ink, font=value_font)
        y += 114

    draw.text((margin, y), 'IIE Rules & Regulations Form', fill=ink, font=title_font)
    y += 58
    if selfie_bytes:
        selfie = Image.open(io.BytesIO(selfie_bytes)).convert('RGB')
        card_width = 230
        card_height = 290
        image_width = 196
        image_height = 220
        card_x = width - margin - card_width
        card_y = margin - 8
        image_x = card_x + 17
        image_y = card_y + 17
        resample = getattr(getattr(Image, 'Resampling', Image), 'LANCZOS', Image.LANCZOS)
        photo = ImageOps.fit(selfie, (image_width, image_height), method=resample)
        draw.rounded_rectangle(
            (card_x, card_y, card_x + card_width, card_y + card_height),
            radius=12,
            fill='white',
            outline=border,
            width=1,
        )
        page.paste(photo, (image_x, image_y))
        draw.rounded_rectangle(
            (image_x, image_y, image_x + image_width, image_y + image_height),
            radius=8,
            outline=border,
            width=1,
        )
        draw.text((image_x, image_y + image_height + 18), 'Student Identity Photo', fill=muted, font=label_font)
        y = max(y, card_y + card_height + 34)
    details = [
        ('Name', enrollment.name),
        ('Phone', enrollment.phone),
        ('Course Enrolled', enrollment.course.name if enrollment.course else ''),
        ('Batch Timing', enrollment.batch_timing or enrollment.get_preferred_timing_display() or ''),
        ('Batch Start Date', rules_pdf_date(enrollment.start_date)),
        ('Duration', rules_duration_label(enrollment.course.duration_months if enrollment.course else None)),
        ('Final Payable Fees', f'Rs {Decimal(str(enrollment_payable_fee(enrollment) or 0)):,.0f}'),
    ]
    write_section('Candidate Details')
    write_detail_grid(details)

    write_section('Payment Schedule')
    for item in build_rules_installment_plan(enrollment):
        write_installment(item)

    write_section('Rules and Regulations')
    for paragraph in extract_rules_template_text():
        write_wrapped(paragraph, current_font=font, fill=ink, spacing=34)
        y += 16

    if y > height - margin - 260:
        add_page()
    y += 18
    draw.text((margin, y), 'Student Signature', fill=ink, font=section_font)
    if signature_bytes:
        signature = Image.open(io.BytesIO(signature_bytes)).convert('RGBA')
        signature.thumbnail((420, 180))
        page.paste(signature, (margin, y + 46), signature)
        submitted_value = submitted_at or timezone.now()
        draw.text((margin, y + 225), f'Submitted At: {timezone.localtime(submitted_value).strftime("%d %b %Y %I:%M %p")}', fill=muted, font=small_font)
    else:
        line_y = y + 150
        draw.line((margin, line_y, margin + 420, line_y), fill=border, width=2)
        draw.text((margin, line_y + 24), 'To be signed by the student', fill=muted, font=small_font)
    pages.append(page)

    output = io.BytesIO()
    pages[0].save(output, format='PDF', save_all=True, append_images=pages[1:])
    output.seek(0)
    return output.read()


def create_enrollment_from_transfer_request(transfer_request, reviewer):
    from serializers import EnrollmentDetailSerializer

    if transfer_request.enrollment_id or hasattr(transfer_request.walkin, 'enrollment'):
        raise ValueError('This walk-in is already enrolled.')

    payload = dict(transfer_request.enrollment_payload or {})
    payload['branch'] = transfer_request.requested_branch_id
    required_fields = [
        'name', 'phone', 'course', 'branch', 'preferred_timing',
        'enrollment_date', 'start_date',
    ]
    missing = [field for field in required_fields if payload.get(field) in (None, '')]
    if missing:
        raise ValueError(f'Please complete all mandatory fields: {", ".join(missing)}.')
    serializer = EnrollmentDetailSerializer(data=payload)
    serializer.is_valid(raise_exception=True)
    enrollment = serializer.save(
        walkin=transfer_request.walkin,
        original_walkin_course=transfer_request.walkin.course,
        final_enrollment_course_id=payload.get('course'),
        enrolled_by=reviewer,
        created_by=reviewer,
        status=Enrollment.Status.PENDING_RULES,
    )

    walkin = transfer_request.walkin
    mark_walkin_enrollment_converted(walkin, enrollment, reviewer, payload)
    clear_follow_up_notifications_for_record(FollowUp.RecordType.WALKIN, walkin.id)
    resolve_public_walkin_notifications(walkin.id)

    transfer_request.enrollment = enrollment
    transfer_request.status = transfer_request.Status.APPROVED
    transfer_request.reviewed_by = reviewer
    transfer_request.reviewed_at = timezone.now()
    transfer_request.save(update_fields=['enrollment', 'status', 'reviewed_by', 'reviewed_at', 'updated_at'])
    notify_admin_enrollment_created(enrollment, reviewer)
    return enrollment


class LoginView(TokenObtainPairView):
    """POST /api/auth/login/ — Returns access + refresh JWT tokens."""
    serializer_class = CustomTokenObtainPairSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        user_data = response.data.get('user') if response.status_code == 200 else None
        if user_data and user_data.get('role') != User.Role.SUPER_ADMIN:
            session_log = UserSessionLog.objects.create(
                user_id=user_data['id'],
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
            )
            response.data['session_log_id'] = session_log.id
        return response


class LogoutView(APIView):
    """POST /api/auth/logout/ — Blacklists the refresh token."""
    permission_classes = [IsStaffOrAdmin]

    def post(self, request):
        if not request.user.is_super_admin:
            session_log_id = request.data.get('session_log_id')
            session_qs = UserSessionLog.objects.filter(
                user=request.user,
                is_active_session=True,
                logout_at__isnull=True,
            )
            if session_log_id:
                session_qs = session_qs.filter(id=session_log_id)
            else:
                latest = session_qs.order_by('-login_at').first()
                session_qs = session_qs.filter(id=latest.id) if latest else session_qs.none()
            session_qs.update(
                logout_at=timezone.now(),
                last_seen_at=timezone.now(),
                is_active_session=False,
            )
        try:
            refresh = RefreshToken(request.data['refresh'])
            refresh.blacklist()
            return Response({'detail': 'Successfully logged out.'}, status=200)
        except Exception:
            return Response({'detail': 'Invalid token.'}, status=400)


class MeView(generics.RetrieveUpdateAPIView):
    """GET/PATCH /api/auth/me/ — Current user profile."""
    permission_classes = [IsStaffOrAdmin]
    serializer_class   = UserSerializer

    def get_object(self):
        touch_user_session(self.request)
        return self.request.user


class ChangePasswordView(APIView):
    """POST /api/auth/change-password/ - lets users clear first-login password change."""
    permission_classes = [IsStaffOrAdmin]

    def post(self, request):
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')

        if not current_password or not user.check_password(current_password):
            return Response({'current_password': 'Current password is incorrect.'}, status=status.HTTP_400_BAD_REQUEST)
        if not new_password:
            return Response({'new_password': 'New password is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_password(new_password, user=user)
        except DjangoValidationError as exc:
            return Response({'new_password': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.must_change_password = False
        user.save(update_fields=['password', 'must_change_password', 'updated_at'])
        return Response({'detail': 'Password changed successfully.'})


class BranchViewSet(viewsets.ModelViewSet):
    """CRUD for branches. Only super admin can write."""
    serializer_class   = BranchSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    filter_backends    = [SearchFilter, OrderingFilter]
    pagination_class   = None
    search_fields      = ['name', 'city']

    def get_queryset(self):
        queryset = Branch.objects.order_by('name')
        if self.request.user.is_super_admin:
            return queryset
        return queryset.filter(is_active=True)


class UserViewSet(viewsets.ModelViewSet):
    """Manage staff users. Only super admin."""
    queryset           = User.objects.all().select_related('branch').order_by('username')
    serializer_class   = UserSerializer
    permission_classes = [IsSuperAdmin]
    filter_backends    = [DjangoFilterBackend, SearchFilter]
    filterset_fields   = ['role', 'branch', 'is_active']
    pagination_class   = None
    search_fields      = ['username', 'email', 'first_name', 'last_name']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        headers = self.get_success_headers(serializer.data)

        try:
            send_login_credentials_email(
                user=user,
                plain_password=user._plain_password,
                login_url=build_login_url(request),
            )
        except Exception:
            return Response(
                {
                    **serializer.data,
                    'detail': 'User created, but credential email failed to send.',
                    'credential_email_sent': False,
                },
                status=status.HTTP_201_CREATED,
                headers=headers,
            )

        return Response(
            {
                **serializer.data,
                'detail': 'User created successfully. Login credentials sent to email.',
                'credential_email_sent': True,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        user = self.get_object()
        new_password = request.data.get('new_password')
        if not new_password or len(new_password) < 8:
            return Response({'error': 'Password must be at least 8 characters.'}, status=400)
        user.set_password(new_password)
        user.save(update_fields=['password', 'updated_at'])
        return Response({'detail': 'Password reset successfully.'})


class UserTargetViewSet(viewsets.ModelViewSet):
    """Set monthly targets per user."""
    queryset           = UserTarget.objects.select_related('user').order_by('-year', '-month')
    serializer_class   = UserTargetSerializer
    permission_classes = [IsSuperAdmin]
    pagination_class   = None
    filterset_fields   = ['user', 'month', 'year']


class BranchTargetViewSet(viewsets.ModelViewSet):
    """Monthly targets per branch. Staff can read only their branch target."""
    serializer_class = BranchTargetSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['branch', 'month', 'year']
    pagination_class = None
    ordering_fields = ['year', 'month', 'branch']
    ordering = ['-year', '-month', 'branch']

    def get_queryset(self):
        queryset = BranchTarget.objects.select_related('branch', 'created_by')
        if self.request.user.is_super_admin:
            return queryset
        return queryset.filter(branch=self.request.user.branch)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def update(self, request, *args, **kwargs):
        return Response(
            {'detail': 'Target records cannot be edited. Delete the existing target and create a new one.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def partial_update(self, request, *args, **kwargs):
        return Response(
            {'detail': 'Target records cannot be edited. Delete the existing target and create a new one.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )


class HistoricalAnalyticsEntryViewSet(viewsets.ModelViewSet):
    """Admin CRUD for manually entered 2023-2025 branch analytics."""
    serializer_class = HistoricalAnalyticsEntrySerializer
    permission_classes = [IsSuperAdmin]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['branch', 'month', 'year']
    pagination_class = None
    ordering_fields = ['year', 'month', 'branch']
    ordering = ['-year', '-month', 'branch']

    def get_queryset(self):
        return HistoricalAnalyticsEntry.objects.select_related('branch', 'created_by')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class UserPerformanceReportView(APIView):
    """GET /api/reports/user-performance/?month=YYYY-MM — super admin user summary."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        month_str = request.query_params.get('month', timezone.now().strftime('%Y-%m'))
        year, month = map(int, month_str.split('-'))

        users = User.objects.exclude(role=User.Role.SUPER_ADMIN).select_related('branch').order_by('username')
        report_rows = []

        def attainment(actual, target):
            actual_value = float(actual or 0)
            target_value = float(target or 0)
            if target_value <= 0:
                return 100.0 if actual_value > 0 else 0.0
            return min((actual_value / target_value) * 100, 100.0)

        for user in users:
            target = BranchTarget.objects.filter(branch=user.branch, year=year, month=month).first()
            row_metrics = scoped_metrics(
                user_id=user.id,
                start_date=date(year, month, 1),
                end_date=date(year, month, monthrange(year, month)[1]),
            )
            leads_count = row_metrics['leads']
            transferred_leads_count = LeadTransferHistory.objects.filter(
                from_user=user, created_at__year=year, created_at__month=month
            ).count()
            received_leads_count = LeadTransferHistory.objects.filter(
                to_user=user, created_at__year=year, created_at__month=month
            ).count()
            walkins_count = row_metrics['walkins']
            enrollments_count = row_metrics['enrollments']
            value_total = row_metrics['revenue']

            lead_target = target.lead_target if target else 0
            walkin_target = target.walkin_target if target else 0
            enroll_target = target.enroll_target if target else 0
            value_target = target.revenue_target if target else 0
            performance_score = round(
                (
                    attainment(leads_count, lead_target)
                    + attainment(walkins_count, walkin_target)
                    + attainment(enrollments_count, enroll_target)
                    + attainment(value_total, value_target)
                ) / 4,
                2,
            )

            report_rows.append({
                'position': 0,
                'user_id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'branch_id': user.branch_id,
                'branch_name': user.branch.name if user.branch else None,
                'target_scope': 'branch',
                'lead_target': lead_target,
                'walkin_target': walkin_target,
                'enroll_target': enroll_target,
                'revenue_target': value_target,
                'value_target': value_target,
                'leads': leads_count,
                'transferred_leads': transferred_leads_count,
                'received_leads': received_leads_count,
                'walkins': walkins_count,
                'enrollments': enrollments_count,
                'revenue': value_total,
                'value': value_total,
                'performance_score': performance_score,
            })

        report_rows.sort(
            key=lambda row: (
                -row['performance_score'],
                -row['value'],
                -row['enrollments'],
                -row['walkins'],
                -row['leads'],
                row['username'],
            )
        )

        for index, row in enumerate(report_rows, start=1):
            row['position'] = index

        serializer = UserPerformanceReportSerializer(report_rows, many=True)
        return Response(serializer.data)


def month_bounds_for(value=None):
    today = timezone.localdate()
    if value:
        try:
            year, month = [int(part) for part in str(value).split('-', 1)]
            start = today.replace(year=year, month=month, day=1)
        except (TypeError, ValueError):
            start = today.replace(day=1)
    else:
        start = today.replace(day=1)
    return start, start.replace(day=monthrange(start.year, start.month)[1])


def previous_month_bounds(start_date):
    previous_end = start_date - timedelta(days=1)
    previous_start = previous_end.replace(day=1)
    return previous_start, previous_end


def date_range_from_request(request):
    quick_range = (
        request.query_params.get('date_range')
        or request.query_params.get('duration')
        or request.query_params.get('range')
        or ''
    )
    quick_range = str(quick_range).strip().lower()
    today = timezone.localdate()
    if quick_range and quick_range != 'custom':
        if quick_range == 'today':
            return today, today
        if quick_range == 'yesterday':
            yesterday = today - timedelta(days=1)
            return yesterday, yesterday
        if quick_range == 'this_month':
            start = today.replace(day=1)
            return start, start.replace(day=monthrange(start.year, start.month)[1])
        if quick_range == 'last_month':
            current_start = today.replace(day=1)
            end = current_start - timedelta(days=1)
            return end.replace(day=1), end
        if quick_range in ('last_3_months', 'last3months'):
            return today - timedelta(days=90), today
        if quick_range in ('last_6_months', 'last6months'):
            return today - timedelta(days=180), today
    start_date = parse_date(str(request.query_params.get('date_from') or '').strip())
    end_date = parse_date(str(request.query_params.get('date_to') or '').strip())
    if start_date and end_date:
        return start_date, end_date
    if start_date:
        return start_date, timezone.localdate()
    if end_date:
        return end_date.replace(day=1), end_date
    return month_bounds_for(request.query_params.get('month'))


def pct_change(current, previous):
    current_value = float(current or 0)
    previous_value = float(previous or 0)
    if previous_value == 0:
        return 100.0 if current_value > 0 else 0.0
    return round(((current_value - previous_value) / previous_value) * 100, 2)


def ratio(numerator, denominator):
    denominator_value = float(denominator or 0)
    if denominator_value <= 0:
        return 0.0
    return round((float(numerator or 0) / denominator_value) * 100, 2)


def bounded_ratio(numerator, denominator):
    return min(ratio(numerator, denominator), 100.0)


def counselor_owner_q(field_name):
    return Q(**{f'{field_name}__role': User.Role.STAFF, f'{field_name}__is_active': True})


def reporting_enrollment_queryset(queryset):
    return enrollment_value_queryset(visible_candidate_queryset(queryset))


def enrollment_counselor_filter_q(user_id, prefix=''):
    base = f'{prefix}__' if prefix else ''
    return (
        Q(**{f'{base}counselor_id': user_id})
        | Q(**{f'{base}counselor__isnull': True, f'{base}walkin__assigned_to_id': user_id})
        | Q(**{f'{base}counselor__isnull': True, f'{base}lead__assigned_to_id': user_id})
        | Q(**{f'{base}counselor__isnull': True, f'{base}created_by_id': user_id})
    )


def candidate_search_query(model, value):
    """Return the common free-text search used by lists and aggregate reports."""
    value = str(value or '').strip()
    if not value:
        return Q()
    query = Q(name__icontains=value) | Q(phone__icontains=value) | Q(email__icontains=value)
    query |= Q(course__name__icontains=value) | Q(branch__name__icontains=value)
    if model is Lead:
        query |= Q(lead_number__icontains=value) | Q(source_description__icontains=value)
        query |= Q(assigned_to__first_name__icontains=value) | Q(assigned_to__last_name__icontains=value) | Q(assigned_to__username__icontains=value)
    elif model is WalkIn:
        query |= Q(candidate_number__icontains=value)
        query |= Q(assigned_to__first_name__icontains=value) | Q(assigned_to__last_name__icontains=value) | Q(assigned_to__username__icontains=value)
        query |= Q(counseling_by__first_name__icontains=value) | Q(counseling_by__last_name__icontains=value) | Q(counseling_by__username__icontains=value)
    else:
        query |= Q(student_number__icontains=value)
        query |= Q(counselor__first_name__icontains=value) | Q(counselor__last_name__icontains=value) | Q(counselor__username__icontains=value)
    return query


def reporting_scope(branch=None, user_id=None, course=None, source=None, status_value=None, search=None):
    leads = visible_candidate_queryset(Lead.objects.all())
    walkins = visible_candidate_queryset(WalkIn.objects.all())
    enrollments = reporting_enrollment_queryset(Enrollment.objects.all())
    followups = FollowUp.objects.all()

    if branch:
        leads = leads.filter(branch_id=branch)
        walkins = walkins.filter(branch_id=branch)
        enrollments = enrollments.filter(branch_id=branch)
        followups = followups.filter(updated_by__branch_id=branch)
    if user_id:
        leads = leads.filter(assigned_to_id=user_id)
        walkins = walkins.filter(assigned_to_id=user_id)
        enrollments = enrollments.filter(enrollment_counselor_filter_q(user_id))
        followups = followups.filter(updated_by_id=user_id)
    if course:
        leads = leads.filter(course_id=course)
        walkins = walkins.filter(course_id=course)
        enrollments = enrollments.filter(course_id=course)
    if source:
        lead_source = normalize_source_filter(source, Lead.Source.choices)
        walkin_source = normalize_source_filter(source, WalkIn.Source.choices)
        enrollment_source = normalize_source_filter(source, Enrollment._meta.get_field('source').choices)
        leads = leads.filter(source=lead_source) if lead_source else leads.none()
        walkins = walkins.filter(source=walkin_source) if walkin_source else walkins.none()
        enrollments = enrollments.filter(source=enrollment_source) if enrollment_source else enrollments.none()
    if status_value:
        leads = filter_candidate_status(leads, status_value, Lead, LEAD_STATUS_FILTER_ALIASES)
        walkins = filter_candidate_status(walkins, status_value, WalkIn, WALKIN_STATUS_FILTER_ALIASES)
        if status_value == Enrollment.Status.ACTIVE:
            enrollments = enrollments.filter(status__in=[Enrollment.Status.ACTIVE, Enrollment.Status.ENROLLED])
        elif status_value == 'pending':
            enrollments = enrollments.filter(status__in=[Enrollment.Status.DRAFT, Enrollment.Status.PENDING_RULES, Enrollment.Status.RULES_SENT, Enrollment.Status.RULES_SUBMITTED])
        elif status_value in {choice[0] for choice in Enrollment.Status.choices}:
            enrollments = enrollments.filter(status=status_value)
        else:
            enrollments = enrollments.none()
    if search:
        leads = leads.filter(candidate_search_query(Lead, search)).distinct()
        walkins = walkins.filter(candidate_search_query(WalkIn, search)).distinct()
        enrollments = enrollments.filter(candidate_search_query(Enrollment, search)).distinct()
    if source or status_value or search:
        followups = followups.filter(
            Q(record_type=FollowUp.RecordType.LEAD, record_id__in=leads.values('id'))
            | Q(record_type=FollowUp.RecordType.WALKIN, record_id__in=walkins.values('id'))
        )

    return {
        'leads': leads,
        'walkins': walkins,
        'enrollments': enrollments,
        'followups': followups,
    }


def scoped_metrics(branch=None, user_id=None, course=None, source=None, status_value=None, search=None, start_date=None, end_date=None):
    scope = reporting_scope(branch=branch, user_id=user_id, course=course, source=source, status_value=status_value, search=search)
    if not start_date or not end_date:
        start_date, end_date = month_bounds_for()
    return performance_metrics(scope, start_date, end_date)


def month_filter_kwargs(date_field, year, month):
    return {f'{date_field}__year': year, f'{date_field}__month': month}


def seconds_to_duration(total_seconds):
    total_seconds = max(int(total_seconds or 0), 0)
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    if days:
        return f'{days}d {hours}h {minutes}m'
    if hours:
        return f'{hours}h {minutes}m'
    return f'{minutes}m'


SESSION_ACTIVITY_GRACE_SECONDS = 90


def session_duration_seconds(session, end_limit=None):
    login_at = session.login_at
    last_seen_at = session.last_seen_at or login_at
    active_until = last_seen_at + timedelta(seconds=SESSION_ACTIVITY_GRACE_SECONDS)
    if session.logout_at:
        end_at = min(session.logout_at, active_until)
    elif session.is_active_session:
        end_at = min(timezone.now(), active_until)
    else:
        end_at = last_seen_at
    if end_limit:
        end_at = min(end_at, end_limit)
    return max(int((end_at - login_at).total_seconds()), 0)


def usage_seconds_for(user, start_date, end_date):
    start_dt = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_dt = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    now = timezone.now()
    effective_end_dt = min(end_dt, now)
    if effective_end_dt <= start_dt:
        return 0, {}
    sessions = UserSessionLog.objects.filter(
        user=user,
        login_at__lte=effective_end_dt,
        last_seen_at__gte=start_dt,
    ).only('login_at', 'logout_at', 'last_seen_at', 'is_active_session').order_by('login_at', 'last_seen_at')
    intervals = []
    for session in sessions:
        if not session.login_at:
            continue
        if not session.last_seen_at:
            continue
        if session.logout_at:
            raw_end = min(session.logout_at, session.last_seen_at + timedelta(seconds=SESSION_ACTIVITY_GRACE_SECONDS))
        elif session.is_active_session:
            raw_end = min(now, session.last_seen_at + timedelta(seconds=SESSION_ACTIVITY_GRACE_SECONDS))
        else:
            raw_end = session.last_seen_at
        if raw_end <= session.login_at:
            continue
        session_start = max(session.login_at, start_dt)
        session_end = min(raw_end, effective_end_dt)
        if session_end <= session_start:
            continue
        intervals.append((session_start, session_end))
    if not intervals:
        return 0, {}

    merged = []
    for interval_start, interval_end in sorted(intervals):
        if not merged or interval_start > merged[-1][1]:
            merged.append([interval_start, interval_end])
        else:
            merged[-1][1] = max(merged[-1][1], interval_end)

    total = 0
    daily = {}
    for interval_start, interval_end in merged:
        cursor = interval_start
        while cursor < interval_end:
            local_cursor = timezone.localtime(cursor)
            day_end = timezone.make_aware(datetime.combine(local_cursor.date(), datetime.max.time()))
            slice_end = min(interval_end, day_end)
            seconds = max(int((slice_end - cursor).total_seconds()), 0)
            if seconds:
                day_key = local_cursor.date().isoformat()
                daily[day_key] = daily.get(day_key, 0) + seconds
                total += seconds
            cursor = slice_end + timedelta(microseconds=1)

    max_window_seconds = max(int((effective_end_dt - start_dt).total_seconds()), 0)
    if total > max_window_seconds:
        total = max_window_seconds
    for day_key, seconds in list(daily.items()):
        daily[day_key] = min(seconds, 86400)
    return total, daily


def count_by_date(queryset, date_field, start_date, end_date):
    rows = (
        queryset
        .filter(**{f'{date_field}__date__gte': start_date, f'{date_field}__date__lte': end_date})
        .annotate(day=TruncDate(date_field))
        .values('day')
        .annotate(count=Count('id'))
        .order_by('day')
    )
    return {row['day'].isoformat(): row['count'] for row in rows if row.get('day')}


def enrollment_value(queryset):
    return queryset.aggregate(total=Sum('net_payable_fee'))['total'] or 0


def performance_metrics(scope, start_date, end_date):
    leads = scope['leads'].filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    walkins = scope['walkins'].filter(visit_date__gte=start_date, visit_date__lte=end_date)
    enrollments = reporting_enrollment_queryset(scope['enrollments']).filter(
        enrollment_date__gte=start_date,
        enrollment_date__lte=end_date,
    )
    followups = scope['followups'].filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    pending_leads = scope['leads'].filter(next_follow_up_date__lt=timezone.localdate()).exclude(status__in=[
        Lead.Status.ENROLLED,
        Lead.Status.CONVERTED,
        Lead.Status.CONVERTED_TO_WALKIN,
        Lead.Status.LOST,
        Lead.Status.DROPPED,
    ])
    pending_walkins = scope['walkins'].filter(follow_up_date__lt=timezone.localdate()).exclude(status__in=[
        WalkIn.Status.CONVERTED,
        WalkIn.Status.NOT_INTERESTED,
    ])
    leads_count = leads.count()
    walkins_count = walkins.count()
    direct_walkins_count = walkins.filter(walk_in_by=WalkIn.WalkInBy.DIRECT, assigned_to__isnull=True).count()
    friends_reference_walkins_count = walkins.filter(
        walk_in_by=WalkIn.WalkInBy.FRIENDS_REFERENCE,
        assigned_to__isnull=True,
    ).count()
    staff_referrals_count = walkins.filter(assigned_to__isnull=False).count()
    enrollments_count = enrollments.count()
    converted_walkins = walkins.filter(status=WalkIn.Status.CONVERTED).count()
    revenue = enrollment_value(enrollments)
    return {
        'leads': leads_count,
        'walkins': walkins_count,
        'direct_walkins': direct_walkins_count,
        'friends_reference_walkins': friends_reference_walkins_count,
        'staff_referrals': staff_referrals_count,
        'staff_generated_walkins': staff_referrals_count,
        'walkins_converted': converted_walkins,
        'enrollments': enrollments_count,
        'conversion_ratio': bounded_ratio(enrollments_count, walkins_count),
        'conversion_basis': 'walkins',
        'pending_followups': pending_leads.count() + pending_walkins.count(),
        'total_followups': followups.count(),
        'revenue': revenue,
    }


def performance_comparison_rows(current, previous, keys):
    rows = []
    for key, label in keys:
        rows.append({
            'key': key,
            'label': label,
            'this_month': current.get(key, 0),
            'last_month': previous.get(key, 0),
            'change_percent': pct_change(current.get(key, 0), previous.get(key, 0)),
        })
    return rows


def daily_usage_points(daily_usage, start_date, end_date, cap_day=True):
    points = []
    cursor = start_date
    while cursor <= end_date:
        key = cursor.isoformat()
        seconds = int((daily_usage or {}).get(key, 0) or 0)
        if cap_day:
            seconds = min(seconds, 86400)
        points.append({
            'date': key,
            'label': cursor.strftime('%d %b'),
            'weekday': cursor.strftime('%A'),
            'seconds': seconds,
            'hours': round(seconds / 3600, 2),
            'display': seconds_to_duration(seconds),
        })
        cursor += timedelta(days=1)
    return points


def usage_summary_for(user, start_date, end_date):
    today = timezone.localdate()
    effective_end = min(end_date, today)
    if effective_end < start_date:
        return 0, {}, [], 0, '0m', 'No activity'
    total_seconds, daily_usage = usage_seconds_for(user, start_date, effective_end)
    days_in_range = max((effective_end - start_date).days + 1, 1)
    total_seconds = min(total_seconds, days_in_range * 86400)
    average_seconds = min(int(total_seconds / days_in_range), 86400)
    points = daily_usage_points(daily_usage, start_date, effective_end)
    active_point = max(points, key=lambda item: item['seconds']) if points else None
    most_active_day = active_point['weekday'] if active_point and active_point['seconds'] > 0 else 'No activity'
    return total_seconds, daily_usage, points, average_seconds, seconds_to_duration(average_seconds), most_active_day


def report_money(value):
    try:
        amount = Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        amount = Decimal('0')
    return f'Rs {amount:,.0f}'


def report_percent(value):
    return f'{float(value or 0):.2f}%'


def build_monthly_consolidated_report_pdf(payload):
    from PIL import Image, ImageDraw, ImageFont

    width, height = 1240, 1754
    margin = 70
    pages = []
    title_font = ImageFont.load_default()
    header_font = ImageFont.load_default()
    normal_font = ImageFont.load_default()
    small_font = ImageFont.load_default()

    def new_page():
        page = Image.new('RGB', (width, height), 'white')
        return page, ImageDraw.Draw(page), margin

    def footer(draw, number):
        draw.line((margin, height - 70, width - margin, height - 70), fill='#e2e8f0', width=2)
        draw.text((margin, height - 50), 'Indra Institute of Education - Management Review Copy', fill='#475569', font=small_font)
        draw.text((width - margin - 80, height - 50), f'Page {number}', fill='#475569', font=small_font)

    def draw_table(draw, y, headers, rows, widths, max_rows=20):
        x = margin
        draw.rectangle((margin, y, width - margin, y + 44), fill='#f1f5f9')
        for index, header in enumerate(headers):
            draw.text((x + 10, y + 14), str(header), fill='#334155', font=small_font)
            x += widths[index]
        y += 44
        for row in rows[:max_rows]:
            x = margin
            draw.line((margin, y, width - margin, y), fill='#e2e8f0', width=1)
            for index, value in enumerate(row):
                draw.text((x + 10, y + 11), str(value)[:32], fill='#0f172a', font=small_font)
                x += widths[index]
            y += 38
        return y + 20

    def draw_trend(draw, y, title, points, color='#334155'):
        draw.text((margin, y), title, fill='#0f172a', font=header_font)
        y += 38
        chart_x, chart_y, chart_w, chart_h = margin, y, width - (margin * 2), 190
        draw.rectangle((chart_x, chart_y, chart_x + chart_w, chart_y + chart_h), outline='#cbd5e1', width=2)
        values = [float(item.get('value') or 0) for item in points]
        if values:
            max_value = max(max(values), 1)
            coords = []
            for index, value in enumerate(values):
                x = chart_x + (index / max(len(values) - 1, 1)) * chart_w
                point_y = chart_y + chart_h - (value / max_value) * (chart_h - 20) - 10
                coords.append((x, point_y))
            for first, second in zip(coords, coords[1:]):
                draw.line((first[0], first[1], second[0], second[1]), fill=color, width=4)
            for x, point_y in coords:
                draw.ellipse((x - 5, point_y - 5, x + 5, point_y + 5), fill=color)
        else:
            draw.text((chart_x + 20, chart_y + 80), 'No data recorded for this period.', fill='#64748b', font=normal_font)
        return y + chart_h + 35

    period = payload.get('period') or {}
    start_date = period.get('start')
    month_text = start_date.strftime('%B %Y') if hasattr(start_date, 'strftime') else str(start_date or '')
    generated_at = timezone.localtime(timezone.now()).strftime('%d %b %Y %I:%M %p')
    metrics = payload.get('metrics') or {}
    payment = payload.get('payment_summary') or {}

    page, draw, y = new_page()
    draw.text((margin, y), 'IIE Monthly Consolidated Report', fill='#0f172a', font=title_font)
    y += 54
    draw.text((margin, y), 'Indra Institute of Education', fill='#334155', font=header_font)
    y += 36
    draw.text((margin, y), f'Month: {month_text}   Generated: {generated_at}', fill='#475569', font=normal_font)
    y += 58
    cards = [
        ('Total Leads', metrics.get('leads', 0)),
        ('Total Walk-ins', metrics.get('walkins', 0)),
        ('Total Enrollments', metrics.get('enrollments', 0)),
        ('Total Revenue', report_money(metrics.get('revenue', 0))),
        ('Conversion Ratio', report_percent(metrics.get('conversion_ratio', 0))),
        ('Total Collection', report_money(payment.get('total_collection', 0))),
        ('Pending Payments', payment.get('pending_payments', 0)),
        ('Overdue Payments', payment.get('overdue_payments', 0)),
        ('Collection Ratio', report_percent(payment.get('collection_ratio', 0))),
    ]
    card_w, card_h = 340, 105
    for index, (label, value) in enumerate(cards):
        col = index % 3
        row = index // 3
        x = margin + col * (card_w + 40)
        cy = y + row * (card_h + 28)
        draw.rectangle((x, cy, x + card_w, cy + card_h), outline='#cbd5e1', width=2)
        draw.text((x + 18, cy + 18), label, fill='#64748b', font=small_font)
        draw.text((x + 18, cy + 54), str(value), fill='#0f172a', font=header_font)
    y += 3 * (card_h + 28) + 25
    draw.text((margin, y), 'Branch Summary', fill='#0f172a', font=header_font)
    y += 44
    branch_rows = [[
        row.get('branch_name', '-'), row.get('leads', 0), row.get('walkins', 0),
        row.get('enrollments', 0), report_money(row.get('revenue', 0)), report_percent(row.get('conversion_ratio', 0)),
    ] for row in payload.get('branch_comparison', [])]
    draw_table(draw, y, ['Branch', 'Leads', 'Walk-ins', 'Enrolls', 'Revenue', 'Conv.'], branch_rows, [260, 130, 150, 150, 210, 140], 12)
    footer(draw, 1)
    pages.append(page)

    page, draw, y = new_page()
    draw.text((margin, y), 'Counselor Summary', fill='#0f172a', font=header_font)
    y += 44
    counselor_rows = [[
        row.get('name', '-'), row.get('leads', 0), row.get('walkins', 0),
        row.get('enrollments', 0), report_percent(row.get('conversion_ratio', 0)), report_money(row.get('revenue', 0)),
    ] for row in payload.get('counselor_comparison', [])]
    y = draw_table(draw, y, ['Counselor', 'Leads', 'Walk-ins', 'Enrolls', 'Conv.', 'Revenue'], counselor_rows, [310, 110, 140, 140, 140, 220], 24)
    y += 15
    draw.text((margin, y), 'Payment Summary', fill='#0f172a', font=header_font)
    y += 42
    draw_table(draw, y, ['Metric', 'Value'], [
        ['Total Collection', report_money(payment.get('total_collection', 0))],
        ['Pending Payments', payment.get('pending_payments', 0)],
        ['Pending Amount', report_money(payment.get('pending_amount', 0))],
        ['Overdue Payments', payment.get('overdue_payments', 0)],
        ['Collection Ratio', report_percent(payment.get('collection_ratio', 0))],
    ], [520, 420], 10)
    footer(draw, 2)
    pages.append(page)

    page, draw, y = new_page()
    draw.text((margin, y), 'Monthly Trends', fill='#0f172a', font=header_font)
    y += 46
    trends = payload.get('trends') or {}
    y = draw_trend(draw, y, 'Leads Trend', trends.get('leads') or [], '#1d4ed8')
    y = draw_trend(draw, y, 'Walk-ins Trend', trends.get('walkins') or [], '#b45309')
    y = draw_trend(draw, y, 'Enrollment Trend', trends.get('enrollments') or [], '#047857')
    draw_trend(draw, y, 'Revenue Trend', trends.get('revenue') or [], '#6d28d9')
    footer(draw, 3)
    pages.append(page)

    buffer = io.BytesIO()
    pages[0].save(buffer, format='PDF', save_all=True, append_images=pages[1:])
    return buffer.getvalue()


def highest_enrollment_day(scope, start_date, end_date):
    rows = (
        scope['enrollments']
        .filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)
        .values('enrollment_date')
        .annotate(count=Count('id'))
        .order_by('-count', 'enrollment_date')
    )
    top = rows.first()
    if not top or not top.get('enrollment_date'):
        return 'No enrollments yet'
    return f"{top['enrollment_date'].strftime('%A')} ({top['count']} enrollments)"


def best_metric_label(rows):
    if not rows:
        return 'No activity yet'
    best = max(rows, key=lambda row: float(row.get('change_percent') or 0))
    return f"{best['label']} ({best['change_percent']:.0f}%)"


def user_scope(user):
    return {
        'leads': Lead.objects.filter(assigned_to=user, is_deleted=False),
        'walkins': WalkIn.objects.filter(assigned_to=user, is_deleted=False),
        'enrollments': reporting_enrollment_queryset(Enrollment.objects.filter(enrollment_counselor_filter_q(user.id))),
        'followups': FollowUp.objects.filter(updated_by=user),
    }


def branch_scope(branch):
    return {
        'leads': Lead.objects.filter(branch=branch, is_deleted=False),
        'walkins': WalkIn.objects.filter(branch=branch, is_deleted=False),
        'enrollments': reporting_enrollment_queryset(Enrollment.objects.filter(branch=branch)),
        'followups': FollowUp.objects.filter(updated_by__branch=branch),
    }


def organization_scope():
    return {
        'leads': Lead.objects.filter(is_deleted=False),
        'walkins': WalkIn.objects.filter(is_deleted=False),
        'enrollments': reporting_enrollment_queryset(Enrollment.objects.all()),
        'followups': FollowUp.objects.all(),
    }


def average_followup_response_seconds(user, start_date, end_date):
    followups = FollowUp.objects.filter(
        updated_by=user,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    ).order_by('record_type', 'record_id', 'created_at')
    first_followups = {}
    for item in followups:
        first_followups.setdefault((item.record_type, item.record_id), item.created_at)
    if not first_followups:
        return 0
    lead_ids = [record_id for record_type, record_id in first_followups if record_type == FollowUp.RecordType.LEAD]
    walkin_ids = [record_id for record_type, record_id in first_followups if record_type == FollowUp.RecordType.WALKIN]
    created_lookup = {
        (FollowUp.RecordType.LEAD, item.id): item.created_at
        for item in Lead.objects.filter(id__in=lead_ids).only('id', 'created_at')
    }
    created_lookup.update({
        (FollowUp.RecordType.WALKIN, item.id): item.created_at
        for item in WalkIn.objects.filter(id__in=walkin_ids).only('id', 'created_at')
    })
    durations = []
    for key, first_at in first_followups.items():
        created_at = created_lookup.get(key)
        if created_at and first_at >= created_at:
            durations.append((first_at - created_at).total_seconds())
    return int(sum(durations) / len(durations)) if durations else 0


class PerformanceHubView(APIView):
    """Role-aware performance analytics for staff and organization reporting for admins."""
    permission_classes = [IsStaffOrAdmin]

    def admin_response(self, request, start_date, end_date, previous_start, previous_end):
        scope = organization_scope()
        current = performance_metrics(scope, start_date, end_date)
        previous = performance_metrics(scope, previous_start, previous_end)
        comparison = performance_comparison_rows(current, previous, [
            ('leads', 'Total Leads'),
            ('walkins', 'Total Walk-ins'),
            ('direct_walkins', 'Direct Walk-ins'),
            ('friends_reference_walkins', 'Friends Reference'),
            ('staff_referrals', 'Staff Referrals'),
            ('enrollments', 'Total Enrollments'),
            ('conversion_ratio', 'Conversion %'),
            ('revenue', 'Revenue Generated'),
        ])

        users = User.objects.exclude(role=User.Role.SUPER_ADMIN).select_related('branch')
        user_rows = []
        total_user_seconds = 0
        daily_totals = {}
        for row_user in users:
            row_metrics = performance_metrics(user_scope(row_user), start_date, end_date)
            usage_seconds, row_daily = usage_seconds_for(row_user, start_date, end_date)
            total_user_seconds += usage_seconds
            for day_key, seconds in row_daily.items():
                daily_totals[day_key] = daily_totals.get(day_key, 0) + seconds
            rating = calculate_user_monthly_rating(row_user, start_date.year, start_date.month)
            user_rows.append({
                'user_id': row_user.id,
                'name': row_user.full_name,
                'branch_name': row_user.branch.name if row_user.branch else '',
                'screen_time_seconds': usage_seconds,
                'screen_time_display': seconds_to_duration(usage_seconds),
                'rating_score': rating.score if rating else 100,
                'rating_stars': rating.stars if rating else 5,
                **row_metrics,
            })
        user_rows.sort(key=lambda item: (-item['enrollments'], -float(item['revenue'] or 0), -item['rating_score'], item['name']))
        ranked_users = [{**row, 'rank': index + 1} for index, row in enumerate(user_rows)]

        branch_rows = []
        for branch in Branch.objects.filter(is_active=True).order_by('name'):
            row_metrics = performance_metrics(branch_scope(branch), start_date, end_date)
            branch_seconds = sum(row['screen_time_seconds'] for row in user_rows if row.get('branch_name') == branch.name)
            branch_rows.append({
                'branch_id': branch.id,
                'branch_name': branch.name,
                'screen_time_seconds': branch_seconds,
                'screen_time_display': seconds_to_duration(branch_seconds),
                **row_metrics,
            })
        branch_rows.sort(key=lambda item: (-item['enrollments'], -float(item['revenue'] or 0), item['branch_name']))

        days_in_range = max((end_date - start_date).days + 1, 1)
        daily_points = daily_usage_points(daily_totals, start_date, end_date, cap_day=False)
        total_screen_seconds = sum(point['seconds'] for point in daily_points)
        average_screen_seconds = int(total_screen_seconds / max(users.count(), 1)) if users.exists() else 0
        average_daily_seconds = int(total_screen_seconds / days_in_range)
        return Response({
            'role': User.Role.SUPER_ADMIN,
            'dashboard_type': 'admin',
            'period': {'start': start_date, 'end': end_date},
            'metrics': current,
            'previous_metrics': previous,
            'monthly_comparison': comparison,
            'changes': {
                key: pct_change(current[key], previous[key])
                for key in ['leads', 'walkins', 'enrollments', 'conversion_ratio', 'revenue']
            },
            'branch_performance': branch_rows,
            'user_performance_ranking': ranked_users,
            'top_performers': ranked_users[:5],
            'lowest_performers': sorted(ranked_users, key=lambda item: (item['enrollments'], float(item['revenue'] or 0), item['rating_score'], item['name']))[:5],
            'screen_time': {
                'total_seconds': total_screen_seconds,
                'total_display': seconds_to_duration(total_screen_seconds),
                'average_user_seconds': average_screen_seconds,
                'average_user_display': seconds_to_duration(average_screen_seconds),
                'average_daily_seconds': average_daily_seconds,
                'average_daily_display': seconds_to_duration(average_daily_seconds),
                'user_wise': [
                    {
                        'user_id': row['user_id'],
                        'name': row['name'],
                        'branch_name': row['branch_name'],
                        'seconds': row['screen_time_seconds'],
                        'display': row['screen_time_display'],
                    }
                    for row in ranked_users
                ],
                'branch_wise': [
                    {
                        'branch_id': row['branch_id'],
                        'branch_name': row['branch_name'],
                        'seconds': row['screen_time_seconds'],
                        'display': row['screen_time_display'],
                    }
                    for row in branch_rows
                ],
                'daily': daily_points,
            },
        })

    def get(self, request):
        user = request.user
        start_date, end_date = date_range_from_request(request)
        previous_start, previous_end = previous_month_bounds(start_date)
        if user.is_super_admin:
            return self.admin_response(request, start_date, end_date, previous_start, previous_end)

        personal_scope = user_scope(user)
        branch = user.branch
        branch_data_scope = branch_scope(branch) if branch else {
            'leads': Lead.objects.none(),
            'walkins': WalkIn.objects.none(),
            'enrollments': Enrollment.objects.none(),
            'followups': FollowUp.objects.none(),
        }
        current = performance_metrics(personal_scope, start_date, end_date)
        previous = performance_metrics(personal_scope, previous_start, previous_end)
        branch_current = performance_metrics(branch_data_scope, start_date, end_date)
        branch_previous = performance_metrics(branch_data_scope, previous_start, previous_end)
        today = timezone.localdate()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        today_seconds, today_daily = usage_seconds_for(user, today, today)
        week_seconds, _ = usage_seconds_for(user, week_start, today)
        month_seconds, month_daily = usage_seconds_for(user, month_start, today)
        range_seconds, daily_usage, usage_points, average_daily_seconds, average_daily_display, most_active_day = usage_summary_for(
            user,
            start_date,
            end_date,
        )
        rating = calculate_user_monthly_rating(user, start_date.year, start_date.month)
        current['crm_usage_time'] = seconds_to_duration(range_seconds)
        current['avg_followup_response_time'] = seconds_to_duration(average_followup_response_seconds(user, start_date, end_date))
        personal_comparison = performance_comparison_rows(current, previous, [
            ('leads', 'Leads'),
            ('walkins', 'Walk-ins'),
            ('walkins_converted', 'Walk-ins Converted'),
            ('enrollments', 'Enrollments'),
            ('conversion_ratio', 'Conversion %'),
            ('revenue', 'Revenue Contribution'),
        ])
        monthly_comparison = performance_comparison_rows(current, previous, [
            ('leads', 'Leads'),
            ('walkins', 'Walk-ins'),
            ('enrollments', 'Enrollments'),
            ('conversion_ratio', 'Conversion %'),
            ('revenue', 'Revenue'),
        ])
        branch_comparison = performance_comparison_rows(branch_current, branch_previous, [
            ('revenue', 'Branch Revenue'),
            ('leads', 'Branch Leads'),
            ('walkins', 'Branch Walk-ins'),
            ('direct_walkins', 'Branch Direct Walk-ins'),
            ('friends_reference_walkins', 'Branch Friends Reference'),
            ('staff_referrals', 'Branch Staff Referrals'),
            ('enrollments', 'Branch Enrollments'),
            ('conversion_ratio', 'Branch Conversion %'),
        ])
        lead_change = pct_change(current['leads'], previous['leads'])
        enrollment_change = pct_change(current['enrollments'], previous['enrollments'])
        usage_change = pct_change(range_seconds, usage_seconds_for(user, previous_start, previous_end)[0])
        usage_direction = 'increased' if usage_change >= 0 else 'decreased'
        insights = [
            {'label': 'Best Performing Metric', 'value': best_metric_label(personal_comparison)},
            {'label': 'Highest Enrollment Day', 'value': highest_enrollment_day(personal_scope, start_date, end_date)},
            {'label': 'Lead Growth', 'value': f'{lead_change:.0f}% vs last month'},
            {'label': 'Enrollment Growth', 'value': f'{enrollment_change:.0f}% vs last month'},
            {'label': 'Usage Trend', 'value': f'Your CRM usage {usage_direction} by {abs(usage_change):.0f}% compared to last month.'},
        ]
        return Response({
            'role': user.role,
            'dashboard_type': 'user',
            'period': {'start': start_date, 'end': end_date},
            'personal': current,
            'previous_personal': previous,
            'my_performance_comparison': personal_comparison,
            'monthly_comparison': monthly_comparison,
            'usage': {
                'today_seconds': today_seconds,
                'today_display': seconds_to_duration(today_seconds),
                'week_seconds': week_seconds,
                'week_display': seconds_to_duration(week_seconds),
                'month_seconds': month_seconds,
                'month_display': seconds_to_duration(month_seconds),
                'monthly_seconds': range_seconds,
                'monthly_display': seconds_to_duration(range_seconds),
                'range_seconds': range_seconds,
                'range_display': seconds_to_duration(range_seconds),
                'average_daily_seconds': average_daily_seconds,
                'average_daily_display': average_daily_display,
                'most_active_day': most_active_day,
                'daily': usage_points or [
                    {'date': key, 'seconds': value, 'display': seconds_to_duration(value)}
                    for key, value in sorted((daily_usage or today_daily or month_daily).items())
                ],
            },
            'monthly_rating': UserMonthlyRatingSerializer(rating).data if rating else None,
            'branch_overview': {
                **branch_current,
                'branch_id': branch.id if branch else None,
                'branch_name': branch.name if branch else '',
                'growth_percent': pct_change(branch_current['enrollments'], branch_previous['enrollments']),
            },
            'previous_branch': branch_previous,
            'branch_performance_comparison': branch_comparison,
            'insights': insights,
        })


class AdminAnalyticsDashboardView(APIView):
    """Filtered business analytics for super admin reports."""
    permission_classes = [IsSuperAdmin]

    def _filtered_scope(self, request):
        start_date, end_date = date_range_from_request(request)
        branch = request.query_params.get('branch')
        user_id = request.query_params.get('user')
        course = request.query_params.get('course')
        source = request.query_params.get('source')
        status_value = request.query_params.get('status')
        search = request.query_params.get('search')
        return start_date, end_date, reporting_scope(branch=branch, user_id=user_id, course=course, source=source, status_value=status_value, search=search)

    def get(self, request):
        start_date, end_date, scope = self._filtered_scope(request)
        previous_start, previous_end = previous_month_bounds(start_date)
        metrics = performance_metrics(scope, start_date, end_date)
        previous = performance_metrics(scope, previous_start, previous_end)
        lead_trend = count_by_date(scope['leads'], 'created_at', start_date, end_date)
        walkin_trend = {
            row['visit_date'].isoformat(): row['count']
            for row in scope['walkins'].filter(visit_date__gte=start_date, visit_date__lte=end_date)
                .values('visit_date').annotate(count=Count('id')).order_by('visit_date')
        }
        enrollment_trend = {
            row['enrollment_date'].isoformat(): row['count']
            for row in scope['enrollments'].filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)
                .values('enrollment_date').annotate(count=Count('id')).order_by('enrollment_date')
        }
        revenue_trend = {
            row['enrollment_date'].isoformat(): row['total'] or 0
            for row in scope['enrollments'].filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)
                .values('enrollment_date').annotate(total=Sum('net_payable_fee')).order_by('enrollment_date')
        }
        branch_filter = request.query_params.get('branch')
        user_filter = request.query_params.get('user')
        course_filter = request.query_params.get('course')
        source_filter = request.query_params.get('source')
        status_filter = request.query_params.get('status')
        search_filter = request.query_params.get('search')
        users = User.objects.filter(is_active=True).exclude(role=User.Role.SUPER_ADMIN).select_related('branch')
        if branch_filter:
            users = users.filter(branch_id=branch_filter)
        if user_filter:
            users = users.filter(id=user_filter)
        counselor_rows = []
        for user in users:
            row_metrics = scoped_metrics(
                branch=branch_filter,
                user_id=user.id,
                course=course_filter,
                source=source_filter,
                status_value=status_filter,
                search=search_filter,
                start_date=start_date,
                end_date=end_date,
            )
            counselor_rows.append({
                'user_id': user.id,
                'name': user.full_name,
                'branch_name': user.branch.name if user.branch else '',
                **row_metrics,
            })
        if not user_filter:
            unassigned_scope = reporting_scope(branch=branch_filter, course=course_filter, source=source_filter, status_value=status_filter, search=search_filter)
            unassigned_scope['leads'] = unassigned_scope['leads'].filter(assigned_to__isnull=True)
            unassigned_scope['walkins'] = unassigned_scope['walkins'].filter(assigned_to__isnull=True)
            unassigned_scope['enrollments'] = unassigned_scope['enrollments'].filter(
                counselor__isnull=True,
                walkin__assigned_to__isnull=True,
                lead__assigned_to__isnull=True,
                created_by__isnull=True,
            )
            unassigned_metrics = performance_metrics(unassigned_scope, start_date, end_date)
            if any(unassigned_metrics.get(key) for key in ('leads', 'walkins', 'enrollments', 'revenue')):
                counselor_rows.append({
                    'user_id': None,
                    'name': 'Unassigned',
                    'branch_name': '',
                    **unassigned_metrics,
                })
        counselor_rows.sort(key=lambda item: (-item['enrollments'], -float(item['revenue'] or 0), item['name']))
        branch_rows = []
        branches = Branch.objects.order_by('name')
        if branch_filter:
            branches = branches.filter(id=branch_filter)
        for branch in branches:
            row_metrics = scoped_metrics(
                branch=branch.id,
                user_id=user_filter,
                course=course_filter,
                source=source_filter,
                status_value=status_filter,
                search=search_filter,
                start_date=start_date,
                end_date=end_date,
            )
            branch_rows.append({
                'branch_id': branch.id,
                'branch_name': branch.name,
                **row_metrics,
            })
        if not branch_filter:
            unassigned_branch_scope = reporting_scope(user_id=user_filter, course=course_filter, source=source_filter, status_value=status_filter, search=search_filter)
            unassigned_branch_scope['leads'] = unassigned_branch_scope['leads'].filter(branch__isnull=True)
            unassigned_branch_scope['walkins'] = unassigned_branch_scope['walkins'].filter(branch__isnull=True)
            unassigned_branch_scope['enrollments'] = unassigned_branch_scope['enrollments'].filter(branch__isnull=True)
            unassigned_branch_metrics = performance_metrics(unassigned_branch_scope, start_date, end_date)
            if any(unassigned_branch_metrics.get(key) for key in ('leads', 'walkins', 'enrollments', 'revenue')):
                branch_rows.append({
                    'branch_id': None,
                    'branch_name': 'Unassigned Branch',
                    **unassigned_branch_metrics,
                })
        followup_total = metrics['total_followups']
        followup_pending = metrics['pending_followups']
        insights = [
            f'{abs(pct_change(metrics["revenue"], previous["revenue"])):.0f}% {"improvement" if pct_change(metrics["revenue"], previous["revenue"]) >= 0 else "decline"} from last month.',
            f'Conversion rate is {metrics["conversion_ratio"]:.2f}% for the selected period.',
        ]
        payments = Payment.objects.select_related('enrollment', 'enrollment__branch', 'enrollment__course').filter(enrollment__is_deleted=False)
        if branch_filter:
            payments = payments.filter(enrollment__branch_id=branch_filter)
        if user_filter:
            payments = payments.filter(
                enrollment_counselor_filter_q(user_filter, prefix='enrollment')
            )
        if course_filter:
            payments = payments.filter(enrollment__course_id=course_filter)
        if source_filter:
            normalized_payment_source = normalize_source_filter(source_filter, Enrollment._meta.get_field('source').choices)
            payments = payments.filter(enrollment__source=normalized_payment_source) if normalized_payment_source else payments.none()
        if status_filter:
            if status_filter == Enrollment.Status.ACTIVE:
                payments = payments.filter(enrollment__status__in=[Enrollment.Status.ACTIVE, Enrollment.Status.ENROLLED])
            elif status_filter == 'pending':
                payments = payments.filter(enrollment__status__in=[Enrollment.Status.DRAFT, Enrollment.Status.PENDING_RULES, Enrollment.Status.RULES_SENT, Enrollment.Status.RULES_SUBMITTED])
            elif status_filter in {choice[0] for choice in Enrollment.Status.choices}:
                payments = payments.filter(enrollment__status=status_filter)
            else:
                payments = payments.none()
        if search_filter:
            payments = payments.filter(
                Q(enrollment__name__icontains=search_filter) | Q(enrollment__phone__icontains=search_filter) | Q(enrollment__student_number__icontains=search_filter) |
                Q(enrollment__course__name__icontains=search_filter) | Q(enrollment__branch__name__icontains=search_filter) |
                Q(enrollment__counselor__first_name__icontains=search_filter) | Q(enrollment__counselor__last_name__icontains=search_filter)
            ).distinct()
        collection = PaymentInstallment.objects.filter(
            payment__in=payments,
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            document_type=PaymentInstallment.DocumentType.RECEIPT,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
        pending_payments = payments.filter(status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL])
        pending_amount = sum((payment.balance for payment in pending_payments), Decimal('0'))
        overdue_payments = pending_payments.filter(next_payment_date__isnull=False, next_payment_date__lte=end_date)
        collection_ratio = ratio(collection, Decimal(str(collection or 0)) + pending_amount)
        payload = {
            'period': {'start': start_date, 'end': end_date},
            'metrics': metrics,
            'changes': {
                key: pct_change(metrics[key], previous[key])
                for key in ['leads', 'walkins', 'enrollments', 'conversion_ratio', 'revenue']
            },
            'trends': {
                'leads': [{'date': key, 'value': value} for key, value in sorted(lead_trend.items())],
                'walkins': [{'date': key, 'value': value} for key, value in sorted(walkin_trend.items())],
                'enrollments': [{'date': key, 'value': value} for key, value in sorted(enrollment_trend.items())],
                'revenue': [{'date': key, 'value': value} for key, value in sorted(revenue_trend.items())],
            },
            'funnel': [
                {'stage': 'Leads', 'count': metrics['leads'], 'conversion_percent': 100},
                {'stage': 'Walk-ins', 'count': metrics['walkins'], 'conversion_percent': ratio(metrics['walkins'], metrics['leads'])},
                {'stage': 'Enrollments', 'count': metrics['enrollments'], 'conversion_percent': ratio(metrics['enrollments'], metrics['leads'])},
            ],
            'followup_efficiency': {
                'total_followups': followup_total,
                'pending_followups': followup_pending,
                'completion_ratio': ratio(max(followup_total - followup_pending, 0), followup_total),
            },
            'counselor_comparison': counselor_rows[:30],
            'branch_comparison': branch_rows,
            'filters': {
                'branches': BranchSerializer(Branch.objects.filter(is_active=True), many=True).data,
                'users': [
                    {'id': user.id, 'name': user.full_name, 'branch_id': user.branch_id}
                    for user in User.objects.filter(is_active=True).exclude(role=User.Role.SUPER_ADMIN).select_related('branch')
                ],
                'courses': [{'id': item.id, 'name': item.name} for item in Course.objects.filter(is_active=True).order_by('name')],
                'sources': [{'value': value, 'label': label} for value, label in Lead.Source.choices],
                'statuses': [
                    {'value': value, 'label': label}
                    for value, label in [*Lead.CounselorStatus.choices, *Enrollment.Status.choices]
                ],
            },
            'insights': insights,
            'payment_summary': {
                'total_collection': collection,
                'pending_payments': pending_payments.count(),
                'pending_amount': pending_amount,
                'overdue_payments': overdue_payments.count(),
                'collection_ratio': collection_ratio,
            },
        }
        if str(request.query_params.get('download') or '').lower() in ('1', 'true', 'pdf'):
            pdf_bytes = build_monthly_consolidated_report_pdf(payload)
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            report_month = start_date.strftime('%Y-%m')
            response['Content-Disposition'] = f'attachment; filename="IIE-monthly-consolidated-report-{report_month}.pdf"'
            return response
        return Response(payload)


class SessionHeartbeatView(APIView):
    """POST /api/auth/heartbeat/ — updates the current user's last seen timestamp."""
    permission_classes = [IsStaffOrAdmin]

    def post(self, request):
        touch_user_session(request)
        return Response({'detail': 'ok', 'last_seen': timezone.now()})


class UserMonitoringView(APIView):
    """GET /api/admin/user-monitoring/ — admin-only session monitoring."""
    permission_classes = [IsSuperAdmin]

    def get_date_range(self, request):
        today = timezone.localdate()
        date_filter = request.query_params.get('date_filter', '').strip()
        from_date = parse_date(request.query_params.get('from_date', '').strip())
        to_date = parse_date(request.query_params.get('to_date', '').strip())

        if date_filter == 'today':
            return today, today
        if date_filter == 'yesterday':
            yesterday = today - timedelta(days=1)
            return yesterday, yesterday
        if date_filter == 'this_week':
            week_start = today - timedelta(days=today.weekday())
            return week_start, today
        if date_filter == 'this_month':
            return today.replace(day=1), today
        if date_filter == 'custom':
            return from_date, to_date
        return None, None

    def apply_filters(self, request, queryset, cutoff):
        search = request.query_params.get('search', '').strip()
        branch = request.query_params.get('branch', '').strip()
        status_filter = request.query_params.get('status', '').strip().lower()
        start_date, end_date = self.get_date_range(request)

        if search:
            queryset = queryset.filter(
                Q(user__username__icontains=search) |
                Q(user__first_name__icontains=search) |
                Q(user__last_name__icontains=search) |
                Q(user__email__icontains=search)
            )
        if branch:
            queryset = queryset.filter(user__branch_id=branch)
        if status_filter == 'online':
            queryset = queryset.filter(is_active_session=True, last_seen_at__gte=cutoff)
        elif status_filter == 'offline':
            queryset = queryset.exclude(is_active_session=True, last_seen_at__gte=cutoff)
        if start_date:
            queryset = queryset.filter(login_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(login_at__date__lte=end_date)
        return queryset

    def get(self, request):
        today = timezone.localdate()
        cutoff = timezone.now() - timedelta(minutes=5)
        sessions = UserSessionLog.objects.filter(user__role=User.Role.STAFF)
        stale_sessions = sessions.filter(is_active_session=True, last_seen_at__lt=cutoff)
        stale_sessions.update(is_active_session=False, logout_at=timezone.now())
        sessions = self.apply_filters(
            request,
            UserSessionLog.objects.filter(user__role=User.Role.STAFF),
            cutoff,
        ).select_related('user__branch').order_by('-login_at', 'user__username')
        serializer = UserMonitoringSerializer(sessions, many=True)
        return Response({
            'total_logins_today': sessions.filter(login_at__date=today).count(),
            'currently_online_users': sessions.filter(
                is_active_session=True,
                last_seen_at__gte=cutoff,
            ).values('user_id').distinct().count(),
            'branches': BranchSerializer(
                Branch.objects.filter(is_active=True).order_by('name'),
                many=True,
            ).data,
            'users': serializer.data,
        })


# ============================================================
# backend/apps/courses/views.py
# ============================================================
from crm.models import Course
from serializers import COURSE_LINKED_DELETE_MESSAGE, CourseSerializer, course_is_linked


class CourseViewSet(viewsets.ModelViewSet):
    """
    Course catalogue.
    - GET: any authenticated user
    - POST/PUT/DELETE: super admin only
    """
    serializer_class   = CourseSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    filter_backends    = [SearchFilter]
    pagination_class   = None
    search_fields      = ['name']

    def get_queryset(self):
        queryset = Course.objects.order_by('name')
        include_inactive = self.request.query_params.get('include_inactive') in ('1', 'true', 'yes')
        if self.request.user.is_super_admin and (include_inactive or self.request.method not in SAFE_METHODS):
            return queryset
        return queryset.filter(is_active=True)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'courses', 'courses-export')

    def destroy(self, request, *args, **kwargs):
        course = self.get_object()
        if course_is_linked(course):
            return Response({'detail': COURSE_LINKED_DELETE_MESSAGE}, status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)


# ============================================================
# backend/apps/discounts/views.py
# ============================================================
from crm.models import Discount
from serializers import DiscountSerializer


class DiscountViewSet(viewsets.ModelViewSet):
    """Admin CRUD for discounts; staff can read currently available discounts."""
    serializer_class = DiscountSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'apply_to_all_courses', 'courses', 'apply_to_all_branches', 'branches']
    search_fields = ['name']
    ordering_fields = ['valid_from', 'valid_to', 'name', 'created_at']
    ordering = ['-valid_to', 'name']

    def get_queryset(self):
        queryset = Discount.objects.prefetch_related('courses', 'branches').select_related('created_by')
        course_id = self.request.query_params.get('course')
        branch_id = self.request.query_params.get('branch')
        available_only = self.request.query_params.get('available') in ('1', 'true', 'yes')
        if available_only or not self.request.user.is_super_admin:
            today = timezone.localdate()
            queryset = queryset.filter(
                is_active=True,
                valid_from__lte=today,
            ).filter(Q(valid_to__isnull=True) | Q(valid_to__gte=today))
        if course_id:
            queryset = queryset.filter(Q(apply_to_all_courses=True) | Q(courses__id=course_id)).distinct()
        if branch_id:
            queryset = queryset.filter(Q(apply_to_all_branches=True) | Q(branches__id=branch_id)).distinct()
        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError as exc:
            logger.exception('Discount create failed because of a database constraint.')
            return Response(
                {
                    'detail': 'Could not create discount because the submitted values conflict with the database schema or an existing record.',
                    'error': str(exc),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        except DRFValidationError:
            raise
        except Exception as exc:
            logger.exception('Discount create failed unexpectedly.')
            return Response(
                {'detail': 'Could not create discount. Please check the submitted discount details and try again.'},
                status=status.HTTP_400_BAD_REQUEST,
            )


# ============================================================
# backend/apps/leads/views.py
# ============================================================
from crm.models import FollowUp, Lead, LeadImportHistory
from serializers import (
    FollowUpSerializer, LeadListSerializer, LeadDetailSerializer,
    LeadStaffUpdateSerializer, LeadImportHistorySerializer, LeadInboxSerializer,
    effective_lead_status,
)
import django_filters
import csv


def create_follow_up_entry(record, record_type, request):
    data = request.data.copy()
    data.pop('close_follow_up', None)
    serializer = FollowUpSerializer(data=data)
    serializer.is_valid(raise_exception=True)
    follow_up = serializer.save(
        record_type=record_type,
        record_id=record.id,
        updated_by=request.user,
    )
    return Response(FollowUpSerializer(follow_up).data, status=status.HTTP_201_CREATED)


class LeadFilter(django_filters.FilterSet):
    status = django_filters.CharFilter(method='filter_status')
    name = django_filters.CharFilter(method='filter_name')
    phone = django_filters.CharFilter(method='filter_phone')
    source = django_filters.CharFilter(method='filter_source')
    follow_up_by = django_filters.NumberFilter(field_name='assigned_to')
    assigned_user = django_filters.NumberFilter(field_name='assigned_to')
    counselor = django_filters.NumberFilter(field_name='assigned_to')
    walkin_date_from = django_filters.DateFilter(field_name='walkin_date', lookup_expr='gte')
    walkin_date_to   = django_filters.DateFilter(field_name='walkin_date', lookup_expr='lte')
    next_follow_up_date_from = django_filters.DateFilter(method='filter_next_follow_up_date_from')
    next_follow_up_date_to   = django_filters.DateFilter(method='filter_next_follow_up_date_to')
    created_from     = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    created_to       = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')
    date_from        = django_filters.DateFilter(field_name='created_at', lookup_expr='date__gte')
    date_to          = django_filters.DateFilter(field_name='created_at', lookup_expr='date__lte')
    date_range       = django_filters.CharFilter(method='filter_date_range')
    important_only   = django_filters.BooleanFilter(method='filter_important_only')

    def filter_name(self, queryset, name, value):
        value = (value or '').strip()
        return queryset.filter(name__icontains=value) if value else queryset

    def filter_phone(self, queryset, name, value):
        value = (value or '').strip()
        return queryset.filter(phone__icontains=value) if value else queryset

    def filter_source(self, queryset, name, value):
        value = (value or '').strip()
        if not value:
            return queryset
        if value == '__unknown__':
            return queryset.filter(Q(source__isnull=True) | Q(source=''))
        if value == 'manual':
            if 'imported_via_csv' in missing_model_columns(Lead, ['imported_via_csv']):
                return queryset.filter(created_by__isnull=False)
            return queryset.filter(created_by__isnull=False, imported_via_csv=False)
        if value == 'csv_import':
            if 'imported_via_csv' in missing_model_columns(Lead, ['imported_via_csv']):
                return queryset.none()
            return queryset.filter(imported_via_csv=True)
        normalized = normalize_source_filter(value, Lead.Source.choices)
        return queryset.filter(source__iexact=normalized) if normalized else queryset.none()

    def filter_status(self, queryset, name, value):
        value = (value or '').strip()
        if not value:
            return queryset
        if value in {choice[0] for choice in Lead.CounselorStatus.choices} or value in LEAD_STATUS_FILTER_ALIASES:
            return filter_candidate_status(queryset, value, Lead, LEAD_STATUS_FILTER_ALIASES)
        enrollment_lookup = Q(enrollments__isnull=False) | Q(walkin__enrollment__isnull=False)
        walkin_lookup = Q(walkin__isnull=False)
        if value in (Lead.Status.ENROLLED, Lead.Status.CONVERTED):
            return queryset.filter(enrollment_lookup).distinct()
        if value in (Lead.Status.CONVERTED_TO_WALKIN, Lead.Status.WALK_IN):
            return queryset.filter(walkin_lookup).exclude(enrollment_lookup).distinct()
        if value == Lead.Status.LOST:
            return queryset.filter(status=Lead.Status.LOST).exclude(enrollment_lookup).exclude(walkin_lookup).distinct()
        return queryset.filter(status=value).exclude(enrollment_lookup).exclude(walkin_lookup).distinct()

    def filter_next_follow_up_date_from(self, queryset, name, value):
        if not value or 'next_follow_up_date' in missing_model_columns(Lead, ['next_follow_up_date']):
            return queryset
        return queryset.filter(next_follow_up_date__gte=value)

    def filter_next_follow_up_date_to(self, queryset, name, value):
        if not value or 'next_follow_up_date' in missing_model_columns(Lead, ['next_follow_up_date']):
            return queryset
        return queryset.filter(next_follow_up_date__lte=value)

    def filter_important_only(self, queryset, name, value):
        return queryset.filter(is_important=True) if value else queryset

    def filter_date_range(self, queryset, name, value):
        start_date, end_date = date_range_from_request(self.request)
        return queryset.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)

    class Meta:
        model  = Lead
        fields = ['status', 'source', 'branch', 'assigned_to', 'counselor', 'course', 'date_range', 'important_only']


class LeadViewSet(viewsets.ModelViewSet):
    """
    Leads module.
    - Staff can create leads and update ONLY walkin_date + next_follow_up_date + remarks + status
    - Super admin has full access
    - Branch-level data isolation for staff
    """
    permission_classes = [IsStaffOrAdmin]
    filter_backends    = [DjangoFilterBackend, CRMSearchFilter, OrderingFilter]
    filterset_class    = LeadFilter
    pagination_class   = None
    search_fields      = [
        'name', 'phone', 'lead_number', 'email', 'source', 'source_description',
        'course__name', 'assigned_to__first_name', 'assigned_to__last_name',
        'assigned_to__username', 'created_by__first_name', 'created_by__last_name',
        'created_by__username',
    ]
    id_search_fields   = ['id']
    ordering_fields    = ['created_at', 'walkin_date', 'name']
    ordering           = ['-created_at']

    def get_queryset(self):
        missing_columns = missing_model_columns(Lead, [
            'qualification', 'degree', 'willing_to_join',
            'external_course_interested', 'external_message',
            'source_description', 'is_duplicate', 'imported_via_csv', 'preferred_timing',
            'next_follow_up_date', 'converted_to_type',
            'converted_record_id', 'converted_at', 'converted_by',
        ])
        if self.action == 'list':
            qs = visible_candidate_queryset(Lead.objects.select_related('course', 'branch', 'assigned_to', 'created_by'))
            latest_follow_up = FollowUp.objects.filter(
                record_type=FollowUp.RecordType.LEAD,
                record_id=OuterRef('pk'),
            ).order_by('-created_at', '-id')
            qs = qs.annotate(
                latest_follow_up_remark=Subquery(latest_follow_up.values('remarks')[:1]),
                latest_follow_up_at=Subquery(latest_follow_up.values('created_at')[:1]),
                has_walkin_record=Exists(
                    WalkIn.objects.filter(lead_id=OuterRef('pk'), is_deleted=False)
                ),
                has_enrollment_record=Exists(
                    Enrollment.objects.filter(
                        Q(lead_id=OuterRef('pk')) | Q(walkin__lead_id=OuterRef('pk')),
                        is_deleted=False,
                    )
                ),
            )
        else:
            related = ['course', 'branch', 'assigned_to', 'created_by']
            if 'converted_by' not in missing_columns:
                related.append('converted_by')
            qs = visible_candidate_queryset(Lead.objects.select_related(*related).prefetch_related('transfer_history'))
        if missing_columns:
            qs = qs.defer(*missing_columns)
        if self.action != 'list':
            missing_user_columns = missing_model_columns(User, ['must_change_password'])
            if missing_user_columns:
                qs = qs.defer(
                    *[f'assigned_to__{field}' for field in missing_user_columns],
                    *[f'created_by__{field}' for field in missing_user_columns],
                    *[f'converted_by__{field}' for field in missing_user_columns],
                )
        # Staff can only see leads currently owned by their branch.
        if not self.request.user.is_super_admin:
            qs = qs.filter(branch=self.request.user.branch)
        if (
            self.request.query_params.get('focus') == 'today-follow-up'
            and 'next_follow_up_date' not in missing_columns
        ):
            qs = active_lead_follow_up_queryset(
                qs,
                'next_follow_up_date',
                timezone.localdate(),
            )
        return qs

    def _reconcile_lead_lifecycle_status(self, lead):
        old_status = lead.status
        next_status = effective_lead_status(lead)
        if next_status not in (Lead.Status.ENROLLED, Lead.Status.CONVERTED_TO_WALKIN) or lead.status == next_status:
            return lead
        lead.status = next_status
        lead.next_follow_up_date = None
        if next_status == Lead.Status.ENROLLED:
            lead.remarks = lead.remarks or 'Joined'
            lead.converted_to_type = lead.converted_to_type or 'enrollment'
        else:
            lead.converted_to_type = lead.converted_to_type or 'walkin'
        lead.save(update_fields=[
            'status', 'next_follow_up_date', 'remarks',
            'converted_to_type', 'updated_at',
        ])
        create_status_history(
            lead,
            CandidateStatusHistory.RecordType.LEAD,
            old_status,
            lead.status,
            remarks=lead.remarks,
        )
        return lead

    @action(detail=True, methods=['post'], url_path='toggle-important')
    def toggle_important(self, request, pk=None):
        lead = self.get_object()
        requested = request.data.get('is_important', None)
        lead.is_important = truthy_query_param(requested) if requested is not None else not lead.is_important
        lead.save(update_fields=['is_important', 'updated_at'])
        return Response({'id': lead.id, 'is_important': lead.is_important})

    def _transfer_lead(self, lead, target_user, transferred_by, note=''):
        if lead.assigned_to_id == target_user.id and lead.branch_id == target_user.branch_id:
            return None
        logger.info(
            'Lead transfer started lead_id=%s source_user_id=%s destination_user_id=%s destination_branch_id=%s transferred_by_id=%s',
            lead.id,
            (lead.assigned_to_id or lead.created_by_id),
            target_user.id,
            target_user.branch_id,
            transferred_by.id if transferred_by else None,
        )
        from_user = lead.assigned_to or lead.created_by
        from_branch = lead.branch
        lead.assigned_to = target_user
        if target_user.branch_id:
            lead.branch = target_user.branch
        lead.save(update_fields=['assigned_to', 'branch', 'updated_at'])
        logger.info(
            'Lead transfer lead assignment saved lead_id=%s from_user_id=%s from_branch_id=%s to_user_id=%s to_branch_id=%s',
            lead.id,
            from_user.id if from_user else None,
            from_branch.id if from_branch else None,
            target_user.id,
            lead.branch_id,
        )
        history = LeadTransferHistory.objects.create(
            lead=lead,
            from_user=from_user,
            to_user=target_user,
            transferred_by=transferred_by,
            from_branch=from_branch,
            to_branch=lead.branch,
            note=str(note or '').strip(),
        )
        logger.info('Lead transfer history created lead_id=%s history_id=%s', lead.id, history.id)
        if target_user.id != transferred_by.id:
            try:
                create_user_notification(
                    target_user,
                    'Lead Transferred To You',
                    f'{lead.name} was transferred to you by {transferred_by.full_name}.',
                    Notification.NType.INFO,
                    f'/leads/{lead.id}',
                )
                logger.info('Lead transfer user notification created lead_id=%s user_id=%s', lead.id, target_user.id)
            except Exception:
                logger.exception(
                    'Lead transfer target notification failed lead_id=%s destination_user_id=%s',
                    lead.id,
                    target_user.id,
                )
        try:
            notify_admin_users(
                'Lead Transferred',
                f'{lead.name} was transferred to {target_user.full_name} by {transferred_by.full_name}.',
                Notification.NType.INFO,
                f'/leads/{lead.id}',
            )
            logger.info('Lead transfer admin notifications processed lead_id=%s', lead.id)
        except Exception:
            logger.exception('Lead transfer admin notification failed lead_id=%s', lead.id)
        return history

    @action(detail=True, methods=['post'], url_path='transfer')
    def transfer(self, request, pk=None):
        lead = self.get_object()
        source_user_id = lead.assigned_to_id or lead.created_by_id
        target_id = request.data.get('transfer_to') or request.data.get('to_user') or request.data.get('user')
        requested_branch_id = request.data.get('branch') or request.data.get('to_branch') or request.data.get('transfer_to_branch')
        logger.info(
            'Lead transfer request received lead_id=%s current_owner_id=%s target_user_id=%s target_branch_id=%s payload=%s requested_by_id=%s',
            lead.id,
            source_user_id,
            target_id,
            requested_branch_id,
            dict(request.data) if hasattr(request.data, 'dict') else request.data,
            request.user.id,
        )
        if not target_id:
            return Response({'transfer_to': 'Select a counselor.'}, status=status.HTTP_400_BAD_REQUEST)
        target_user = User.objects.filter(pk=target_id, is_active=True).exclude(role=User.Role.SUPER_ADMIN).select_related('branch').first()
        if not target_user:
            logger.warning(
                'Lead transfer invalid destination lead_id=%s source_user_id=%s destination_user_id=%s transferred_by_id=%s',
                lead.id,
                source_user_id,
                target_id,
                request.user.id,
            )
            return Response({
                'success': False,
                'message': 'Lead transfer failed',
                'error': 'Destination user does not exist or is not an active counselor.',
            }, status=status.HTTP_400_BAD_REQUEST)
        if not target_user.branch_id:
            logger.warning(
                'Lead transfer destination user has no branch lead_id=%s source_user_id=%s destination_user_id=%s transferred_by_id=%s',
                lead.id,
                source_user_id,
                target_user.id,
                request.user.id,
            )
            return Response({
                'success': False,
                'message': 'Lead transfer failed',
                'error': 'Destination user does not have an assigned branch.',
            }, status=status.HTTP_400_BAD_REQUEST)
        if requested_branch_id and str(target_user.branch_id) != str(requested_branch_id):
            logger.warning(
                'Lead transfer target user branch mismatch lead_id=%s source_user_id=%s destination_user_id=%s requested_branch_id=%s actual_branch_id=%s',
                lead.id,
                source_user_id,
                target_user.id,
                requested_branch_id,
                target_user.branch_id,
            )
            return Response({
                'success': False,
                'message': 'Selected user does not belong to the selected branch.',
            }, status=status.HTTP_400_BAD_REQUEST)
        try:
            with transaction.atomic():
                lead = Lead.objects.select_for_update().get(pk=lead.pk)
                history = self._transfer_lead(lead, target_user, request.user, request.data.get('note'))
            lead.refresh_from_db()
            logger.info(
                'Lead transfer completed lead_id=%s source_user_id=%s destination_user_id=%s destination_branch_id=%s history_created=%s',
                lead.id,
                source_user_id,
                target_user.id,
                target_user.branch_id,
                bool(history),
            )
            data = self.get_serializer(lead).data
            data['success'] = True
            data['transfer_created'] = bool(history)
            data['detail'] = f'Lead transferred to {target_user.full_name}.'
            return Response(data)
        except Exception as exc:
            logger.exception(
                f'Lead transfer failed for lead {lead.id} source_user_id={source_user_id} destination_user_id={target_user.id if target_user else target_id} destination_branch_id={target_user.branch_id if target_user else None} transferred_by_id={request.user.id}'
            )
            logger.error(
                'Lead transfer failure context lead_id=%s source_user_id=%s destination_user_id=%s destination_branch_id=%s transferred_by_id=%s',
                lead.id,
                source_user_id,
                target_user.id if target_user else target_id,
                target_user.branch_id if target_user else None,
                request.user.id,
                exc_info=False,
            )
            return Response({
                'success': False,
                'message': 'Lead transfer failed',
                'error': str(exc),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'leads', 'leads-export')

    def get_serializer_class(self):
        if self.action == 'list':
            return LeadListSerializer
        # Staff can only patch walkin_date + next_follow_up_date + remarks
        if self.action in ('partial_update',) and not self.request.user.is_super_admin:
            return LeadDetailSerializer
        return LeadDetailSerializer

    def partial_update(self, request, *args, **kwargs):
        if 'course' in request.data:
            lead = self.get_object()
            has_enrollment = Enrollment.objects.filter(
                Q(lead_id=lead.id) | Q(walkin__lead_id=lead.id),
                is_deleted=False,
            ).exists()
            converted = bool(lead.converted_to_type) or lead.status in {
                Lead.Status.CONVERTED,
                Lead.Status.CONVERTED_TO_WALKIN,
                Lead.Status.ENROLLED,
            }
            if has_enrollment or converted:
                return Response(
                    {'detail': 'Course can be edited only while this record is still a lead. Use the enrollment course change workflow after enrollment.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        return super().partial_update(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='request-counselor-change')
    def request_counselor_change(self, request, pk=None):
        lead = self.get_object()
        counselor_id = request.data.get('counselor') or request.data.get('new_counselor') or request.data.get('user')
        if not counselor_id:
            return Response({'counselor': 'Select a counselor.'}, status=status.HTTP_400_BAD_REQUEST)
        counselor = User.objects.filter(pk=counselor_id, role=User.Role.STAFF, is_active=True).first()
        if not counselor:
            return Response({'counselor': 'Select an active staff counselor.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            change_request = create_counselor_change_request(
                lead,
                CounselorChangeRequest.RecordType.LEAD,
                counselor,
                request.user,
                request.data.get('reason'),
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CounselorChangeRequestSerializer(change_request, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @staticmethod
    def _lead_source_to_walkin_source(lead_source):
        walkin_sources = {value for value, _ in WalkIn.Source.choices}
        if lead_source in walkin_sources:
            return lead_source
        if lead_source == Lead.Source.MANUAL:
            return WalkIn.Source.DIRECT
        return WalkIn.Source.LEAD_CONVERSION

    @staticmethod
    def _phone_numbers_match(left, right):
        left = normalize_phone_number(left)
        right = normalize_phone_number(right)
        if not left or not right:
            return False
        left_variants = {left}
        right_variants = {right}
        if left.startswith('91') and len(left) == 12:
            left_variants.add(left[2:])
        if left.startswith('0') and len(left) == 11:
            left_variants.add(left[1:])
        if right.startswith('91') and len(right) == 12:
            right_variants.add(right[2:])
        if right.startswith('0') and len(right) == 11:
            right_variants.add(right[1:])
        return bool(left_variants & right_variants)

    def _find_existing_walkin_for_lead_conversion(self, lead, phone):
        existing_walkin = None
        if lead.converted_to_type == 'walkin' and lead.converted_record_id:
            existing_walkin = WalkIn.objects.select_for_update().filter(pk=lead.converted_record_id).first()
        existing_walkin = existing_walkin or WalkIn.objects.select_for_update().filter(lead=lead).first()
        if existing_walkin:
            return existing_walkin

        normalized_phone = normalize_phone_number(phone)
        if not normalized_phone:
            return None
        walkins = visible_candidate_queryset(
            WalkIn.objects.select_for_update().exclude(phone__isnull=True).exclude(phone='')
        )
        for walkin in walkins:
            if self._phone_numbers_match(walkin.phone, normalized_phone):
                return walkin
        return None

    def _convert_lead_to_walkin(self, request, lead, raw_data=None):
        data = (raw_data or {}).copy()
        data.setdefault('name', lead.name)
        data.setdefault('phone', lead.phone)
        data.setdefault('dob', lead.dob)
        data.setdefault('email', lead.email)
        data.setdefault('location', lead.location)
        data.setdefault('pincode', lead.pincode)
        data.setdefault('course', lead.course_id)
        data.setdefault('branch', lead.branch_id)
        data.setdefault('preferred_timing', lead.preferred_timing)
        data.setdefault('qualification', lead.qualification)
        data.setdefault('degree', lead.degree)
        for field_name in QUALIFICATION_KPI_FIELDS:
            data.setdefault(field_name, getattr(lead, field_name, ''))
        data.setdefault('remarks', lead.remarks)
        data.setdefault('follow_up_date', lead.next_follow_up_date)
        data.setdefault('visit_date', data.get('conversion_date') or lead.walkin_date or timezone.localdate())
        if not request.user.is_super_admin:
            if not request.user.branch_id:
                return None, None, Response(
                    {'detail': 'Your account is not assigned to a branch.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            data['branch'] = request.user.branch_id

        required_fields = ['name', 'phone', 'course', 'branch', 'visit_date']
        missing = [field for field in required_fields if data.get(field) in (None, '')]
        if missing:
            return None, None, Response(
                {'detail': 'Please complete all mandatory fields before converting to walk-in.', 'missing_fields': missing},
                status=status.HTTP_400_BAD_REQUEST,
            )
        branch = Branch.objects.filter(pk=data.get('branch'), is_active=True).first()
        if not branch:
            return None, None, Response({'branch': 'Please select a valid active branch.'}, status=status.HTTP_400_BAD_REQUEST)
        course = Course.objects.filter(pk=data.get('course'), is_active=True).first()
        if not course:
            return None, None, Response({'course': 'Please select a valid active course.'}, status=status.HTTP_400_BAD_REQUEST)

        year_of_passing = None
        if data.get('year_of_passing') not in (None, ''):
            try:
                year_of_passing = int(data.get('year_of_passing'))
            except (TypeError, ValueError):
                return None, None, Response({'year_of_passing': 'Enter a valid passed out year.'}, status=status.HTTP_400_BAD_REQUEST)
            if year_of_passing < 1900 or year_of_passing > 2100:
                return None, None, Response({'year_of_passing': 'Enter a valid passed out year.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                lead = Lead.objects.select_for_update().get(pk=lead.pk)
                old_status = lead.status
                if lead.converted_to_type == 'enrollment' or lead.enrollments.exists():
                    return None, None, Response(
                        {'detail': 'This record has already been converted.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                name = str(data.get('name') or '').strip()
                phone = str(data.get('phone') or '').strip()
                existing_walkin = self._find_existing_walkin_for_lead_conversion(lead, phone)
                if existing_walkin:
                    if existing_walkin.lead_id and existing_walkin.lead_id != lead.id:
                        return None, None, Response(
                            {'detail': 'A walk-in already exists for this phone number and is linked to another lead.'},
                            status=status.HTTP_400_BAD_REQUEST,
                        )
                    walkin_update_fields = []
                    if existing_walkin.lead_id != lead.id:
                        existing_walkin.lead = lead
                        walkin_update_fields.append('lead')
                    mapped_source = self._lead_source_to_walkin_source(lead.source)
                    if existing_walkin.source in ('', WalkIn.Source.LEAD_CONVERSION) and existing_walkin.source != mapped_source:
                        existing_walkin.source = mapped_source
                        walkin_update_fields.append('source')
                    if 'source_description' in data:
                        source_description = str(data.get('source_description') or '').strip()
                        if existing_walkin.source_description != source_description:
                            existing_walkin.source_description = source_description
                            walkin_update_fields.append('source_description')
                    for field_name in QUALIFICATION_KPI_FIELDS:
                        value = data.get(field_name) or getattr(lead, field_name, '')
                        if value and not getattr(existing_walkin, field_name, ''):
                            setattr(existing_walkin, field_name, value)
                            walkin_update_fields.append(field_name)
                    if walkin_update_fields:
                        existing_walkin.save(update_fields=[*walkin_update_fields, 'updated_at'])
                    walkin = existing_walkin
                    response_status = status.HTTP_200_OK
                else:
                    walkin = WalkIn.objects.create(
                        lead=lead,
                        branch=branch,
                        course=course,
                        assigned_to=lead.assigned_to,
                        created_by=request.user,
                        name=name,
                        phone=phone,
                        dob=data.get('dob') or None,
                        email=str(data.get('email') or '').strip(),
                        location=str(data.get('location') or '').strip(),
                        pincode=str(data.get('pincode') or '').strip(),
                        qualification=data.get('qualification') or '',
                        degree=data.get('degree') or '',
                        expected_course_budget=data.get('expected_course_budget') or '',
                        planned_joining_time=data.get('planned_joining_time') or '',
                        primary_goal=data.get('primary_goal') or '',
                        other_institutes_considering=data.get('other_institutes_considering') or '',
                        counselor_status=data.get('counselor_status') or '',
                        competitor_status=data.get('competitor_status') or '',
                        follow_up_priority=data.get('follow_up_priority') or '',
                        conversion_probability=data.get('conversion_probability') or '',
                        year_of_passing=year_of_passing,
                        college_company=data.get('college_company') or '',
                        preferred_timing=data.get('preferred_timing') or '',
                        visit_date=data.get('visit_date'),
                        follow_up_date=data.get('follow_up_date') or None,
                        remarks=str(data.get('remarks') or '').strip(),
                        source=self._lead_source_to_walkin_source(lead.source),
                        source_description=str(data.get('source_description') or '').strip(),
                        demo_class=data.get('demo_class', False),
                        interested_global_certification=data.get('interested_global_certification', False),
                        is_important=lead.is_important,
                    )
                    lead_follow_ups = FollowUp.objects.filter(
                        record_type=FollowUp.RecordType.LEAD,
                        record_id=lead.id,
                    ).order_by('created_at', 'id')
                    FollowUp.objects.bulk_create([
                        FollowUp(
                            record_type=FollowUp.RecordType.WALKIN,
                            record_id=walkin.id,
                            follow_up_date=follow_up.follow_up_date,
                            next_follow_up_date=follow_up.next_follow_up_date,
                            remarks=follow_up.remarks,
                            updated_by=follow_up.updated_by,
                        )
                        for follow_up in lead_follow_ups
                    ])
                    response_status = status.HTTP_201_CREATED

                lead.name = name
                lead.phone = phone
                lead.dob = data.get('dob') or None
                lead.email = str(data.get('email') or '').strip()
                lead.location = str(data.get('location') or '').strip()
                lead.pincode = str(data.get('pincode') or '').strip()
                lead.preferred_timing = data.get('preferred_timing') or ''
                lead.qualification = data.get('qualification') or ''
                lead.degree = data.get('degree') or ''
                for field_name in QUALIFICATION_KPI_FIELDS:
                    setattr(lead, field_name, data.get(field_name) or '')
                lead.branch = branch
                lead.course = course
                lead.walkin_date = data.get('visit_date')
                lead.next_follow_up_date = None
                lead.status = Lead.Status.CONVERTED_TO_WALKIN
                lead.converted_to_type = 'walkin'
                lead.converted_record_id = walkin.id
                lead.converted_at = lead.converted_at or timezone.now()
                lead.converted_by = lead.converted_by or request.user
                lead.save(update_fields=[
                    'name', 'phone', 'dob', 'email', 'location', 'pincode', 'preferred_timing', 'qualification',
                    'degree', *QUALIFICATION_KPI_FIELDS, 'branch', 'course', 'walkin_date', 'next_follow_up_date', 'status',
                    'converted_to_type', 'converted_record_id', 'converted_at', 'converted_by', 'updated_at',
                ])
                create_status_history(
                    lead,
                    CandidateStatusHistory.RecordType.LEAD,
                    old_status,
                    lead.status,
                    request.user,
                    lead.remarks,
                )
                walkin.refresh_from_db()
        except IntegrityError:
            return None, None, Response(
                {'detail': 'Walk-in conversion failed. No lead changes were saved.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        clear_follow_up_notifications_for_record(FollowUp.RecordType.LEAD, lead.id)
        return walkin, lead, response_status

    @action(detail=True, methods=['post'], url_path='send-follow-up-whatsapp')
    def send_follow_up_whatsapp(self, request, pk=None):
        _, response = send_candidate_template(
            self.get_object(),
            WhatsAppTemplate.TemplateType.LEAD_FOLLOW_UP,
            WhatsAppMessage.MsgType.FOLLOW_UP,
            request,
            'lead',
        )
        return response

    @action(detail=True, methods=['post'], url_path='send-offer-whatsapp')
    def send_offer_whatsapp(self, request, pk=None):
        _, response = send_candidate_template(
            self.get_object(),
            WhatsAppTemplate.TemplateType.OFFER_MESSAGE,
            WhatsAppMessage.MsgType.OFFER_MESSAGE,
            request,
            'lead',
        )
        return response

    def perform_create(self, serializer):
        # Auto-set branch + created_by for staff
        transfer_to_id = self.request.data.get('transfer_to') or self.request.data.get('transfer_to_user')
        transfer_user = None
        if transfer_to_id:
            transfer_user = User.objects.filter(
                pk=transfer_to_id,
                is_active=True,
                role=User.Role.STAFF,
            ).select_related('branch').first()
            if not transfer_user:
                raise DRFValidationError({'transfer_to': 'Select an active counselor.'})
        branch = self.request.user.branch if not self.request.user.is_super_admin else serializer.validated_data.get('branch')
        assigned_to = transfer_user or serializer.validated_data.get('assigned_to') or self.request.user
        if transfer_user and transfer_user.branch_id:
            branch = transfer_user.branch
        lead = serializer.save(
            created_by=self.request.user,
            branch=branch,
            assigned_to=assigned_to,
            status=Lead.Status.FOLLOW_UP,
        )
        if transfer_user and transfer_user.id != self.request.user.id:
            LeadTransferHistory.objects.create(
                lead=lead,
                from_user=self.request.user,
                to_user=transfer_user,
                transferred_by=self.request.user,
                from_branch=self.request.user.branch,
                to_branch=lead.branch,
                note='Transferred during lead creation.',
            )
            create_user_notification(
                transfer_user,
                'Lead Transferred To You',
                f'New lead {lead.name} transferred to you by {self.request.user.full_name}.',
                Notification.NType.INFO,
                f'/leads/{lead.id}',
            )
            notify_admin_users(
                'Lead Transferred',
                f'New lead {lead.name} was transferred to {transfer_user.full_name} by {self.request.user.full_name}.',
                Notification.NType.INFO,
                f'/leads/{lead.id}',
            )

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except IntegrityError as exc:
            logger.exception('Lead create failed because of a database constraint.')
            return Response(
                {
                    'detail': 'Could not create lead because one of the generated or selected values conflicts with an existing record.',
                    'error': str(exc),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=['get'], url_path='staff-options')
    def staff_options(self, request):
        users = User.objects.filter(is_active=True).select_related('branch').order_by(
            'first_name', 'last_name', 'username'
        )
        if not request.user.is_super_admin:
            users = users.filter(branch=request.user.branch).exclude(role=User.Role.SUPER_ADMIN)
        branch_id = request.query_params.get('branch')
        if branch_id:
            users = users.filter(Q(branch_id=branch_id) | Q(branch__isnull=True))
        rows = []
        seen = set()
        for user in users:
            if user.id in seen:
                continue
            seen.add(user.id)
            rows.append({
                'id': user.id,
                'name': user.full_name or user.username,
                'full_name': user.full_name,
                'username': user.username,
                'email': user.email,
                'branch_id': user.branch_id,
                'branch_name': user.branch.name if user.branch else '',
                'identity_color': user.identity_color or '',
            })
        return Response(rows)

    @action(detail=False, methods=['get'], url_path='transfer-options')
    def transfer_options(self, request):
        users = User.objects.filter(
            is_active=True,
            role=User.Role.STAFF,
        ).select_related('branch').order_by('branch__name', 'first_name', 'last_name', 'username')
        return Response([
            {
                'id': user.id,
                'name': user.full_name or user.username,
                'full_name': user.full_name,
                'username': user.username,
                'branch_id': user.branch_id,
                'branch_name': user.branch.name if user.branch else '',
                'identity_color': user.identity_color or '',
            }
            for user in users
        ])

    @action(detail=False, methods=['get'], url_path='source-options')
    def source_options(self, request):
        queryset = visible_candidate_queryset(Lead.objects.all())
        if not request.user.is_super_admin:
            queryset = queryset.filter(branch=request.user.branch)
        branch_id = request.query_params.get('branch')
        if branch_id and request.user.is_super_admin:
            queryset = queryset.filter(branch_id=branch_id)

        choices = dict(Lead.Source.choices)
        rows = []
        seen = set()
        for source in queryset.order_by('source').values_list('source', flat=True):
            value = (source or '').strip()
            key = value.lower()
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                'value': value or '__unknown__',
                'label': choices.get(value, value or 'Unknown'),
            })
        return Response(rows)

    def update(self, request, *args, **kwargs):
        # Staff can update candidate details and follow-up fields only.
        lead = self.get_object()
        old_status = tracked_crm_status(lead)
        old_next_follow_up_date = lead.next_follow_up_date
        data = request.data.copy()
        if not request.user.is_super_admin:
            if 'branch' in request.data:
                return Response({'detail': 'Only admin can change a lead branch.'}, status=status.HTTP_403_FORBIDDEN)
            allowed = {
                'name', 'phone', 'dob', 'email', 'location', 'pincode',
                'course', 'preferred_timing', 'qualification', 'degree', 'walkin_date',
                'next_follow_up_date', 'remarks', 'status', 'assigned_to', 'follow_up_by',
                'source', 'source_description', *QUALIFICATION_KPI_FIELDS,
            }
            if set(request.data.keys()) - allowed:
                return Response(
                    {'error': 'Staff can only update candidate, follow-up, status, and Follow-up By fields.'},
                    status=status.HTTP_403_FORBIDDEN
                )
        if request.user.is_super_admin and 'branch' in data:
            branch = Branch.objects.filter(pk=data.get('branch'), is_active=True).first()
            if not branch and data.get('branch') not in (None, ''):
                return Response({'detail': 'Select a valid active branch.'}, status=status.HTTP_400_BAD_REQUEST)
        kwargs['partial'] = kwargs.get('partial', False)
        serializer = self.get_serializer(lead, data=data, partial=kwargs['partial'])
        serializer.is_valid(raise_exception=True)
        if serializer.validated_data.get('status') == Lead.Status.CONVERTED_TO_WALKIN:
            walkin, converted_lead, result = self._convert_lead_to_walkin(request, lead, data)
            if isinstance(result, Response):
                return result
            return Response(self.get_serializer(converted_lead).data)
        with transaction.atomic():
            self.perform_update(serializer)
            lead.refresh_from_db()
            create_status_history(
                lead,
                CandidateStatusHistory.RecordType.LEAD,
                old_status,
                tracked_crm_status(lead),
                request.user,
                lead.remarks,
            )
            self._reconcile_lead_lifecycle_status(lead)
            lead.refresh_from_db()
            if (
                old_next_follow_up_date != lead.next_follow_up_date
                or lead.status in LEAD_CLOSED_FOLLOW_UP_STATUSES
            ):
                clear_follow_up_notifications_for_record(FollowUp.RecordType.LEAD, lead.id)
        lead.refresh_from_db()
        response = Response(self.get_serializer(lead).data)
        return response

    @action(detail=True, methods=['post'], url_path='convert-to-walkin')
    def convert_to_walkin(self, request, pk=None):
        """Create a walk-in entry from this lead after required details are completed."""
        from serializers import WalkInDetailSerializer
        lead = self.get_object()
        walkin, _, result = self._convert_lead_to_walkin(request, lead, request.data)
        if isinstance(result, Response):
            return result
        return Response(WalkInDetailSerializer(walkin).data, status=result)

    @action(detail=True, methods=['post'], url_path='convert-to-enrollment')
    def convert_to_enrollment(self, request, pk=None):
        """Create an enrollment directly from this lead after required details are completed."""
        lead = self.get_object()
        from crm.models import Course, Enrollment, WalkIn
        from serializers import EnrollmentDetailSerializer
        existing_enrollment = None
        if lead.converted_to_type == 'enrollment' and lead.converted_record_id:
            existing_enrollment = Enrollment.objects.filter(pk=lead.converted_record_id).first()
        existing_enrollment = existing_enrollment or Enrollment.objects.filter(lead=lead).first()
        if existing_enrollment:
            if (
                lead.status != Lead.Status.CONVERTED
                or lead.converted_to_type != 'enrollment'
                or lead.converted_record_id != existing_enrollment.id
                or lead.next_follow_up_date is not None
                or lead.remarks != 'Joined'
            ):
                mark_lead_enrollment_converted(lead, existing_enrollment, request.user)
            clear_follow_up_notifications_for_record(FollowUp.RecordType.LEAD, lead.id)
            return Response(EnrollmentDetailSerializer(existing_enrollment).data, status=200)
        if lead.converted_to_type == 'walkin' or hasattr(lead, 'walkin'):
            return Response({'detail': 'This record has already been converted.'}, status=status.HTTP_400_BAD_REQUEST)
        data = request.data.copy()
        data.setdefault('name', lead.name)
        data.setdefault('phone', lead.phone)
        data.setdefault('dob', lead.dob)
        data.setdefault('email', lead.email)
        data.setdefault('location', lead.location)
        data.setdefault('pincode', lead.pincode)
        data.setdefault('course', lead.course_id)
        data.setdefault('branch', lead.branch_id)
        data.setdefault('preferred_timing', lead.preferred_timing)
        data.setdefault('qualification', lead.qualification)
        data.setdefault('degree', lead.degree)
        for field_name in QUALIFICATION_KPI_FIELDS:
            data.setdefault(field_name, getattr(lead, field_name, ''))
        if not request.user.is_super_admin:
            if not request.user.branch_id:
                return Response(
                    {'detail': 'Your account is not assigned to a branch.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            data['branch'] = request.user.branch_id
        required_fields = [
            'name', 'phone', 'course', 'branch', 'preferred_timing',
            'enrollment_date', 'start_date',
        ]
        missing = [field for field in required_fields if data.get(field) in (None, '')]
        if missing:
            return Response(
                {'detail': 'Please complete all mandatory fields.', 'missing_fields': missing},
                status=status.HTTP_400_BAD_REQUEST,
            )

        course = Course.objects.filter(pk=data.get('course')).first()
        if not course:
            return Response({'detail': 'Please select a valid course.'}, status=status.HTTP_400_BAD_REQUEST)
        data['actual_fees'] = data.get('actual_fees') or course.actual_fees
        try:
            apply_enrollment_discount(data, course, data.get('branch'))
            apply_spot_conversion_discount(data, lead.walkin_date)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        enrollment_source_values = {choice[0] for choice in Enrollment._meta.get_field('source').choices}
        source = lead.source if lead.source in enrollment_source_values else WalkIn.Source.GOOGLE

        with transaction.atomic():
            enrollment = Enrollment.objects.create(
                lead=lead,
                branch_id=data.get('branch'),
                course_id=data.get('course'),
                original_walkin_course_id=lead.course_id,
                final_enrollment_course_id=data.get('course'),
                enrolled_by=request.user,
                created_by=request.user,
                counselor=lead.assigned_to or lead.created_by or request.user,
                name=data.get('name'),
                phone=data.get('phone'),
                dob=data.get('dob'),
                email=data.get('email'),
                location=data.get('location'),
                pincode=data.get('pincode'),
                qualification=data.get('qualification', ''),
                degree=data.get('degree', ''),
                expected_course_budget=data.get('expected_course_budget') or '',
                planned_joining_time=data.get('planned_joining_time') or '',
                primary_goal=data.get('primary_goal') or '',
                other_institutes_considering=data.get('other_institutes_considering') or '',
                counselor_status=data.get('counselor_status') or '',
                competitor_status=data.get('competitor_status') or '',
                follow_up_priority=data.get('follow_up_priority') or '',
                conversion_probability=data.get('conversion_probability') or '',
                preferred_timing=data.get('preferred_timing'),
                enrollment_date=data.get('enrollment_date'),
                source=source,
                actual_fees=data.get('actual_fees'),
                discount_amount=data.get('discount_amount') or 0,
                discount_reason=data.get('discount_reason') or '',
                discount_id=data.get('discount') or None,
                spot_conversion_discount_applied=data.get('spot_conversion_discount_applied', False),
                start_date=data.get('start_date') or None,
                batch_timing=data.get('batch_timing') or '',
                demo_class=data.get('demo_class', False),
                interested_global_certification=data.get('interested_global_certification', False),
                status=Enrollment.Status.PENDING_RULES,
                is_important=lead.is_important,
            )
            enrollment.refresh_from_db()

            mark_lead_enrollment_converted(lead, enrollment, request.user, data)
        clear_follow_up_notifications_for_record(FollowUp.RecordType.LEAD, lead.id)
        notify_admin_enrollment_created(enrollment, request.user)
        return Response(EnrollmentDetailSerializer(enrollment).data, status=201)

    @action(detail=True, methods=['post'], url_path='follow-ups')
    def add_follow_up(self, request, pk=None):
        lead = self.get_object()
        response = create_follow_up_entry(lead, FollowUp.RecordType.LEAD, request)
        if response.status_code >= 400:
            return response

        lead.next_follow_up_date = response.data['next_follow_up_date']
        lead.remarks = response.data.get('remarks') or lead.remarks
        close_follow_up = request.data.get('close_follow_up') in (True, 'true', '1', 1)
        if close_follow_up:
            lead.next_follow_up_date = None
        lead.save(update_fields=['next_follow_up_date', 'remarks', 'updated_at'])
        clear_follow_up_notifications_for_record(FollowUp.RecordType.LEAD, lead.id)
        return response

    @action(detail=False, methods=['get'], url_path='admin-inbox', permission_classes=[IsSuperAdmin])
    def admin_inbox(self, request):
        queryset = Lead.objects.filter(branch__isnull=True).select_related('course', 'created_by').order_by('-created_at')
        return Response(LeadInboxSerializer(queryset, many=True).data)

    @action(detail=True, methods=['post'], url_path='assign-branch')
    def assign_branch(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can assign lead branches.'}, status=status.HTTP_403_FORBIDDEN)
        lead = self.get_object()
        branch = Branch.objects.filter(pk=request.data.get('branch'), is_active=True).first()
        if not branch:
            return Response({'detail': 'Select a valid active branch.'}, status=status.HTTP_400_BAD_REQUEST)
        lead.branch = branch
        lead.save(update_fields=['branch', 'updated_at'])
        notify_branch_users(
            branch,
            'New lead assigned',
            f'{lead.name} has been assigned to {branch.name}.',
            Notification.NType.INFO,
            f'/leads/{lead.id}',
        )
        notify_admin_users(
            'Lead Assigned',
            f'{lead.name} has been assigned to {branch.name}.',
            Notification.NType.INFO,
            f'/leads/{lead.id}',
        )
        return Response(LeadDetailSerializer(lead, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='duplicate-check')
    def duplicate_check(self, request):
        phone = request.query_params.get('phone', '').strip()
        if not phone:
            return Response({'duplicate': False, 'records': []})
        normalized_phone = normalize_phone_number(phone)
        if normalized_phone.startswith('91') and len(normalized_phone) == 12:
            normalized_phone = normalized_phone[2:]
        if normalized_phone.startswith('0') and len(normalized_phone) == 11:
            normalized_phone = normalized_phone[1:]
        records = []
        matched_leads = [
            lead for lead in Lead.objects.exclude(phone='').select_related('branch', 'course')
            if self._phone_numbers_match(lead.phone, normalized_phone)
        ][:10]
        for lead in matched_leads:
            records.append({
                'type': 'Lead',
                'id': lead.id,
                'name': lead.name,
                'phone': lead.phone,
                'branch_name': lead.branch.name if lead.branch else 'Unassigned',
                'course_name': lead.course.name if lead.course else '',
                'status': automated_lead_status_display(lead),
                'url': f'/leads/{lead.id}',
            })
        matched_walkins = [
            walkin for walkin in WalkIn.objects.exclude(phone='').select_related('branch', 'course')
            if self._phone_numbers_match(walkin.phone, normalized_phone)
        ][:10]
        for walkin in matched_walkins:
            records.append({
                'type': 'Walk-in',
                'id': walkin.id,
                'name': walkin.name,
                'phone': walkin.phone,
                'branch_name': walkin.branch.name if walkin.branch else '',
                'course_name': walkin.course.name if walkin.course else '',
                'status': automated_walkin_status_display(walkin),
                'url': f'/walkins/{walkin.id}',
            })
        matched_enrollments = [
            enrollment for enrollment in Enrollment.objects.exclude(phone='').select_related('branch', 'course')
            if self._phone_numbers_match(enrollment.phone, normalized_phone)
        ][:10]
        for enrollment in matched_enrollments:
            records.append({
                'type': 'Student',
                'id': enrollment.id,
                'name': enrollment.name,
                'phone': enrollment.phone,
                'branch_name': enrollment.branch.name if enrollment.branch else '',
                'course_name': enrollment.course.name if enrollment.course else '',
                'status': enrollment.get_status_display(),
                'url': f'/students/{enrollment.id}',
            })
        return Response({'duplicate': bool(records), 'warning': 'Duplicate lead found' if records else '', 'records': records})

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can delete leads.'}, status=status.HTTP_403_FORBIDDEN)
        lead = self.get_object()
        FollowUp.objects.filter(record_type=FollowUp.RecordType.LEAD, record_id=lead.id).delete()
        return super().destroy(request, *args, **kwargs)


REQUIRED_LEAD_IMPORT_HEADINGS = [
    'Candidate Name',
    'Phone Number',
]
OPTIONAL_LEAD_IMPORT_HEADINGS = [
    'Course Interested',
    'Course Name',
    'Course Names',
    'Branch',
    'Email',
    'Source',
    'Lead Source',
    'Source Name',
    'Source Names',
    'How They Know IIE',
    'Follow-up Date',
    'Follow Up Date',
    'Remarks',
    'Location',
    'Qualification',
    'Source Description',
]


def normalise_import_cell(value):
    if value is None:
        return ''
    if hasattr(value, 'date') and not isinstance(value, str):
        return value.date().isoformat()
    return str(value).strip()


def parse_import_followup_date(value):
    value = normalise_import_cell(value)
    if not value:
        return None
    parsed = parse_date(value)
    if parsed:
        return parsed
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return timezone.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def first_import_value(row, *headings):
    for heading in headings:
        value = normalise_import_cell(row.get(heading))
        if value:
            return value
    return ''


def read_lead_import_rows(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        text = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return reader.fieldnames or [], [dict(row) for row in reader]
    if name.endswith('.xlsx'):
        import openpyxl
        workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        raw_headings = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        headings = [normalise_import_cell(value) for value in raw_headings]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(normalise_import_cell(value) for value in values):
                continue
            rows.append({
                heading: normalise_import_cell(values[index] if index < len(values) else '')
                for index, heading in enumerate(headings)
            })
        return headings, rows
    raise ValueError('Only CSV and .xlsx files are supported.')


class LeadImportView(APIView):
    """Counselor-only CSV/XLSX lead import."""
    permission_classes = [IsStaffOrAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if request.user.is_super_admin:
            return Response({'detail': 'Lead import is available only for users/counselors.'}, status=403)
        if not request.user.branch_id:
            return Response({'detail': 'Your account is not assigned to a branch.'}, status=400)

        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'detail': 'Upload a CSV or Excel file.'}, status=400)

        history = LeadImportHistory.objects.create(
            uploaded_by=request.user,
            branch=request.user.branch,
            file_name=uploaded_file.name,
        )

        try:
            headings, rows = read_lead_import_rows(uploaded_file)
        except Exception as exc:
            history.error_log = [{'row': None, 'error': str(exc)}]
            history.status = LeadImportHistory.Status.FAILED
            history.save(update_fields=['error_log', 'status', 'updated_at'])
            return Response({'detail': str(exc), 'history': LeadImportHistorySerializer(history).data}, status=400)

        missing = [heading for heading in REQUIRED_LEAD_IMPORT_HEADINGS if heading not in headings]
        if missing:
            errors = []
            errors.append({'row': None, 'error': f'Missing columns: {", ".join(missing)}'})
            history.total_rows = len(rows)
            history.failed_count = len(rows)
            history.status = LeadImportHistory.Status.FAILED
            history.error_log = errors
            history.save(update_fields=['total_rows', 'failed_count', 'status', 'error_log', 'updated_at'])
            return Response({
                'detail': 'Import failed. File headings do not match the required format.',
                'missing_columns': missing,
                'history': LeadImportHistorySerializer(history).data,
            }, status=400)

        course_map = {course.name.strip().lower(): course for course in Course.objects.filter(is_active=True)}
        source_map = {label.lower(): value for value, label in Lead.Source.choices}
        source_map.update({value.lower(): value for value, _label in Lead.Source.choices})
        existing_phones = {
            normalize_phone_number(phone)
            for phone in list(Lead.objects.values_list('phone', flat=True))
            + list(WalkIn.objects.values_list('phone', flat=True))
            + list(Enrollment.objects.values_list('phone', flat=True))
            if normalize_phone_number(phone)
        }
        seen_phones = set()
        errors = []
        imported = 0
        duplicates = 0
        duplicate_rows = []

        for index, row in enumerate(rows, start=2):
            name = normalise_import_cell(row.get('Candidate Name'))
            raw_phone = normalise_import_cell(row.get('Phone Number'))
            phone = valid_import_phone(raw_phone)
            course_name = first_import_value(row, 'Course Interested', 'Course Name', 'Course Names')
            source_label = first_import_value(row, 'How They Know IIE', 'Source', 'Lead Source', 'Source Name', 'Source Names')
            source_description = normalise_import_cell(row.get('Source Description'))
            followup_raw = first_import_value(row, 'Follow-up Date', 'Follow Up Date')
            followup_date = parse_import_followup_date(followup_raw)
            remarks = normalise_import_cell(row.get('Remarks'))
            email = normalise_import_cell(row.get('Email'))
            location = normalise_import_cell(row.get('Location'))
            qualification = normalise_import_cell(row.get('Qualification'))

            row_errors = []
            if not name:
                row_errors.append('Candidate Name is required.')
            if not raw_phone:
                row_errors.append('Phone Number is required.')
            elif not phone:
                row_errors.append('Phone Number is invalid.')
            if phone and (phone in existing_phones or phone in seen_phones):
                duplicates += 1
                duplicate_rows.append({'row': index, 'phone': phone})
                seen_phones.add(phone)
                continue
            course = course_map.get(course_name.lower())
            branch = request.user.branch
            source = source_map.get(source_label.lower()) if source_label else Lead.Source.MANUAL
            if not source:
                source = Lead.Source.OTHERS
                if source_label and source_label.lower() not in source_description.lower():
                    source_description = f'{source_label} - {source_description}' if source_description else source_label
            if followup_raw and not followup_date:
                row_errors.append('Follow-up Date is invalid. Use YYYY-MM-DD or DD-MM-YYYY.')

            if row_errors:
                errors.append({'row': index, 'phone': phone, 'error': ' '.join(row_errors)})
                if phone:
                    seen_phones.add(phone)
                continue

            lead_kwargs = {
                'name': name,
                'phone': phone,
                'course': course,
                'external_course_interested': course_name,
                'branch': branch,
                'source': source,
                'source_description': source_description,
                'next_follow_up_date': followup_date,
                'remarks': remarks,
                'email': email,
                'location': location,
                'qualification': qualification,
                'created_by': request.user,
                'assigned_to': request.user,
            }
            if 'imported_via_csv' not in missing_model_columns(Lead, ['imported_via_csv']):
                lead_kwargs['imported_via_csv'] = True
            Lead.objects.create(**lead_kwargs)
            imported += 1
            seen_phones.add(phone)
            existing_phones.add(phone)

        failed = len(errors)
        if imported and failed:
            import_status = LeadImportHistory.Status.PARTIAL
        elif imported or (duplicates and not failed):
            import_status = LeadImportHistory.Status.SUCCESS
        else:
            import_status = LeadImportHistory.Status.FAILED

        history.total_rows = len(rows)
        history.success_count = imported
        history.failed_count = failed
        history.duplicate_count = duplicates
        history.status = import_status
        history.error_log = errors
        history.save()

        response_status = status.HTTP_201_CREATED if imported else (
            status.HTTP_200_OK if duplicates and not failed else status.HTTP_400_BAD_REQUEST
        )
        return Response({
            'total_rows': history.total_rows,
            'successfully_imported': imported,
            'failed_rows': failed,
            'duplicate_rows': duplicates,
            'duplicate_rows_detail': duplicate_rows,
            'errors': errors,
            'status': history.status,
            'import_summary': {
                'imported_to': [{'branch': request.user.branch.name, 'count': imported}],
                'leads_added': imported,
                'enrollments_added': 0,
                'payments_updated': 0,
                'duplicates_skipped': duplicates,
                'invalid_rows': failed,
            },
            'history': LeadImportHistorySerializer(history).data,
        }, status=response_status)


class LeadImportTemplateView(APIView):
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Leads Template'
        sheet.append([
            'Candidate Name',
            'Phone Number',
            'Course Interested',
            'Branch',
            'How They Know IIE',
            'Follow-up Date',
            'Remarks',
            'Source Description',
        ])
        sheet.append([
            'Sample Candidate',
            '9876543210',
            'Python',
            request.user.branch.name if getattr(request.user, 'branch', None) else 'Gandhipuram',
            'Google',
            timezone.localdate().isoformat(),
            'Follow up after demo enquiry.',
            'Google Ads campaign',
        ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="leads-import-template.xlsx"'
        workbook.save(response)
        return response


class LeadImportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = LeadImportHistorySerializer
    permission_classes = [IsSuperAdmin]
    pagination_class = None

    def get_queryset(self):
        return LeadImportHistory.objects.select_related('uploaded_by', 'branch').order_by('-created_at')


def import_cell(value):
    if value is None:
        return ''
    if hasattr(value, 'date') and not isinstance(value, str):
        return value.date().isoformat()
    if hasattr(value, 'isoformat') and not isinstance(value, str):
        return value.isoformat()
    return str(value).strip()


def parse_excel_date(value):
    value = import_cell(value)
    if not value:
        return None, ''
    parsed = parse_date(value)
    if parsed:
        return parsed, ''
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return timezone.datetime.strptime(value, fmt).date(), ''
        except ValueError:
            continue
    return None, 'Invalid date format. Use YYYY-MM-DD or DD-MM-YYYY.'


def parse_excel_amount(value):
    value = import_cell(value).replace(',', '')
    if value == '':
        return None, ''
    try:
        amount = Decimal(value)
    except Exception:
        return None, 'Amount must be numeric.'
    if amount < 0:
        return None, 'Amount cannot be negative.'
    return amount, ''


def read_xlsx_rows(uploaded_file):
    name = str(uploaded_file.name).lower()
    if name.endswith('.csv'):
        text = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        headers = [import_cell(header) for header in (reader.fieldnames or []) if import_cell(header)]
        rows = []
        for row_number, row in enumerate(reader, start=2):
            if not any(import_cell(value) for value in row.values()):
                continue
            rows.append({'_row_number': row_number, **{header: import_cell(row.get(header, '')) for header in headers}})
        return headers, rows
    if not name.endswith('.xlsx'):
        raise ValueError('Only Excel .xlsx or CSV files are supported.')
    import openpyxl
    workbook = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
    headers = [import_cell(value) for value in raw_headers if import_cell(value)]
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(import_cell(value) for value in values):
            continue
        rows.append({
            '_row_number': row_number,
            **{
                headers[index]: import_cell(values[index] if index < len(values) else '')
                for index in range(len(headers))
            },
        })
    return headers, rows


def active_model_fields(model):
    ignored = {'id', 'created_at', 'updated_at'}
    return [
        field.name
        for field in model._meta.fields
        if field.name not in ignored and not field.auto_created
    ]


ADMIN_IMPORT_SPECS = {
    'leads': {
        'model': Lead,
        'required': ['name', 'phone', 'branch', 'course'],
        'fields': [
            {'field': 'name', 'label': 'Name'},
            {'field': 'phone', 'label': 'Phone'},
            {'field': 'course', 'label': 'Course'},
            {'field': 'branch', 'label': 'Branch'},
            {'field': 'source', 'label': 'Source'},
            {'field': 'source_description', 'label': 'Source Description'},
            {'field': 'next_follow_up_date', 'label': 'Follow Up Date'},
            {'field': 'remarks', 'label': 'Remarks'},
            {'field': 'assigned_to', 'label': 'Assigned Counselor'},
            {'field': 'dob', 'label': 'DOB'},
            {'field': 'email', 'label': 'Email'},
            {'field': 'location', 'label': 'Location'},
            {'field': 'pincode', 'label': 'Pincode'},
            {'field': 'qualification', 'label': 'Qualification'},
            {'field': 'degree', 'label': 'Degree'},
            {'field': 'preferred_timing', 'label': 'Preferred Timing'},
            {'field': 'walkin_date', 'label': 'Walkin Date'},
        ],
    },
    'walkins': {
        'model': WalkIn,
        'required': ['name', 'phone', 'branch', 'course', 'visit_date'],
        'fields': [
            {'field': 'name', 'label': 'Name'},
            {'field': 'phone', 'label': 'Phone'},
            {'field': 'course', 'label': 'Course'},
            {'field': 'branch', 'label': 'Branch'},
            {'field': 'source', 'label': 'Source'},
            {'field': 'visit_date', 'label': 'Walk-in Date'},
            {'field': 'follow_up_date', 'label': 'Follow Up Date'},
            {'field': 'remarks', 'label': 'Remarks'},
            {'field': 'assigned_to', 'label': 'Assigned Counselor'},
            {'field': 'dob', 'label': 'DOB'},
            {'field': 'email', 'label': 'Email'},
            {'field': 'location', 'label': 'Location'},
            {'field': 'pincode', 'label': 'Pincode'},
            {'field': 'qualification', 'label': 'Qualification'},
            {'field': 'degree', 'label': 'Degree'},
            {'field': 'preferred_timing', 'label': 'Preferred Timing'},
            {'field': 'demo_class', 'label': 'Demo Class'},
            {'field': 'status', 'label': 'Status'},
        ],
    },
    'courses': {
        'model': Course,
        'required': ['name', 'actual_fees'],
        'fields': [
            {'field': 'name', 'label': 'Course Name'},
            {'field': 'duration_months', 'label': 'Duration Months'},
            {'field': 'actual_fees', 'label': 'Actual Fees'},
            {'field': 'discount_amount', 'label': 'Discount Amount'},
            {'field': 'is_active', 'label': 'Status'},
        ],
    },
    'enrollments': {
        'model': Enrollment,
        'required': ['name', 'phone', 'branch', 'course', 'actual_fees', 'enrollment_date'],
        'fields': [
            {'field': 'student_number', 'label': 'Student ID'},
            {'field': 'name', 'label': 'Student Name'},
            {'field': 'phone', 'label': 'Phone'},
            {'field': 'course', 'label': 'Course'},
            {'field': 'branch', 'label': 'Branch'},
            {'field': 'enrollment_date', 'label': 'Enrollment Date'},
            {'field': 'actual_fees', 'label': 'Final Fee'},
            {'field': 'assigned_to', 'label': 'Counselor'},
            {'field': 'status', 'label': 'Status'},
            {'field': 'email', 'label': 'Email'},
            {'field': 'dob', 'label': 'DOB'},
            {'field': 'location', 'label': 'Location'},
            {'field': 'pincode', 'label': 'Pincode'},
            {'field': 'qualification', 'label': 'Qualification'},
            {'field': 'degree', 'label': 'Degree'},
            {'field': 'source', 'label': 'Source'},
            {'field': 'preferred_timing', 'label': 'Preferred Timing'},
            {'field': 'discount_amount', 'label': 'Discount Amount'},
            {'field': 'discount_reason', 'label': 'Discount Reason'},
            {'field': 'start_date', 'label': 'Start Date'},
            {'field': 'batch_timing', 'label': 'Batch Timing'},
        ],
    },
    'students': {
        'model': Enrollment,
        'required': ['name', 'phone', 'branch', 'course', 'actual_fees', 'enrollment_date'],
        'fields': [
            {'field': 'student_number', 'label': 'Student ID'},
            {'field': 'name', 'label': 'Student Name'},
            {'field': 'phone', 'label': 'Phone'},
            {'field': 'course', 'label': 'Course'},
            {'field': 'branch', 'label': 'Branch'},
            {'field': 'enrollment_date', 'label': 'Enrollment Date'},
            {'field': 'actual_fees', 'label': 'Final Fee'},
            {'field': 'assigned_to', 'label': 'Counselor'},
            {'field': 'status', 'label': 'Status'},
            {'field': 'email', 'label': 'Email'},
            {'field': 'dob', 'label': 'DOB'},
            {'field': 'start_date', 'label': 'Start Date'},
        ],
    },
    'payments': {
        'model': PaymentInstallment,
        'required': ['amount', 'payment_date'],
        'fields': [
            {'field': 'student_number', 'label': 'Student ID'},
            {'field': 'name', 'label': 'Student Name'},
            {'field': 'phone', 'label': 'Phone'},
            {'field': 'branch', 'label': 'Branch'},
            {'field': 'amount', 'label': 'Amount'},
            {'field': 'payment_date', 'label': 'Payment Date'},
            {'field': 'payment_mode', 'label': 'Payment Mode'},
            {'field': 'reference_number', 'label': 'Reference Number'},
            {'field': 'notes', 'label': 'Notes'},
        ],
    },
}


def import_spec_fields(import_type):
    spec = ADMIN_IMPORT_SPECS[import_type]
    model_fields = set(active_model_fields(spec['model']))
    related = {'branch', 'course', 'assigned_to', 'student_number', 'phone', 'name'}
    return [
        item for item in spec['fields']
        if item['field'] in model_fields or item['field'] in related
    ]


def import_template_headers(import_type):
    return [item['label'] for item in import_spec_fields(import_type)]


def import_template_mapping(import_type):
    return {item['field']: item['label'] for item in import_spec_fields(import_type)}


def lookup_branch(value):
    value = import_cell(value)
    if not value:
        return None
    if value.isdigit():
        found = Branch.objects.filter(pk=value, is_active=True).first()
        if found:
            return found
    return Branch.objects.filter(name__iexact=value, is_active=True).first()


def lookup_course(value):
    value = import_cell(value)
    if not value:
        return None
    if value.isdigit():
        found = Course.objects.filter(pk=value, is_active=True).first()
        if found:
            return found
    return Course.objects.filter(name__iexact=value, is_active=True).first()


def lookup_user(value, branch=None):
    value = import_cell(value)
    if not value:
        return None
    users = User.objects.filter(is_active=True)
    if branch:
        users = users.filter(Q(branch=branch) | Q(branch__isnull=True))
    if value.isdigit():
        found = users.filter(pk=value).first()
        if found:
            return found
    found = users.filter(
        Q(username__iexact=value) | Q(email__iexact=value) |
        Q(first_name__iexact=value) | Q(last_name__iexact=value)
    ).first()
    if found:
        return found
    value_lower = value.lower()
    return next((user for user in users if user.full_name.lower() == value_lower), None)


def choice_value(value, choices, default=''):
    raw = import_cell(value)
    if not raw:
        return default
    raw_key = raw.lower().replace(' ', '_').replace('-', '_')
    for choice_value_item, label in choices:
        if raw_key == choice_value_item.lower() or raw.lower() == str(label).lower():
            return choice_value_item
    return None


def normalized_phone_exists(model, phone):
    normalized = normalize_phone_number(phone)
    if not normalized:
        return False
    return any(
        normalize_phone_number(existing) == normalized
        for existing in model.objects.exclude(phone='').values_list('phone', flat=True)
    )


def enrollment_duplicate_exists(student_number='', phone=''):
    student_number = import_cell(student_number)
    if student_number and Enrollment.objects.filter(student_number__iexact=student_number).exists():
        return True
    return normalized_phone_exists(Enrollment, phone)


def lead_duplicate_exists(phone=''):
    normalized = normalize_phone_number(phone)
    if not normalized:
        return False
    return normalized_phone_exists(Lead, normalized) or normalized_phone_exists(Enrollment, normalized)


def payment_installment_duplicate_exists(payment, amount, payment_date, payment_mode, reference_number=''):
    if not payment or amount is None or not payment_date:
        return False
    existing_installments = PaymentInstallment.objects.filter(
        payment=payment,
        amount=amount,
        payment_date=payment_date,
        payment_mode=payment_mode,
    )
    reference_number = import_cell(reference_number)
    if reference_number:
        existing_installments = existing_installments.filter(reference_number__iexact=reference_number)
    return existing_installments.exists()


def course_duplicate_exists(name):
    name = import_cell(name)
    return bool(name and Course.objects.filter(name__iexact=name).exists())


def walkin_duplicate_exists(phone='', course=None):
    normalized = normalize_phone_number(phone)
    if not normalized:
        return False
    queryset = WalkIn.objects.exclude(phone='')
    if course:
        queryset = queryset.filter(course=course)
    return any(normalize_phone_number(existing) == normalized for existing in queryset.values_list('phone', flat=True))


def valid_import_phone(value):
    phone = normalize_phone_number(value)
    if phone.startswith('91') and len(phone) == 12:
        phone = phone[2:]
    if phone.startswith('0') and len(phone) == 11:
        phone = phone[1:]
    return phone if len(phone) == 10 else ''


def format_export_value(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if hasattr(value, 'strftime'):
        if hasattr(value, 'hour'):
            return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S') if timezone.is_aware(value) else value.strftime('%Y-%m-%d %H:%M:%S')
        return value.strftime('%Y-%m-%d')
    return value


def export_tabular_response(filename, headers, rows, file_format='xlsx', sheet_title='Export'):
    safe_format = 'csv' if file_format == 'csv' else 'xlsx'
    if safe_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([format_export_value(value) for value in row])
        return response

    try:
        import openpyxl
    except ImportError:
        return export_tabular_response(filename, headers, rows, 'csv')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    safe_title = ''.join(char for char in str(sheet_title or 'Export') if char not in '[]:*?/\\')[:31] or 'Export'
    sheet.title = safe_title
    sheet.append([format_export_value(value) for value in headers])
    for row in rows or []:
        sheet.append([format_export_value(value) for value in row])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response


def requested_export_format(request):
    return 'csv' if (request.query_params.get('format') or '').lower() == 'csv' else 'xlsx'


def validate_admin_import(import_type, headers, rows, mapping):
    spec = ADMIN_IMPORT_SPECS[import_type]
    required = [field for field in spec['required'] if field in {item['field'] for item in import_spec_fields(import_type)}]
    expected_headers = import_template_headers(import_type)
    required_labels = {item['field']: item['label'] for item in import_spec_fields(import_type) if item['field'] in required}
    missing = [label for field, label in required_labels.items() if label not in headers]
    missing_optional = [label for label in expected_headers if label not in headers and label not in missing]
    extra = [header for header in headers if header and header not in expected_headers]
    comparable_headers = [header for header in headers if header in expected_headers]
    expected_present_headers = [header for header in expected_headers if header in headers]
    invalid_order = bool(comparable_headers and comparable_headers != expected_present_headers)
    column_results = [
        {
            'field': item['field'],
            'label': item['label'],
            'header': item['label'] if item['label'] in headers else '',
            'status': 'matched' if item['label'] in headers else ('missing' if item['field'] in required else 'optional'),
        }
        for item in import_spec_fields(import_type)
    ]
    ready, skipped, failed = [], [], []
    seen_phones, seen_students, seen_payments = set(), set(), set()

    for row in rows:
        row_number = row['_row_number']
        def value(field):
            return import_cell(row.get(mapping.get(field, ''), ''))

        errors = []
        skip_reason = ''
        payload = {}
        for field in required:
            if not value(field):
                errors.append(f'{required_labels.get(field, field)} is required.')

        if import_type == 'leads':
            phone = normalize_phone_number(value('phone'))
            if value('phone') and not valid_import_phone(value('phone')):
                errors.append('Phone must be a valid 10 digit Indian mobile number.')
            if phone and (lead_duplicate_exists(phone) or phone in seen_phones):
                skip_reason = 'Lead already exists - skipped safely.'
            if phone:
                seen_phones.add(phone)
            branch = lookup_branch(value('branch'))
            course = lookup_course(value('course'))
            if value('branch') and not branch:
                errors.append('Branch does not exist.')
            if value('course') and not course:
                errors.append('Course does not exist.')
            counsellor = lookup_user(value('assigned_to'), branch) if value('assigned_to') else None
            if value('assigned_to') and not counsellor:
                errors.append('Counsellor does not exist.')
            source = choice_value(value('source'), Lead.Source.choices, Lead.Source.MANUAL)
            if source is None:
                errors.append('Source is invalid.')
            for date_field in ('dob', 'walkin_date', 'next_follow_up_date'):
                parsed, error = parse_excel_date(value(date_field))
                if error:
                    errors.append(f'{date_field}: {error}')
                payload[date_field] = parsed
            payload.update({
                'name': value('name'), 'phone': value('phone'), 'branch_id': branch.id if branch else None,
                'branch_name': branch.name if branch else '',
                'course_id': course.id if course else None, 'assigned_to_id': counsellor.id if counsellor else None,
                'email': value('email'), 'location': value('location'), 'pincode': value('pincode'),
                'qualification': value('qualification'), 'degree': value('degree'),
                'preferred_timing': value('preferred_timing'), 'source': source,
                'remarks': value('remarks'),
            })

        elif import_type == 'walkins':
            phone = normalize_phone_number(value('phone'))
            if value('phone') and not valid_import_phone(value('phone')):
                errors.append('Phone must be a valid 10 digit Indian mobile number.')
            branch = lookup_branch(value('branch'))
            course = lookup_course(value('course'))
            if value('branch') and not branch:
                errors.append('Branch does not exist.')
            if value('course') and not course:
                errors.append('Course does not exist.')
            if phone and (walkin_duplicate_exists(phone, course) or (course and (phone, course.id) in seen_phones)):
                skip_reason = 'Walk-in already exists - skipped safely.'
            if phone and course:
                seen_phones.add((phone, course.id))
            counsellor = lookup_user(value('assigned_to'), branch) if value('assigned_to') else None
            if value('assigned_to') and not counsellor:
                errors.append('Counselor does not exist.')
            source = choice_value(value('source'), WalkIn.Source.choices, WalkIn.Source.DIRECT)
            if source is None:
                errors.append('Source is invalid.')
            status_value = choice_value(value('status'), WalkIn.Status.choices, WalkIn.Status.NEW)
            if status_value is None:
                errors.append('Status is invalid.')
            for date_field in ('dob', 'visit_date', 'follow_up_date'):
                parsed, error = parse_excel_date(value(date_field))
                if error:
                    errors.append(f'{date_field}: {error}')
                payload[date_field] = parsed
            payload.update({
                'name': value('name'), 'phone': value('phone'), 'branch_id': branch.id if branch else None,
                'branch_name': branch.name if branch else '',
                'course_id': course.id if course else None, 'assigned_to_id': counsellor.id if counsellor else None,
                'email': value('email'), 'location': value('location'), 'pincode': value('pincode'),
                'qualification': value('qualification'), 'degree': value('degree'),
                'preferred_timing': value('preferred_timing'), 'source': source,
                'remarks': value('remarks'), 'status': status_value,
                'demo_class': str(value('demo_class')).strip().lower() in ('1', 'true', 'yes', 'y'),
            })

        elif import_type == 'courses':
            if course_duplicate_exists(value('name')):
                skip_reason = 'Course already exists - skipped safely.'
            actual_fees, error = parse_excel_amount(value('actual_fees'))
            if error:
                errors.append(f'actual_fees: {error}')
            discount_amount, error = parse_excel_amount(value('discount_amount'))
            if error:
                errors.append(f'discount_amount: {error}')
            duration_months = value('duration_months')
            if duration_months:
                try:
                    duration_months = int(Decimal(str(duration_months)))
                except Exception:
                    errors.append('Duration Months must be numeric.')
                    duration_months = None
            status_raw = value('is_active').strip().lower()
            is_active = status_raw not in ('inactive', 'false', '0', 'no', 'n')
            payload.update({
                'name': value('name'),
                'duration_months': duration_months or None,
                'actual_fees': actual_fees,
                'discount_amount': discount_amount or Decimal('0'),
                'is_active': is_active,
            })

        elif import_type in ('enrollments', 'students'):
            student_number = value('student_number')
            phone = normalize_phone_number(value('phone'))
            if value('phone') and not valid_import_phone(value('phone')):
                errors.append('Phone must be a valid 10 digit Indian mobile number.')
            student_key = student_number.lower() if student_number else ''
            if student_number and (Enrollment.objects.filter(student_number__iexact=student_number).exists() or student_key in seen_students):
                skip_reason = 'Student already exists - skipped safely.'
            if phone and not skip_reason:
                if normalized_phone_exists(Enrollment, phone) or phone in seen_phones:
                    skip_reason = 'Student already exists - skipped safely.'
            if phone:
                seen_phones.add(phone)
            if student_key:
                seen_students.add(student_key)
            branch = lookup_branch(value('branch'))
            course = lookup_course(value('course'))
            if value('branch') and not branch:
                errors.append('Branch does not exist.')
            if value('course') and not course:
                errors.append('Course does not exist.')
            counsellor = lookup_user(value('assigned_to'), branch) if value('assigned_to') else None
            if value('assigned_to') and not counsellor:
                errors.append('Counselor does not exist.')
            actual_fees, error = parse_excel_amount(value('actual_fees'))
            if error:
                errors.append(f'actual_fees: {error}')
            discount_amount, error = parse_excel_amount(value('discount_amount'))
            if error:
                errors.append(f'discount_amount: {error}')
            for date_field in ('dob', 'enrollment_date', 'start_date'):
                parsed, error = parse_excel_date(value(date_field))
                if error:
                    errors.append(f'{date_field}: {error}')
                payload[date_field] = parsed
            source = choice_value(value('source'), WalkIn.Source.choices, '')
            if value('source') and source is None:
                errors.append('Source is invalid.')
            status_value = choice_value(value('status'), Enrollment.Status.choices, Enrollment.Status.ACTIVE)
            if status_value is None:
                errors.append('Status is invalid.')
            payload.update({
                'student_number': student_number or None, 'name': value('name'), 'phone': value('phone'),
                'branch_id': branch.id if branch else None, 'course_id': course.id if course else None,
                'branch_name': branch.name if branch else '',
                'assigned_to_id': counsellor.id if counsellor else None,
                'email': value('email'), 'location': value('location'), 'pincode': value('pincode'),
                'qualification': value('qualification'), 'degree': value('degree'),
                'preferred_timing': value('preferred_timing'), 'source': source,
                'actual_fees': actual_fees, 'discount_amount': discount_amount or Decimal('0'),
                'discount_reason': value('discount_reason'), 'batch_timing': value('batch_timing'),
                'status': status_value,
            })

        elif import_type == 'payments':
            student_number = value('student_number')
            phone = normalize_phone_number(value('phone'))
            if value('phone') and not valid_import_phone(value('phone')):
                errors.append('Phone must be a valid 10 digit Indian mobile number.')
            branch = lookup_branch(value('branch')) if value('branch') else None
            enrollment = None
            if not student_number and not phone:
                errors.append('Stud_Id or phone is required for payment matching.')
            if student_number:
                enrollment = Enrollment.objects.filter(student_number__iexact=student_number).first()
            if not enrollment and phone:
                for candidate in Enrollment.objects.filter(status__in=Enrollment.FINAL_STATUSES).select_related('payment'):
                    if normalize_phone_number(candidate.phone) == phone:
                        enrollment = candidate
                        break
            if not enrollment:
                errors.append('Existing student/enrollment not found by Student ID or Phone Number.')
            elif enrollment.status not in Enrollment.FINAL_STATUSES:
                errors.append('Enrollment is not an active/student record yet.')
            payment = None
            if enrollment and not errors:
                try:
                    payment = enrollment.payment
                except Payment.DoesNotExist:
                    skip_reason = 'Payment record does not exist - skipped safely.'
            if value('branch') and not branch:
                errors.append('Branch does not exist.')
            if branch and enrollment and enrollment.branch_id != branch.id:
                errors.append('Branch does not match the matched student/enrollment.')
            if enrollment and phone and normalize_phone_number(enrollment.phone) != phone:
                errors.append('Phone number does not match the matched Student ID.')
            amount, error = parse_excel_amount(value('amount'))
            if error:
                errors.append(error)
            if amount is not None and amount <= 0:
                errors.append('Amount must be greater than zero.')
            payment_date, error = parse_excel_date(value('payment_date'))
            if error:
                errors.append(f'payment_date: {error}')
            payment_mode = choice_value(value('payment_mode'), PaymentInstallment.Mode.choices, PaymentInstallment.Mode.CASH)
            if payment_mode is None:
                errors.append('Payment mode is invalid.')
            if payment_mode and payment_mode != PaymentInstallment.Mode.CASH and not value('reference_number'):
                errors.append('Reference number is required for non-cash payments.')
            if enrollment and payment and amount is not None:
                pending_balance = payment.balance
                if amount > pending_balance:
                    errors.append('Payment amount exceeds pending balance.')
                duplicate_key = (
                    enrollment.id,
                    str(amount),
                    payment_date.isoformat() if payment_date else '',
                    payment_mode or '',
                    value('reference_number'),
                )
                if duplicate_key in seen_payments:
                    skip_reason = 'Duplicate payment entry in this file.'
                seen_payments.add(duplicate_key)
                if payment and payment_installment_duplicate_exists(
                    payment,
                    amount,
                    payment_date,
                    payment_mode,
                    value('reference_number'),
                ):
                    skip_reason = 'Payment already exists - skipped safely.'
            payload.update({
                'name': value('name'),
                'branch_id': branch.id if branch else (enrollment.branch_id if enrollment else None),
                'branch_name': branch.name if branch else (enrollment.branch.name if enrollment and enrollment.branch else ''),
                'enrollment_id': enrollment.id if enrollment else None,
                'amount': amount,
                'payment_date': payment_date,
                'payment_mode': payment_mode,
                'reference_number': value('reference_number'),
                'notes': value('notes'),
            })

        preview = {field: value(field) for field in mapping}
        item = {'row': row_number, 'preview': preview, 'errors': errors, 'payload': make_json_safe_payload(payload)}
        if errors:
            failed.append(item)
        elif skip_reason:
            item['skip_reason'] = skip_reason
            skipped.append(item)
        else:
            ready.append(item)

    return {
        'column_results': column_results,
        'missing_columns': missing,
        'missing_optional_columns': missing_optional,
        'extra_columns': extra,
        'invalid_order': invalid_order,
        'ready_rows': ready,
        'skipped_rows': skipped,
        'failed_rows': failed,
        'blocked': bool(missing),
    }


class AdminDataImportView(APIView):
    permission_classes = [IsSuperAdmin]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get(self, request):
        history = DataImportHistory.objects.select_related('imported_by').order_by('-created_at')[:100]
        return Response({
            'types': {
                key: {
                    'required': spec['required'],
                    'fields': import_spec_fields(key),
                    'model_fields': active_model_fields(spec['model']),
                }
                for key, spec in ADMIN_IMPORT_SPECS.items()
            },
            'history': DataImportHistorySerializer(history, many=True).data,
        })

    def post(self, request):
        action_name = request.data.get('action') or 'preview'
        if action_name == 'confirm':
            return self.confirm(request)
        return self.preview(request)

    def preview(self, request):
        import_type = request.data.get('import_type')
        if import_type not in ADMIN_IMPORT_SPECS:
            return Response({'detail': 'Select a valid import type.'}, status=400)
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'detail': 'Upload an Excel .xlsx or CSV file.'}, status=400)
        try:
            headers, rows = read_xlsx_rows(uploaded_file)
        except Exception as exc:
            return Response({'detail': str(exc)}, status=400)
        resolved_mapping = import_template_mapping(import_type)
        result = validate_admin_import(import_type, headers, rows, resolved_mapping)
        history = DataImportHistory.objects.create(
            imported_by=request.user,
            file_name=uploaded_file.name,
            import_type=import_type,
            rows_skipped=len(result['skipped_rows']),
            rows_failed=len(result['failed_rows']),
            status=DataImportHistory.Status.FAILED if result['blocked'] else DataImportHistory.Status.PREVIEWED,
            error_log=result['failed_rows'][:50] + result['skipped_rows'][:50],
        )
        token = None
        if not result['blocked']:
            token = str(uuid.uuid4())
            cache.set(f'admin-import:{token}', {
                'history_id': history.id,
                'import_type': import_type,
                'ready_rows': result['ready_rows'],
            }, timeout=60 * 30)
        return Response({
            'preview_token': token,
            'headers': headers,
            'history': DataImportHistorySerializer(history).data,
            'summary': {
                'ready_to_import': len(result['ready_rows']),
                'skipped_duplicates': len(result['skipped_rows']),
                'invalid_rows': len(result['failed_rows']),
            },
            **result,
        }, status=400 if result['blocked'] else 200)

    def confirm(self, request):
        token = request.data.get('preview_token')
        cached = cache.get(f'admin-import:{token}') if token else None
        if not cached:
            return Response({'detail': 'Preview expired. Upload and validate the file again.'}, status=400)
        import_type = cached['import_type']
        rows = cached['ready_rows']
        history = DataImportHistory.objects.get(pk=cached['history_id'])
        imported, skipped, failed, import_summary = self._create_records(import_type, rows, request.user)
        history.rows_imported = imported
        history.rows_skipped = history.rows_skipped + len(skipped)
        history.rows_failed = history.rows_failed + len(failed)
        history.status = DataImportHistory.Status.PARTIAL if (failed or skipped) else DataImportHistory.Status.SUCCESS
        if not imported and failed and not skipped:
            history.status = DataImportHistory.Status.FAILED
        history.error_log = (history.error_log or []) + skipped[:100] + failed[:100]
        history.save()
        cache.delete(f'admin-import:{token}')
        response_status = 201 if imported else (200 if skipped and not failed else 400)
        return Response({
            'history': DataImportHistorySerializer(history).data,
            'rows_imported': imported,
            'rows_skipped': len(skipped),
            'rows_failed': len(failed),
            'errors': failed,
            'import_summary': import_summary,
        }, status=response_status)

    def _create_records(self, import_type, rows, user):
        imported = 0
        skipped = []
        failed = []
        branch_counts = {}
        summary = {
            'imported_to': [],
            'leads_added': 0,
            'enrollments_added': 0,
            'payments_added': 0,
            'payments_updated': 0,
            'new_records_added': 0,
            'duplicates_skipped': 0,
            'invalid_rows': 0,
            'failed_rows': 0,
        }

        def note_branch(payload):
            branch_name = payload.get('branch_name') or ''
            if not branch_name and payload.get('branch_id'):
                branch = Branch.objects.filter(pk=payload.get('branch_id')).first()
                branch_name = branch.name if branch else ''
            if branch_name:
                branch_counts[branch_name] = branch_counts.get(branch_name, 0) + 1

        for row in rows:
            payload = row['payload']
            try:
                with transaction.atomic():
                    if import_type == 'leads':
                        if lead_duplicate_exists(payload.get('phone')):
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Lead already exists - skipped safely.'})
                            continue
                        Lead.objects.create(
                            name=payload['name'],
                            phone=payload['phone'],
                            branch_id=payload['branch_id'],
                            course_id=payload['course_id'],
                            assigned_to_id=payload.get('assigned_to_id') or user.id,
                            created_by=user,
                            dob=payload.get('dob') or None,
                            email=payload.get('email') or '',
                            location=payload.get('location') or '',
                            pincode=payload.get('pincode') or '',
                            qualification=payload.get('qualification') or '',
                            degree=payload.get('degree') or '',
                            preferred_timing=payload.get('preferred_timing') or '',
                            walkin_date=payload.get('walkin_date') or None,
                            next_follow_up_date=payload.get('next_follow_up_date') or None,
                            source=payload.get('source') or Lead.Source.MANUAL,
                            remarks=payload.get('remarks') or '',
                            imported_via_csv=True,
                        )
                        summary['leads_added'] += 1
                        summary['new_records_added'] += 1
                        note_branch(payload)
                    elif import_type == 'walkins':
                        course = Course.objects.filter(pk=payload.get('course_id')).first()
                        if walkin_duplicate_exists(payload.get('phone'), course):
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Walk-in already exists - skipped safely.'})
                            continue
                        WalkIn.objects.create(
                            name=payload['name'],
                            phone=payload['phone'],
                            branch_id=payload['branch_id'],
                            course_id=payload['course_id'],
                            assigned_to_id=payload.get('assigned_to_id') or user.id,
                            created_by=user,
                            dob=payload.get('dob') or None,
                            email=payload.get('email') or '',
                            location=payload.get('location') or '',
                            pincode=payload.get('pincode') or '',
                            qualification=payload.get('qualification') or '',
                            degree=payload.get('degree') or '',
                            preferred_timing=payload.get('preferred_timing') or '',
                            visit_date=payload.get('visit_date') or timezone.localdate(),
                            follow_up_date=payload.get('follow_up_date') or None,
                            source=payload.get('source') or WalkIn.Source.DIRECT,
                            remarks=payload.get('remarks') or '',
                            demo_class=bool(payload.get('demo_class')),
                            status=payload.get('status') or WalkIn.Status.NEW,
                        )
                        summary['new_records_added'] += 1
                        note_branch(payload)
                    elif import_type == 'courses':
                        if course_duplicate_exists(payload.get('name')):
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Course already exists - skipped safely.'})
                            continue
                        Course.objects.create(
                            name=payload['name'],
                            duration_months=payload.get('duration_months'),
                            actual_fees=Decimal(str(payload.get('actual_fees') or 0)),
                            discount_amount=Decimal(str(payload.get('discount_amount') or 0)),
                            is_active=bool(payload.get('is_active', True)),
                            created_by=user,
                        )
                        summary['new_records_added'] += 1
                    elif import_type in ('enrollments', 'students'):
                        if enrollment_duplicate_exists(payload.get('student_number'), payload.get('phone')):
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Student already exists - skipped safely.'})
                            continue
                        enrollment = Enrollment(
                            student_number=payload.get('student_number') or None,
                            name=payload['name'],
                            phone=payload['phone'],
                            branch_id=payload['branch_id'],
                            course_id=payload['course_id'],
                            final_enrollment_course_id=payload['course_id'],
                            enrolled_by=user,
                            created_by=user,
                            counselor_id=payload.get('assigned_to_id') or user.id,
                            dob=payload.get('dob') or None,
                            email=payload.get('email') or '',
                            location=payload.get('location') or '',
                            pincode=payload.get('pincode') or '',
                            qualification=payload.get('qualification') or '',
                            degree=payload.get('degree') or '',
                            preferred_timing=payload.get('preferred_timing') or '',
                            source=payload.get('source') or '',
                            actual_fees=Decimal(str(payload.get('actual_fees') or 0)),
                            discount_amount=Decimal(str(payload.get('discount_amount') or 0)),
                            discount_reason=payload.get('discount_reason') or '',
                            enrollment_date=payload.get('enrollment_date') or timezone.localdate(),
                            start_date=payload.get('start_date') or None,
                            batch_timing=payload.get('batch_timing') or '',
                            status=payload.get('status') or Enrollment.Status.ACTIVE,
                        )
                        enrollment.save()
                        enrollment.payment_schedule = serialize_enrollment_payment_schedule(enrollment)
                        enrollment.payment_schedule_locked = enrollment.status in Enrollment.FINAL_STATUSES
                        enrollment.payment_schedule_finalized_at = timezone.now() if enrollment.payment_schedule_locked else None
                        enrollment.save(update_fields=['payment_schedule', 'payment_schedule_locked', 'payment_schedule_finalized_at', 'updated_at'])
                        Payment.objects.get_or_create(
                            enrollment=enrollment,
                            defaults={
                                'total_fees': enrollment_payable_fee(enrollment),
                                'manual_installment_schedule': serialize_enrollment_payment_schedule(enrollment),
                            },
                        )
                        summary['enrollments_added'] += 1
                        summary['new_records_added'] += 1
                        note_branch(payload)
                    elif import_type == 'payments':
                        enrollment = Enrollment.objects.get(pk=payload['enrollment_id'])
                        try:
                            payment = enrollment.payment
                        except Payment.DoesNotExist:
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Payment record does not exist - skipped safely.'})
                            continue
                        amount = Decimal(str(payload.get('amount') or 0))
                        if amount > payment.balance:
                            raise ValueError('Payment amount exceeds pending balance.')
                        payment_date = payload.get('payment_date') or timezone.localdate()
                        payment_mode = payload.get('payment_mode') or PaymentInstallment.Mode.CASH
                        if payment_installment_duplicate_exists(
                            payment,
                            amount,
                            payment_date,
                            payment_mode,
                            payload.get('reference_number'),
                        ):
                            skipped.append({'row': row.get('row'), 'skip_reason': 'Payment already exists - skipped safely.'})
                            continue
                        allocation = PaymentInstallmentViewSet()._payment_allocation(payment, amount)
                        active = allocation[0] if allocation else {'index': 1, 'label': 'Imported Payment'}
                        reference_number = payload.get('reference_number') or ''
                        if not reference_number and payment_mode == PaymentInstallment.Mode.CASH:
                            student_id = enrollment.student_number or f'ENR{enrollment.id}'
                            reference_number = f'{student_id}-P{payment.installments.count() + 1:02d}'
                        PaymentInstallment.objects.create(
                            payment=payment,
                            enrollment=enrollment,
                            amount=amount,
                            payment_date=payment_date,
                            payment_mode=payment_mode,
                            reference_number=reference_number,
                            notes=payload.get('notes') or '',
                            collected_by=user,
                            installment_index=active['index'],
                            installment_label=active['label'],
                            document_type=PaymentInstallment.DocumentType.RECEIPT,
                        )
                        payment.refresh_from_db()
                        resolve_payment_due_notifications_if_inactive(payment)
                        summary['payments_added'] += 1
                        summary['payments_updated'] += 1
                        summary['new_records_added'] += 1
                        note_branch({
                            **payload,
                            'branch_id': enrollment.branch_id,
                            'branch_name': enrollment.branch.name if enrollment.branch else '',
                        })
                    imported += 1
            except Exception as exc:
                failed.append({'row': row.get('row'), 'error': str(exc)})
        summary['duplicates_skipped'] = len(skipped)
        summary['invalid_rows'] = len(failed)
        summary['failed_rows'] = len(failed)
        summary['imported_to'] = [
            {'branch': branch, 'count': count}
            for branch, count in sorted(branch_counts.items())
        ]
        return imported, skipped, failed, summary


class AdminDataImportTemplateView(APIView):
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        import_type = request.query_params.get('type')
        if import_type not in ADMIN_IMPORT_SPECS:
            return Response({'detail': 'Select a valid import type.'}, status=400)
        import openpyxl
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f'{import_type.title()} Template'
        fields = import_spec_fields(import_type)
        sheet.append([item['label'] for item in fields])
        sample = []
        for item in fields:
            field = item['field']
            if field == 'student_number':
                sample.append('STU2022-145' if import_type == 'payments' else '')
            elif field in ('name',):
                sample.append('Sample Student')
            elif field == 'phone':
                sample.append('9876543210')
            elif field == 'branch':
                sample.append('Gandhipuram')
            elif field == 'course':
                sample.append('Python')
            elif field in ('dob', 'walkin_date', 'next_follow_up_date', 'enrollment_date', 'start_date', 'payment_date'):
                sample.append(timezone.localdate().isoformat())
            elif field in ('actual_fees', 'amount'):
                sample.append('10000')
            elif field == 'discount_amount':
                sample.append('0')
            elif field == 'payment_mode':
                sample.append('cash')
            elif field == 'is_active':
                sample.append('active')
            elif field == 'status':
                sample.append('active')
            elif field == 'duration_months':
                sample.append('3')
            else:
                sample.append('')
        sheet.append(sample)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{import_type}-import-template.xlsx"'
        workbook.save(response)
        return response


DATA_EXPORT_TYPES = {
    'leads': 'Leads',
    'walkins': 'Walkins',
    'enrollments': 'Enrollments',
    'students': 'Students',
    'payments': 'Payments',
    'courses': 'Courses',
    'users': 'Users Report',
    'users_report': 'Users Report',
}


DATA_EXPORT_TYPE_ALIASES = {
    'users_report': 'users_report',
    'users': 'users_report',
}


def normalize_data_export_type(export_type):
    export_type = str(export_type or '').strip()
    return DATA_EXPORT_TYPE_ALIASES.get(export_type, export_type)


DATA_EXPORT_PERIODS = {
    'last_1_month': 'Last 1 Month',
    'last_3_months': 'Last 3 Months',
    'last_6_months': 'Last 6 Months',
    'last_1_year': 'Last 1 Year',
    'last_2_years': 'Last 2 Years',
    'last_3_years': 'Last 3 Years',
}


DATA_EXPORT_PERIOD_ALIASES = {
    'last_3_month': 'last_3_months',
    'last_6_month': 'last_6_months',
    'last_2_year': 'last_2_years',
    'last_3_year': 'last_3_years',
}


def normalize_data_export_period(period):
    period = str(period or 'last_1_month').strip()
    return DATA_EXPORT_PERIOD_ALIASES.get(period, period)


def add_months(value, months):
    month = value.month + months
    year = value.year + (month - 1) // 12
    month = ((month - 1) % 12) + 1
    day = min(value.day, monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def data_export_date_bounds(period, date_from=None, date_to=None):
    period = normalize_data_export_period(period)
    today = timezone.localdate()
    relative_periods = {
        'last_1_month': (add_months(today, -1), today),
        'last_3_months': (add_months(today, -3), today),
        'last_6_months': (add_months(today, -6), today),
        'last_1_year': (add_months(today, -12), today),
        'last_2_years': (add_months(today, -24), today),
        'last_3_years': (add_months(today, -36), today),
    }
    if period in relative_periods:
        return relative_periods[period]
    if period == 'custom':
        try:
            start_date = parse_date(date_from or '') if date_from else None
            end_date = parse_date(date_to or '') if date_to else None
        except (TypeError, ValueError):
            logger.warning('Invalid admin data export custom date range: from=%s to=%s', date_from, date_to)
            return None, None
        if start_date and end_date and start_date > end_date:
            return end_date, start_date
        return start_date, end_date
    return None, None


def filter_by_date_range(queryset, field_name, start_date, end_date, is_datetime=False):
    if start_date:
        lookup = f'{field_name}__date__gte' if is_datetime else f'{field_name}__gte'
        queryset = queryset.filter(**{lookup: start_date})
    if end_date:
        lookup = f'{field_name}__date__lte' if is_datetime else f'{field_name}__lte'
        queryset = queryset.filter(**{lookup: end_date})
    return queryset


def data_export_filename(export_type, period, branch_id=None):
    parts = [export_type]
    if branch_id:
        try:
            branch = Branch.objects.filter(pk=int(branch_id)).first()
        except (TypeError, ValueError):
            branch = None
        if branch:
            parts.append(branch.name.lower().replace(' ', '_'))
    parts.append(period or timezone.localdate().strftime('%b_%Y').lower())
    return '_'.join(parts).replace('-', '_')


def admin_template_export_response(request, export_type, filename):
    headers, rows, _total = AdminDataExportView().export_rows(request, export_type)
    return export_tabular_response(
        filename,
        headers,
        rows,
        requested_export_format(request),
        DATA_EXPORT_TYPES.get(export_type, 'Export'),
    )


class AdminDataExportView(APIView):
    permission_classes = [IsSuperAdmin]
    preview_limit = 10

    def request_payload(self, request):
        query_params = getattr(request, 'query_params', {})
        request_data = getattr(request, 'data', {})
        return {
            'request_data': dict(request_data.items()) if hasattr(request_data, 'items') else request_data,
            'query_params': dict(query_params.items()) if hasattr(query_params, 'items') else query_params,
        }

    def request_value(self, request, *keys, default=''):
        sources = []
        for source in (getattr(request, 'data', None), getattr(request, 'query_params', None)):
            if source is not None and source not in sources:
                sources.append(source)
        for source in sources:
            for key in keys:
                value = source.get(key)
                if value not in (None, ''):
                    return value
        return default

    def export_request_context(self, request):
        download = str(self.request_value(request, 'download', default='')).lower() in ('1', 'true', 'yes')
        payload = self.request_payload(request)
        return {
            'export_type': normalize_data_export_type(self.request_value(request, 'exportType', 'export_type', 'type', default='')),
            'period': normalize_data_export_period(self.request_value(request, 'dateFilter', 'date_filter', 'period', default='')),
            'branch_id': self.safe_pk(self.request_value(request, 'branch', 'branchId', 'branch_id', default=''), 'branch'),
            'user_id': self.safe_pk(self.request_value(request, 'userId', 'user_id', 'user', default=''), 'user'),
            'download': download,
            'raw': {
                **payload,
                'exportType': self.request_value(request, 'exportType', 'export_type', 'type', default=''),
                'dateFilter': self.request_value(request, 'dateFilter', 'date_filter', 'period', default=''),
                'branch': self.request_value(request, 'branch', 'branchId', 'branch_id', default=''),
                'userId': self.request_value(request, 'userId', 'user_id', 'user', default=''),
            },
        }

    def validate_export_context(self, context):
        errors = {}
        if not context['raw'].get('exportType'):
            errors['exportType'] = 'Export type is required.'
        elif context['export_type'] not in DATA_EXPORT_TYPES:
            errors['exportType'] = 'Select a valid export type.'
        if not context['raw'].get('dateFilter'):
            errors['dateFilter'] = 'Date filter is required.'
        elif context['period'] not in DATA_EXPORT_PERIODS and context['period'] != 'custom':
            errors['dateFilter'] = 'Invalid date filter.'
        if context['branch_id'] == 'invalid':
            errors['branch'] = 'Select a valid branch.'
        elif context['branch_id'] and not Branch.objects.filter(pk=context['branch_id']).exists():
            errors['branch'] = 'Selected branch was not found.'
        if context['user_id'] == 'invalid':
            errors['userId'] = 'Select a valid user.'
        elif context['user_id'] and not User.objects.filter(pk=context['user_id']).exists():
            errors['userId'] = 'Selected user was not found.'
        return errors

    def get(self, request):
        context = self.export_request_context(request)
        export_type = context['export_type']
        period = context['period']
        logger.info('Admin data export request payload=%s', context['raw'])
        logger.info('Admin data export type=%s', export_type or '(missing)')
        errors = self.validate_export_context(context)
        if errors:
            logger.warning(
                'Invalid admin data export parameters type=%s period=%s branch=%s user=%s payload=%s errors=%s',
                export_type,
                period,
                context['raw'].get('branch') or '',
                context['raw'].get('userId') or '',
                context['raw'],
                errors,
            )
            return Response({
                'success': False,
                'message': 'Invalid export parameters.',
                'errors': errors,
            }, status=400)
        try:
            logger.info(
                'Admin data export requested type=%s download=%s period=%s branch=%s user=%s payload=%s',
                export_type,
                context['download'],
                period,
                context['branch_id'] or 'all',
                context['user_id'] or 'all',
                context['raw'],
            )
            if context['download']:
                return self.download_response(request, context)
            return self.preview_response(request, context)
        except Exception as exc:
            logger.exception('Unable to generate admin data export for type=%s filters=%s', export_type, context)
            label = DATA_EXPORT_TYPES.get(export_type, 'data')
            return Response({
                'success': False,
                'message': f'Unable to generate {label} export: {exc}',
                'error': str(exc),
            }, status=400)

    def post(self, request):
        return self.get(request)

    def preview_response(self, request, context):
        export_type = context['export_type']
        headers, rows, total = self.export_rows(request, export_type, context)
        logger.info('Admin data export preview ready type=%s count=%s', export_type, total)
        return Response({
            'success': True,
            'type': export_type,
            'label': DATA_EXPORT_TYPES[export_type],
            'headers': headers,
            'rows': rows[:self.preview_limit],
            'total': total,
            'filters': self.filter_summary(request, context),
            'message': '' if total else 'No records found for selected filters',
        })

    def download_response(self, request, context):
        export_type = context['export_type']
        headers, rows, total = self.export_rows(request, export_type, context)
        logger.info('Admin data export download rows ready type=%s count=%s', export_type, total)
        filename = data_export_filename(export_type, context['period'], context['branch_id'] or '')
        response = export_tabular_response(filename, headers, rows, 'xlsx', DATA_EXPORT_TYPES[export_type])
        logger.info('Admin data export Excel generated type=%s filename=%s', export_type, filename)
        return response

    def safe_pk(self, value, label):
        if value in (None, ''):
            return ''
        try:
            return int(value)
        except (TypeError, ValueError):
            logger.warning('Invalid admin data export %s filter: %s', label, value)
            return 'invalid'

    def filter_summary(self, request, context=None):
        context = context or self.export_request_context(request)
        period = context['period']
        branch_id = context['branch_id']
        user_id = context['user_id']
        branch = Branch.objects.filter(pk=branch_id).first() if branch_id else None
        user = User.objects.filter(pk=user_id).first() if user_id else None
        start_date, end_date = data_export_date_bounds(
            period,
            self.request_value(request, 'dateFrom', 'date_from', default=''),
            self.request_value(request, 'dateTo', 'date_to', default=''),
        )
        return {
            'period': DATA_EXPORT_PERIODS.get(period, (period or 'all_dates').replace('_', ' ').title()),
            'date_from': format_export_value(start_date),
            'date_to': format_export_value(end_date),
            'branch': branch.name if branch else 'All branches',
            'user': user.full_name if user else 'All users',
        }

    def filtered_queryset(self, request, export_type, context=None):
        context = context or self.export_request_context(request)
        branch_id = context['branch_id']
        user_id = context['user_id']
        period = context['period']
        start_date, end_date = data_export_date_bounds(
            period,
            self.request_value(request, 'dateFrom', 'date_from', default=''),
            self.request_value(request, 'dateTo', 'date_to', default=''),
        )
        logger.info(
            'Admin data export filters type=%s start=%s end=%s branch=%s user=%s',
            export_type,
            start_date,
            end_date,
            branch_id or 'all',
            user_id or 'all',
        )
        if export_type == 'leads':
            qs = visible_candidate_queryset(Lead.objects.select_related('branch', 'course', 'assigned_to', 'created_by'))
            qs = filter_by_date_range(qs, 'created_at', start_date, end_date, is_datetime=True)
            if branch_id:
                qs = qs.filter(branch_id=branch_id)
            if user_id:
                qs = qs.filter(Q(assigned_to_id=user_id) | Q(created_by_id=user_id))
            return qs.order_by('-created_at')
        if export_type == 'walkins':
            qs = visible_candidate_queryset(WalkIn.objects.select_related('branch', 'course', 'assigned_to', 'created_by'))
            qs = filter_by_date_range(qs, 'visit_date', start_date, end_date)
            if branch_id:
                qs = qs.filter(branch_id=branch_id)
            if user_id:
                qs = qs.filter(assigned_to_id=user_id)
            return qs.order_by('-visit_date', '-created_at')
        if export_type in ('enrollments', 'students'):
            qs = visible_candidate_queryset(Enrollment.objects.select_related('branch', 'course', 'counselor', 'enrolled_by', 'created_by', 'payment'))
            qs = filter_by_date_range(qs, 'enrollment_date', start_date, end_date)
            if export_type == 'students':
                qs = qs.filter(status__in=Enrollment.FINAL_STATUSES)
            if branch_id:
                qs = qs.filter(branch_id=branch_id)
            if user_id:
                qs = qs.filter(enrollment_counselor_filter_q(user_id))
            return qs.order_by('-enrollment_date', '-created_at')
        if export_type == 'payments':
            qs = PaymentInstallment.objects.select_related('payment', 'enrollment__branch', 'enrollment__course', 'collected_by')
            qs = filter_by_date_range(qs, 'payment_date', start_date, end_date)
            if branch_id:
                qs = qs.filter(enrollment__branch_id=branch_id)
            if user_id:
                qs = qs.filter(enrollment_counselor_filter_q(user_id, prefix='enrollment') | Q(collected_by_id=user_id)).distinct()
            return qs.order_by('-payment_date', '-id')
        if export_type == 'courses':
            qs = Course.objects.all()
            return qs.order_by('name')
        qs = User.objects.filter(is_active=True).exclude(role=User.Role.SUPER_ADMIN).select_related('branch').order_by('first_name', 'username')
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if user_id:
            qs = qs.filter(id=user_id)
        return qs

    def export_rows(self, request, export_type, context=None):
        queryset = self.filtered_queryset(request, export_type, context)
        count = queryset.count()
        logger.info('Admin data export queryset count type=%s count=%s', export_type, count)
        if export_type in ADMIN_IMPORT_SPECS:
            headers = import_template_headers(export_type)
            fields = [item['field'] for item in import_spec_fields(export_type)]
            rows = [
                [self.safe_cell(self.import_compatible_value(export_type, record, field)) for field in fields]
                for record in queryset
            ]
            logger.info('Admin data export serializer status type=%s rows=%s', export_type, len(rows))
            return headers, rows, count
        headers, rows = self.user_report_rows(request, queryset, context)
        logger.info('Admin data export serializer status type=%s rows=%s', export_type, len(rows))
        return headers, rows, count

    def safe_cell(self, value, numeric_default=None):
        if value is None:
            return 0 if numeric_default == 0 else ''
        return format_export_value(value)

    def export_attr(self, record, field, default=''):
        return getattr(record, field, default)

    def export_related_attr(self, record, relation, field, default=''):
        related = getattr(record, relation, None)
        return getattr(related, field, default) if related else default

    def export_display_value(self, record, field):
        value = getattr(record, field, '')
        if value in (None, ''):
            return ''
        display_method = getattr(record, f'get_{field}_display', None)
        if callable(display_method):
            try:
                return display_method()
            except AttributeError:
                return value
        return value

    def import_compatible_value(self, export_type, record, field):
        if export_type == 'leads':
            values = {
                'name': self.export_attr(record, 'name'),
                'phone': self.export_attr(record, 'phone'),
                'course': self.export_related_attr(record, 'course', 'name'),
                'branch': self.export_related_attr(record, 'branch', 'name'),
                'source': self.export_display_value(record, 'source'),
                'source_description': self.export_attr(record, 'source_description'),
                'next_follow_up_date': self.export_attr(record, 'next_follow_up_date'),
                'remarks': self.export_attr(record, 'remarks'),
                'assigned_to': self.export_related_attr(record, 'assigned_to', 'full_name'),
                'dob': self.export_attr(record, 'dob'),
                'email': self.export_attr(record, 'email'),
                'location': self.export_attr(record, 'location'),
                'pincode': self.export_attr(record, 'pincode'),
                'qualification': self.export_attr(record, 'qualification'),
                'degree': self.export_attr(record, 'degree'),
                'preferred_timing': self.export_display_value(record, 'preferred_timing'),
                'walkin_date': self.export_attr(record, 'walkin_date'),
            }
            return values.get(field, '')
        if export_type == 'walkins':
            values = {
                'name': self.export_attr(record, 'name'),
                'phone': self.export_attr(record, 'phone'),
                'course': self.export_related_attr(record, 'course', 'name'),
                'branch': self.export_related_attr(record, 'branch', 'name'),
                'source': self.export_display_value(record, 'source'),
                'visit_date': self.export_attr(record, 'visit_date'),
                'follow_up_date': self.export_attr(record, 'follow_up_date'),
                'remarks': self.export_attr(record, 'remarks'),
                'assigned_to': self.export_related_attr(record, 'assigned_to', 'full_name'),
                'dob': self.export_attr(record, 'dob'),
                'email': self.export_attr(record, 'email'),
                'location': self.export_attr(record, 'location'),
                'pincode': self.export_attr(record, 'pincode'),
                'qualification': self.export_display_value(record, 'qualification'),
                'degree': self.export_attr(record, 'degree'),
                'preferred_timing': self.export_display_value(record, 'preferred_timing'),
                'demo_class': 'Yes' if self.export_attr(record, 'demo_class') else 'No',
                'status': self.export_display_value(record, 'status'),
            }
            return values.get(field, '')
        if export_type in ('enrollments', 'students'):
            values = {
                'student_number': self.export_attr(record, 'student_number'),
                'name': self.export_attr(record, 'name'),
                'phone': self.export_attr(record, 'phone'),
                'course': self.export_related_attr(record, 'course', 'name'),
                'branch': self.export_related_attr(record, 'branch', 'name'),
                'enrollment_date': self.export_attr(record, 'enrollment_date'),
                'actual_fees': enrollment_payable_fee(record),
                'assigned_to': self.export_related_attr(record, 'enrolled_by', 'full_name'),
                'status': self.export_display_value(record, 'status'),
                'email': self.export_attr(record, 'email'),
                'dob': self.export_attr(record, 'dob'),
                'location': self.export_attr(record, 'location'),
                'pincode': self.export_attr(record, 'pincode'),
                'qualification': self.export_attr(record, 'qualification'),
                'degree': self.export_attr(record, 'degree'),
                'source': self.export_display_value(record, 'source'),
                'preferred_timing': self.export_display_value(record, 'preferred_timing'),
                'discount_amount': self.export_attr(record, 'discount_amount'),
                'discount_reason': self.export_attr(record, 'discount_reason'),
                'start_date': self.export_attr(record, 'start_date'),
                'batch_timing': self.export_attr(record, 'batch_timing'),
            }
            return values.get(field, '')
        if export_type == 'payments':
            enrollment = getattr(record, 'enrollment', None)
            values = {
                'student_number': getattr(enrollment, 'student_number', '') if enrollment else '',
                'name': getattr(enrollment, 'name', '') if enrollment else '',
                'phone': getattr(enrollment, 'phone', '') if enrollment else '',
                'branch': getattr(getattr(enrollment, 'branch', None), 'name', '') if enrollment else '',
                'amount': self.export_attr(record, 'amount'),
                'payment_date': self.export_attr(record, 'payment_date'),
                'payment_mode': self.export_display_value(record, 'payment_mode'),
                'reference_number': self.export_attr(record, 'reference_number'),
                'notes': self.export_attr(record, 'notes'),
            }
            return values.get(field, '')
        if export_type == 'courses':
            values = {
                'name': self.export_attr(record, 'name'),
                'duration_months': self.export_attr(record, 'duration_months'),
                'actual_fees': self.export_attr(record, 'actual_fees'),
                'discount_amount': self.export_attr(record, 'discount_amount'),
                'is_active': 'Active' if self.export_attr(record, 'is_active') else 'Inactive',
            }
            return values.get(field, '')
        return ''

    def user_report_rows(self, request, users, context=None):
        context = context or self.export_request_context(request)
        period = context['period']
        start_date, end_date = data_export_date_bounds(
            period,
            self.request_value(request, 'dateFrom', 'date_from', default=''),
            self.request_value(request, 'dateTo', 'date_to', default=''),
        )
        headers = ['User', 'Branch', 'Leads Handled', 'Walkins Handled', 'Conversions', 'Enrollments', 'Payments Collected', 'Pending Followups']
        rows = []
        for user in users:
            leads = filter_by_date_range(visible_candidate_queryset(Lead.objects.filter(Q(assigned_to=user) | Q(created_by=user))), 'created_at', start_date, end_date, is_datetime=True)
            walkins = filter_by_date_range(visible_candidate_queryset(WalkIn.objects.filter(assigned_to=user)), 'visit_date', start_date, end_date)
            enrollments = filter_by_date_range(visible_candidate_queryset(Enrollment.objects.filter(enrollment_counselor_filter_q(user.id))), 'enrollment_date', start_date, end_date)
            installments = filter_by_date_range(PaymentInstallment.objects.filter(collected_by=user), 'payment_date', start_date, end_date)
            pending_leads = pending_follow_up_queryset(leads, FollowUp.RecordType.LEAD, 'next_follow_up_date', LEAD_CLOSED_FOLLOW_UP_STATUSES).exclude(next_follow_up_date__isnull=True).count()
            pending_walkins = pending_follow_up_queryset(walkins, FollowUp.RecordType.WALKIN, 'follow_up_date', WALKIN_CLOSED_FOLLOW_UP_STATUSES).exclude(follow_up_date__isnull=True).count()
            conversions = leads.filter(converted_to_type__in=['walkin', 'enrollment']).count() + walkins.filter(converted_to_type='enrollment').count()
            rows.append([
                self.safe_cell(user.full_name),
                self.safe_cell(user.branch.name if user.branch else ''),
                leads.count(),
                walkins.count(),
                conversions,
                enrollments.count(),
                self.safe_cell(installments.aggregate(total=Sum('amount'))['total'] or 0, numeric_default=0),
                pending_leads + pending_walkins,
            ])
        return headers, rows


class ExternalLeadCaptureView(APIView):
    """Public/API endpoint for auto-captured leads. Leads remain unassigned for admin review."""
    permission_classes = [AllowAny]

    def post(self, request):
        configured_key = getattr(settings, 'LEAD_CAPTURE_API_KEY', '')
        supplied_key = request.headers.get('X-API-KEY') or request.query_params.get('key')
        if not configured_key or supplied_key != configured_key:
            logger.warning(
                'Rejected external lead capture attempt from %s. API key missing or invalid.',
                get_client_ip(request),
            )
            return Response({'detail': 'Invalid API key.'}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        raw_phone = str(data.get('phone_number') or data.get('phone') or '').strip()
        phone_digits = normalize_phone_number(raw_phone)
        if phone_digits.startswith('0') and len(phone_digits) == 11:
            phone_digits = phone_digits[1:]
        if phone_digits.startswith('91') and len(phone_digits) == 12:
            phone_digits = phone_digits[2:]
        name = str(data.get('candidate_name') or data.get('name') or '').strip()
        if not name or not phone_digits:
            return Response({'detail': 'Candidate name and phone number are required.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(phone_digits) != 10:
            return Response({'detail': 'Phone number must be a valid 10 digit Indian mobile number.'}, status=status.HTTP_400_BAD_REQUEST)

        course = None
        course_name = str(data.get('course_interested') or data.get('course') or data.get('course_name') or '').strip()
        if course_name:
            course = Course.objects.filter(name__iexact=course_name, is_active=True).first()
        source_label = str(data.get('source') or 'website').strip().lower()
        source_map = {
            'website': Lead.Source.WEBSITE,
            'google': Lead.Source.GOOGLE,
            'justdial': Lead.Source.JUSTDIAL,
        }
        source_value = source_map.get(source_label)
        if not source_value:
            return Response({'detail': 'Source must be Website, Google, or Justdial.'}, status=status.HTTP_400_BAD_REQUEST)

        duplicate_records = matching_candidate_phone_records(phone_digits)
        message = str(data.get('message') or '').strip()
        duplicate_note = ''
        if duplicate_records:
            duplicate_note = 'Duplicate phone found in CRM: ' + '; '.join(
                f"{item['type']} #{item['id']} {item['name']} ({item['branch_name']})"
                for item in duplicate_records[:5]
            )
        remarks = '\n'.join(part for part in [message, duplicate_note] if part)

        lead = Lead.objects.create(
            name=name,
            phone=phone_digits,
            email=data.get('email', ''),
            location=data.get('location', ''),
            course=course,
            source=source_value,
            status=Lead.Status.NEW,
            branch=None,
            remarks=remarks,
            external_course_interested=course_name,
            external_message=message,
            is_duplicate=bool(duplicate_records),
            assigned_to=None,
            created_by=None,
        )
        for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
            create_user_notification(
                admin_user,
                'New unassigned lead',
                f'{lead.name} was captured from an external source.',
                Notification.NType.INFO,
                f'/leads/{lead.id}',
            )
        return Response({
            'id': lead.id,
            'lead_number': lead.lead_number,
            'duplicate': bool(duplicate_records),
            'duplicate_records': duplicate_records,
            'detail': 'Lead captured and added to Admin Lead Inbox.',
        }, status=status.HTTP_201_CREATED)


class PublicLeadFormThrottle(ScopedRateThrottle):
    scope = 'public_lead_form'


class PublicLeadFormView(APIView):
    """GET/POST /api/public/lead-form/ - public website lead capture."""
    permission_classes = [AllowAny]
    throttle_classes = [PublicLeadFormThrottle]

    willing_to_join_options = [
        {'value': Lead.WillingToJoin.WITHIN_MONTH, 'label': 'Within a month'},
        {'value': Lead.WillingToJoin.MONTH_LATER, 'label': 'A month later'},
        {'value': Lead.WillingToJoin.JUST_ENQUIRY, 'label': 'Just enquiry'},
    ]
    qualification_options = [
        {'value': Lead.Qualification.SCHOOL_STUDENT, 'label': 'School Student'},
        {'value': Lead.Qualification.COLLEGE_STUDENT, 'label': 'College Student'},
        {'value': Lead.Qualification.GRADUATE, 'label': 'Graduate'},
        {'value': Lead.Qualification.HOUSEWIFE, 'label': 'Housewife'},
        {'value': Lead.Qualification.WORKING_PROFESSIONAL, 'label': 'Working Professional'},
    ]
    preferred_timing_options = [
        {'value': Lead.PreferredTiming.MORNING, 'label': 'Morning'},
        {'value': Lead.PreferredTiming.AFTERNOON, 'label': 'Afternoon'},
        {'value': Lead.PreferredTiming.EVENING, 'label': 'Evening'},
        {'value': Lead.PreferredTiming.WEEKEND, 'label': 'Weekend'},
    ]

    def get(self, request):
        branches = Branch.objects.filter(
            is_active=True,
            name__in=['Gandhipuram', 'Hopes', 'Kuniyamuthur'],
        ).order_by('name')
        return Response({
            'branches': BranchSerializer(branches, many=True).data,
            'courses': [{'id': name, 'name': name} for name in PUBLIC_LEAD_COURSE_NAMES],
            'willing_to_join_options': self.willing_to_join_options,
            'qualification_options': self.qualification_options,
            'preferred_timing_options': self.preferred_timing_options,
        })

    def post(self, request):
        if request.data.get('company'):
            return Response({'detail': 'Spam detected.'}, status=status.HTTP_400_BAD_REQUEST)

        name = str(request.data.get('full_name') or '').strip()
        phone = normalize_phone_number(request.data.get('mobile_number'))
        preferred_timing = str(request.data.get('preferred_timing') or '').strip()
        branch_id = request.data.get('branch')
        course_value = str(request.data.get('course_interested') or '').strip()

        if not name or not phone:
            return Response({'detail': 'Full Name and Mobile Number are required.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(phone) == 12 and phone.startswith('91'):
            phone = phone[2:]
        if len(phone) != 10:
            return Response({'detail': 'Mobile Number must be a valid 10 digit Indian number.'}, status=status.HTTP_400_BAD_REQUEST)

        if branch_id in (None, ''):
            return Response({'detail': 'Please select a valid branch.'}, status=status.HTTP_400_BAD_REQUEST)
        if not course_value:
            return Response({'detail': 'Please select a valid course.'}, status=status.HTTP_400_BAD_REQUEST)

        branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
        if not branch:
            return Response({'detail': 'Please select a valid branch.'}, status=status.HTTP_400_BAD_REQUEST)

        course = None
        selected_course_name = course_value
        if course_value.isdigit():
            course = Course.objects.filter(pk=course_value, is_active=True).first()
            selected_course_name = course.name if course else course_value
        else:
            course = Course.objects.filter(name__iexact=course_value, is_active=True).first()
        valid_public_course_names = {name.lower() for name in PUBLIC_LEAD_COURSE_NAMES}
        if selected_course_name.lower() not in valid_public_course_names and not course:
            return Response({'detail': 'Please select a valid course.'}, status=status.HTTP_400_BAD_REQUEST)
        if not course:
            selected_course_name = next(
                (name for name in PUBLIC_LEAD_COURSE_NAMES if name.lower() == selected_course_name.lower()),
                selected_course_name,
            )

        valid_timing = {item['value'] for item in self.preferred_timing_options}
        if preferred_timing not in valid_timing:
            return Response({'detail': 'Please select preferred timing.'}, status=status.HTTP_400_BAD_REQUEST)

        duplicate_records = matching_candidate_phone_records(phone)
        existing_lead = Lead.objects.filter(phone=phone).order_by('-updated_at', '-created_at').first()
        if existing_lead:
            existing_lead.next_follow_up_date = timezone.localdate()
            existing_lead.is_duplicate = True
            existing_lead.external_message = 'Duplicate public website lead submission.'
            existing_lead.external_course_interested = selected_course_name
            existing_lead.course = course or existing_lead.course
            existing_lead.preferred_timing = preferred_timing
            existing_lead.save(update_fields=[
                'next_follow_up_date', 'is_duplicate', 'external_message',
                'external_course_interested', 'course', 'preferred_timing', 'updated_at',
            ])
            return Response({
                'duplicate': True,
                'detail': 'Thank you! Our team will contact you shortly.',
                'lead_id': existing_lead.id,
            }, status=status.HTTP_200_OK)

        if duplicate_records:
            return Response({
                'duplicate': True,
                'detail': 'Thank you! Our team will contact you shortly.',
                'records': duplicate_records,
            }, status=status.HTTP_200_OK)

        lead = Lead.objects.create(
            name=name,
            phone=phone,
            course=course,
            branch=branch,
            status=Lead.Status.NEW,
            source=Lead.Source.WEBSITE,
            preferred_timing=preferred_timing,
            remarks='Public website lead form submission.',
            external_course_interested=selected_course_name,
            created_by=None,
            assigned_to=None,
        )
        notify_branch_users(
            branch,
            'New website lead assigned',
            f'{lead.name} submitted a website enquiry for {selected_course_name}.',
            Notification.NType.INFO,
            f'/leads/{lead.id}',
        )
        return Response({
            'id': lead.id,
            'detail': 'Thank you! Our team will contact you shortly.',
        }, status=status.HTTP_201_CREATED)


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsStaffOrAdmin]
    pagination_class = None

    def get_queryset(self):
        generate_smart_notifications(self.request.user)
        queryset = Notification.objects.filter(user=self.request.user).order_by('-created_at')
        if self.request.user.is_super_admin:
            queryset = queryset.filter(admin_action_required_filter())
        scope = self.request.query_params.get('scope', 'active')
        status_filter = self.request.query_params.get('status', '').strip()
        state_filter = self.request.query_params.get('state', '').strip()
        date_filter = parse_date(self.request.query_params.get('date', '').strip())

        if status_filter in dict(Notification.Status.choices):
            queryset = queryset.filter(status=status_filter)
        if state_filter == 'active':
            queryset = queryset.filter(status=Notification.Status.UNREAD, is_read=False)
        elif state_filter == 'done':
            queryset = queryset.filter(status=Notification.Status.RESOLVED)
        if date_filter:
            queryset = queryset.filter(created_at__date=date_filter)
        if scope != 'all':
            queryset = queryset.filter(status=Notification.Status.UNREAD, is_read=False)
        return queryset

    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        self.get_queryset().exclude(status=Notification.Status.RESOLVED).update(
            is_read=True,
            status=Notification.Status.READ,
        )
        return Response({'detail': 'Notifications marked as read.'})


class PendingManagementView(APIView):
    permission_classes = [IsStaffOrAdmin]

    def base_leads(self, request):
        qs = visible_candidate_queryset(Lead.objects.select_related('branch', 'course', 'assigned_to'))
        latest_follow_up = FollowUp.objects.filter(
            record_type=FollowUp.RecordType.LEAD,
            record_id=OuterRef('pk'),
        ).order_by('-created_at', '-id')
        qs = qs.annotate(
            latest_follow_up_remark=Subquery(latest_follow_up.values('remarks')[:1]),
            latest_follow_up_at=Subquery(latest_follow_up.values('created_at')[:1]),
        )
        if not request.user.is_super_admin:
            qs = qs.filter(branch=request.user.branch)
        elif request.query_params.get('branch'):
            qs = qs.filter(branch_id=request.query_params.get('branch'))
        staff_id = pending_staff_filter(request)
        if staff_id:
            qs = qs.filter(assigned_to_id=staff_id)
        if truthy_query_param(request.query_params.get('important_only')):
            qs = qs.filter(is_important=True)
        qs = apply_text_search(qs, request.query_params.get('search'), [
            'lead_number', 'name', 'phone', 'email', 'source', 'source_description',
            'course__name', 'assigned_to__first_name', 'assigned_to__last_name',
            'assigned_to__username',
        ], id_fields=['id'])
        qs = pending_follow_up_queryset(
            qs,
            FollowUp.RecordType.LEAD,
            'next_follow_up_date',
            LEAD_CLOSED_FOLLOW_UP_STATUSES,
        ).exclude(next_follow_up_date__isnull=True)
        return apply_overdue_follow_up_date_filter(qs, 'next_follow_up_date', request).order_by('next_follow_up_date', 'name')

    def base_walkins(self, request):
        qs = visible_candidate_queryset(WalkIn.objects.select_related('branch', 'course', 'assigned_to'))
        latest_follow_up = FollowUp.objects.filter(
            record_type=FollowUp.RecordType.WALKIN,
            record_id=OuterRef('pk'),
        ).order_by('-created_at', '-id')
        qs = qs.annotate(
            latest_follow_up_remark=Subquery(latest_follow_up.values('remarks')[:1]),
            latest_follow_up_at=Subquery(latest_follow_up.values('created_at')[:1]),
        )
        if not request.user.is_super_admin:
            qs = qs.filter(branch=request.user.branch)
        elif request.query_params.get('branch'):
            qs = qs.filter(branch_id=request.query_params.get('branch'))
        staff_id = pending_staff_filter(request)
        if staff_id:
            qs = qs.filter(assigned_to_id=staff_id)
        if truthy_query_param(request.query_params.get('important_only')):
            qs = qs.filter(is_important=True)
        qs = apply_text_search(qs, request.query_params.get('search'), [
            'candidate_number', 'name', 'phone', 'email', 'source',
            'course__name', 'assigned_to__first_name', 'assigned_to__last_name',
            'assigned_to__username',
        ], id_fields=['id'])
        qs = pending_follow_up_queryset(
            qs,
            FollowUp.RecordType.WALKIN,
            'follow_up_date',
            WALKIN_CLOSED_FOLLOW_UP_STATUSES,
        ).exclude(follow_up_date__isnull=True)
        return apply_overdue_follow_up_date_filter(qs, 'follow_up_date', request).order_by('follow_up_date', 'name')

    def base_payments(self, request):
        qs = visible_payment_queryset(Payment.objects.select_related(
            'enrollment__branch', 'enrollment__course', 'enrollment__counselor',
            'enrollment__enrolled_by', 'enrollment__created_by',
        ).prefetch_related('installments').filter(
            enrollment__status__in=Enrollment.FINAL_STATUSES,
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            paid_amount__lt=F('total_fees'),
            next_payment_date__isnull=False,
        ))
        if not request.user.is_super_admin:
            qs = qs.filter(enrollment__branch=request.user.branch)
        elif request.query_params.get('branch'):
            qs = qs.filter(enrollment__branch_id=request.query_params.get('branch'))
        staff_id = pending_staff_filter(request)
        if staff_id:
            qs = qs.filter(enrollment_counselor_filter_q(staff_id, prefix='enrollment'))
        status_filter = request.query_params.get('status') or ''
        if status_filter:
            if status_filter == 'pending':
                qs = qs.filter(status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL])
            else:
                qs = qs.filter(status=status_filter)
        if truthy_query_param(request.query_params.get('important_only')):
            qs = qs.filter(enrollment__is_important=True)
        qs = apply_text_search(qs, request.query_params.get('search'), [
            'enrollment__student_number', 'enrollment__name',
            'enrollment__phone', 'enrollment__email', 'enrollment__source',
            'enrollment__course__name', 'enrollment__enrolled_by__first_name',
            'enrollment__enrolled_by__last_name', 'enrollment__enrolled_by__username',
            'enrollment__created_by__first_name', 'enrollment__created_by__last_name',
            'enrollment__created_by__username',
        ], id_fields=['id', 'enrollment__id'])
        qs = apply_pending_date_filter(qs, 'next_payment_date', request)
        return qs.order_by('next_payment_date', 'enrollment__name')

    def lead_row(self, lead):
        latest_remark = getattr(lead, 'latest_follow_up_remark', None)
        return {
            'id': lead.id,
            'name': lead.name,
            'phone': lead.phone,
            'course_name': lead.course.name if lead.course else '',
            'branch_name': lead.branch.name if lead.branch else '',
            'status': lead.status,
            'status_display': automated_lead_status_display(lead),
            'due_date': lead.next_follow_up_date,
            'remarks': latest_remark if latest_remark not in (None, '') else lead.remarks,
            'latest_remark': latest_remark if latest_remark not in (None, '') else lead.remarks,
            'latest_follow_up_at': getattr(lead, 'latest_follow_up_at', None),
            'assigned_to_name': lead.assigned_to.full_name if lead.assigned_to else '',
            'assigned_user': user_identity_payload(lead.assigned_to),
            'is_important': lead.is_important,
            'detail_url': f'/leads/{lead.id}',
        }

    def walkin_row(self, walkin):
        latest_remark = getattr(walkin, 'latest_follow_up_remark', None)
        return {
            'id': walkin.id,
            'name': walkin.name,
            'phone': walkin.phone,
            'course_name': walkin.course.name if walkin.course else '',
            'branch_name': walkin.branch.name if walkin.branch else '',
            'status': walkin.status,
            'status_display': automated_walkin_status_display(walkin),
            'due_date': walkin.follow_up_date,
            'remarks': latest_remark if latest_remark not in (None, '') else walkin.remarks,
            'latest_remark': latest_remark if latest_remark not in (None, '') else walkin.remarks,
            'latest_follow_up_at': getattr(walkin, 'latest_follow_up_at', None),
            'assigned_to_name': walkin.assigned_to.full_name if walkin.assigned_to else '',
            'assigned_user': user_identity_payload(walkin.assigned_to),
            'is_important': walkin.is_important,
            'detail_url': f'/walkins/{walkin.id}',
        }

    def payment_row(self, payment):
        summary = payment_installment_summary(payment)
        due = next((item for item in summary if Decimal(str(item.get('pending_amount') or 0)) > 0), None)
        enrollment = payment.enrollment
        counselor = enrollment.counselor or enrollment.created_by or enrollment.enrolled_by
        return {
            'id': payment.id,
            'student_name': enrollment.name,
            'student_number': enrollment.student_number,
            'phone': enrollment.phone,
            'course_name': enrollment.course.name if enrollment.course else '',
            'branch_name': enrollment.branch.name if enrollment.branch else '',
            'due_date': payment.next_payment_date,
            'due_amount': due.get('pending_amount') if due else payment.balance,
            'pending_balance': payment.balance,
            'installment_label': due.get('label') if due else 'Pending Payment',
            'installment_status': due.get('status') if due else payment.status,
            'payment_status': payment.status,
            'counselor_name': counselor.full_name if counselor else '',
            'counselor_user': user_identity_payload(counselor),
            'is_important': enrollment.is_important,
            'detail_url': f'/payments/{payment.id}',
        }

    def get(self, request, section='summary'):
        if section == 'summary':
            leads = self.base_leads(request).count()
            walkins = self.base_walkins(request).count()
            payments = self.base_payments(request).count()
            return Response({
                'lead_pending': leads,
                'walkin_pending': walkins,
                'payment_pending': payments,
                'total_pending': leads + walkins + payments,
            })
        if section == 'leads':
            rows = [self.lead_row(lead) for lead in self.base_leads(request)[:500]]
        elif section == 'walkins':
            rows = [self.walkin_row(walkin) for walkin in self.base_walkins(request)[:500]]
        elif section == 'payments':
            rows = [self.payment_row(payment) for payment in self.base_payments(request)[:500]]
        else:
            return Response({'detail': 'Invalid pending module.'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'results': rows, 'count': len(rows)})


class TeamNoticeViewSet(viewsets.ModelViewSet):
    serializer_class = TeamNoticeSerializer
    permission_classes = [IsStaffOrAdmin]
    pagination_class = None

    def visible_queryset(self):
        queryset = TeamNotice.objects.select_related('branch', 'created_by').prefetch_related(
            'replies__replied_by',
            'replies__branch',
        )
        user = self.request.user
        if user.is_super_admin:
            return queryset
        return queryset.filter(
            Q(audience_type=TeamNotice.AudienceType.ALL_BRANCHES)
            | Q(audience_type=TeamNotice.AudienceType.SPECIFIC_BRANCH, branch=user.branch)
        )

    def get_queryset(self):
        queryset = self.visible_queryset().order_by('-created_at')
        status_filter = self.request.query_params.get('status', '').strip()
        branch_id = self.request.query_params.get('branch', '').strip()

        if branch_id and self.request.user.is_super_admin:
            queryset = queryset.filter(audience_type=TeamNotice.AudienceType.SPECIFIC_BRANCH, branch_id=branch_id)
        if status_filter in dict(TeamNotice.Status.choices):
            queryset = queryset.filter(status=status_filter)
        return queryset

    def perform_create(self, serializer):
        notice = serializer.save(created_by=self.request.user)
        recipients = self.notice_recipients(notice)
        for user in recipients:
            create_user_notification(
                user,
                'Team Board Notice',
                notice.title,
                Notification.NType.INFO,
                f'/team-board?notice={notice.id}',
            )

    def create(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can create Team Board notices.'}, status=403)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can update Team Board notices.'}, status=403)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can update Team Board notices.'}, status=403)
        return super().partial_update(request, *args, **kwargs)

    def notice_recipients(self, notice):
        users = User.objects.filter(is_active=True).exclude(role=User.Role.SUPER_ADMIN)
        if notice.audience_type == TeamNotice.AudienceType.SPECIFIC_BRANCH:
            users = users.filter(branch=notice.branch)
        return users.select_related('branch')

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        notices = self.visible_queryset().filter(status=TeamNotice.Status.ACTIVE).order_by('-created_at')[:3]
        unread_count = Notification.objects.filter(
            user=request.user,
            status=Notification.Status.UNREAD,
            related_url__startswith='/team-board?notice=',
        ).count()
        return Response({
            'unread_count': unread_count,
            'notices': TeamNoticeSerializer(notices, many=True).data,
        })

    @action(detail=True, methods=['post'], url_path='reply')
    def reply(self, request, pk=None):
        if request.user.is_super_admin:
            return Response({'detail': 'Admin users cannot reply to Team Board notices.'}, status=403)
        notice = self.get_object()
        if notice.status != TeamNotice.Status.ACTIVE:
            return Response({'detail': 'Replies are allowed only on active notices.'}, status=400)
        message = str(request.data.get('reply_message') or '').strip()
        if not message:
            return Response({'reply_message': 'Reply message is required.'}, status=400)
        reply = TeamNoticeReply.objects.create(
            notice=notice,
            reply_message=message,
            replied_by=request.user,
            branch=request.user.branch,
        )
        for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
            create_user_notification(
                admin_user,
                'Team Board Reply',
                f'{request.user.full_name} replied to Team Board notice.',
                Notification.NType.INFO,
                f'/team-board?notice={notice.id}',
            )
        return Response(TeamNoticeReplySerializer(reply).data, status=201)

    @action(detail=True, methods=['post'], url_path='close')
    def close(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can close Team Board notices.'}, status=403)
        notice = self.get_object()
        notice.status = TeamNotice.Status.CLOSED
        notice.closed_at = timezone.now()
        notice.save(update_fields=['status', 'closed_at', 'updated_at'])
        return Response(TeamNoticeSerializer(notice).data)

    @action(detail=True, methods=['post'], url_path='archive')
    def archive(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can archive Team Board notices.'}, status=403)
        notice = self.get_object()
        notice.status = TeamNotice.Status.ARCHIVED
        notice.archived_at = timezone.now()
        notice.save(update_fields=['status', 'archived_at', 'updated_at'])
        return Response(TeamNoticeSerializer(notice).data)

    def retrieve(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        Notification.objects.filter(
            user=request.user,
            related_url=f'/team-board?notice={kwargs.get("pk")}',
            status=Notification.Status.UNREAD,
        ).update(status=Notification.Status.READ, is_read=True)
        return response


class WhatsAppTemplateViewSet(viewsets.ModelViewSet):
    serializer_class = WhatsAppTemplateSerializer
    permission_classes = [IsSuperAdminOrReadOnly]
    pagination_class = None
    filterset_fields = ['template_type', 'is_active']

    def get_queryset(self):
        qs = WhatsAppTemplate.objects.order_by('template_type', 'name')
        if not self.request.user.is_super_admin:
            qs = qs.filter(is_active=True)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# ============================================================
# backend/apps/walkins/views.py
# ============================================================
from crm.models import WalkIn, WalkInBranchChangeHistory
from serializers import WalkInListSerializer, WalkInDetailSerializer
import django_filters


class WalkInFilter(django_filters.FilterSet):
    name = django_filters.CharFilter(field_name='name', lookup_expr='icontains')
    phone = django_filters.CharFilter(field_name='phone', lookup_expr='icontains')
    source = django_filters.CharFilter(method='filter_source')
    branch = django_filters.ModelChoiceFilter(queryset=Branch.objects.all(), method='filter_branch')
    created_by = django_filters.ModelChoiceFilter(queryset=User.objects.all(), method='filter_created_by')
    assigned_to = django_filters.ModelChoiceFilter(queryset=User.objects.all(), method='filter_assigned_to')
    counselor = django_filters.ModelChoiceFilter(queryset=User.objects.all(), method='filter_counselor')
    follow_up_by = django_filters.ModelChoiceFilter(queryset=User.objects.all(), method='filter_assigned_to')
    status = django_filters.CharFilter(method='filter_status')
    visit_date_from = django_filters.DateFilter(field_name='visit_date', lookup_expr='gte')
    visit_date_to   = django_filters.DateFilter(field_name='visit_date', lookup_expr='lte')
    date_from = django_filters.DateFilter(method='filter_activity_date_from')
    date_to = django_filters.DateFilter(method='filter_activity_date_to')
    date_range = django_filters.CharFilter(method='filter_date_range')
    follow_up_date_from = django_filters.DateFilter(field_name='follow_up_date', lookup_expr='gte')
    follow_up_date_to   = django_filters.DateFilter(field_name='follow_up_date', lookup_expr='lte')
    important_only = django_filters.BooleanFilter(method='filter_important_only')

    def filter_branch(self, queryset, name, value):
        if not self.request or not self.request.user.is_super_admin:
            return queryset
        return queryset.filter(branch=value)

    def filter_created_by(self, queryset, name, value):
        if not self.request or not self.request.user.is_super_admin:
            return queryset
        return queryset.filter(created_by=value)

    def filter_assigned_to(self, queryset, name, value):
        if not value:
            return queryset
        if not self.request or not self.request.user.is_super_admin:
            if value.branch_id != self.request.user.branch_id:
                return queryset.none()
        return queryset.filter(assigned_to=value)

    def filter_counselor(self, queryset, name, value):
        if not value:
            return queryset
        if not self.request or not self.request.user.is_super_admin:
            if value.branch_id != self.request.user.branch_id:
                return queryset.none()
        return queryset.filter(Q(counseling_by=value) | Q(assigned_to=value))

    def filter_source(self, queryset, name, value):
        value = (value or '').strip()
        if not value:
            return queryset
        normalized = normalize_source_filter(value, WalkIn.Source.choices)
        return queryset.filter(source=normalized) if normalized else queryset.none()

    def filter_status(self, queryset, name, value):
        return filter_candidate_status(queryset, value, WalkIn, WALKIN_STATUS_FILTER_ALIASES)

    def filter_activity_date_from(self, queryset, name, value):
        if not value:
            return queryset
        return queryset.filter(created_at__date__gte=value)

    def filter_activity_date_to(self, queryset, name, value):
        if not value:
            return queryset
        return queryset.filter(created_at__date__lte=value)

    def filter_important_only(self, queryset, name, value):
        return queryset.filter(is_important=True) if value else queryset

    def filter_date_range(self, queryset, name, value):
        start_date, end_date = date_range_from_request(self.request)
        return queryset.filter(visit_date__gte=start_date, visit_date__lte=end_date)

    class Meta:
        model  = WalkIn
        fields = ['status', 'branch', 'created_by', 'assigned_to', 'counselor', 'course', 'source', 'demo_class', 'date_range', 'important_only']


class WalkInViewSet(viewsets.ModelViewSet):
    """Walk-in management. Staff sees own branch; admin sees all."""
    permission_classes = [IsStaffOrAdmin]
    filter_backends    = [DjangoFilterBackend, CRMSearchFilter, OrderingFilter]
    filterset_class    = WalkInFilter
    pagination_class   = None
    search_fields      = [
        'name', 'phone', 'candidate_number', 'email', 'source', 'source_description',
        'course__name', 'assigned_to__first_name', 'assigned_to__last_name',
        'assigned_to__username', 'created_by__first_name', 'created_by__last_name',
        'created_by__username',
    ]
    id_search_fields   = ['id']
    ordering_fields    = ['visit_date', 'created_at']
    ordering           = ['-visit_date']

    def get_queryset(self):
        if self.action == 'list':
            qs = visible_candidate_queryset(WalkIn.objects.select_related('course', 'branch', 'assigned_to', 'created_by', 'enrollment'))
            latest_follow_up = FollowUp.objects.filter(
                record_type=FollowUp.RecordType.WALKIN,
                record_id=OuterRef('pk'),
            ).order_by('-created_at', '-id')
            qs = qs.annotate(
                latest_follow_up_remark=Subquery(latest_follow_up.values('remarks')[:1]),
                latest_follow_up_at=Subquery(latest_follow_up.values('created_at')[:1]),
            )
        else:
            qs = visible_candidate_queryset(WalkIn.objects.select_related('course','branch','assigned_to','created_by','lead','enrollment'))
        missing_columns = missing_model_columns(WalkIn, [
            'qualification', 'degree', 'profession', 'year_of_passing',
            'college_company', 'preferred_timing', 'interested_global_certification',
            'source_description', 'walk_in_by', 'follow_up_date', 'converted_to_type',
            'converted_record_id', 'converted_at', 'converted_by', 'counseling_by',
        ])
        if missing_columns:
            qs = qs.defer(*missing_columns)
        if not self.request.user.is_super_admin:
            qs = qs.filter(branch=self.request.user.branch)
        if self.request.query_params.get('focus') == 'today-follow-up':
            qs = active_walkin_follow_up_queryset(
                qs,
                'follow_up_date',
                timezone.localdate(),
            )
        return qs

    @action(detail=True, methods=['post'], url_path='toggle-important')
    def toggle_important(self, request, pk=None):
        walkin = self.get_object()
        requested = request.data.get('is_important', None)
        walkin.is_important = truthy_query_param(requested) if requested is not None else not walkin.is_important
        walkin.save(update_fields=['is_important', 'updated_at'])
        return Response({'id': walkin.id, 'is_important': walkin.is_important})

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'walkins', 'walkins-export')

    @action(detail=True, methods=['post'], url_path='send-follow-up-whatsapp')
    def send_follow_up_whatsapp(self, request, pk=None):
        _, response = send_candidate_template(
            self.get_object(),
            WhatsAppTemplate.TemplateType.WALKIN_FOLLOW_UP,
            WhatsAppMessage.MsgType.WALKIN_REMIND,
            request,
            'walkin',
        )
        return response

    @action(detail=True, methods=['post'], url_path='send-offer-whatsapp')
    def send_offer_whatsapp(self, request, pk=None):
        _, response = send_candidate_template(
            self.get_object(),
            WhatsAppTemplate.TemplateType.OFFER_MESSAGE,
            WhatsAppMessage.MsgType.OFFER_MESSAGE,
            request,
            'walkin',
        )
        return response

    def get_serializer_class(self):
        return WalkInListSerializer if self.action == 'list' else WalkInDetailSerializer

    def list(self, request, *args, **kwargs):
        if request.query_params.get('sectioned') not in ('1', 'true', 'yes'):
            return super().list(request, *args, **kwargs)
        return self._sectioned_response()

    def retrieve(self, request, *args, **kwargs):
        walkin = self.get_object()
        if clear_stale_walkin_conversion(walkin):
            walkin.refresh_from_db()
        mark_public_walkin_notifications_read(walkin.id)
        serializer = self.get_serializer(walkin)
        return Response(serializer.data)

    def _sectioned_response(self):
        queryset = self.filter_queryset(self.get_queryset())
        now = timezone.localtime(timezone.now())
        current_month_filter = {
            'created_at__year': now.year,
            'created_at__month': now.month,
        }
        current_month_walkins = queryset.filter(**current_month_filter)
        other_walkins = queryset.exclude(**current_month_filter)

        current_serializer = self.get_serializer(current_month_walkins, many=True)
        other_serializer = self.get_serializer(other_walkins, many=True)
        return Response({
            'current_month_walkins': current_serializer.data,
            'other_walkins': other_serializer.data,
            'current_month_count': current_month_walkins.count(),
            'other_walkins_count': other_walkins.count(),
        })

    @action(detail=False, methods=['get'], url_path='sectioned')
    def sectioned(self, request):
        return self._sectioned_response()

    def perform_create(self, serializer):
        branch = self.request.user.branch if not self.request.user.is_super_admin else None
        assigned_to = serializer.validated_data.get('assigned_to')
        fixed_walk_in_by_selected = serializer.validated_data.get('walk_in_by') in (
            WalkIn.WalkInBy.DIRECT,
            WalkIn.WalkInBy.FRIENDS_REFERENCE,
        )
        if not assigned_to and not fixed_walk_in_by_selected and not self.request.user.is_super_admin:
            assigned_to = self.request.user
        serializer.save(
            created_by=self.request.user,
            branch=branch or serializer.validated_data.get('branch'),
            assigned_to=assigned_to,
            status=WalkIn.Status.NEW,
            converted_to_type='',
            converted_record_id=None,
            converted_at=None,
            converted_by=None,
        )

    @action(detail=False, methods=['get'], url_path='staff-options')
    def staff_options(self, request):
        users = User.objects.filter(
            is_active=True,
            role=User.Role.STAFF,
            is_superuser=False,
        ).select_related('branch').order_by('first_name', 'last_name', 'username')
        if request.user.is_super_admin:
            branch_id = request.query_params.get('branch')
            if branch_id:
                users = users.filter(branch_id=branch_id)
        else:
            users = users.filter(branch=request.user.branch)
        seen = set()
        rows = []
        for user in users:
            if user.id in seen:
                continue
            seen.add(user.id)
            rows.append({
                'id': user.id,
                'name': user.full_name or user.username,
                'full_name': user.full_name,
                'username': user.username,
                'email': user.email,
                'branch_id': user.branch_id,
                'branch_name': user.branch.name if user.branch else '',
                'identity_color': user.identity_color or '',
            })
        return Response(rows)

    @action(detail=False, methods=['get'], url_path='walk-in-by-options')
    def walk_in_by_options(self, request):
        users = User.objects.filter(
            is_active=True,
            role=User.Role.STAFF,
            is_superuser=False,
        ).select_related('branch').order_by(
            'first_name', 'last_name', 'username'
        )
        rows = [{
            'id': WalkIn.WalkInBy.DIRECT,
            'name': 'Direct',
            'is_fixed': True,
        }, {
            'id': WalkIn.WalkInBy.FRIENDS_REFERENCE,
            'name': 'Friends Reference',
            'is_fixed': True,
        }]
        seen = set()
        for user in users:
            if user.id in seen:
                continue
            seen.add(user.id)
            rows.append({
                'id': user.id,
                'name': user.full_name or user.username,
                'full_name': user.full_name,
                'username': user.username,
                'email': user.email,
                'branch_id': user.branch_id,
                'branch_name': user.branch.name if user.branch else '',
                'identity_color': user.identity_color or '',
            })
        return Response(rows)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        walkin = self.get_object()
        old_status = tracked_crm_status(walkin)
        old_follow_up_date = walkin.follow_up_date
        data = request.data.copy()
        branch_changed = False
        old_branch = walkin.branch
        new_branch = None

        if 'branch' in request.data:
            if not request.user.is_super_admin:
                return Response({'detail': 'Only admin can change a walk-in branch.'}, status=status.HTTP_403_FORBIDDEN)
            new_branch = Branch.objects.filter(pk=request.data.get('branch'), is_active=True).first()
            if not new_branch:
                return Response({'detail': 'Select a valid active branch.'}, status=status.HTTP_400_BAD_REQUEST)
            branch_changed = walkin.branch_id != new_branch.id
            if branch_changed and walkin.counseling_by_id and walkin.counseling_by.branch_id != new_branch.id:
                data['counseling_by'] = None
        data.pop('branch_change_reason', None)

        requested_fixed_walk_in_by = data.get('walk_in_by') in (
            WalkIn.WalkInBy.DIRECT,
            WalkIn.WalkInBy.FRIENDS_REFERENCE,
        )
        assigned_to = data.get('assigned_to')
        if assigned_to in ('', None) and 'assigned_to' in data:
            assigned_to = None
        if requested_fixed_walk_in_by:
            if walkin.assigned_to_id or (
                walkin.walk_in_by in (
                    WalkIn.WalkInBy.DIRECT,
                    WalkIn.WalkInBy.FRIENDS_REFERENCE,
                )
                and walkin.walk_in_by != data.get('walk_in_by')
            ):
                return Response({'detail': 'Walk-in By is locked. Submit a change request.'}, status=status.HTTP_403_FORBIDDEN)
            data['assigned_to'] = None
            assigned_to = None
        if assigned_to not in (None, ''):
            user_qs = User.objects.filter(
                pk=assigned_to,
                is_active=True,
                role=User.Role.STAFF,
                is_superuser=False,
            )
            if not user_qs.exists():
                return Response({'detail': 'Select a valid active staff user for Walk-in By.'}, status=status.HTTP_400_BAD_REQUEST)
            if (
                (walkin.assigned_to_id or walkin.walk_in_by in (
                    WalkIn.WalkInBy.DIRECT,
                    WalkIn.WalkInBy.FRIENDS_REFERENCE,
                ))
                and int(assigned_to) != walkin.assigned_to_id
            ):
                return Response({'detail': 'Walk-in By is locked. Submit a change request.'}, status=status.HTTP_403_FORBIDDEN)
            data['walk_in_by'] = ''
        elif 'assigned_to' in data and walkin.assigned_to_id and not requested_fixed_walk_in_by:
            return Response({'detail': 'Walk-in By is locked. Submit a change request.'}, status=status.HTTP_403_FORBIDDEN)

        counseling_by = data.get('counseling_by')
        if counseling_by in ('', None) and 'counseling_by' in data:
            counseling_by = None
        if counseling_by not in (None, ''):
            counseling_branch = new_branch or walkin.branch
            user_qs = User.objects.filter(
                pk=counseling_by,
                is_active=True,
                role=User.Role.STAFF,
                is_superuser=False,
                branch=counseling_branch,
            )
            if not user_qs.exists():
                return Response({'detail': 'Select an active user from the candidate branch for Counseling By.'}, status=status.HTTP_400_BAD_REQUEST)
            if (
                walkin.counseling_by_id
                and int(counseling_by) != walkin.counseling_by_id
            ):
                return Response({'detail': 'Counseling By is locked. Submit a change request.'}, status=status.HTTP_403_FORBIDDEN)
        elif 'counseling_by' in data and walkin.counseling_by_id:
            return Response({'detail': 'Counseling By is locked. Submit a change request.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = self.get_serializer(walkin, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            self.perform_update(serializer)
            walkin.refresh_from_db()
            create_status_history(
                walkin,
                CandidateStatusHistory.RecordType.WALKIN,
                old_status,
                tracked_crm_status(walkin),
                request.user,
                walkin.remarks,
            )
            if branch_changed:
                WalkInBranchChangeHistory.objects.create(
                    walkin=walkin,
                    old_branch=old_branch,
                    new_branch=new_branch,
                    changed_by=request.user,
                    reason='Updated from walk-in detail edit.',
                )
            mark_public_walkin_notifications_read(walkin.id)
            if (
                old_follow_up_date != walkin.follow_up_date
                or walkin.status in WALKIN_CLOSED_FOLLOW_UP_STATUSES
            ):
                clear_follow_up_notifications_for_record(FollowUp.RecordType.WALKIN, walkin.id)
        walkin.refresh_from_db()
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='request-assignment-change')
    def request_assignment_change(self, request, pk=None):
        walkin = self.get_object()
        field_type = request.data.get('field_type')
        requested_user_id = request.data.get('requested_user')
        requested_walk_in_by = request.data.get('requested_walk_in_by', '')
        reason = request.data.get('reason', '')
        requested_fixed_walk_in_by = (
            field_type == WalkInAssignmentChangeRequest.FieldType.WALK_IN_BY
            and requested_walk_in_by in (
                WalkIn.WalkInBy.DIRECT,
                WalkIn.WalkInBy.FRIENDS_REFERENCE,
            )
        )
        if not requested_user_id and not requested_fixed_walk_in_by:
            return Response({'detail': 'Select the requested user.'}, status=status.HTTP_400_BAD_REQUEST)
        requested_user = User.objects.filter(pk=requested_user_id, is_active=True).first() if requested_user_id else None
        if not requested_user and not requested_fixed_walk_in_by:
            return Response({'detail': 'Select a valid active user.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            change_request = create_walkin_assignment_change_request(
                walkin,
                field_type,
                requested_user,
                request.user,
                reason,
                requested_walk_in_by=requested_walk_in_by,
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(
            WalkInAssignmentChangeRequestSerializer(change_request, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=['post'], url_path='change-branch')
    def change_branch(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can change a walk-in branch.'}, status=status.HTTP_403_FORBIDDEN)

        walkin = self.get_object()
        branch_id = request.data.get('branch')
        reason = request.data.get('reason', '')
        if not branch_id:
            return Response({'detail': 'Branch is required.'}, status=status.HTTP_400_BAD_REQUEST)

        new_branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
        if not new_branch:
            return Response({'detail': 'Select a valid active branch.'}, status=status.HTTP_400_BAD_REQUEST)

        old_branch = walkin.branch
        if old_branch_id := getattr(old_branch, 'id', None):
            if old_branch_id == new_branch.id:
                return Response(WalkInDetailSerializer(walkin, context={'request': request}).data)

        WalkInBranchChangeHistory.objects.create(
            walkin=walkin,
            old_branch=old_branch,
            new_branch=new_branch,
            changed_by=request.user,
            reason=reason,
        )
        walkin.branch = new_branch
        update_fields = ['branch', 'updated_at']
        if walkin.counseling_by_id and walkin.counseling_by.branch_id != new_branch.id:
            walkin.counseling_by = None
            update_fields.append('counseling_by')
        walkin.save(update_fields=update_fields)
        mark_public_walkin_notifications_read(walkin.id)
        return Response(WalkInDetailSerializer(walkin, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='follow-ups')
    def add_follow_up(self, request, pk=None):
        walkin = self.get_object()
        response = create_follow_up_entry(walkin, FollowUp.RecordType.WALKIN, request)
        if response.status_code >= 400:
            return response

        walkin.follow_up_date = response.data['next_follow_up_date']
        close_follow_up = request.data.get('close_follow_up') in (True, 'true', '1', 1)
        if close_follow_up:
            walkin.follow_up_date = None
        walkin.save(update_fields=['follow_up_date', 'updated_at'])
        clear_follow_up_notifications_for_record(FollowUp.RecordType.WALKIN, walkin.id)
        return response

    @action(detail=True, methods=['post'], url_path='convert-to-enrollment')
    def convert_to_enrollment(self, request, pk=None):
        """Convert walk-in to confirmed enrollment."""
        walkin = self.get_object()
        from serializers import EnrollmentDetailSerializer
        data = request.data.copy()
        existing_enrollment = valid_walkin_enrollment(walkin)
        if existing_enrollment:
            if (
                walkin.status != WalkIn.Status.CONVERTED
                or walkin.converted_to_type != 'enrollment'
                or walkin.converted_record_id != existing_enrollment.id
                or walkin.follow_up_date is not None
            ):
                with transaction.atomic():
                    mark_walkin_enrollment_converted(walkin, existing_enrollment, request.user)
            clear_follow_up_notifications_for_record(FollowUp.RecordType.WALKIN, walkin.id)
            return Response({'detail': 'This record has already been converted.'}, status=status.HTTP_400_BAD_REQUEST)
        clear_stale_walkin_conversion(walkin)
        walkin.refresh_from_db()

        data.setdefault('name',        walkin.name)
        data.setdefault('phone',       walkin.phone)
        data.setdefault('email',       walkin.email)
        data.setdefault('dob',         walkin.dob)
        data.setdefault('location',    walkin.location)
        data.setdefault('pincode',     walkin.pincode)
        data.setdefault('source',      walkin.source)
        data.setdefault('preferred_timing', walkin.preferred_timing)
        data.setdefault('qualification', walkin.qualification)
        data.setdefault('degree', walkin.degree)
        for field_name in QUALIFICATION_KPI_FIELDS:
            data.setdefault(field_name, getattr(walkin, field_name, ''))
        data.setdefault('demo_class',  walkin.demo_class)
        data.setdefault('interested_global_certification', walkin.interested_global_certification)
        data.setdefault('branch',      walkin.branch_id)
        data.setdefault('course',      walkin.course_id)
        required_fields = [
            'name', 'phone', 'course', 'branch', 'preferred_timing',
            'enrollment_date', 'start_date',
        ]
        missing = [field for field in required_fields if data.get(field) in (None, '')]
        if missing:
            return Response(
                {'detail': 'Please complete all mandatory fields.', 'missing_fields': missing},
                status=status.HTTP_400_BAD_REQUEST,
            )
        course = Course.objects.filter(pk=data.get('course')).first()
        if not course:
            return Response({'detail': 'Please select a valid course.'}, status=status.HTTP_400_BAD_REQUEST)
        data['actual_fees'] = data.get('actual_fees') or course.actual_fees
        try:
            apply_enrollment_discount(data, course, data.get('branch'))
            apply_spot_conversion_discount(data, walkin.visit_date)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = EnrollmentDetailSerializer(data=data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            enrollment = serializer.save(
                walkin=walkin,
                original_walkin_course=walkin.course,
                final_enrollment_course=course,
                enrolled_by=request.user,
                created_by=request.user,
                counselor=walkin.assigned_to or walkin.counseling_by or walkin.created_by or request.user,
                status=Enrollment.Status.PENDING_RULES,
            )
            enrollment.refresh_from_db()
            mark_walkin_enrollment_converted(walkin, enrollment, request.user, data)
        clear_follow_up_notifications_for_record(FollowUp.RecordType.WALKIN, walkin.id)
        resolve_public_walkin_notifications(walkin.id)
        notify_admin_enrollment_created(enrollment, request.user)
        return Response(EnrollmentDetailSerializer(enrollment).data, status=201)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can delete walk-ins.'}, status=status.HTTP_403_FORBIDDEN)
        walkin = self.get_object()
        with transaction.atomic():
            if walkin.lead_id:
                Lead.objects.filter(
                    pk=walkin.lead_id,
                    converted_to_type='walkin',
                    converted_record_id=walkin.id,
                ).update(
                    status=Lead.Status.NEW,
                    converted_to_type='',
                    converted_record_id=None,
                    converted_at=None,
                    converted_by=None,
                )
            FollowUp.objects.filter(record_type=FollowUp.RecordType.WALKIN, record_id=walkin.id).delete()
            self.perform_destroy(walkin)
        return Response(status=status.HTTP_204_NO_CONTENT)


from serializers import BranchTransferRequestSerializer


class WalkInAssignmentChangeRequestFilter(django_filters.FilterSet):
    date = django_filters.DateFilter(field_name='created_at', lookup_expr='date')

    class Meta:
        model = WalkInAssignmentChangeRequest
        fields = ['status', 'field_type', 'branch', 'requested_by']


class WalkInAssignmentChangeRequestViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WalkInAssignmentChangeRequestSerializer
    permission_classes = [IsStaffOrAdmin]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = WalkInAssignmentChangeRequestFilter
    search_fields = [
        'candidate_name', 'candidate_phone',
        'previous_user__first_name', 'previous_user__last_name', 'previous_user__username',
        'requested_user__first_name', 'requested_user__last_name', 'requested_user__username',
    ]
    ordering_fields = ['created_at', 'reviewed_at', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = WalkInAssignmentChangeRequest.objects.select_related(
            'walkin', 'branch', 'previous_user', 'requested_user', 'requested_by',
            'counselor_reviewed_by', 'reviewed_by',
        )
        if self.request.user.is_super_admin:
            return queryset
        return queryset.filter(
            Q(requested_by=self.request.user)
            | Q(requested_user=self.request.user)
            | Q(walkin__branch=self.request.user.branch)
        )

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        change_request = self.get_object()
        remarks = request.data.get('admin_remarks') or request.data.get('remarks') or ''
        try:
            with transaction.atomic():
                change_request = WalkInAssignmentChangeRequest.objects.select_for_update().get(pk=change_request.pk)
                if change_request.status == WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR:
                    if change_request.requested_user_id != request.user.id:
                        return Response({'detail': 'Only the requested counselor can approve this request.'}, status=status.HTTP_403_FORBIDDEN)
                    approve_walkin_assignment_by_counselor(change_request, request.user, remarks)
                elif change_request.status == WalkInAssignmentChangeRequest.Status.PENDING_ADMIN:
                    if not request.user.is_super_admin:
                        return Response({'detail': 'Only admin can give final approval.'}, status=status.HTTP_403_FORBIDDEN)
                    apply_walkin_assignment_change(change_request, request.user, remarks)
                else:
                    return Response({'detail': 'Only pending requests can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        change_request.refresh_from_db()
        return Response(self.get_serializer(change_request).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        change_request = self.get_object()
        if change_request.status not in (
            WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR,
            WalkInAssignmentChangeRequest.Status.PENDING_ADMIN,
        ):
            return Response({'detail': 'Only pending requests can be rejected.'}, status=status.HTTP_400_BAD_REQUEST)
        if change_request.status == WalkInAssignmentChangeRequest.Status.PENDING_COUNSELOR:
            if change_request.requested_user_id != request.user.id:
                return Response({'detail': 'Only the requested counselor can reject this request.'}, status=status.HTTP_403_FORBIDDEN)
            change_request.counselor_reviewed_by = request.user
            change_request.counselor_reviewed_at = timezone.now()
            change_request.counselor_remarks = request.data.get('counselor_remarks') or request.data.get('remarks') or ''
            update_fields = [
                'status', 'counselor_reviewed_by', 'counselor_reviewed_at',
                'counselor_remarks', 'updated_at',
            ]
        else:
            if not request.user.is_super_admin:
                return Response({'detail': 'Only admin can reject requests awaiting final approval.'}, status=status.HTTP_403_FORBIDDEN)
            change_request.reviewed_by = request.user
            change_request.reviewed_at = timezone.now()
            change_request.admin_remarks = request.data.get('admin_remarks') or request.data.get('remarks') or ''
            update_fields = ['status', 'reviewed_by', 'reviewed_at', 'admin_remarks', 'updated_at']
        change_request.status = WalkInAssignmentChangeRequest.Status.REJECTED
        change_request.save(update_fields=update_fields)
        create_user_notification(
            change_request.requested_by,
            'Walk-in Assignment Change Rejected',
            f'{change_request.get_field_type_display()} change for {change_request.candidate_name} was rejected.',
            Notification.NType.ERROR,
            f'/walkins/{change_request.walkin_id}',
        )
        return Response(self.get_serializer(change_request).data)


class BranchTransferRequestViewSet(viewsets.ModelViewSet):
    """Branch-change approval workflow for walk-in to enrollment conversion."""
    serializer_class = BranchTransferRequestSerializer
    permission_classes = [IsStaffOrAdmin]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['status', 'walkin', 'current_branch', 'requested_branch', 'course', 'requested_by']
    search_fields = ['candidate_name', 'phone']
    ordering_fields = ['created_at', 'reviewed_at', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = BranchTransferRequest.objects.select_related(
            'walkin', 'current_branch', 'requested_branch', 'course',
            'requested_by', 'reviewed_by', 'enrollment',
        )
        if not self.request.user.is_super_admin:
            queryset = queryset.filter(requested_by=self.request.user)
        return queryset

    def create(self, request, *args, **kwargs):
        return Response(
            {'detail': 'Transfer requests are created from walk-in enrollment conversion.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def update(self, request, *args, **kwargs):
        return Response({'detail': 'Use approve or reject actions.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def partial_update(self, request, *args, **kwargs):
        return Response({'detail': 'Use approve or reject actions.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can delete transfer requests.'}, status=status.HTTP_403_FORBIDDEN)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can approve transfer requests.'}, status=status.HTTP_403_FORBIDDEN)

        transfer_request = self.get_object()
        if transfer_request.status != BranchTransferRequest.Status.PENDING:
            return Response({'detail': 'Only pending transfer requests can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
        if BranchTransferRequest.objects.filter(
            walkin=transfer_request.walkin,
            status=BranchTransferRequest.Status.APPROVED,
        ).exclude(pk=transfer_request.pk).exists():
            return Response({'detail': 'This walk-in already has an approved transfer.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                create_enrollment_from_transfer_request(transfer_request, request.user)
                transfer_request.review_remarks = request.data.get('review_remarks') or request.data.get('remarks') or ''
                transfer_request.save(update_fields=['review_remarks', 'updated_at'])
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        transfer_request.refresh_from_db()
        return Response(self.get_serializer(transfer_request).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can reject transfer requests.'}, status=status.HTTP_403_FORBIDDEN)

        transfer_request = self.get_object()
        if transfer_request.status != BranchTransferRequest.Status.PENDING:
            return Response({'detail': 'Only pending transfer requests can be rejected.'}, status=status.HTTP_400_BAD_REQUEST)

        transfer_request.status = BranchTransferRequest.Status.REJECTED
        transfer_request.reviewed_by = request.user
        transfer_request.reviewed_at = timezone.now()
        transfer_request.review_remarks = request.data.get('review_remarks') or request.data.get('remarks') or ''
        transfer_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_remarks', 'updated_at'])
        return Response(self.get_serializer(transfer_request).data)


from serializers import PublicWalkInCreateSerializer


class PublicWalkInFormView(APIView):
    """POST /api/public/walkin/ — public walk-in capture."""
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request):
        allowed_branch_names = ['Gandhipuram', 'Hopes', 'Kuniyamuthur']
        branch_rows = [
            {'id': branch.id, 'name': branch.name}
            for branch in Branch.objects.filter(
                is_active=True,
                name__in=allowed_branch_names,
            ).order_by('name')
        ]
        course_rows = [
            {'id': course.id, 'name': course.name}
            for course in Course.objects.filter(is_active=True).order_by('name')
        ]
        return Response({
            'branches': branch_rows,
            'courses': course_rows,
            'preferred_timing_options': [
                {'value': value, 'label': label}
                for value, label in WalkIn.PreferredTiming.choices
            ],
            'qualification_options': [
                {'value': value, 'label': label}
                for value, label in WalkIn.Qualification.choices
            ],
            'source_options': [
                {'value': value, 'label': label}
                for value, label in WalkIn.Source.choices
            ],
        })

    def post(self, request):
        data = request.data.copy()
        phone = normalize_phone_number(data.get('phone'))
        if len(phone) == 12 and phone.startswith('91'):
            phone = phone[2:]
        data['phone'] = phone

        errors = {}
        if not phone:
            errors['phone'] = 'Phone number is required.'
        elif len(phone) != 10:
            errors['phone'] = 'Phone number should be 10 digits.'

        branch_id = data.get('branch')
        course_id = data.get('course')
        branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
        if not branch:
            errors['branch'] = 'Please select a valid active branch.'
        course = Course.objects.filter(pk=course_id, is_active=True).first()
        if not course:
            errors['course'] = 'Please select a valid active course.'
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = PublicWalkInCreateSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                duplicate_records = matching_candidate_phone_records(phone)
                existing_walkin = None
                active_walkins = visible_candidate_queryset(WalkIn.objects.all()).exclude(
                    status=WalkIn.Status.CONVERTED,
                ).filter(
                    Q(converted_to_type__isnull=True) | Q(converted_to_type='')
                ).order_by('-updated_at', '-created_at')
                for walkin in active_walkins:
                    if normalize_phone_number(walkin.phone) == phone:
                        existing_walkin = walkin
                        break

                save_kwargs = {
                    'status': WalkIn.Status.NEW,
                    'branch': branch,
                    'course': course,
                    'converted_to_type': '',
                    'converted_record_id': None,
                    'converted_at': None,
                    'converted_by': None,
                }
                if existing_walkin:
                    walkin = serializer.update(existing_walkin, {**serializer.validated_data, **save_kwargs})
                else:
                    walkin = serializer.save(**save_kwargs)
            logger.info(
                'Public walk-in saved successfully: walkin_id=%s candidate_number=%s branch_id=%s updated=%s',
                walkin.id,
                walkin.candidate_number,
                walkin.branch_id,
                bool(existing_walkin),
            )
        except Exception as exc:
            logger.exception(
                'Public walk-in save failed: phone=%s branch_id=%s course_id=%s error=%s',
                phone,
                getattr(branch, 'id', None),
                getattr(course, 'id', None),
                exc.__class__.__name__,
            )
            return Response(
                {
                    'detail': 'We could not submit your enquiry right now. Please try again or contact the selected branch.',
                    'error': exc.__class__.__name__,
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            notify_branch_users(
                branch,
                'New public walk-in submitted',
                f'{walkin.name} submitted the public walk-in form for {course.name}.',
                Notification.NType.INFO,
                f'/walkins/{walkin.id}',
            )
            for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
                create_user_notification(
                    admin_user,
                    'New public walk-in submitted',
                    f'{walkin.name} submitted the public walk-in form for {branch.name}.',
                    Notification.NType.INFO,
                    f'/walkins/{walkin.id}',
                )
        except Exception:
            logger.exception(
                'Public walk-in notification creation failed after save: walkin_id=%s branch_id=%s',
                walkin.id,
                walkin.branch_id,
            )

        return Response(
            {
                'detail': 'Thanks for filling out the form.',
                'candidate_number': walkin.candidate_number,
                'id': walkin.id,
                'duplicate': bool(duplicate_records),
                'updated': bool(existing_walkin),
            },
            status=status.HTTP_200_OK if existing_walkin else status.HTTP_201_CREATED
        )


class PublicRulesSigningView(APIView):
    """Public token endpoint for rules signing."""
    permission_classes = [AllowAny]
    authentication_classes = []

    def get_signing(self, token):
        return RulesSigningRequest.objects.defer(
            'selfie_image_file',
            'signature_image_file',
            'signed_pdf_file',
            'submitted_ip',
            'submitted_user_agent',
        ).select_related(
            'enrollment__course',
        ).filter(token=token).first()

    def public_pdf_url(self, request, signing):
        if signing.signed_pdf:
            return request.build_absolute_uri(f'{app_url(f"public/rules-signed-pdf/{signing.token}")}/')
        try:
            has_pdf = RulesSigningRequest.objects.filter(
                pk=signing.pk,
                signed_pdf_file__isnull=False,
            ).exists()
        except (OperationalError, ProgrammingError):
            has_pdf = False
        if not has_pdf:
            return None
        return request.build_absolute_uri(f'{app_url(f"public/rules-signed-pdf/{signing.token}")}/')

    def get(self, request, token):
        signing = self.get_signing(token)
        if not signing:
            return Response({'detail': 'Invalid signing link.'}, status=404)
        if signing.status == RulesSigningRequest.Status.SENT:
            signing.status = RulesSigningRequest.Status.VIEWED
            signing.save(update_fields=['status', 'updated_at'])
        enrollment = signing.enrollment
        installments = build_rules_installment_plan(enrollment)
        rules_paragraphs = extract_rules_template_text()
        return Response({
            'status': signing.status,
            'candidate': {
                'name': enrollment.name,
                'phone': enrollment.phone,
                'course': enrollment.course.name if enrollment.course else '',
                'course_enrolled': enrollment.course.name if enrollment.course else '',
                'batch_timing': enrollment.batch_timing or enrollment.get_preferred_timing_display() or '',
                'batch_start_date': enrollment.start_date,
                'duration': enrollment.course.duration_months if enrollment.course else None,
                'final_payable_fees': enrollment_payable_fee(enrollment),
                'total_course_fee': enrollment_payable_fee(enrollment),
                'enrollment_date': enrollment.enrollment_date,
            },
            'installments': installments,
            'rules_paragraphs': rules_paragraphs,
            'rules_content': rules_paragraphs,
            'signed_pdf_url': self.public_pdf_url(request, signing),
            'selfie_url': self.file_url(request, signing, 'selfie_image'),
        })

    def file_url(self, request, signing, field_name):
        try:
            field = getattr(signing, field_name)
            return request.build_absolute_uri(field.url) if field else None
        except (OperationalError, ProgrammingError, ValueError):
            return None

    def post(self, request, token):
        signing = self.get_signing(token)
        if not signing:
            return Response({'detail': 'Invalid signing link.'}, status=404)
        if signing.status == RulesSigningRequest.Status.SUBMITTED:
            return Response({'detail': 'Rules & Regulation form has already been submitted.'}, status=400)

        try:
            selfie_bytes = validate_data_image(request.data.get('selfie'), 'Identity photo')
            signature_bytes = validate_data_image(request.data.get('signature'), 'Signature')
        except RuntimeError as exc:
            logger.exception('Rules signing image validation failed for token=%s.', token)
            return Response({'detail': FORM_PROCESSING_ERROR_MESSAGE}, status=503)
        except ValueError as exc:
            detail = str(exc)
            if detail == 'Identity photo is required.':
                detail = 'Identity photo is required before signing the form.'
            return Response({'detail': detail}, status=400)

        enrollment = signing.enrollment
        submitted_at = timezone.now()
        try:
            try:
                pdf_bytes = build_signed_rules_pdf(enrollment, signature_bytes, selfie_bytes, submitted_at)
            except RuntimeError as exc:
                logger.exception(
                    'Rules signing PDF generation failed for token=%s enrollment_id=%s.',
                    token,
                    enrollment.id,
                )
                return Response({'detail': FORM_PROCESSING_ERROR_MESSAGE}, status=503)
            selfie_extension = image_storage_extension(selfie_bytes)
            signature_extension = image_storage_extension(signature_bytes)
            save_rules_proof_file(
                signing.selfie_image,
                proof_storage_name(enrollment, 'selfie', selfie_extension),
                selfie_bytes,
            )
            save_rules_proof_file(
                signing.signature_image,
                proof_storage_name(enrollment, 'signature', signature_extension),
                signature_bytes,
            )
            save_rules_proof_file(
                signing.signed_pdf,
                proof_storage_name(enrollment, 'signed', 'pdf'),
                pdf_bytes,
            )
            signing.selfie_image_file = None
            signing.signature_image_file = None
            signing.signed_pdf_file = None
            signing.status = RulesSigningRequest.Status.SUBMITTED
            signing.submitted_at = submitted_at
            signing.submitted_ip = get_client_ip(request)
            signing.submitted_user_agent = request.META.get('HTTP_USER_AGENT', '')[:2000]
            signing.save(update_fields=[
                'selfie_image',
                'signature_image',
                'signed_pdf',
                'selfie_image_file',
                'signature_image_file',
                'signed_pdf_file',
                'status',
                'submitted_at',
                'submitted_ip',
                'submitted_user_agent',
                'updated_at',
            ])
        except (OperationalError, ProgrammingError):
            logger.exception(
                'Rules signing storage failed for token=%s enrollment_id=%s.',
                token,
                enrollment.id,
            )
            return Response({'detail': FORM_PROCESSING_ERROR_MESSAGE}, status=503)
        except Exception:
            logger.exception(
                'Rules signing submission failed for token=%s enrollment_id=%s.',
                token,
                enrollment.id,
            )
            return Response({'detail': FORM_PROCESSING_ERROR_MESSAGE}, status=503)
        if enrollment.status != Enrollment.Status.ENROLLED:
            old_status = enrollment.status
            enrollment.status = Enrollment.Status.RULES_SUBMITTED
            enrollment.save(update_fields=['status', 'updated_at'])
            create_status_history(
                enrollment,
                CandidateStatusHistory.RecordType.ENROLLMENT,
                old_status,
                enrollment.status,
                remarks='Rules form submitted.',
            )
        notify_rules_signed(enrollment, signing.submitted_at)
        return Response({
            'detail': 'Rules & Regulation form submitted successfully.',
            'status': signing.status,
            'signed_pdf_url': self.public_pdf_url(request, signing),
            'selfie_url': None,
            'submitted_at': signing.submitted_at,
        })


def proof_filename(enrollment):
    student_id = enrollment.student_number or enrollment.id
    return f'IIE-Rules-Regulations-{student_id}.pdf'


def proof_unavailable_response():
    return Response(
        {'detail': 'Signed PDF is not available. Please resend and collect the signed form again.'},
        status=404,
    )


def binary_or_legacy_file(signing, binary_field, legacy_field):
    try:
        binary_value = getattr(signing, binary_field, None)
    except (OperationalError, ProgrammingError):
        binary_value = None
    if binary_value:
        return bytes(binary_value)
    try:
        field = getattr(signing, legacy_field, None)
        if field:
            with field.open('rb') as handle:
                return handle.read()
    except Exception:
        return None
    return None


def proof_storage_name(enrollment, suffix, extension):
    student_id = enrollment.student_number or enrollment.id
    safe_student_id = re.sub(r'[^A-Za-z0-9_.-]', '-', str(student_id)).strip('-') or str(enrollment.id)
    return f'rules_proofs/{enrollment.id}/IIE-Rules-Regulations-{safe_student_id}-{suffix}.{extension}'


def save_rules_proof_file(field, storage_name, file_bytes):
    target_name = field.field.generate_filename(field.instance, storage_name)
    if field and field.name and field.name != target_name:
        try:
            field.delete(save=False)
        except Exception:
            logger.warning('Unable to delete replaced rules proof file path=%s.', field.name, exc_info=True)
    if default_storage.exists(target_name):
        default_storage.delete(target_name)
    field.save(storage_name, ContentFile(file_bytes), save=False)


def file_response_from_field(field, content_type, filename=None):
    if not field:
        return None
    try:
        handle = field.open('rb')
    except Exception:
        return None
    response = FileResponse(handle, content_type=content_type)
    if filename:
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


def signing_image_content_type(signing, field_name):
    field = getattr(signing, field_name, None)
    guessed_type = mimetypes.guess_type(getattr(field, 'name', '') or '')[0]
    return guessed_type or 'image/jpeg'


def latest_signed_rules_pdf(signing):
    return binary_or_legacy_file(signing, 'signed_pdf_file', 'signed_pdf')


class RulesSignedPdfView(APIView):
    permission_classes = [IsStaffOrAdmin]

    def get(self, request, enrollment_id):
        signing = RulesSigningRequest.objects.defer(
            'selfie_image_file',
            'signature_image',
            'signature_image_file',
            'signed_pdf_file',
            'submitted_ip',
            'submitted_user_agent',
        ).select_related('enrollment').filter(enrollment_id=enrollment_id).first()
        if not signing:
            return proof_unavailable_response()
        enrollment = signing.enrollment
        if not request.user.is_super_admin and enrollment.branch_id != request.user.branch_id:
            return Response({'detail': 'You do not have permission to view this signed PDF.'}, status=403)
        stored_response = file_response_from_field(
            signing.signed_pdf,
            'application/pdf',
            proof_filename(enrollment),
        )
        if stored_response:
            resolve_rules_signed_notifications(enrollment.id)
            return stored_response
        pdf_bytes = latest_signed_rules_pdf(signing)
        if not pdf_bytes:
            return proof_unavailable_response()
        resolve_rules_signed_notifications(enrollment.id)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{proof_filename(enrollment)}"'
        return response


class RulesSelfieView(APIView):
    permission_classes = [IsStaffOrAdmin]

    def get(self, request, enrollment_id):
        signing = RulesSigningRequest.objects.defer(
            'selfie_image_file',
            'signature_image',
            'signature_image_file',
            'signed_pdf',
            'signed_pdf_file',
            'submitted_ip',
            'submitted_user_agent',
        ).select_related('enrollment').filter(enrollment_id=enrollment_id).first()
        if not signing:
            return Response({'detail': 'Selfie is not available.'}, status=404)
        enrollment = signing.enrollment
        if not request.user.is_super_admin and enrollment.branch_id != request.user.branch_id:
            return Response({'detail': 'You do not have permission to view this selfie.'}, status=403)
        stored_response = file_response_from_field(
            signing.selfie_image,
            signing_image_content_type(signing, 'selfie_image'),
        )
        if stored_response:
            return stored_response
        image_bytes = binary_or_legacy_file(signing, 'selfie_image_file', 'selfie_image')
        if not image_bytes:
            return Response({'detail': 'Selfie is not available.'}, status=404)
        return HttpResponse(image_bytes, content_type='image/jpeg')


class PublicRulesSignedPdfView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request, token):
        signing = RulesSigningRequest.objects.defer(
            'selfie_image_file',
            'signature_image',
            'signature_image_file',
            'signed_pdf_file',
            'submitted_ip',
            'submitted_user_agent',
        ).select_related('enrollment').filter(
            token=token,
            status=RulesSigningRequest.Status.SUBMITTED,
        ).first()
        if not signing:
            return proof_unavailable_response()
        stored_response = file_response_from_field(
            signing.signed_pdf,
            'application/pdf',
            proof_filename(signing.enrollment),
        )
        if stored_response:
            return stored_response
        pdf_bytes = latest_signed_rules_pdf(signing)
        if not pdf_bytes:
            return proof_unavailable_response()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{proof_filename(signing.enrollment)}"'
        return response


class PublicRulesReviewPdfView(APIView):
    permission_classes = [AllowAny]
    authentication_classes = []

    def get(self, request, token):
        signing = RulesSigningRequest.objects.select_related(
            'enrollment__course',
            'enrollment__branch',
        ).filter(token=token).first()
        if not signing:
            return Response({'detail': 'Invalid signing link.'}, status=404)
        try:
            pdf_bytes = build_signed_rules_pdf(signing.enrollment)
        except Exception:
            logger.exception('Rules review PDF generation failed for token=%s.', token)
            return Response({'detail': 'Unable to generate Rules & Regulation PDF at the moment.'}, status=503)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{proof_filename(signing.enrollment)}"'
        return response


# ============================================================
# backend/apps/enrollments/views.py
# ============================================================
from crm.models import Enrollment
from serializers import EnrollmentListSerializer, EnrollmentDetailSerializer
import django_filters


class EnrollmentFilter(django_filters.FilterSet):
    status = django_filters.CharFilter(method='filter_status')
    queue = django_filters.CharFilter(method='filter_queue')
    enrolled_from = django_filters.DateFilter(field_name='enrollment_date', lookup_expr='gte')
    enrolled_to   = django_filters.DateFilter(field_name='enrollment_date', lookup_expr='lte')
    date_from = django_filters.DateFilter(field_name='enrollment_date', lookup_expr='gte')
    date_to = django_filters.DateFilter(field_name='enrollment_date', lookup_expr='lte')
    counselor = django_filters.NumberFilter(method='filter_counselor')
    source = django_filters.CharFilter(method='filter_source')
    date_range = django_filters.CharFilter(method='filter_date_range')
    important_only = django_filters.BooleanFilter(method='filter_important_only')

    def filter_status(self, queryset, name, value):
        if value == Enrollment.Status.ACTIVE:
            return queryset.filter(status__in=[Enrollment.Status.ACTIVE, Enrollment.Status.ENROLLED])
        if value == 'pending':
            return queryset.filter(status__in=[Enrollment.Status.DRAFT, Enrollment.Status.PENDING_RULES, Enrollment.Status.RULES_SENT, Enrollment.Status.RULES_SUBMITTED])
        return queryset.filter(status=value) if value else queryset

    def filter_counselor(self, queryset, name, value):
        if not value:
            return queryset
        return queryset.filter(enrollment_counselor_filter_q(value))

    def filter_source(self, queryset, name, value):
        normalized = normalize_source_filter(value, Enrollment._meta.get_field('source').choices)
        return queryset.filter(source=normalized) if normalized else (queryset if not value else queryset.none())

    def filter_queue(self, queryset, name, value):
        if value == 'yet_to_enroll':
            return queryset.exclude(status__in=Enrollment.FINAL_STATUSES)
        if value == 'enrolled':
            return queryset.filter(status__in=Enrollment.FINAL_STATUSES)
        return queryset

    def filter_important_only(self, queryset, name, value):
        return queryset.filter(is_important=True) if value else queryset

    def filter_date_range(self, queryset, name, value):
        start_date, end_date = date_range_from_request(self.request)
        return queryset.filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)

    class Meta:
        model  = Enrollment
        fields = ['status', 'queue', 'branch', 'course', 'counselor', 'source', 'date_range', 'important_only']


def serialize_installment_schedule(schedule):
    rows = []
    for item in schedule or []:
        due_date = item.get('due_date')
        rows.append({
            **item,
            'due_date': due_date.isoformat() if hasattr(due_date, 'isoformat') else due_date,
        })
    return rows


def serialize_enrollment_payment_schedule(enrollment):
    return serialize_installment_schedule(get_enrollment_installment_schedule(enrollment))


def decimal_to_schedule_amount(value):
    value = Decimal(str(value or 0))
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def course_change_installment_label(index):
    known = {
        1: '1st Installment',
        2: '2nd Installment',
        3: '3rd Installment',
    }
    if index in known:
        return known[index]
    if 10 <= index % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(index % 10, 'th')
    return f'{index}{suffix} Installment'


def schedule_total(schedule):
    return sum(Decimal(str(item.get('amount') or 0)) for item in schedule or [])


def adjust_course_change_schedule_total(schedule, course_fee, locked_count=0, enrollment=None):
    rows = [dict(item) for item in (schedule or [])]
    expected_total = Decimal(str(course_fee or 0))
    current_total = schedule_total(rows)
    difference = expected_total - current_total
    if difference == 0:
        return rows, difference

    adjustable_index = len(rows) - 1
    if adjustable_index < locked_count:
        due_date = getattr(enrollment, 'start_date', None) or getattr(enrollment, 'enrollment_date', None) or timezone.localdate()
        rows.append({
            'label': course_change_installment_label(len(rows) + 1),
            'amount': 0,
            'due_date': due_date.isoformat() if hasattr(due_date, 'isoformat') else due_date,
        })
        adjustable_index = len(rows) - 1

    current_amount = Decimal(str(rows[adjustable_index].get('amount') or 0))
    adjusted_amount = current_amount + difference
    if adjusted_amount < 0:
        adjusted_amount = Decimal('0')
    rows[adjustable_index]['amount'] = decimal_to_schedule_amount(adjusted_amount)

    remaining_difference = expected_total - schedule_total(rows)
    if remaining_difference != 0:
        rows[adjustable_index]['amount'] = decimal_to_schedule_amount(
            Decimal(str(rows[adjustable_index].get('amount') or 0)) + remaining_difference
        )
    return rows, difference


def paid_numbered_installment_count(locked_items):
    count = 0
    for item in locked_items or []:
        label = str(item.get('label') or '').strip().lower()
        if label in {'enrollment', 'course fee', 'full payment'}:
            continue
        count += 1
    return count


def log_decimal(value):
    return '' if value is None else str(value)


def course_change_log_context(change_request=None, enrollment=None, old_course=None, new_course=None, payment=None, **extra):
    enrollment = enrollment or getattr(change_request, 'enrollment', None)
    new_course = new_course or getattr(change_request, 'requested_course', None)
    old_course = old_course or getattr(change_request, 'old_course', None) or getattr(enrollment, 'course', None)
    branch = getattr(enrollment, 'branch', None)
    context = {
        'request_id': getattr(change_request, 'id', None),
        'enrollment_id': getattr(enrollment, 'id', None) or getattr(change_request, 'enrollment_id', None),
        'old_course_id': getattr(old_course, 'id', None),
        'old_course': getattr(old_course, 'name', None),
        'new_course_id': getattr(new_course, 'id', None),
        'new_course': getattr(new_course, 'name', None),
        'branch_id': getattr(branch, 'id', None),
        'branch': getattr(branch, 'name', None),
        'old_fee': log_decimal(extra.pop('old_fee', None)),
        'new_fee': log_decimal(extra.pop('new_fee', None)),
        'course_actual_fees': log_decimal(getattr(new_course, 'actual_fees', None)),
        'course_discount_amount': log_decimal(getattr(new_course, 'discount_amount', None)),
        'payment_id': getattr(payment, 'id', None),
        'payment_total_fees': log_decimal(getattr(payment, 'total_fees', None)),
        'payment_paid_amount': log_decimal(getattr(payment, 'paid_amount', None)),
    }
    context.update(extra)
    return context


def rebuild_pending_installment_schedule(enrollment, payment):
    new_total = Decimal(str(enrollment_payable_fee(enrollment) or 0))
    paid_amount = Decimal(str(payment.paid_amount or 0))
    if new_total <= 0:
        logger.info(
            'Course change payment schedule calculation: no payable fee context=%s',
            course_change_log_context(enrollment=enrollment, payment=payment, new_fee=new_total),
        )
        return []

    default_schedule = serialize_enrollment_payment_schedule(enrollment)
    if paid_amount <= 0:
        logger.info(
            'Course change payment schedule calculation: unpaid schedule rebuilt context=%s',
            course_change_log_context(
                enrollment=enrollment,
                payment=payment,
                new_fee=new_total,
                default_schedule=default_schedule,
                rebuilt_schedule=default_schedule,
                rebuilt_schedule_total=str(schedule_total(default_schedule)),
            ),
        )
        return default_schedule

    if paid_amount >= new_total:
        rebuilt_schedule = [{
            'label': 'Course Fee',
            'amount': decimal_to_schedule_amount(new_total),
            'due_date': enrollment.enrollment_date.isoformat() if enrollment.enrollment_date else None,
        }]
        logger.info(
            'Course change payment schedule calculation: already fully paid context=%s',
            course_change_log_context(
                enrollment=enrollment,
                payment=payment,
                new_fee=new_total,
                default_schedule=default_schedule,
                rebuilt_schedule=rebuilt_schedule,
                rebuilt_schedule_total=str(schedule_total(rebuilt_schedule)),
            ),
        )
        return rebuilt_schedule

    locked_items = []
    locked_total = Decimal('0')
    paid_installments = payment.installments.order_by('payment_date', 'created_at', 'pk')
    existing_schedule = get_payment_installment_schedule(payment)
    for installment in paid_installments:
        item_amount = Decimal(str(installment.amount or 0))
        if item_amount <= 0:
            continue
        if locked_total + item_amount > paid_amount:
            break
        if locked_total + item_amount > new_total:
            break
        schedule_item = {}
        if 0 < int(installment.installment_index or 0) <= len(existing_schedule):
            schedule_item = existing_schedule[int(installment.installment_index) - 1]
        due_date = installment.payment_date or schedule_item.get('due_date')
        locked_items.append({
            'label': installment.installment_label or schedule_item.get('label') or f'{len(locked_items) + 1} Installment',
            'amount': decimal_to_schedule_amount(item_amount),
            'due_date': due_date.isoformat() if hasattr(due_date, 'isoformat') else due_date,
        })
        locked_total += item_amount

    remaining_total = new_total - locked_total
    if remaining_total <= 0:
        adjusted_schedule, correction = adjust_course_change_schedule_total(
            locked_items,
            new_total,
            locked_count=len(locked_items),
            enrollment=enrollment,
        )
        logger.info(
            'Course change payment schedule calculation: locked schedule adjusted context=%s',
            course_change_log_context(
                enrollment=enrollment,
                payment=payment,
                new_fee=new_total,
                paid_amount=str(paid_amount),
                remaining_balance='0',
                locked_total=str(locked_total),
                rebuilt_schedule=adjusted_schedule,
                schedule_correction=str(correction),
                final_total=str(schedule_total(adjusted_schedule)),
            ),
        )
        return adjusted_schedule

    pending_count = max(len(default_schedule) - len(locked_items), 1)
    base_amount = remaining_total // pending_count
    amounts = [base_amount for _ in range(pending_count)]
    amounts[-1] += remaining_total - (base_amount * pending_count)
    due_dates = [item.get('due_date') for item in default_schedule[len(locked_items):]]
    while len(due_dates) < pending_count:
        due_dates.append(enrollment.start_date or enrollment.enrollment_date)

    future_items = []
    paid_numbered_count = paid_numbered_installment_count(locked_items)
    for offset, amount in enumerate(amounts, start=1):
        index = len(locked_items) + offset
        default_item = default_schedule[index - 1] if index - 1 < len(default_schedule) else {}
        label = default_item.get('label') or course_change_installment_label(index)
        if locked_items:
            label = course_change_installment_label(paid_numbered_count + offset)
        future_items.append({
            'label': label,
            'amount': decimal_to_schedule_amount(amount),
            'due_date': due_dates[offset - 1].isoformat() if hasattr(due_dates[offset - 1], 'isoformat') else due_dates[offset - 1],
        })
    rebuilt_schedule = locked_items + future_items
    rebuilt_schedule, correction = adjust_course_change_schedule_total(
        rebuilt_schedule,
        new_total,
        locked_count=len(locked_items),
        enrollment=enrollment,
    )
    logger.info(
        'Course change payment schedule calculation: partial payment schedule rebuilt context=%s',
        course_change_log_context(
            enrollment=enrollment,
            payment=payment,
            new_fee=new_total,
            paid_amount=str(paid_amount),
            locked_total=str(locked_total),
            remaining_balance=str(remaining_total),
            pending_count=pending_count,
            default_schedule=default_schedule,
            locked_schedule=locked_items,
            rebuilt_schedule=rebuilt_schedule,
            schedule_correction=str(correction),
            final_total=str(schedule_total(rebuilt_schedule)),
        ),
    )
    return rebuilt_schedule


def reset_rules_for_course_change(enrollment, user, reason, previous_schedule, previous_schedule_locked):
    signing = RulesSigningRequest.objects.select_for_update().filter(enrollment=enrollment).first()
    previous_rules_status = signing.status if signing else RulesSigningRequest.Status.PENDING
    previous_token = str(signing.token) if signing else ''
    previous_signed = bool(
        signing and (
            signing.status == RulesSigningRequest.Status.SUBMITTED
            or signing.signed_pdf
            or signing.signed_pdf_file
        )
    )

    reset_history = EnrollmentRulesResetHistory.objects.create(
        enrollment=enrollment,
        reset_by=user,
        reason=reason,
        previous_rules_status=previous_rules_status,
        previous_signing_token=previous_token,
        previous_schedule_locked=previous_schedule_locked,
        previous_payment_schedule=previous_schedule or [],
        previous_signed=previous_signed,
        previous_signed_pdf=signing.signed_pdf.name if signing and signing.signed_pdf else None,
        previous_signed_pdf_file=signing.signed_pdf_file if signing and signing.signed_pdf_file else None,
    )

    if signing:
        signing.token = uuid.uuid4()
        signing.status = RulesSigningRequest.Status.PENDING
        signing.sent_at = None
        signing.submitted_at = None
        signing.sent_by = None
        signing.submitted_ip = None
        signing.submitted_user_agent = ''
        signing.selfie_image = None
        signing.signature_image = None
        signing.signed_pdf = None
        signing.selfie_image_file = None
        signing.signature_image_file = None
        signing.signed_pdf_file = None
        signing.save(update_fields=[
            'token', 'status', 'sent_at', 'submitted_at', 'sent_by',
            'submitted_ip', 'submitted_user_agent', 'selfie_image',
            'signature_image', 'signed_pdf', 'selfie_image_file',
            'signature_image_file', 'signed_pdf_file', 'updated_at',
        ])
    else:
        RulesSigningRequest.objects.create(enrollment=enrollment, status=RulesSigningRequest.Status.PENDING)
    return reset_history


def notify_admin_course_change_completed(enrollment):
    notify_admin_users(
        'Course Changed',
        f'Course changed for {enrollment.name}. Payment schedule regenerated. Awaiting Rules & Regulations re-sign.',
        Notification.NType.WARNING,
        f'/enrollments/{enrollment.id}',
        category=Notification.Category.APPROVAL,
    )


def apply_enrollment_course_change(enrollment, new_course, user, reason='', effective_date=None):
    effective_date = effective_date or timezone.localdate()
    old_course = enrollment.course
    old_fee = enrollment_payable_fee(enrollment)
    old_status = enrollment.status
    previous_payment_schedule = enrollment.payment_schedule or []
    previous_schedule_locked = enrollment.payment_schedule_locked
    logger.info(
        'Course change apply starting context=%s',
        course_change_log_context(
            enrollment=enrollment,
            old_course=old_course,
            new_course=new_course,
            old_fee=old_fee,
            course_fee=str(getattr(new_course, 'actual_fees', '') or ''),
            course_discount=str(getattr(new_course, 'discount_amount', '') or ''),
            effective_date=effective_date.isoformat() if hasattr(effective_date, 'isoformat') else effective_date,
        ),
    )

    enrollment.course = new_course
    enrollment.final_enrollment_course = new_course
    enrollment.actual_fees = new_course.actual_fees
    enrollment.discount_amount = new_course.discount_amount
    enrollment.custom_payable_fee = None
    enrollment.save(update_fields=[
        'course', 'final_enrollment_course', 'actual_fees', 'discount_amount',
        'custom_payable_fee', 'final_fees', 'net_payable_fee', 'spot_conversion_discount_amount', 'updated_at',
    ])
    new_fee = enrollment_payable_fee(enrollment)
    logger.info(
        'Course change enrollment fee recalculated context=%s',
        course_change_log_context(
            enrollment=enrollment,
            old_course=old_course,
            new_course=new_course,
            old_fee=old_fee,
            new_fee=new_fee,
            enrollment_actual_fees=str(enrollment.actual_fees),
            enrollment_discount_amount=str(enrollment.discount_amount),
            enrollment_final_fees=str(enrollment.final_fees),
            enrollment_net_payable_fee=str(enrollment.net_payable_fee),
        ),
    )

    CourseChangeHistory.objects.create(
        enrollment=enrollment,
        old_course=old_course,
        new_course=new_course,
        changed_by=user,
        old_fee=old_fee,
        new_fee=new_fee,
        reason=reason,
        effective_date=effective_date,
    )

    payment = getattr(enrollment, 'payment', None)
    if payment:
        payment = Payment.objects.select_for_update().get(pk=payment.pk)
        payment.paid_amount = payment.installments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        payment.total_fees = new_fee
        payment.manual_installment_schedule = rebuild_pending_installment_schedule(enrollment, payment)
        rebuilt_total = schedule_total(payment.manual_installment_schedule)
        if rebuilt_total != Decimal(str(new_fee or 0)):
            payment.manual_installment_schedule, correction = adjust_course_change_schedule_total(
                payment.manual_installment_schedule,
                new_fee,
                enrollment=enrollment,
            )
            rebuilt_total = schedule_total(payment.manual_installment_schedule)
            logger.info(
                'Course change payment schedule post-rebuild total auto-corrected context=%s',
                course_change_log_context(
                    enrollment=enrollment,
                    old_course=old_course,
                    new_course=new_course,
                    payment=payment,
                    old_fee=old_fee,
                    new_fee=new_fee,
                    schedule_correction=str(correction),
                    rebuilt_schedule=payment.manual_installment_schedule,
                    final_total=str(rebuilt_total),
                ),
            )
        payment.save(update_fields=[
            'total_fees', 'paid_amount', 'manual_installment_schedule',
            'status', 'next_payment_date', 'updated_at',
        ])
        enrollment.payment_schedule = payment.manual_installment_schedule
        enrollment.payment_schedule_locked = False
        enrollment.payment_schedule_finalized_at = None
        enrollment.status = Enrollment.Status.PENDING_RULES
        enrollment.save(update_fields=[
            'payment_schedule', 'payment_schedule_locked',
            'payment_schedule_finalized_at', 'status', 'updated_at',
        ])
        create_status_history(
            enrollment,
            CandidateStatusHistory.RecordType.ENROLLMENT,
            old_status,
            enrollment.status,
            user,
            reason or 'Course or fee changed.',
        )
        logger.info(
            'Course change payment schedule saved context=%s',
            course_change_log_context(
                enrollment=enrollment,
                old_course=old_course,
                new_course=new_course,
                payment=payment,
                old_fee=old_fee,
                new_fee=new_fee,
                remaining_balance=str(Decimal(str(new_fee or 0)) - Decimal(str(payment.paid_amount or 0))),
                rebuilt_schedule=payment.manual_installment_schedule,
                final_total=str(rebuilt_total),
                payment_status=payment.status,
                next_payment_date=payment.next_payment_date.isoformat() if hasattr(payment.next_payment_date, 'isoformat') else payment.next_payment_date,
                installment_count=payment.installments.count(),
            ),
        )
    else:
        enrollment.payment_schedule = serialize_enrollment_payment_schedule(enrollment)
        enrollment.payment_schedule_locked = False
        enrollment.payment_schedule_finalized_at = None
        enrollment.status = Enrollment.Status.PENDING_RULES
        enrollment.save(update_fields=[
            'payment_schedule', 'payment_schedule_locked',
            'payment_schedule_finalized_at', 'status', 'updated_at',
        ])
        create_status_history(
            enrollment,
            CandidateStatusHistory.RecordType.ENROLLMENT,
            old_status,
            enrollment.status,
            user,
            reason or 'Course or fee changed.',
        )
        logger.info(
            'Course change has no payment record to rebuild context=%s',
            course_change_log_context(
                enrollment=enrollment,
                old_course=old_course,
                new_course=new_course,
                old_fee=old_fee,
                new_fee=new_fee,
            ),
        )
    reset_rules_for_course_change(
        enrollment,
        user,
        reason or 'Course or fee changed. Previous Rules & Regulations PDF archived and re-sign required.',
        previous_payment_schedule,
        previous_schedule_locked,
    )
    resolve_rules_signed_notifications(enrollment.id)
    notify_admin_course_change_completed(enrollment)
    return old_course, old_fee, new_fee


class EnrollmentViewSet(viewsets.ModelViewSet):
    """Enrollment records. Staff sees own branch."""
    permission_classes = [IsStaffOrAdmin]
    filter_backends    = [DjangoFilterBackend, CRMSearchFilter, OrderingFilter]
    filterset_class    = EnrollmentFilter
    pagination_class   = None
    search_fields      = [
        'name', 'phone', 'student_number', 'email', 'source',
        'course__name', 'counselor__first_name', 'counselor__last_name', 'counselor__username',
        'enrolled_by__first_name', 'enrolled_by__last_name',
        'enrolled_by__username', 'created_by__first_name', 'created_by__last_name',
        'created_by__username',
    ]
    id_search_fields   = ['id']
    ordering_fields    = ['enrollment_date', 'name']
    ordering           = ['-enrollment_date']

    def get_queryset(self):
        qs = visible_candidate_queryset(Enrollment.objects.select_related(
            'course','branch','counselor','enrolled_by','created_by','lead','walkin','walkin__lead','rules_signing'
        ).prefetch_related(
            'payment__installments',
            'course_change_history',
            'counselor_change_history__old_counselor',
            'counselor_change_history__new_counselor',
            'counselor_change_history__changed_by',
        ))
        if not self.request.user.is_super_admin:
            qs = qs.filter(branch=self.request.user.branch)
        if getattr(self, 'action', None) == 'list' and self.request.query_params.get('queue') != 'yet_to_enroll':
            qs = official_enrollment_queryset(qs)
        return qs

    @action(detail=True, methods=['post'], url_path='toggle-important')
    def toggle_important(self, request, pk=None):
        enrollment = self.get_object()
        requested = request.data.get('is_important', None)
        enrollment.is_important = truthy_query_param(requested) if requested is not None else not enrollment.is_important
        enrollment.save(update_fields=['is_important', 'updated_at'])
        return Response({'id': enrollment.id, 'is_important': enrollment.is_important})

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        export_kind = request.query_params.get('kind') or 'enrollments'
        export_type = 'students' if export_kind == 'students' else 'enrollments'
        return admin_template_export_response(request, export_type, f'{export_kind}-export')

    def get_serializer_class(self):
        return EnrollmentListSerializer if self.action == 'list' else EnrollmentDetailSerializer

    def _schedule_from_split_count(self, enrollment, split_count):
        schedule = get_default_installment_schedule(enrollment, split_count=split_count)
        total = sum(int(item.get('amount') or 0) for item in schedule)
        expected_total = int(round(float(enrollment_payable_fee(enrollment) or 0)))
        if total != expected_total:
            return None, 'Payment schedule total must match net payable fee.'
        return normalize_installment_schedule(schedule), ''

    def _clean_manual_payment_schedule(self, enrollment, schedule):
        if not isinstance(schedule, list) or not schedule:
            return None, 'Payment schedule is required.'
        cleaned = []
        total = Decimal('0')
        expected_total = Decimal(str(enrollment_payable_fee(enrollment) or 0)).quantize(Decimal('0.01'))
        if Decimal('0') < expected_total <= Decimal(str(LOW_FEE_SINGLE_PAYMENT_MAX_COURSE_FEE)):
            raw_due_date = schedule[0].get('due_date') if isinstance(schedule[0], dict) else None
            due_date = parse_date(str(raw_due_date)) if raw_due_date else enrollment.enrollment_date
            if not due_date:
                return None, 'Each installment needs a valid due date.'
            return normalize_installment_schedule([{
                'label': 'Single Payment',
                'amount': decimal_to_schedule_amount(expected_total),
                'due_date': due_date.isoformat(),
            }]), ''
        if expected_total < Decimal(str(ENROLLMENT_PAYMENT_AMOUNT)):
            return None, 'Course fee must be at least Rs 5,000.'
        for index, item in enumerate(schedule, start=1):
            if not isinstance(item, dict):
                return None, 'Each payment schedule row must be valid.'
            raw_amount = item.get('amount')
            raw_due_date = item.get('due_date')
            if raw_amount in (None, ''):
                return None, 'Each installment needs an amount.'
            if not raw_due_date:
                return None, 'Each installment needs a due date.'
            try:
                amount = Decimal(str(raw_amount))
            except (InvalidOperation, TypeError, ValueError):
                return None, 'Installment amount must be numeric.'
            if not amount.is_finite() or amount <= 0:
                return None, 'Installment amount must be greater than zero.'
            if index == 1 and amount != Decimal(str(ENROLLMENT_PAYMENT_AMOUNT)):
                return None, 'Enrollment Fee must be Rs 5,000.'
            if index > 1 and index < len(schedule) and amount < Decimal(str(MIN_INSTALLMENT_AMOUNT)):
                return None, 'Each installment must be Rs 5,000 or above.'
            if index > 1 and index == len(schedule) and amount < Decimal(str(MIN_INSTALLMENT_AMOUNT)):
                pending_before_final = expected_total - total
                if amount != pending_before_final:
                    return None, 'Each installment must be Rs 5,000 or above.'
            due_date = parse_date(str(raw_due_date))
            if not due_date:
                return None, 'Each installment needs a valid due date.'
            total += amount
            amount = amount.quantize(Decimal('0.01'))
            cleaned.append({
                'label': 'Enrollment' if index == 1 else (str(item.get('label') or '').strip() or f'{index - 1} Installment'),
                'amount': int(amount) if amount == amount.to_integral_value() else float(amount),
                'due_date': due_date.isoformat(),
            })
        if total.quantize(Decimal('0.01')) > expected_total:
            return None, 'Installment total cannot exceed course fee.'
        if total.quantize(Decimal('0.01')) != expected_total:
            return None, 'Payment schedule total must match net payable fee.'
        return normalize_installment_schedule(cleaned), ''

    def _ensure_enrollment_schedule(self, enrollment, split_count=None, lock=False):
        payable_fee = int(round(float(enrollment_payable_fee(enrollment) or 0)))
        if split_count is not None or not enrollment.payment_schedule:
            schedule, error = self._schedule_from_split_count(enrollment, split_count or 2)
            if error:
                raise ValueError(error)
            enrollment.payment_schedule = schedule
        else:
            schedule, error = self._clean_manual_payment_schedule(enrollment, enrollment.payment_schedule)
            if error:
                raise ValueError(error)
            enrollment.payment_schedule = schedule
        update_fields = ['payment_schedule', 'updated_at']
        if lock and not enrollment.payment_schedule_locked:
            enrollment.payment_schedule_locked = True
            enrollment.payment_schedule_finalized_at = timezone.now()
            update_fields.extend(['payment_schedule_locked', 'payment_schedule_finalized_at'])
        enrollment.save(update_fields=update_fields)
        return enrollment.payment_schedule

    def perform_create(self, serializer):
        enrollment = serializer.save(
            enrolled_by=self.request.user,
            created_by=self.request.user,
            counselor=serializer.validated_data.get('counselor') or self.request.user,
        )
        notify_admin_enrollment_created(enrollment, self.request.user)

    def update(self, request, *args, **kwargs):
        if 'branch' in request.data and not request.user.is_super_admin:
            return Response({'detail': 'Only admin can change an enrollment branch.'}, status=status.HTTP_403_FORBIDDEN)
        enrollment = self.get_object()
        old_status = tracked_crm_status(enrollment)
        with transaction.atomic():
            response = super().update(request, *args, **kwargs)
            enrollment.refresh_from_db()
            create_status_history(
                enrollment,
                CandidateStatusHistory.RecordType.ENROLLMENT,
                old_status,
                tracked_crm_status(enrollment),
                request.user,
                getattr(enrollment, 'remarks', ''),
            )
            if enrollment.status in Enrollment.FINAL_STATUSES:
                resolve_rules_signed_notifications(enrollment.id)
        enrollment.refresh_from_db()
        return response

    @action(detail=True, methods=['post'], url_path='change-course')
    def change_course(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can directly change course. Submit a course change request instead.'}, status=status.HTTP_403_FORBIDDEN)
        enrollment = self.get_object()
        course_id = request.data.get('course') or request.data.get('new_course')
        if not course_id:
            return Response({'course': 'Select a new course.'}, status=status.HTTP_400_BAD_REQUEST)

        new_course = Course.objects.filter(pk=course_id, is_active=True).first()
        if not new_course:
            return Response({'course': 'Select a valid active course.'}, status=status.HTTP_400_BAD_REQUEST)
        if enrollment.course_id == new_course.id:
            return Response({'course': 'New course must be different from the current course.'}, status=status.HTTP_400_BAD_REQUEST)

        effective_date_raw = request.data.get('effective_date')
        effective_date = parse_date(effective_date_raw) if effective_date_raw else timezone.localdate()
        if not effective_date:
            return Response({'effective_date': 'Enter a valid effective date.'}, status=status.HTTP_400_BAD_REQUEST)

        reason = str(request.data.get('reason') or '').strip()

        with transaction.atomic():
            enrollment = (
                Enrollment.objects
                .select_for_update()
                .get(pk=enrollment.pk)
            )
            apply_enrollment_course_change(enrollment, new_course, request.user, reason, effective_date)

        enrollment.refresh_from_db()
        return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='request-course-change')
    def request_course_change(self, request, pk=None):
        enrollment = self.get_object()
        course_id = request.data.get('requested_course') or request.data.get('course') or request.data.get('new_course')
        if not course_id:
            return Response({'requested_course': 'Select a new course.'}, status=status.HTTP_400_BAD_REQUEST)

        requested_course = Course.objects.filter(pk=course_id, is_active=True).first()
        if not requested_course:
            return Response({'requested_course': 'Select a valid active course.'}, status=status.HTTP_400_BAD_REQUEST)
        if enrollment.course_id == requested_course.id:
            return Response({'requested_course': 'Requested course must be different from the current course.'}, status=status.HTTP_400_BAD_REQUEST)

        requested_batch_date_raw = request.data.get('requested_batch_date') or request.data.get('preferred_batch_start_date')
        requested_batch_date = parse_date(requested_batch_date_raw) if requested_batch_date_raw else None
        if requested_batch_date_raw and not requested_batch_date:
            return Response({'requested_batch_date': 'Enter a valid preferred batch start date.'}, status=status.HTTP_400_BAD_REQUEST)

        reason = str(request.data.get('reason') or '').strip()
        if not reason:
            return Response({'reason': 'Reason for change is required.'}, status=status.HTTP_400_BAD_REQUEST)

        old_fee = enrollment_payable_fee(enrollment)
        new_fee = max(Decimal(str(requested_course.actual_fees or 0)) - Decimal(str(requested_course.discount_amount or 0)), Decimal('0'))
        if enrollment.spot_conversion_discount_applied:
            new_fee = max(new_fee - Decimal('2000'), Decimal('0'))

        with transaction.atomic():
            existing = CourseChangeRequest.objects.select_for_update().filter(
                enrollment=enrollment,
                status=CourseChangeRequest.Status.PENDING,
            ).first()
            if existing:
                return Response({'detail': 'This enrollment already has a pending course change request.'}, status=status.HTTP_400_BAD_REQUEST)
            change_request = CourseChangeRequest.objects.create(
                student=enrollment,
                enrollment=enrollment,
                old_course=enrollment.course,
                requested_course=requested_course,
                requested_batch_date=requested_batch_date,
                reason=reason,
                requested_by=request.user,
                requested_at=timezone.now(),
                status=CourseChangeRequest.Status.PENDING,
                old_fee=old_fee,
                new_fee=new_fee,
            )
            message = f'Course change request submitted for {enrollment.name}'
            for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
                create_user_notification(
                    admin_user,
                    'Course Change Request',
                    message,
                    Notification.NType.INFO,
                    f'/admin/course-change-requests?request={change_request.id}',
                )

        return Response(CourseChangeRequestSerializer(change_request, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='reassign-counselor')
    def reassign_counselor(self, request, pk=None):
        enrollment = self.get_object()
        counselor_id = request.data.get('counselor') or request.data.get('new_counselor') or request.data.get('user')
        if not counselor_id:
            return Response({'counselor': 'Select a counselor.'}, status=status.HTTP_400_BAD_REQUEST)

        counselor = User.objects.filter(
            pk=counselor_id,
            role=User.Role.STAFF,
            is_active=True,
        ).select_related('branch').first()
        if not counselor:
            return Response({'counselor': 'Select an active staff counselor.'}, status=status.HTTP_400_BAD_REQUEST)
        if enrollment.branch_id and counselor.branch_id != enrollment.branch_id:
            return Response({'counselor': 'Select a counselor from the enrollment branch.'}, status=status.HTTP_400_BAD_REQUEST)

        if request.data.get('force_transfer'):
            if not request.user.is_super_admin:
                return Response({'detail': 'Only admin can force transfer counselors.'}, status=status.HTTP_403_FORBIDDEN)
            reason = str(request.data.get('reason') or '').strip()
            if not reason:
                return Response({'reason': 'Reason for transfer is required.'}, status=status.HTTP_400_BAD_REQUEST)
            with transaction.atomic():
                change_request = CounselorChangeRequest.objects.create(
                    record_type=CounselorChangeRequest.RecordType.ENROLLMENT,
                    enrollment=enrollment,
                    branch=enrollment.branch,
                    candidate_name=enrollment.name,
                    candidate_phone=enrollment.phone,
                    current_counselor=enrollment.counselor or enrollment.created_by or enrollment.enrolled_by,
                    requested_counselor=counselor,
                    requested_by=request.user,
                    reason=reason,
                    status=CounselorChangeRequest.Status.APPROVED,
                    counselor_decision_by=request.user,
                    counselor_decision_at=timezone.now(),
                    counselor_remarks='Super admin force transfer.',
                    admin_remarks=str(request.data.get('admin_remarks') or '').strip(),
                )
                apply_counselor_change(change_request, request.user, force=True)
                change_request.save(update_fields=[
                    'status', 'counselor_decision_by', 'counselor_decision_at',
                    'counselor_remarks', 'admin_decision_by', 'admin_decision_at',
                    'admin_remarks', 'force_transfer',
                ])
            enrollment.refresh_from_db()
            notify_admin_decision(change_request, True)
            return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)

        try:
            change_request = create_counselor_change_request(
                enrollment,
                CounselorChangeRequest.RecordType.ENROLLMENT,
                counselor,
                request.user,
                request.data.get('reason'),
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(CounselorChangeRequestSerializer(change_request, context={'request': request}).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='request-counselor-change')
    def request_counselor_change(self, request, pk=None):
        return self.reassign_counselor(request, pk)

    @action(detail=True, methods=['post'], url_path='force-counselor-transfer')
    def force_counselor_transfer(self, request, pk=None):
        data = request.data.copy()
        data['force_transfer'] = True
        request._full_data = data
        return self.reassign_counselor(request, pk)

    @action(detail=True, methods=['post'], url_path='direct-reassign-counselor-disabled')
    def direct_reassign_counselor_disabled(self, request, pk=None):
        return Response({'detail': 'Use counselor change request workflow.'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

    def _legacy_direct_reassign_removed(self, request, enrollment, counselor):
        reason = str(request.data.get('reason') or '').strip()
        with transaction.atomic():
            enrollment = (
                Enrollment.objects
                .select_for_update()
                .get(pk=enrollment.pk)
            )
            old_counselor = enrollment.counselor or enrollment.created_by or enrollment.enrolled_by
            if old_counselor and old_counselor.id == counselor.id:
                return Response({'detail': 'This counselor is already assigned.'}, status=status.HTTP_400_BAD_REQUEST)

            EnrollmentCounselorChangeHistory.objects.create(
                enrollment=enrollment,
                old_counselor=old_counselor,
                new_counselor=counselor,
                changed_by=request.user,
                reason=reason,
            )
            enrollment.counselor = counselor
            enrollment.save(update_fields=['counselor', 'updated_at'])

        enrollment.refresh_from_db()
        return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='payment-schedule')
    def payment_schedule(self, request, pk=None):
        enrollment = self.get_object()
        signing_status = getattr(getattr(enrollment, 'rules_signing', None), 'status', RulesSigningRequest.Status.PENDING)
        rules_already_sent = signing_status in {
            RulesSigningRequest.Status.SENT,
            RulesSigningRequest.Status.VIEWED,
            RulesSigningRequest.Status.SUBMITTED,
        }
        if enrollment.payment_schedule_locked and rules_already_sent:
            return Response({'detail': 'Payment schedule is locked after Rules & Regulation sending.'}, status=status.HTTP_403_FORBIDDEN)
        if enrollment.status in Enrollment.FINAL_STATUSES and not request.user.is_super_admin:
            return Response({'detail': 'Only admin can override payment schedules after enrollment confirmation.'}, status=status.HTTP_403_FORBIDDEN)
        if not enrollment.start_date:
            return Response({'detail': 'Course start date is required before finalizing payment schedule.'}, status=status.HTTP_400_BAD_REQUEST)
        if request.data.get('payment_schedule') is not None:
            schedule, error = self._clean_manual_payment_schedule(enrollment, request.data.get('payment_schedule'))
            if error:
                return Response({'detail': error}, status=status.HTTP_400_BAD_REQUEST)
            enrollment.payment_schedule = schedule
            enrollment.payment_schedule_locked = False
            enrollment.payment_schedule_finalized_at = None
            enrollment.save(update_fields=[
                'payment_schedule',
                'payment_schedule_locked',
                'payment_schedule_finalized_at',
                'updated_at',
            ])
            return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)
        try:
            split_count = int(request.data.get('split_count') or 2)
        except (TypeError, ValueError):
            split_count = 2
        split_count = min(max(split_count, 2), 12)
        try:
            self._ensure_enrollment_schedule(enrollment, split_count=split_count, lock=False)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)

    def _reset_enrollment_workflow(self, request, enrollment_pk):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can reset the enrollment workflow.'}, status=status.HTTP_403_FORBIDDEN)
        reason = str(request.data.get('reason') or '').strip()
        if not reason:
            return Response({'reason': 'Reset reason is required.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            enrollment = (
                Enrollment.objects
                .select_for_update()
                .select_related('branch', 'course')
                .get(pk=enrollment_pk)
            )
            if enrollment.status in Enrollment.FINAL_STATUSES:
                return Response({'detail': 'Cannot reset enrollment workflow after enrollment is completed.'}, status=status.HTTP_400_BAD_REQUEST)

            signing = RulesSigningRequest.objects.select_for_update().filter(enrollment=enrollment).first()
            previous_rules_status = signing.status if signing else RulesSigningRequest.Status.PENDING
            previous_token = str(signing.token) if signing else ''
            previous_signed = bool(
                signing and (
                    signing.status == RulesSigningRequest.Status.SUBMITTED
                    or signing.signed_pdf
                    or signing.signed_pdf_file
                )
            )
            EnrollmentRulesResetHistory.objects.create(
                enrollment=enrollment,
                reset_by=request.user,
                reason=reason,
                previous_rules_status=previous_rules_status,
                previous_signing_token=previous_token,
                previous_schedule_locked=enrollment.payment_schedule_locked,
                previous_payment_schedule=enrollment.payment_schedule or [],
                previous_signed=previous_signed,
            )
            if signing:
                signing.delete()

            enrollment.payment_schedule_locked = False
            enrollment.payment_schedule_finalized_at = None
            old_status = enrollment.status
            enrollment.status = Enrollment.Status.PENDING_RULES
            enrollment.save(update_fields=[
                'payment_schedule_locked',
                'payment_schedule_finalized_at',
                'status',
                'updated_at',
            ])
            create_status_history(
                enrollment,
                CandidateStatusHistory.RecordType.ENROLLMENT,
                old_status,
                enrollment.status,
                request.user,
                reason,
            )

        enrollment.refresh_from_db()
        return Response({
            'detail': 'Enrollment workflow reset. Edit the payment schedule and send a fresh Rules & Regulation form.',
            'enrollment': EnrollmentDetailSerializer(enrollment, context={'request': request}).data,
        })

    @action(detail=True, methods=['post'], url_path='reset-rules-process')
    def reset_rules_process(self, request, pk=None):
        return self._reset_enrollment_workflow(request, self.get_object().pk)

    @action(detail=True, methods=['post'], url_path='reset-enrollment-workflow')
    def reset_enrollment_workflow(self, request, pk=None):
        return self._reset_enrollment_workflow(request, self.get_object().pk)

    @action(detail=True, methods=['post'], url_path='send-rules-form')
    def send_rules_form(self, request, pk=None):
        enrollment = self.get_object()
        errors = {}
        if not enrollment.start_date:
            errors['start_date'] = 'Course start date is required.'
        if not str(enrollment.batch_timing or '').strip():
            errors['batch_timing'] = 'Batch timing is required.'
        if errors:
            return Response(
                {'detail': 'Course start date and batch timing are required before sending the Rules & Regulation form.', **errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            self._ensure_enrollment_schedule(enrollment, lock=False)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        signing, _ = RulesSigningRequest.objects.get_or_create(enrollment=enrollment)
        signing_path = f'{app_url(f"rules-sign/{signing.token}")}/'
        signing_link = request.build_absolute_uri(signing_path)
        message = (
            f'Dear {enrollment.name},\n\n'
            'Welcome to Indra Institute of Education.\n\n'
            'Please review and sign your Rules & Regulations form using the link below:\n\n'
            f'{signing_link}\n\n'
            'Thank you.\n'
            '-Team IIE'
        )

        logger.info(
            'Rules & Regulations link generated: student_id=%s enrollment_id=%s token=%s url=%s',
            enrollment.student_number or enrollment.id,
            enrollment.id,
            signing.token,
            signing_link,
        )
        try:
            log = send_whatsapp_message(
                candidate_name=enrollment.name,
                phone=enrollment.phone,
                message_type=WhatsAppMessage.MsgType.RULES_FORM_LINK,
                message_body=message,
                sent_by=request.user,
                related_model='enrollment',
                related_id=enrollment.id,
            )
        except Exception as exc:
            logger.exception(
                'Rules & Regulations WhatsApp Web handoff crashed: candidate_id=%s candidate_name=%s candidate_phone=%s enrollment_id=%s',
                enrollment.student_number or enrollment.id,
                enrollment.name,
                enrollment.phone,
                enrollment.id,
            )
            return Response({
                'detail': f'Unable to prepare WhatsApp Web message: {exc}',
                'signing_link': signing_link,
                'whatsapp_message': message,
                'phone': enrollment.phone,
                'status': signing.status,
                'enrollment_status': enrollment.status,
            }, status=status.HTTP_503_SERVICE_UNAVAILABLE)

        logger.info(
            'Rules & Regulations WhatsApp message status: student_id=%s token=%s url=%s whatsapp_log_id=%s whatsapp_status=%s whatsapp_error=%s provider_response=%s',
            enrollment.student_number or enrollment.id,
            signing.token,
            signing_link,
            log.id,
            log.status,
            log.error_message,
            log.provider_response,
        )
        if log.status != WhatsAppMessage.MsgStatus.SENT:
            detail = log.error_message or 'Unable to prepare WhatsApp Web message.'
            logger.error(
                'Rules & Regulations WhatsApp Web handoff failed: candidate_id=%s candidate_name=%s candidate_phone=%s whatsapp_log_id=%s error=%s response=%s',
                enrollment.student_number or enrollment.id,
                enrollment.name,
                enrollment.phone,
                log.id,
                detail,
                log.provider_response,
            )
            return Response({
                'detail': detail,
                'signing_link': signing_link,
                'whatsapp_message': message,
                'phone': enrollment.phone,
                'status': signing.status,
                'enrollment_status': enrollment.status,
                **whatsapp_send_payload(log),
            }, status=status.HTTP_400_BAD_REQUEST)
        signing.status = RulesSigningRequest.Status.SENT
        signing.sent_at = timezone.now()
        signing.sent_by = request.user
        signing.save(update_fields=['status', 'sent_at', 'sent_by', 'updated_at'])
        if enrollment.status != Enrollment.Status.ENROLLED:
            old_status = enrollment.status
            enrollment.status = Enrollment.Status.RULES_SENT
            enrollment.save(update_fields=['status', 'updated_at'])
            create_status_history(
                enrollment,
                CandidateStatusHistory.RecordType.ENROLLMENT,
                old_status,
                enrollment.status,
                request.user,
                'Rules form sent.',
            )
        self._ensure_enrollment_schedule(enrollment, lock=True)
        return Response({
            'detail': 'Rules & Regulations link generated successfully.',
            'signing_link': signing_link,
            'whatsapp_message': message,
            'phone': enrollment.phone,
            'status': signing.status,
            'enrollment_status': enrollment.status,
            'enrollment': EnrollmentDetailSerializer(enrollment, context={'request': request}).data,
            **whatsapp_send_payload(log),
        })

    @action(detail=True, methods=['post'], url_path='enroll-student')
    def enroll_student(self, request, pk=None):
        enrollment = self.get_object()
        signing = getattr(enrollment, 'rules_signing', None)
        if not signing or signing.status != RulesSigningRequest.Status.SUBMITTED:
            return Response({'detail': 'Rules & Regulation form must be signed before enrollment can proceed.'}, status=400)
        with transaction.atomic():
            if enrollment.status not in Enrollment.FINAL_STATUSES:
                old_status = enrollment.status
                enrollment.status = Enrollment.Status.ACTIVE
                enrollment.enrolled_by = request.user
                enrollment.counselor = enrollment.counselor or enrollment.created_by or request.user
                enrollment.save(update_fields=['status', 'enrolled_by', 'counselor', 'student_number', 'final_fees', 'net_payable_fee', 'spot_conversion_discount_amount', 'updated_at'])
                create_status_history(
                    enrollment,
                    CandidateStatusHistory.RecordType.ENROLLMENT,
                    old_status,
                    enrollment.status,
                    request.user,
                    'Student enrollment confirmed.',
                )
            resolve_rules_signed_notifications(enrollment.id)
            payment, created = Payment.objects.get_or_create(
                enrollment=enrollment,
                defaults={
                    'total_fees': enrollment_payable_fee(enrollment),
                    'status': Payment.Status.UNPAID,
                    'manual_installment_schedule': serialize_enrollment_payment_schedule(enrollment),
                },
            )
            if not created:
                update_fields = []
                if payment.total_fees != enrollment_payable_fee(enrollment):
                    payment.total_fees = enrollment_payable_fee(enrollment)
                    update_fields.append('total_fees')
                if not payment.manual_installment_schedule:
                    payment.manual_installment_schedule = serialize_enrollment_payment_schedule(enrollment)
                    update_fields.append('manual_installment_schedule')
                if update_fields:
                    update_fields.extend(['paid_amount', 'status', 'next_payment_date', 'updated_at'])
                    payment.save(update_fields=update_fields)
        return Response(EnrollmentDetailSerializer(enrollment, context={'request': request}).data)

    @action(detail=True, methods=['post'], url_path='add-to-payment')
    def add_to_payment(self, request, pk=None):
        enrollment = self.get_object()
        if enrollment.status not in {
            Enrollment.Status.ENROLLED,
            Enrollment.Status.ACTIVE,
            Enrollment.Status.COMPLETED,
        }:
            return Response(
                {'detail': 'Only enrolled or completed enrollments can be added to payments.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not enrollment.start_date:
            return Response(
                {'detail': 'Course start date is required to create payment schedule.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if Payment.objects.filter(enrollment=enrollment).exists():
            return Response(
                {'detail': 'Payment record already exists for this enrollment.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            payment = Payment.objects.create(
                enrollment=enrollment,
                total_fees=enrollment_payable_fee(enrollment),
                status=Payment.Status.UNPAID,
                manual_installment_schedule=serialize_enrollment_payment_schedule(enrollment),
            )

        enrollment.refresh_from_db()
        return Response({
            'detail': 'Payment record created.',
            'payment': PaymentSerializer(payment, context={'request': request}).data,
            'enrollment': EnrollmentDetailSerializer(enrollment, context={'request': request}).data,
        }, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can delete enrollments and students.'}, status=status.HTTP_403_FORBIDDEN)
        enrollment = self.get_object()
        with transaction.atomic():
            if enrollment.lead_id:
                Lead.objects.filter(
                    pk=enrollment.lead_id,
                    converted_to_type='enrollment',
                    converted_record_id=enrollment.id,
                ).update(
                    status=Lead.Status.NEW,
                    converted_to_type='',
                    converted_record_id=None,
                    converted_at=None,
                    converted_by=None,
                )
            if enrollment.walkin_id:
                WalkIn.objects.filter(
                    pk=enrollment.walkin_id,
                    converted_to_type='enrollment',
                    converted_record_id=enrollment.id,
                ).update(
                    status=WalkIn.Status.NEW,
                    converted_to_type='',
                    converted_record_id=None,
                    converted_at=None,
                    converted_by=None,
                )
            self.perform_destroy(enrollment)
        return Response(status=status.HTTP_204_NO_CONTENT)


class CourseChangeRequestFilter(django_filters.FilterSet):
    branch = django_filters.NumberFilter(field_name='enrollment__branch_id')
    date = django_filters.DateFilter(field_name='requested_at', lookup_expr='date')

    class Meta:
        model = CourseChangeRequest
        fields = ['status', 'branch', 'date']


class CourseChangeRequestViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CourseChangeRequestSerializer
    permission_classes = [IsStaffOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = CourseChangeRequestFilter
    search_fields = ['enrollment__name', 'enrollment__phone', 'old_course__name', 'requested_course__name']
    ordering_fields = ['requested_at', 'reviewed_at']
    ordering = ['-requested_at']
    pagination_class = None

    def get_queryset(self):
        queryset = CourseChangeRequest.objects.select_related(
            'enrollment__branch', 'student', 'old_course', 'requested_course', 'requested_by', 'reviewed_by'
        )
        if self.request.user.is_super_admin:
            return queryset
        return queryset.filter(Q(requested_by=self.request.user) | Q(enrollment__branch=self.request.user.branch))

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can approve course change requests.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            change_request = self.get_object()
            if change_request.status != CourseChangeRequest.Status.PENDING:
                return Response({'detail': 'Only pending course change requests can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
            remarks = str(request.data.get('admin_remarks') or request.data.get('remarks') or '').strip()

            logger.info(
                'Course change approval requested context=%s',
                course_change_log_context(change_request=change_request),
            )

            with transaction.atomic():
                change_request = CourseChangeRequest.objects.select_for_update().get(pk=change_request.pk)

                if change_request.status != CourseChangeRequest.Status.PENDING:
                    return Response({'detail': 'Only pending course change requests can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
                if not change_request.enrollment_id:
                    return Response({'detail': 'Course change request is missing an enrollment.'}, status=status.HTTP_400_BAD_REQUEST)
                requested_course = Course.objects.filter(pk=change_request.requested_course_id).first()
                if not requested_course:
                    return Response({'detail': 'Requested course no longer exists.'}, status=status.HTTP_400_BAD_REQUEST)
                if requested_course.actual_fees is None:
                    return Response({'detail': 'Course fee is missing for requested course.'}, status=status.HTTP_400_BAD_REQUEST)

                try:
                    enrollment = Enrollment.objects.select_for_update().get(pk=change_request.enrollment_id)
                except Enrollment.DoesNotExist:
                    return Response({'detail': 'Enrollment for this course change request no longer exists.'}, status=status.HTTP_400_BAD_REQUEST)

                logger.info(
                    'Course change approval validation passed context=%s',
                    course_change_log_context(
                        change_request=change_request,
                        enrollment=enrollment,
                        old_course=enrollment.course,
                        new_course=requested_course,
                        old_fee=enrollment_payable_fee(enrollment),
                        new_fee=max(
                            Decimal(str(requested_course.actual_fees or 0))
                            - Decimal(str(requested_course.discount_amount or 0)),
                            Decimal('0'),
                        ),
                    ),
                )

                if enrollment.course_id == change_request.requested_course_id:
                    return Response({'detail': 'Enrollment is already in the requested course.'}, status=status.HTTP_400_BAD_REQUEST)
                old_course, old_fee, new_fee = apply_enrollment_course_change(
                    enrollment,
                    requested_course,
                    request.user,
                    change_request.reason,
                    change_request.requested_batch_date or timezone.localdate(),
                )
                change_request.old_course = old_course
                change_request.old_fee = old_fee
                change_request.new_fee = new_fee
                change_request.status = CourseChangeRequest.Status.APPROVED
                change_request.reviewed_by = request.user
                change_request.reviewed_at = timezone.now()
                change_request.admin_remarks = remarks
                change_request.save(update_fields=[
                    'old_course', 'old_fee', 'new_fee', 'status', 'reviewed_by',
                    'reviewed_at', 'admin_remarks', 'updated_at',
                ])
                mark_notifications_terminal(
                    Notification.objects.filter(
                        title='Course Change Request',
                        related_url=f'/admin/course-change-requests?request={change_request.id}',
                    ),
                    Notification.Status.APPROVED,
                )
                create_user_notification(
                    change_request.requested_by,
                    'Course Change Request Approved',
                    'Your course change request has been approved.',
                    Notification.NType.SUCCESS,
                    f'/enrollments/{change_request.enrollment_id}',
                )

            logger.info(
                'Course change approval completed context=%s',
                course_change_log_context(
                    change_request=change_request,
                    old_course=change_request.old_course,
                    new_course=change_request.requested_course,
                    old_fee=change_request.old_fee,
                    new_fee=change_request.new_fee,
                ),
            )
            return Response(self.get_serializer(change_request).data)
        except Exception as exc:
            logger.exception(
                'Course approval failed context=%s',
                {
                    'request_id': pk,
                    'admin_user_id': getattr(request.user, 'id', None),
                    'payload': dict(request.data) if hasattr(request.data, 'items') else request.data,
                },
            )
            return Response({
                'success': False,
                'detail': str(exc),
                'message': str(exc),
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can reject course change requests.'}, status=status.HTTP_403_FORBIDDEN)
        change_request = self.get_object()
        if change_request.status != CourseChangeRequest.Status.PENDING:
            return Response({'detail': 'Only pending course change requests can be rejected.'}, status=status.HTTP_400_BAD_REQUEST)
        remarks = str(request.data.get('admin_remarks') or request.data.get('remarks') or '').strip()

        change_request.status = CourseChangeRequest.Status.REJECTED
        change_request.reviewed_by = request.user
        change_request.reviewed_at = timezone.now()
        change_request.admin_remarks = remarks
        change_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'admin_remarks', 'updated_at'])
        mark_notifications_terminal(
            Notification.objects.filter(
                title='Course Change Request',
                related_url=f'/admin/course-change-requests?request={change_request.id}',
            ),
            Notification.Status.REJECTED,
        )
        create_user_notification(
            change_request.requested_by,
            'Course Change Request Rejected',
            'Your course change request has been rejected.',
            Notification.NType.ERROR,
            f'/enrollments/{change_request.enrollment_id}',
        )
        return Response(self.get_serializer(change_request).data)


class CounselorChangeRequestFilter(django_filters.FilterSet):
    branch = django_filters.NumberFilter(field_name='branch_id')
    date = django_filters.DateFilter(field_name='requested_at', lookup_expr='date')

    class Meta:
        model = CounselorChangeRequest
        fields = ['status', 'branch', 'date', 'record_type']


class CounselorChangeRequestViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CounselorChangeRequestSerializer
    permission_classes = [IsStaffOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = CounselorChangeRequestFilter
    search_fields = ['candidate_name', 'candidate_phone', 'current_counselor__username', 'requested_counselor__username']
    ordering_fields = ['requested_at', 'counselor_decision_at', 'admin_decision_at']
    ordering = ['-requested_at']
    pagination_class = None

    def get_queryset(self):
        queryset = CounselorChangeRequest.objects.select_related(
            'lead', 'enrollment', 'branch', 'current_counselor',
            'requested_counselor', 'requested_by', 'counselor_decision_by',
            'admin_decision_by',
        )
        if self.request.user.is_super_admin:
            return queryset
        return queryset.filter(
            Q(requested_by=self.request.user)
            | Q(current_counselor=self.request.user)
            | Q(requested_counselor=self.request.user)
        )

    @action(detail=True, methods=['post'], url_path='counselor-approve')
    def counselor_approve(self, request, pk=None):
        change_request = self.get_object()
        if change_request.current_counselor_id != request.user.id:
            return Response({'detail': 'Only the current counselor can approve this request.'}, status=status.HTTP_403_FORBIDDEN)
        if change_request.status != CounselorChangeRequest.Status.PENDING_COUNSELOR:
            return Response({'detail': 'Only requests pending counselor approval can be approved.'}, status=status.HTTP_400_BAD_REQUEST)
        change_request.status = CounselorChangeRequest.Status.PENDING_ADMIN
        change_request.counselor_decision_by = request.user
        change_request.counselor_decision_at = timezone.now()
        change_request.counselor_remarks = str(request.data.get('remarks') or '').strip()
        change_request.save(update_fields=['status', 'counselor_decision_by', 'counselor_decision_at', 'counselor_remarks'])
        notify_counselor_decision(change_request, True)
        return Response(self.get_serializer(change_request).data)

    @action(detail=True, methods=['post'], url_path='counselor-reject')
    def counselor_reject(self, request, pk=None):
        change_request = self.get_object()
        if change_request.current_counselor_id != request.user.id:
            return Response({'detail': 'Only the current counselor can reject this request.'}, status=status.HTTP_403_FORBIDDEN)
        if change_request.status != CounselorChangeRequest.Status.PENDING_COUNSELOR:
            return Response({'detail': 'Only requests pending counselor approval can be rejected.'}, status=status.HTTP_400_BAD_REQUEST)
        change_request.status = CounselorChangeRequest.Status.REJECTED
        change_request.counselor_decision_by = request.user
        change_request.counselor_decision_at = timezone.now()
        change_request.counselor_remarks = str(request.data.get('remarks') or '').strip()
        change_request.save(update_fields=['status', 'counselor_decision_by', 'counselor_decision_at', 'counselor_remarks'])
        notify_counselor_decision(change_request, False)
        return Response(self.get_serializer(change_request).data)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can approve counselor change requests.'}, status=status.HTTP_403_FORBIDDEN)
        change_request = self.get_object()
        if change_request.status != CounselorChangeRequest.Status.PENDING_ADMIN:
            return Response({'detail': 'Counselor approval is required before admin approval.'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            change_request = CounselorChangeRequest.objects.select_for_update().get(pk=change_request.pk)
            apply_counselor_change(change_request, request.user)
            change_request.admin_remarks = str(request.data.get('admin_remarks') or request.data.get('remarks') or '').strip()
            change_request.save(update_fields=['status', 'admin_decision_by', 'admin_decision_at', 'admin_remarks', 'force_transfer'])
        notify_admin_decision(change_request, True)
        return Response(self.get_serializer(change_request).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can reject counselor change requests.'}, status=status.HTTP_403_FORBIDDEN)
        change_request = self.get_object()
        if change_request.status not in [
            CounselorChangeRequest.Status.PENDING_COUNSELOR,
            CounselorChangeRequest.Status.PENDING_ADMIN,
        ]:
            return Response({'detail': 'Only pending counselor change requests can be rejected.'}, status=status.HTTP_400_BAD_REQUEST)
        change_request.status = CounselorChangeRequest.Status.REJECTED
        change_request.admin_decision_by = request.user
        change_request.admin_decision_at = timezone.now()
        change_request.admin_remarks = str(request.data.get('admin_remarks') or request.data.get('remarks') or '').strip()
        change_request.save(update_fields=['status', 'admin_decision_by', 'admin_decision_at', 'admin_remarks'])
        notify_admin_decision(change_request, False)
        return Response(self.get_serializer(change_request).data)

    @action(detail=True, methods=['post'], url_path='force-approve')
    def force_approve(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can force transfer counselors.'}, status=status.HTTP_403_FORBIDDEN)
        change_request = self.get_object()
        if change_request.status not in [
            CounselorChangeRequest.Status.PENDING_COUNSELOR,
            CounselorChangeRequest.Status.PENDING_ADMIN,
        ]:
            return Response({'detail': 'Only pending requests can be force transferred.'}, status=status.HTTP_400_BAD_REQUEST)
        with transaction.atomic():
            change_request = CounselorChangeRequest.objects.select_for_update().get(pk=change_request.pk)
            apply_counselor_change(change_request, request.user, force=True)
            change_request.admin_remarks = str(request.data.get('admin_remarks') or request.data.get('remarks') or '').strip()
            change_request.save(update_fields=['status', 'admin_decision_by', 'admin_decision_at', 'admin_remarks', 'force_transfer'])
        notify_admin_decision(change_request, True)
        return Response(self.get_serializer(change_request).data)


# ============================================================
# backend/apps/payments/views.py
# ============================================================
from crm.models import Payment, PaymentInstallment, get_payment_installment_schedule
from serializers import PaymentSerializer, PaymentInstallmentSerializer, payment_installment_summary


class PaymentFilter(django_filters.FilterSet):
    status = django_filters.CharFilter(method='filter_status')
    branch = django_filters.NumberFilter(field_name='enrollment__branch_id')
    user = django_filters.NumberFilter(method='filter_user')
    counselor = django_filters.NumberFilter(method='filter_user')
    important_only = django_filters.BooleanFilter(method='filter_important_only')
    duration = django_filters.CharFilter(method='filter_duration')
    date_from = django_filters.DateFilter(method='filter_date_from')
    date_to = django_filters.DateFilter(method='filter_date_to')
    due_this_week = django_filters.BooleanFilter(method='filter_due_this_week')
    next_payment_from = django_filters.DateFilter(field_name='next_payment_date', lookup_expr='gte')
    next_payment_to = django_filters.DateFilter(field_name='next_payment_date', lookup_expr='lte')

    class Meta:
        model = Payment
        fields = [
            'status', 'branch', 'enrollment__branch', 'user', 'counselor', 'important_only',
            'duration', 'date_from', 'date_to', 'due_this_week', 'next_payment_from', 'next_payment_to',
        ]

    def filter_status(self, queryset, name, value):
        if value == 'weekly_pending':
            today = timezone.localdate()
            week_start = today - timedelta(days=(today.weekday() + 1) % 7)
            week_end = week_start + timedelta(days=6)
            return queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                paid_amount__lt=F('total_fees'),
                next_payment_date__isnull=False,
                next_payment_date__gte=week_start,
                next_payment_date__lte=week_end,
            )
        if value in ('due', 'pending_today'):
            today = timezone.localdate()
            return queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                paid_amount__lt=F('total_fees'),
                next_payment_date__isnull=False,
                next_payment_date__lte=today,
            )
        if value == 'pending':
            return queryset.filter(status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL])
        return queryset.filter(status=value)

    def filter_user(self, queryset, name, value):
        if not value:
            return queryset
        if self.request and not self.request.user.is_super_admin:
            allowed = User.objects.filter(
                id=value,
                branch_id=self.request.user.branch_id,
                is_active=True,
            ).exclude(role=User.Role.SUPER_ADMIN).exists()
            if not allowed:
                return queryset.none()
        return queryset.filter(enrollment_counselor_filter_q(value, prefix='enrollment'))

    def filter_important_only(self, queryset, name, value):
        return queryset.filter(enrollment__is_important=True) if value else queryset

    def filter_duration(self, queryset, name, value):
        today = timezone.localdate()
        if value == 'overdue':
            return queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                paid_amount__lt=F('total_fees'),
                next_payment_date__isnull=False,
                next_payment_date__lt=today,
            )
        if value == 'month':
            start = today.replace(day=1)
            end = start.replace(day=monthrange(start.year, start.month)[1])
        else:
            start, end = pending_duration_bounds(self.request)
        if value in ('today', 'tomorrow', 'this_week', 'month'):
            queryset = queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                paid_amount__lt=F('total_fees'),
                next_payment_date__isnull=False,
            )
        if start:
            queryset = queryset.filter(next_payment_date__gte=start)
        if end:
            queryset = queryset.filter(next_payment_date__lte=end)
        return queryset

    def filter_date_from(self, queryset, name, value):
        return queryset.filter(next_payment_date__gte=value) if value else queryset

    def filter_date_to(self, queryset, name, value):
        return queryset.filter(next_payment_date__lte=value) if value else queryset

    def filter_due_this_week(self, queryset, name, value):
        if not value:
            return queryset
        today = timezone.localdate()
        week_start = today - timedelta(days=(today.weekday() + 1) % 7)
        week_end = week_start + timedelta(days=6)
        return queryset.filter(
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            paid_amount__lt=F('total_fees'),
            next_payment_date__gte=week_start,
            next_payment_date__lte=week_end,
        )


class PaymentReasonRequestViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentReasonRequestSerializer
    permission_classes = [IsStaffOrAdmin]
    pagination_class = None

    ACTIVE_STATUSES = [
        PaymentReasonRequest.Status.PENDING_RESPONSE,
        PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL,
    ]

    def get_queryset(self):
        queryset = PaymentReasonRequest.objects.select_related(
            'payment__enrollment__branch',
            'payment__enrollment__course',
            'admin_user',
            'branch_staff',
        ).prefetch_related('messages__sender').order_by('-created_at')
        user = self.request.user
        if user.is_super_admin:
            return queryset
        return queryset.filter(branch_staff=user)

    def _installment_item(self, payment, installment_index):
        for item in payment_installment_summary(payment):
            if int(item.get('index') or 0) == int(installment_index or 0):
                return item
        return None

    def _assigned_branch_staff(self, payment):
        enrollment = payment.enrollment
        candidates = [getattr(enrollment, 'enrolled_by', None), getattr(enrollment, 'created_by', None)]
        for user in candidates:
            if (
                user
                and user.is_active
                and not user.is_super_admin
                and user.branch_id == enrollment.branch_id
            ):
                return user
        return User.objects.filter(
            branch=enrollment.branch,
            is_active=True,
        ).exclude(role=User.Role.SUPER_ADMIN).order_by('first_name', 'last_name', 'username').first()

    def _serialize(self, reason_request, status_code=200):
        reason_request = PaymentReasonRequest.objects.select_related(
            'payment__enrollment__branch',
            'payment__enrollment__course',
            'admin_user',
            'branch_staff',
        ).prefetch_related('messages__sender').get(pk=reason_request.pk)
        return Response(PaymentReasonRequestSerializer(reason_request).data, status=status_code)

    def _message_role(self, user):
        return PaymentReasonMessage.SenderRole.ADMIN if user.is_super_admin else PaymentReasonMessage.SenderRole.USER

    def _append_message(self, reason_request, user, message, promised_payment_date=None, sender_role=None):
        return PaymentReasonMessage.objects.create(
            reason_request=reason_request,
            sender=user,
            sender_role=sender_role or self._message_role(user),
            message=str(message or '').strip(),
            status=reason_request.status,
            promised_payment_date=promised_payment_date,
        )

    def _notify_conversation_target(self, reason_request, sender):
        if sender.is_super_admin:
            create_user_notification(
                reason_request.branch_staff,
                'Payment Reason Conversation',
                'Admin sent a payment reason message.',
                Notification.NType.INFO,
                f'/payments?reason_request={reason_request.id}',
            )
            return
        create_user_notification(
            reason_request.admin_user,
            'Payment Reason Conversation',
            f'{sender.full_name or sender.username} replied in payment reason conversation.',
            Notification.NType.INFO,
            f'/payments?reason_request={reason_request.id}',
        )

    def create(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can request a pending payment reason.'}, status=403)
        payment = Payment.objects.select_related('enrollment__branch', 'enrollment__course').filter(
            pk=request.data.get('payment'),
        ).first()
        if not payment:
            return Response({'detail': 'Payment record not found.'}, status=404)
        if payment.status not in [Payment.Status.UNPAID, Payment.Status.PARTIAL] or payment.balance <= 0:
            return Response({'detail': 'Reason can be requested only for pending or partial payments.'}, status=400)
        try:
            installment_index = int(request.data.get('installment_index') or 0)
        except (TypeError, ValueError):
            installment_index = 0
        installment = self._installment_item(payment, installment_index)
        if not installment or Decimal(str(installment.get('pending_amount') or 0)) <= 0:
            return Response({'detail': 'Select a pending installment for the reason request.'}, status=400)
        existing = PaymentReasonRequest.objects.filter(
            payment=payment,
            installment_index=installment_index,
            status__in=self.ACTIVE_STATUSES,
        ).prefetch_related('messages__sender').first()
        if existing:
            message = str(request.data.get('message') or request.data.get('question') or '').strip()
            if message:
                self._append_message(existing, request.user, message)
                self._notify_conversation_target(existing, request.user)
            return self._serialize(existing, status_code=200)
        branch_staff = self._assigned_branch_staff(payment)
        if not branch_staff:
            return Response({'detail': 'No active branch staff user found for this payment.'}, status=400)
        reason_request = PaymentReasonRequest.objects.create(
            payment=payment,
            installment_index=installment_index,
            installment_label=installment.get('label') or f'{installment_index} Installment',
            installment_due_date=parse_date(installment.get('due_date')) if isinstance(installment.get('due_date'), str) else installment.get('due_date'),
            admin_user=request.user,
            branch_staff=branch_staff,
            question=str(request.data.get('message') or request.data.get('question') or 'Why is this payment still pending?').strip() or 'Why is this payment still pending?',
        )
        self._append_message(reason_request, request.user, reason_request.question)
        create_user_notification(
            branch_staff,
            'Payment Reason Requested',
            'Admin requested reason for pending payment.',
            Notification.NType.WARNING,
            f'/payments?reason_request={reason_request.id}',
        )
        return self._serialize(reason_request, status_code=201)

    @action(detail=True, methods=['post'], url_path='messages')
    def messages(self, request, pk=None):
        reason_request = self.get_object()
        if not (request.user.is_super_admin or reason_request.branch_staff_id == request.user.id):
            return Response({'detail': 'You do not have access to this conversation.'}, status=403)
        message = str(request.data.get('message') or request.data.get('staff_response') or '').strip()
        if not message:
            return Response({'message': 'Message is required.'}, status=400)

        promised_payment_date = None
        raw_promised_date = str(request.data.get('promised_payment_date') or '').strip()
        if raw_promised_date:
            promised_payment_date = parse_date(raw_promised_date)
            if not promised_payment_date:
                return Response({'promised_payment_date': 'Enter a valid promised payment date.'}, status=400)
            if promised_payment_date < timezone.localdate():
                return Response({'promised_payment_date': 'Promised Payment Date cannot be in the past.'}, status=400)

        update_fields = ['updated_at']
        if not request.user.is_super_admin:
            if reason_request.status == PaymentReasonRequest.Status.PENDING_RESPONSE and not promised_payment_date:
                return Response({'promised_payment_date': 'Promised Payment Date is required.'}, status=400)
            if reason_request.status == PaymentReasonRequest.Status.PENDING_RESPONSE:
                reason_request.status = PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL
                reason_request.responded_at = timezone.now()
                update_fields += ['status', 'responded_at']
            if not reason_request.staff_response:
                reason_request.staff_response = message
                update_fields.append('staff_response')
            if promised_payment_date:
                reason_request.promised_payment_date = promised_payment_date
                update_fields.append('promised_payment_date')
            reason_request.save(update_fields=list(dict.fromkeys(update_fields)))
            resolve_notifications(Notification.objects.filter(
                user=request.user,
                title='Payment Reason Requested',
                related_url=f'/payments?reason_request={reason_request.id}',
            ))
        else:
            reason_request.save(update_fields=['updated_at'])

        self._append_message(reason_request, request.user, message, promised_payment_date)
        self._notify_conversation_target(reason_request, request.user)
        return self._serialize(reason_request)

    @action(detail=True, methods=['post'], url_path='respond')
    def respond(self, request, pk=None):
        reason_request = self.get_object()
        if request.user.is_super_admin or reason_request.branch_staff_id != request.user.id:
            return Response({'detail': 'Only the assigned branch staff user can respond.'}, status=403)
        if reason_request.status != PaymentReasonRequest.Status.PENDING_RESPONSE:
            return Response({'detail': 'This request is not waiting for a staff response.'}, status=400)
        response_text = str(request.data.get('staff_response') or '').strip()
        promised_payment_date = parse_date(str(request.data.get('promised_payment_date') or '').strip())
        if not response_text:
            return Response({'staff_response': 'Response / Reason is required.'}, status=400)
        if not promised_payment_date:
            return Response({'promised_payment_date': 'Promised Payment Date is required.'}, status=400)
        if promised_payment_date < timezone.localdate():
            return Response({'promised_payment_date': 'Promised Payment Date cannot be in the past.'}, status=400)
        reason_request.staff_response = response_text
        reason_request.promised_payment_date = promised_payment_date
        reason_request.status = PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL
        reason_request.responded_at = timezone.now()
        reason_request.save(update_fields=[
            'staff_response', 'promised_payment_date', 'status', 'responded_at', 'updated_at',
        ])
        self._append_message(reason_request, request.user, response_text, promised_payment_date)
        resolve_notifications(Notification.objects.filter(
            user=request.user,
            title='Payment Reason Requested',
            related_url=f'/payments?reason_request={reason_request.id}',
        ))
        create_user_notification(
            reason_request.admin_user,
            'Payment Reason Response Submitted',
            f'{request.user.full_name} responded to pending payment reason request.',
            Notification.NType.INFO,
            f'/payments?reason_request={reason_request.id}',
        )
        return self._serialize(reason_request)

    @action(detail=True, methods=['post'], url_path='mark-resolved')
    def mark_resolved(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can mark payment reason requests resolved.'}, status=403)
        reason_request = self.get_object()
        if reason_request.status == PaymentReasonRequest.Status.RESOLVED:
            return Response(PaymentReasonRequestSerializer(reason_request).data)
        if reason_request.status not in [
            PaymentReasonRequest.Status.APPROVED,
            PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL,
        ]:
            return Response({'detail': 'Only approved or review-ready payment reason requests can be resolved.'}, status=400)
        reason_request.status = PaymentReasonRequest.Status.RESOLVED
        reason_request.resolved_at = timezone.now()
        reason_request.save(update_fields=['status', 'resolved_at', 'updated_at'])
        self._append_message(
            reason_request,
            request.user,
            str(request.data.get('message') or 'Marked resolved.').strip() or 'Marked resolved.',
        )
        mark_notifications_terminal(
            Notification.objects.filter(
                related_url=f'/payments?reason_request={reason_request.id}',
            ),
            Notification.Status.RESOLVED,
        )
        create_user_notification(
            reason_request.branch_staff,
            'Payment Reason Request Resolved',
            'Admin marked the payment reason request as resolved.',
            Notification.NType.SUCCESS,
            f'/payments?reason_request={reason_request.id}',
        )
        return self._serialize(reason_request)

    def _update_requested_due_date(self, reason_request):
        payment = reason_request.payment
        schedule = [
            dict(item)
            for item in (payment.manual_installment_schedule or get_payment_installment_schedule(payment))
        ]
        target_index = reason_request.installment_index - 1
        if target_index < 0 or target_index >= len(schedule):
            raise ValueError('Requested installment is no longer available.')
        schedule[target_index]['due_date'] = reason_request.promised_payment_date.isoformat()
        payment.manual_installment_schedule = schedule
        payment.save(update_fields=['manual_installment_schedule', 'paid_amount', 'status', 'next_payment_date', 'updated_at'])

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can approve payment reason responses.'}, status=403)
        reason_request = self.get_object()
        if reason_request.status != PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL:
            return Response({'detail': 'This request is not waiting for admin approval.'}, status=400)
        try:
            self._update_requested_due_date(reason_request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)
        reason_request.status = PaymentReasonRequest.Status.APPROVED
        reason_request.approved_at = timezone.now()
        reason_request.save(update_fields=['status', 'approved_at', 'updated_at'])
        self._append_message(
            reason_request,
            request.user,
            str(request.data.get('message') or 'Approved promise.').strip() or 'Approved promise.',
        )
        mark_notifications_terminal(
            Notification.objects.filter(
                user=reason_request.admin_user,
                title='Payment Reason Response Submitted',
                related_url=f'/payments?reason_request={reason_request.id}',
            ),
            Notification.Status.APPROVED,
        )
        return self._serialize(reason_request)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can reject payment reason responses.'}, status=403)
        reason_request = self.get_object()
        if reason_request.status != PaymentReasonRequest.Status.PENDING_ADMIN_APPROVAL:
            return Response({'detail': 'This request is not waiting for admin approval.'}, status=400)
        reason_request.status = PaymentReasonRequest.Status.REJECTED
        reason_request.rejected_at = timezone.now()
        reason_request.save(update_fields=['status', 'rejected_at', 'updated_at'])
        self._append_message(
            reason_request,
            request.user,
            str(request.data.get('message') or 'Rejected response.').strip() or 'Rejected response.',
        )
        mark_notifications_terminal(
            Notification.objects.filter(
                user=reason_request.admin_user,
                title='Payment Reason Response Submitted',
                related_url=f'/payments?reason_request={reason_request.id}',
            ),
            Notification.Status.REJECTED,
        )
        notification = create_user_notification(
            reason_request.branch_staff,
            'Payment Reason Response Rejected',
            'Your payment reason response was rejected by Admin.',
            Notification.NType.ERROR,
            f'/payments?reason_request={reason_request.id}',
        )
        if notification:
            notification.status = Notification.Status.REJECTED
            notification.is_read = True
            notification.resolved_at = timezone.now()
            notification.save(update_fields=['status', 'is_read', 'resolved_at'])
        return self._serialize(reason_request)


class PaymentViewSet(mixins.DestroyModelMixin, viewsets.ReadOnlyModelViewSet):
    """Read payment records (aggregate)."""
    permission_classes = [IsStaffOrAdmin]
    serializer_class   = PaymentSerializer
    filter_backends    = [DjangoFilterBackend, CRMSearchFilter]
    filterset_class    = PaymentFilter
    pagination_class   = None
    search_fields      = [
        'enrollment__name', 'enrollment__student_number',
        'enrollment__phone', 'enrollment__email', 'enrollment__source',
        'enrollment__course__name', 'enrollment__counselor__first_name',
        'enrollment__counselor__last_name', 'enrollment__counselor__username',
        'enrollment__enrolled_by__first_name',
        'enrollment__enrolled_by__last_name', 'enrollment__enrolled_by__username',
        'enrollment__created_by__first_name', 'enrollment__created_by__last_name',
        'enrollment__created_by__username',
    ]
    id_search_fields   = ['id', 'enrollment__id']

    def _month_bounds(self):
        raw_month = self.request.query_params.get('month')
        today = timezone.localdate()
        if raw_month:
            try:
                year, month = [int(part) for part in raw_month.split('-', 1)]
                start = today.replace(year=year, month=month, day=1)
            except (TypeError, ValueError):
                start = today.replace(day=1)
        else:
            start = today.replace(day=1)
        end = start.replace(day=monthrange(start.year, start.month)[1])
        return start, end

    def get_queryset(self):
        qs = visible_payment_queryset(Payment.objects.select_related(
            'enrollment__branch','enrollment__course','enrollment__counselor','enrollment__enrolled_by','enrollment__created_by'
        ).prefetch_related('installments').filter(enrollment__status__in=Enrollment.FINAL_STATUSES))
        if not self.request.user.is_super_admin:
            qs = qs.filter(enrollment__branch=self.request.user.branch)
        elif self.request.query_params.get('branch'):
            qs = qs.filter(enrollment__branch_id=self.request.query_params.get('branch'))
        has_date_filter = any(self.request.query_params.get(key) for key in ('duration', 'date_from', 'date_to', 'next_payment_from', 'next_payment_to'))
        if (
            self.action in ('list', 'export')
            and self.request.query_params.get('status') not in ('due', 'pending_today', 'weekly_pending')
            and not has_date_filter
        ):
            month_start, month_end = self._month_bounds()
            qs = qs.filter(
                Q(installments__payment_date__gte=month_start, installments__payment_date__lte=month_end)
                | Q(next_payment_date__gte=month_start, next_payment_date__lte=month_end)
                | Q(status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL], enrollment__enrollment_date__lte=month_end)
            ).distinct().order_by('enrollment__name', 'id')
        return qs

    def get_summary_queryset(self):
        qs = visible_payment_queryset(Payment.objects.select_related('enrollment__branch').filter(
            enrollment__status__in=Enrollment.FINAL_STATUSES,
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            paid_amount__lt=F('total_fees'),
        ))
        if not self.request.user.is_super_admin:
            qs = qs.filter(enrollment__branch=self.request.user.branch)
        elif self.request.query_params.get('branch'):
            qs = qs.filter(enrollment__branch_id=self.request.query_params.get('branch'))
        return qs

    def pending_amount_for_month(self, queryset, month_start, month_end):
        month_queryset = queryset.filter(
            next_payment_date__gte=month_start,
            next_payment_date__lte=month_end,
        ).distinct()
        return sum((payment.balance for payment in month_queryset), Decimal('0'))

    def pending_summary(self):
        today = timezone.localdate()
        this_month_start = today.replace(day=1)
        this_month_end = this_month_start.replace(day=monthrange(this_month_start.year, this_month_start.month)[1])
        last_month_end = this_month_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        next_month_seed = this_month_end + timedelta(days=1)
        next_month_start = next_month_seed.replace(day=1)
        next_month_end = next_month_start.replace(day=monthrange(next_month_start.year, next_month_start.month)[1])
        queryset = self.get_summary_queryset()
        return {
            'total_pending_amount': sum((payment.balance for payment in queryset), Decimal('0')),
            'this_month_pending': self.pending_amount_for_month(queryset, this_month_start, this_month_end),
            'last_month_pending': self.pending_amount_for_month(queryset, last_month_start, last_month_end),
            'next_month_pending': self.pending_amount_for_month(queryset, next_month_start, next_month_end),
        }

    def _summary_bounds(self):
        duration = self.request.query_params.get('duration') or ''
        if duration:
            start, end = pending_duration_bounds(self.request)
            today = timezone.localdate()
            if duration == 'overdue':
                return None, today - timedelta(days=1)
            return start, end
        if self.request.query_params.get('date_from') or self.request.query_params.get('date_to'):
            return date_range_from_request(self.request)
        return self._month_bounds()

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        month_start, month_end = self._summary_bounds()
        today = timezone.localdate()
        installment_queryset = PaymentInstallment.objects.filter(payment__in=queryset)
        if month_start:
            installment_queryset = installment_queryset.filter(payment_date__gte=month_start)
        if month_end:
            installment_queryset = installment_queryset.filter(payment_date__lte=month_end)
        collection = installment_queryset.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        partial_payments = installment_queryset.filter(document_type=PaymentInstallment.DocumentType.RECEIPT).count()
        completed_installments = installment_queryset.filter(document_type=PaymentInstallment.DocumentType.BILL).count()
        pending = sum((payment.balance for payment in queryset), Decimal('0'))
        due_queryset = queryset.filter(
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            next_payment_date__isnull=False,
        )
        if month_start:
            due_queryset = due_queryset.filter(next_payment_date__gte=month_start)
        if month_end:
            due_queryset = due_queryset.filter(next_payment_date__lte=month_end)
        total_due = sum((payment.balance for payment in due_queryset), Decimal('0'))
        range_includes_today = (month_start is None or month_start <= today) and (month_end is None or today <= month_end)
        if range_includes_today:
            upcoming = due_queryset.filter(next_payment_date__gte=today).count()
            overdue = queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                next_payment_date__lt=today,
            ).count()
        else:
            upcoming = due_queryset.count()
            overdue = queryset.filter(
                status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
                **({'next_payment_date__lte': month_end} if month_end else {}),
            ).count()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'results': serializer.data,
            'summary': {
                'total_records': queryset.count(),
                'total_collection': collection,
                'total_due': total_due,
                'total_balance': pending,
                'pending_amount': pending,
                'partial_payments': partial_payments,
                'completed_installments': completed_installments,
                'upcoming_payments': upcoming,
                'overdue_payments': overdue,
                **self.pending_summary(),
            },
            'filters': {
                'month': month_start.strftime('%Y-%m') if month_start else '',
                'month_start': month_start,
                'month_end': month_end,
            },
        })

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can delete payments.'}, status=status.HTTP_403_FORBIDDEN)
        payment = self.get_object()
        has_generated_documents = payment.installments.filter(
            Q(receipt_number__gt='') | Q(bill_number__gt='') | Q(bill_generated_at__isnull=False)
        ).exists()
        if has_generated_documents:
            return Response(
                {'detail': 'This payment has generated documents and cannot be deleted. Use Void/Cancel instead.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        enrollment = payment.enrollment
        logger.warning(
            'Admin %s deleted payment %s for enrollment %s. Reason: %s',
            request.user.id,
            payment.id,
            enrollment.id,
            str(request.data.get('reason') or 'Admin confirmed permanent delete.').strip(),
        )
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'payments', 'payments-export')

    def _worksheet_export_rows(self, queryset):
        max_installments = max([payment.installments.count() for payment in queryset] or [3])
        headers = [
            'Student Name', 'Student ID', 'Course', 'Total Fees', 'First Class Date',
            '1st Payment Amount', '1st Payment Date',
            '2nd Payment Amount', '2nd Payment Date',
            '3rd Payment Amount', '3rd Payment Date',
            'Balance Amount', 'Payment Status',
        ]
        for index in range(4, max_installments + 1):
            headers.insert(-2, f'{index}th Payment Amount')
            headers.insert(-2, f'{index}th Payment Date')
        rows = []
        for payment in queryset:
            installments = list(payment.installments.all().order_by('payment_date', 'id'))
            row = [
                payment.enrollment.name,
                payment.enrollment.student_number,
                payment.enrollment.course.name if payment.enrollment.course else '',
                payment.total_fees,
                payment.enrollment.start_date,
            ]
            for index in range(max_installments):
                installment = installments[index] if index < len(installments) else None
                row.extend([
                    installment.amount if installment else '',
                    installment.payment_date if installment else '',
                ])
            row.extend([payment.balance, payment.get_status_display()])
            rows.append(row)
        return headers, rows

    def _export_csv(self):
        import csv
        queryset = self.filter_queryset(self.get_queryset())
        headers, rows = self._worksheet_export_rows(queryset)
        month_start, _ = self._month_bounds()
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="payment-worksheet-{month_start:%Y-%m}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    @action(detail=True, methods=['post'], url_path='send-reminder')
    def send_reminder(self, request, pk=None):
        payment = self.get_object()
        if payment.status == Payment.Status.PAID or payment.balance <= 0:
            return Response({'detail': 'This payment has no pending balance.'}, status=400)
        enrollment = payment.enrollment
        template_id = request.data.get('template_id')
        template = None
        if template_id:
            template = WhatsAppTemplate.objects.filter(
                pk=template_id,
                template_type=WhatsAppTemplate.TemplateType.PAYMENT_REMINDER,
                is_active=True,
            ).first()
        template = template or active_whatsapp_template(WhatsAppTemplate.TemplateType.PAYMENT_REMINDER)
        default_message = (
            'Hi {{student_name}},\n\n'
            'This is a gentle reminder regarding your pending course fee payment.\n\n'
            'Total Fee: {{total_fee}}\n'
            'Paid: {{paid_amount}}\n'
            'Balance: {{pending_amount}}\n\n'
            'Kindly complete the pending payment at your earliest convenience.\n\n'
            '-Team IIE'
        )
        values = {
            'student_name': enrollment.name,
            'candidate_name': enrollment.name,
            'course_name': enrollment.course.name if enrollment.course else '',
            'branch_name': enrollment.branch.name if enrollment.branch else '',
            'phone_number': enrollment.phone,
            'total_fee': whatsapp_currency(payment.total_fees),
            'paid_amount': whatsapp_currency(payment.paid_amount),
            'pending_amount': whatsapp_currency(payment.balance),
            'next_payment_date': whatsapp_date(payment.next_payment_date),
            'institute_name': 'IIE',
        }
        body = template.message_body if template else default_message
        log = send_candidate_message(
            candidate_name=enrollment.name,
            phone=enrollment.phone,
            message_type=WhatsAppMessage.MsgType.PAYMENT_REMINDER,
            message_body=body,
            template=template,
            values=values,
            sent_by=request.user,
            related_model='payment',
            related_id=payment.id,
            dedupe=False,
        )
        return Response({
            'detail': 'Payment reminder processed.',
            'phone': enrollment.phone,
            'whatsapp_message': render_whatsapp_template(body, values),
            **whatsapp_send_payload(log),
        })

    @action(detail=True, methods=['post'], url_path='update-schedule')
    def update_schedule(self, request, pk=None):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can override payment schedules.'}, status=403)
        payment = self.get_object()
        schedule = request.data.get('payment_schedule')
        if not isinstance(schedule, list) or not schedule:
            return Response({'detail': 'Payment schedule is required.'}, status=400)
        cleaned = []
        schedule_total = Decimal('0')
        expected_total = Decimal(str(payment.total_fees or 0)).quantize(Decimal('0.01'))
        if Decimal('0') < expected_total <= Decimal(str(LOW_FEE_SINGLE_PAYMENT_MAX_COURSE_FEE)):
            raw_due_date = schedule[0].get('due_date') if isinstance(schedule[0], dict) else None
            parsed_due_date = parse_date(str(raw_due_date)) if raw_due_date else payment.enrollment.enrollment_date
            if not parsed_due_date:
                return Response({'detail': 'Each installment needs a valid due date.'}, status=400)
            cleaned = [{
                'label': 'Single Payment',
                'amount': decimal_to_schedule_amount(expected_total),
                'due_date': parsed_due_date.isoformat(),
            }]
            with transaction.atomic():
                payment = (
                    Payment.objects
                    .select_for_update()
                    .select_related('enrollment')
                    .get(pk=payment.pk)
                )
                payment.enrollment.payment_schedule = cleaned
                payment.enrollment.payment_schedule_locked = True
                payment.enrollment.payment_schedule_finalized_at = payment.enrollment.payment_schedule_finalized_at or timezone.now()
                payment.enrollment.save(update_fields=[
                    'payment_schedule', 'payment_schedule_locked',
                    'payment_schedule_finalized_at', 'updated_at',
                ])
                payment.paid_amount = payment.installments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                payment.manual_installment_schedule = cleaned
                payment.save(update_fields=['manual_installment_schedule', 'paid_amount', 'status', 'next_payment_date', 'updated_at'])
            resolve_payment_due_notifications_if_inactive(payment)
            return Response(PaymentSerializer(payment, context={'request': request}).data)
        summary = payment_installment_summary(payment)
        locked_rows = {
            item['index']: item
            for item in summary
            if Decimal(str(item.get('paid_amount') or 0)) > 0
        }
        for index, item in enumerate(schedule, start=1):
            amount = item.get('amount')
            due_date = item.get('due_date')
            if amount in (None, '') or not due_date:
                return Response({'detail': 'Each installment needs amount and due date.'}, status=400)
            try:
                amount = Decimal(str(amount))
            except (InvalidOperation, TypeError, ValueError):
                return Response({'detail': 'Installment amount must be numeric.'}, status=400)
            if not amount.is_finite():
                return Response({'detail': 'Installment amount must be numeric.'}, status=400)
            if amount <= 0:
                return Response({'detail': 'Installment amount must be greater than zero.'}, status=400)
            amount = amount.quantize(Decimal('0.01'))
            if index == 1 and amount != Decimal(str(ENROLLMENT_PAYMENT_AMOUNT)):
                return Response({'detail': 'Enrollment Fee must be Rs 5,000.'}, status=400)
            if index > 1 and index < len(schedule) and amount < Decimal(str(MIN_INSTALLMENT_AMOUNT)):
                return Response({'detail': 'Each installment must be Rs 5,000 or above.'}, status=400)
            if index > 1 and index == len(schedule) and amount < Decimal(str(MIN_INSTALLMENT_AMOUNT)):
                if amount != expected_total - schedule_total:
                    return Response({'detail': 'Each installment must be Rs 5,000 or above.'}, status=400)
            parsed_due_date = parse_date(str(due_date))
            if not parsed_due_date:
                return Response({'detail': 'Each installment needs a valid due date.'}, status=400)
            schedule_total += amount
            cleaned_row = {
                'label': item.get('label') or ('Enrollment' if index == 1 else f'{index - 1} Installment'),
                'amount': int(amount) if amount == amount.to_integral_value() else float(amount),
                'due_date': parsed_due_date.isoformat(),
            }
            locked = locked_rows.get(index)
            if locked:
                current = get_payment_installment_schedule(payment)[index - 1]
                current_due_date = current.get('due_date')
                current_due_date = current_due_date.isoformat() if hasattr(current_due_date, 'isoformat') else current_due_date
                if (
                    Decimal(str(current.get('amount') or 0)).quantize(Decimal('0.01')) != amount
                    or str(current.get('label') or '') != str(cleaned_row['label'])
                    or str(current_due_date or '') != str(cleaned_row['due_date'] or '')
                ):
                    return Response({'detail': 'Existing paid installments cannot be modified.'}, status=400)
            cleaned.append(cleaned_row)
        if schedule_total != expected_total:
            return Response({'detail': 'Payment schedule total must match course fee.'}, status=400)
        with transaction.atomic():
            payment = (
                Payment.objects
                .select_for_update()
                .select_related('enrollment')
                .get(pk=payment.pk)
            )
            payment.enrollment.payment_schedule = cleaned
            payment.enrollment.payment_schedule_locked = True
            payment.enrollment.payment_schedule_finalized_at = payment.enrollment.payment_schedule_finalized_at or timezone.now()
            payment.enrollment.save(update_fields=[
                'payment_schedule', 'payment_schedule_locked',
                'payment_schedule_finalized_at', 'updated_at',
            ])
            payment.paid_amount = payment.installments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            payment.manual_installment_schedule = cleaned
            payment.save(update_fields=['manual_installment_schedule', 'paid_amount', 'status', 'next_payment_date', 'updated_at'])
        resolve_payment_due_notifications_if_inactive(payment)
        return Response(PaymentSerializer(payment, context={'request': request}).data)


class PaymentInstallmentViewSet(viewsets.ModelViewSet):
    """Add/view payment installments."""
    permission_classes = [IsStaffOrAdmin]
    serializer_class   = PaymentInstallmentSerializer
    pagination_class   = None
    filterset_fields   = ['enrollment', 'payment_mode', 'payment_date']

    def get_queryset(self):
        qs = PaymentInstallment.objects.select_related(
            'payment',
            'enrollment__branch',
            'enrollment__course',
            'collected_by',
            'bill_generated_by',
        )
        if not self.request.user.is_super_admin:
            qs = qs.filter(enrollment__branch=self.request.user.branch)
        return qs

    def create(self, request, *args, **kwargs):
        payment_id = request.data.get('payment')
        enrollment_id = request.data.get('enrollment')
        if not payment_id:
            return Response({'detail': 'Payment record is required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            payment = Payment.objects.select_related('enrollment__branch').get(pk=payment_id)
        except Payment.DoesNotExist:
            return Response({'detail': 'Payment record not found.'}, status=404)
        if not request.user.is_super_admin and payment.enrollment.branch_id != request.user.branch_id:
            return Response({'detail': 'You do not have access to this payment.'}, status=403)
        if enrollment_id and str(payment.enrollment_id) != str(enrollment_id):
            return Response(
                {'detail': 'Installment payment and enrollment do not match.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            amount = Decimal(str(request.data.get('amount') or '0'))
        except (TypeError, ValueError):
            return Response({'detail': 'Payment amount must be numeric.'}, status=status.HTTP_400_BAD_REQUEST)
        if amount <= 0:
            return Response({'detail': 'Payment amount must be greater than zero.'}, status=status.HTTP_400_BAD_REQUEST)

        if amount > payment.balance:
            return Response({'detail': 'Payment amount cannot exceed the pending total fee balance.'}, status=status.HTTP_400_BAD_REQUEST)

        allocation = self._payment_allocation(payment, amount)
        if not allocation:
            return Response({'detail': 'All installments are already completed.'}, status=status.HTTP_400_BAD_REQUEST)
        active = allocation[0]
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        save_kwargs = {
            'payment': payment,
            'enrollment': payment.enrollment,
            'collected_by': request.user,
            'installment_index': active['index'],
            'installment_label': self._allocation_label(allocation),
            'document_type': PaymentInstallment.DocumentType.RECEIPT,
        }
        self.perform_create(serializer, save_kwargs)
        payment.refresh_from_db()
        resolve_payment_due_notifications_if_inactive(payment)
        headers = self.get_success_headers(serializer.data)
        message = 'Payment entry saved and marked Pending Approval.'
        self._notify_admins_payment_added(serializer.instance)
        return Response(
            {
                'detail': message,
                'installment': serializer.data,
                'allocation': allocation,
            },
            status=status.HTTP_201_CREATED,
            headers=headers,
        )

    def perform_create(self, serializer, save_kwargs=None):
        serializer.save(**(save_kwargs or {'collected_by': self.request.user}))

    def _notify_admins_payment_added(self, installment):
        creator_name = notification_actor_name(getattr(installment, 'collected_by', None))
        title = f'Installment Added by {creator_name}'
        message = f'{installment.enrollment.name} payment installment added and awaiting approval.'
        for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
            create_user_notification(
                admin_user,
                title,
                message,
                Notification.NType.INFO,
                f'/payments/{installment.payment_id}',
                category=Notification.Category.APPROVAL,
            )

    def _notify_admins_installment_changed(self, installment, title):
        message = f'{installment.enrollment.name} payment installment was {title.replace("Installment ", "").replace(" by User", "").lower()} by {self.request.user.full_name or self.request.user.username}.'
        for admin_user in User.objects.filter(role=User.Role.SUPER_ADMIN, is_active=True):
            create_user_notification(
                admin_user,
                title,
                message,
                Notification.NType.WARNING,
                f'/payments/{installment.payment_id}',
                category=Notification.Category.APPROVAL,
            )

    def _notify_collector_document_generated(self, installment):
        if not installment.collected_by_id:
            return
        document_label = 'Bill' if installment.bill_number else 'Receipt'
        create_notification_once(
            installment.collected_by,
            f'{document_label} Generated',
            f'{document_label} generated for {installment.enrollment.name}.',
            Notification.NType.SUCCESS,
            f'/payments/{installment.payment_id}',
        )

    def update(self, request, *args, **kwargs):
        installment = self.get_object()
        if self._document_number(installment) or installment.bill_generated_at:
            return Response({'detail': 'Generated payment documents cannot be edited.'}, status=400)
        response = super().update(request, *args, **kwargs)
        if not request.user.is_super_admin:
            installment.refresh_from_db()
            self._notify_admins_installment_changed(installment, 'Installment Edited by User')
        installment.payment.refresh_from_db()
        resolve_payment_due_notifications_if_inactive(installment.payment)
        return response

    def destroy(self, request, *args, **kwargs):
        installment = self.get_object()
        if self._document_number(installment) or installment.bill_generated_at:
            return Response({'detail': 'Generated payment documents cannot be deleted.'}, status=400)
        payment_id = installment.payment_id
        if not request.user.is_super_admin:
            self._notify_admins_installment_changed(installment, 'Installment Deleted by User')
        response = super().destroy(request, *args, **kwargs)
        payment = Payment.objects.filter(pk=payment_id).first()
        if payment:
            payment.recalculate_from_installments()
            resolve_payment_due_notifications_if_inactive(payment)
        return response

    def _active_installment(self, payment):
        summary = payment_installment_summary(payment)
        if not summary:
            return {
                'index': 1,
                'label': '1 Installment',
                'required_amount': payment.total_fees,
                'paid_amount': Decimal('0'),
                'pending_amount': payment.total_fees,
                'status': 'pending',
            }
        for item in summary:
            if item['status'] != 'paid':
                return item
        return summary[-1]

    def _payment_allocation(self, payment, amount):
        remaining = Decimal(str(amount or 0))
        allocation = []
        for item in payment_installment_summary(payment):
            if remaining <= 0:
                break
            pending_amount = Decimal(str(item.get('pending_amount') or 0))
            if pending_amount <= 0:
                continue
            allocated_amount = min(remaining, pending_amount)
            allocation.append({
                'index': item['index'],
                'label': item['label'],
                'amount': allocated_amount,
            })
            remaining -= allocated_amount
        return allocation

    def _allocation_label(self, allocation):
        labels = [str(item.get('label') or '').strip() for item in allocation if item.get('label')]
        if not labels:
            return ''
        if len(labels) == 1:
            return labels[0]
        return ' + '.join(labels)

    def _document_number(self, installment):
        if installment.bill_number:
            return installment.bill_number
        return installment.receipt_number

    def _public_bill_path(self, installment):
        document_number = self._document_number(installment)
        return f'/public/bills/{installment.id}/{document_number}/'

    def _public_bill_url(self, request, installment):
        return request.build_absolute_uri(self._public_bill_path(installment))

    def _stored_bill_pdf_path(self, installment):
        document_number = self._document_number(installment) or f'bill-{installment.id}'
        safe_document_number = re.sub(r'[^A-Za-z0-9_.-]', '-', str(document_number)).strip('-') or f'bill-{installment.id}'
        return f'bills/{installment.id}/{safe_document_number}.pdf'

    def _store_bill_pdf(self, request, installment, pdf_bytes):
        storage_path = self._stored_bill_pdf_path(installment)
        if default_storage.exists(storage_path):
            default_storage.delete(storage_path)
        saved_path = default_storage.save(storage_path, ContentFile(pdf_bytes))
        return request.build_absolute_uri(default_storage.url(saved_path))

    def _stored_bill_image_path(self, installment):
        document_number = self._document_number(installment) or f'bill-{installment.id}'
        safe_document_number = re.sub(r'[^A-Za-z0-9_.-]', '-', str(document_number)).strip('-') or f'bill-{installment.id}'
        return f'bills/{installment.id}/{safe_document_number}.png'

    def _store_bill_image(self, request, installment, image_bytes):
        storage_path = self._stored_bill_image_path(installment)
        if default_storage.exists(storage_path):
            default_storage.delete(storage_path)
        saved_path = default_storage.save(storage_path, ContentFile(image_bytes))
        return request.build_absolute_uri(default_storage.url(saved_path))

    def _bill_whatsapp_message(self, enrollment, installment):
        snapshot = installment.document_snapshot or {}
        course_name = snapshot.get('course_name') or (enrollment.course.name if enrollment.course else '')
        installment_name = snapshot.get('installment_label') or installment.installment_label or str(installment.installment_index)
        paid_amount = Decimal(str(snapshot.get('payment_amount') or installment.amount or 0))
        return (
            f'Dear {enrollment.name},\n\n'
            'Your payment receipt has been generated.\n\n'
            f'Course: {course_name}\n'
            f'Installment: {installment_name}\n'
            f'Paid Amount: \u20b9{paid_amount:,.2f}\n\n'
            '-Team IIE'
        )

    def _bill_api_delivery_enabled(self):
        return bool(
            getattr(settings, 'WHATSCHIMP_ENABLED', False)
            or getattr(settings, 'BILL_WHATSAPP_API_ENABLED', False)
        )

    def _receipt_date(self, value):
        if not value:
            return 'Not set'
        if isinstance(value, str):
            original = value
            value = parse_date(value[:10])
            if not value:
                return original
        if hasattr(value, 'date') and not hasattr(value, 'day'):
            value = value.date()
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y')
        return str(value)

    def _snapshot_decimal(self, value):
        return f'{Decimal(str(value or 0)):.2f}'

    def _snapshot_schedule_date(self, value):
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        return value or ''

    def _default_bill_address_lines(self):
        return ['Branch address not set']

    def _branch_address_lines(self, branch):
        if not branch:
            return self._default_bill_address_lines()

        lines = [line.strip() for line in str(branch.address or '').splitlines() if line.strip()]
        city_line_parts = [
            str(branch.city or '').strip(),
            str(branch.state or '').strip(),
            str(branch.pincode or '').strip(),
        ]
        city_line = ', '.join([part for part in city_line_parts if part])
        if city_line:
            lines.append(city_line)
        return lines or self._default_bill_address_lines()

    def _branch_header_payload(self, branch):
        return {
            'branch_name': branch.name if branch and branch.name else 'Branch not set',
            'branch_address_lines': self._branch_address_lines(branch),
            'branch_phone': branch.phone if branch and branch.phone else 'Phone number not set',
        }

    def _installment_summary_for_paid_amount(self, payment, paid_amount):
        remaining_paid = Decimal(str(paid_amount or 0))
        summary = []
        for index, item in enumerate(get_payment_installment_schedule(payment), start=1):
            required_amount = Decimal(str(item.get('amount') or 0))
            item_paid_amount = min(remaining_paid, required_amount)
            remaining_paid = max(remaining_paid - item_paid_amount, Decimal('0'))
            pending_amount = max(required_amount - item_paid_amount, Decimal('0'))
            if item_paid_amount >= required_amount and required_amount > 0:
                item_status = 'paid'
            elif item_paid_amount > 0:
                item_status = 'partial'
            else:
                item_status = 'pending'
            due_date = item.get('due_date')
            summary.append({
                'index': index,
                'label': item.get('label') or f'{index} Installment',
                'required_amount': required_amount,
                'paid_amount': item_paid_amount,
                'pending_amount': pending_amount,
                'status': item_status,
                'due_date': due_date.isoformat() if hasattr(due_date, 'isoformat') else due_date,
            })
        return summary

    def _next_payment_schedule(self, payment, summary=None):
        for item in summary or payment_installment_summary(payment):
            pending_amount = Decimal(str(item.get('pending_amount') or 0))
            if pending_amount > 0:
                return {
                    'label': item.get('label') or f"{item.get('index') or ''} Installment".strip(),
                    'due_date': self._receipt_date(item.get('due_date')),
                    'pending_amount': self._snapshot_decimal(pending_amount),
                }
        return None

    def _build_document_snapshot(self, installment):
        enrollment = installment.enrollment
        payment = installment.payment
        branch_header = self._branch_header_payload(enrollment.branch)
        is_bill = bool(installment.bill_number) or installment.document_type == PaymentInstallment.DocumentType.BILL
        generated_at = installment.bill_generated_at or timezone.now()
        cutoff = generated_at if is_bill else timezone.now()
        historical_installments = payment.installments.filter(created_at__lte=cutoff)
        current_payment_amount = Decimal(str(installment.amount or 0))
        paid_amount = historical_installments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        paid_amount = Decimal(str(paid_amount or 0))
        if installment.pk and not historical_installments.filter(pk=installment.pk).exists():
            paid_amount += current_payment_amount
        total_fees = Decimal(str(payment.total_fees or 0))
        balance = max(total_fees - paid_amount, Decimal('0'))
        included = [installment] if is_bill else [installment]
        installment_summary = self._installment_summary_for_paid_amount(payment, paid_amount)
        next_schedule = self._next_payment_schedule(payment, installment_summary)
        return {
            'calculation_version': 'bill_amount_v2',
            'document_type': 'bill' if is_bill else 'receipt',
            'document_number': self._document_number(installment),
            'generated_at': generated_at.isoformat() if generated_at else '',
            'generated_by': installment.bill_generated_by.full_name if installment.bill_generated_by_id and installment.bill_generated_by else '',
            'student_name': enrollment.name,
            'student_number': enrollment.student_number or '',
            'course_name': enrollment.course.name if enrollment.course else '',
            'branch_name': branch_header['branch_name'],
            'branch_address_lines': branch_header['branch_address_lines'],
            'branch_phone': branch_header['branch_phone'],
            'payment_mode': installment.get_payment_mode_display(),
            'reference_number': installment.reference_number or 'Not provided',
            'payment_date': self._receipt_date(installment.payment_date),
            'installment_label': installment.installment_label or str(installment.installment_index),
            'payment_amount': self._snapshot_decimal(current_payment_amount),
            'total_fees': self._snapshot_decimal(total_fees),
            'paid_amount': self._snapshot_decimal(paid_amount),
            'balance': self._snapshot_decimal(balance),
            'pending_amount': self._snapshot_decimal(balance),
            'next_payment_amount': next_schedule.get('pending_amount') if next_schedule else self._snapshot_decimal(0),
            'next_payment_date': next_schedule.get('due_date') if next_schedule else '',
            'bill_total': self._snapshot_decimal(sum((item.amount for item in included), Decimal('0'))),
            'payment_schedule': [
                {
                    'label': item.get('label') or f"{item.get('index') or ''} Installment".strip(),
                    'due_date': self._snapshot_schedule_date(item.get('due_date')),
                    'amount': self._snapshot_decimal(item.get('required_amount') or 0),
                    'paid_amount': self._snapshot_decimal(item.get('paid_amount') or 0),
                    'pending_amount': self._snapshot_decimal(item.get('pending_amount') or 0),
                    'status': str(item.get('status') or 'pending').replace('_', ' ').title(),
                }
                for item in installment_summary
            ],
            'included_installments': [
                {
                    'id': item.id,
                    'label': item.installment_label or str(item.installment_index),
                    'payment_date': self._receipt_date(item.payment_date),
                    'amount': self._snapshot_decimal(item.amount),
                    'payment_mode': item.get_payment_mode_display(),
                    'reference_number': item.reference_number or 'Not provided',
                }
                for item in included
            ],
            'next_payment_schedule': next_schedule,
        }

    def _build_bill_html(self, installment, snapshot=None):
        snapshot = snapshot or {}
        enrollment = installment.enrollment
        payment = installment.payment
        branch = enrollment.branch
        branch_header = self._branch_header_payload(branch)
        branch_address_lines = snapshot.get('branch_address_lines') or branch_header['branch_address_lines']
        branch_phone = snapshot.get('branch_phone') or branch_header['branch_phone']
        branch_address_html = ''.join(f'<p>{escape(str(line))}</p>' for line in branch_address_lines)
        branch_phone_html = f'<p>{escape(str(branch_phone))}</p>' if branch_phone else ''
        schedule_rows = []
        schedule = snapshot.get('payment_schedule') or []
        if not schedule:
            schedule = [
                {
                    'label': item.get('label') or 'Installment',
                    'due_date': self._receipt_date(item.get('due_date')),
                    'amount': self._snapshot_decimal(item.get('amount') or 0),
                    'status': 'Upcoming',
                }
                for item in get_payment_installment_schedule(payment)
            ]
        for item in schedule:
            label = item.get('label') or 'Installment'
            due_date_display = self._receipt_date(item.get('due_date')) if item.get('due_date') else 'Not set'
            amount = Decimal(str(item.get('amount') or 0))
            row_status = item.get('status') or 'Upcoming'
            status_class = 'paid' if row_status == 'Paid' else 'upcoming'
            schedule_rows.append(
                f"""
                <tr>
                  <td>{escape(str(label))}</td>
                  <td>{escape(due_date_display)}</td>
                  <td class="amount">Rs {amount:,.2f}</td>
                  <td class="status-cell"><span class="badge {status_class}">{escape(row_status)}</span></td>
                </tr>
                """
            )
        logo_src = ''
        logo_candidates = [
            Path(settings.BASE_DIR) / 'frontend' / 'public' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'brand-logo.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'brand-logo.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'receipt-logo.png',
        ]
        for logo_path in logo_candidates:
            if logo_path.exists():
                logo_src = f"data:image/png;base64,{base64.b64encode(logo_path.read_bytes()).decode('ascii')}"
                break
        is_bill = bool(installment.bill_number) or installment.document_type == PaymentInstallment.DocumentType.BILL
        document_title = 'Official Payment Bill' if is_bill else 'Payment Receipt'
        document_number_label = 'Bill No' if is_bill else 'Receipt No'
        document_number = snapshot.get('document_number') or self._document_number(installment)
        partial_note = '' if is_bill else '<div class="partial-note">Partial Payment Received</div>'
        student_name = snapshot.get('student_name') or enrollment.name
        student_number = snapshot.get('student_number') or enrollment.student_number or ''
        course_name = snapshot.get('course_name') or (enrollment.course.name if enrollment.course else '')
        branch_name = snapshot.get('branch_name') or branch_header['branch_name']
        payment_mode = snapshot.get('payment_mode') or installment.get_payment_mode_display()
        reference_number = snapshot.get('reference_number') or installment.reference_number or 'Not provided'
        payment_date = snapshot.get('payment_date') or self._receipt_date(installment.payment_date)
        installment_label = snapshot.get('installment_label') or installment.installment_label or str(installment.installment_index)
        payment_amount = Decimal(str(snapshot.get('payment_amount') or installment.amount or 0))
        total_fees = Decimal(str(snapshot.get('total_fees') or payment.total_fees or 0))
        paid_amount = Decimal(str(snapshot.get('paid_amount') or payment.paid_amount or 0))
        balance = Decimal(str(snapshot.get('balance') or payment.balance or 0))
        generated_at = snapshot.get('generated_at') or installment.bill_generated_at
        return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>{escape(document_number or document_title)}</title>
    <style>
      * {{ box-sizing: border-box; }}
      body {{ font-family: Libertine, "Linux Libertine", "Libertinus Serif", Georgia, "Times New Roman", serif; color: #111827; margin: 14px; background: #F8FAFC; }}
      .sheet {{ max-width: 780px; margin: 0 auto; border: 1px solid #CBD5E1; background: white; box-shadow: 0 10px 30px rgba(15, 23, 42, 0.08); }}
      .header {{ position: relative; display: flex; align-items: center; justify-content: center; min-height: 126px; padding: 12px 24px; background: #1E3A5F; color: white; }}
      .logo {{ position: absolute; left: 22px; top: 18px; display: flex; align-items: center; justify-content: flex-start; }}
      .logo img {{ width: auto; height: 48px; object-fit: contain; display: block; }}
      .brand {{ width: 100%; min-width: 0; padding: 0 86px; text-align: center; }}
      .brand h1 {{ margin: 0 0 4px; font-size: 20px; font-weight: 800; letter-spacing: 0.08em; text-transform: uppercase; color: white; line-height: 1.15; }}
      .brand .tagline {{ margin: 0 0 7px; font-size: 12.5px; font-weight: 600; color: #F8FAFC; line-height: 1.2; }}
      .brand .address {{ margin: 0; }}
      .brand .address p {{ margin: 1px 0; font-size: 10.8px; line-height: 1.24; color: #F8FAFC; }}
      .receipt-bar {{ display: flex; align-items: center; justify-content: space-between; gap: 16px; min-height: 38px; padding: 9px 18px; border-bottom: 1px solid #CBD5E1; background: #ffffff; }}
      .receipt-title {{ font-size: 14px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.14em; color: #111827; line-height: 1.1; }}
      .receipt-no {{ font-size: 12px; font-weight: 800; color: #1E3A5F; line-height: 1.1; }}
      .partial-note {{ padding: 7px 18px; border-bottom: 1px solid #CBD5E1; background: #FFF7ED; color: #9A3412; font-size: 12px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.08em; }}
      .section {{ padding: 12px 18px; border-bottom: 1px solid #CBD5E1; background: #ffffff; }}
      .grid {{ display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); column-gap: 24px; row-gap: 9px; }}
      .field {{ display: grid; grid-template-columns: 150px minmax(0, 220px); gap: 10px; align-items: baseline; min-height: 22px; }}
      .label {{ font-size: 9.8px; text-transform: uppercase; letter-spacing: 0.07em; color: #334155; line-height: 1.2; }}
      .value {{ font-size: 12.2px; font-weight: 800; color: #111827; line-height: 1.25; text-align: left; }}
      .field .amount {{ text-align: left; color: #111827; }}
      .section h2 {{ margin: 0 0 8px; font-size: 12.5px; text-transform: uppercase; letter-spacing: 0.12em; color: #111827; }}
      table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
      th, td {{ padding: 6px 9px; border: 1px solid #CBD5E1; text-align: left; vertical-align: middle; }}
      th {{ background: #F8FAFC; color: #334155; text-transform: uppercase; letter-spacing: 0.08em; font-size: 9.5px; }}
      tr:nth-child(even) td {{ background: #F8FAFC; }}
      table .amount {{ text-align: right; font-weight: 800; color: #1E3A5F; }}
      td.status-cell {{ text-align: center; }}
      .badge {{ display: inline-block; min-width: 70px; border-radius: 999px; padding: 2px 8px; text-align: center; font-size: 9.5px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.06em; border: 1px solid #CBD5E1; line-height: 1.25; }}
      .badge.paid {{ background: #DCFCE7; color: #334155; border-color: #86EFAC; }}
      .badge.upcoming {{ background: #FFF7ED; color: #334155; border-color: #FED7AA; }}
      .generated-info {{ padding: 18px 24px; background: white; color: #334155; font-size: 11.5px; line-height: 1.8; }}
      .generated-info p {{ margin: 0; }}
      .generated-info .info-label {{ font-weight: 700; }}
      .generated-info .info-value {{ font-weight: 400; color: #111827; }}
      .bottom {{ display: flex; align-items: center; justify-content: space-between; gap: 24px; min-height: 52px; margin-top: 8px; padding: 10px 24px; background: #1E3A5F; border-top: 1px solid #d6dce5; color: white; font-size: 13px; line-height: 1.35; }}
      @media print {{
        body {{ margin: 0; background: white; }}
        .sheet {{ border: 1px solid #CBD5E1; max-width: none; box-shadow: none; }}
        .header {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
        .bottom, th, tr:nth-child(even) td {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
      }}
    </style>
  </head>
  <body>
    <div class="sheet">
      <div class="header">
        <div class="logo">
          {'<img src="' + logo_src + '" alt="Indra Institute of Education logo">' if logo_src else ''}
        </div>
        <div class="brand">
          <h1>INDRA INSTITUTE OF EDUCATION</h1>
          <div class="tagline">IT Training &amp; Testing Services</div>
          <div class="address">
            {branch_address_html}
            {branch_phone_html}
          </div>
        </div>
      </div>
      <div class="receipt-bar">
        <div class="receipt-title">{escape(document_title)}</div>
        <div class="receipt-no">{escape(document_number_label)}: {escape(document_number or '')}</div>
      </div>
      {partial_note}
      <div class="section">
        <div class="grid">
          <div class="field"><div class="label">STUDENT NAME</div><div class="value">{escape(student_name)}</div></div>
          <div class="field"><div class="label">STUDENT ID</div><div class="value">{escape(student_number)}</div></div>
          <div class="field"><div class="label">COURSE</div><div class="value">{escape(course_name)}</div></div>
          <div class="field"><div class="label">BRANCH</div><div class="value">{escape(branch_name)}</div></div>
          <div class="field"><div class="label">PAYMENT MODE</div><div class="value">{escape(payment_mode)}</div></div>
          <div class="field"><div class="label">REFERENCE NO</div><div class="value">{escape(reference_number)}</div></div>
          <div class="field"><div class="label">PAYMENT DATE</div><div class="value">{escape(payment_date)}</div></div>
          <div class="field"><div class="label">INSTALLMENT</div><div class="value">{escape(installment_label)}</div></div>
          <div class="field"><div class="label">PAID AMOUNT</div><div class="value amount">Rs {payment_amount:,.2f}</div></div>
          <div class="field"><div class="label">COURSE FEE</div><div class="value amount">Rs {total_fees:,.2f}</div></div>
          <div class="field"><div class="label">TOTAL PAYMENT PAID</div><div class="value amount">Rs {paid_amount:,.2f}</div></div>
          <div class="field"><div class="label">PENDING AMOUNT</div><div class="value amount">Rs {balance:,.2f}</div></div>
        </div>
      </div>
      <div class="section">
        <h2>Payment Schedule</h2>
        <table>
          <thead>
            <tr>
              <th>INSTALLMENT</th>
              <th>DUE DATE</th>
              <th class="amount">Amount</th>
              <th>STATUS</th>
            </tr>
          </thead>
          <tbody>
            {''.join(schedule_rows)}
          </tbody>
        </table>
      </div>
      <div class="generated-info">
        <p><span class="info-label">GENERATED BY:</span> <span class="info-value">Indra Institute of Education</span></p>
        <p><span class="info-label">GENERATED ON:</span> <span class="info-value">{escape(self._receipt_date(generated_at))}</span></p>
      </div>
      <div class="bottom">
        <div>Fees once paid cannot be refunded.</div>
        <div>This is a computer-generated document, no signature required.</div>
      </div>
    </div>
  </body>
</html>"""

    def _branch_receipt_code(self, branch):
        if not branch:
            return 'GEN'
        raw_code = str(branch.branch_code or '').strip().upper()
        if raw_code and not raw_code.isdigit():
            return re.sub(r'[^A-Z0-9]', '', raw_code)[:6] or 'GEN'
        branch_name_key = re.sub(r'[^a-z0-9]', '', branch.name or '').lower()
        known_codes = {
            'kuniyamuthur': 'KUN',
            'gandhipuram': 'GDP',
            'hopes': 'HOP',
        }
        if branch_name_key in known_codes:
            return known_codes[branch_name_key]
        words = re.findall(r'[A-Za-z0-9]+', branch.name or '')
        if len(words) >= 2:
            return ''.join(word[0] for word in words).upper()[:3]
        return re.sub(r'[^A-Z0-9]', '', (branch.name or 'GEN').upper())[:3] or 'GEN'

    def _generate_document_number(self, branch, document_type):
        year = timezone.localdate().year
        branch_code = self._branch_receipt_code(branch)
        prefix = 'BILL' if document_type == PaymentInstallment.DocumentType.BILL else 'RCPT'
        number_prefix = f'{prefix}-{branch_code}-{year}-'
        field = 'bill_number__startswith' if document_type == PaymentInstallment.DocumentType.BILL else 'receipt_number__startswith'
        generated_count = PaymentInstallment.objects.filter(enrollment__branch=branch, **{field: number_prefix}).count()
        return f'{number_prefix}{generated_count + 1:04d}'

    @action(detail=True, methods=['post'], url_path='generate-bill')
    def generate_bill(self, request, pk=None):
        installment = self.get_object()
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can generate bills.'}, status=403)
        if installment.bill_number:
            return Response({'detail': 'Official bill already generated for this payment entry.'}, status=400)

        installment_status = PaymentInstallmentSerializer(installment).data.get('installment_status')
        if installment_status != 'paid':
            return Response({'detail': 'Official bill can be generated only after the required installment amount is fully paid.'}, status=400)
        installment.bill_number = self._generate_document_number(installment.enrollment.branch, PaymentInstallment.DocumentType.BILL)
        installment.document_type = PaymentInstallment.DocumentType.BILL
        installment.bill_generated_at = timezone.now()
        installment.bill_generated_by = request.user
        installment.bill_total = installment.amount
        installment.document_snapshot = self._build_document_snapshot(installment)
        installment.document_html = self._build_bill_html(installment, installment.document_snapshot)
        installment.save(update_fields=[
            'document_type', 'bill_number', 'bill_generated_at', 'bill_generated_by',
            'bill_total', 'document_snapshot', 'document_html',
        ])
        self._notify_collector_document_generated(installment)
        return Response(PaymentInstallmentSerializer(installment).data)

    @action(detail=True, methods=['post'], url_path='generate-receipt')
    def generate_receipt(self, request, pk=None):
        installment = self.get_object()
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can generate receipts.'}, status=403)
        if installment.bill_number:
            return Response({'detail': 'Official bill already generated for this payment entry.'}, status=400)
        if not installment.receipt_number:
            installment.receipt_number = self._generate_document_number(installment.enrollment.branch, PaymentInstallment.DocumentType.RECEIPT)
        installment.document_type = PaymentInstallment.DocumentType.RECEIPT
        installment.bill_generated_at = timezone.now()
        installment.bill_generated_by = request.user
        installment.bill_total = installment.amount
        installment.document_snapshot = self._build_document_snapshot(installment)
        installment.document_html = self._build_bill_html(installment, installment.document_snapshot)
        installment.save(update_fields=[
            'document_type', 'receipt_number', 'bill_generated_at', 'bill_generated_by',
            'bill_total', 'document_snapshot', 'document_html',
        ])
        self._notify_collector_document_generated(installment)
        return Response(PaymentInstallmentSerializer(installment).data)

    def _immutable_document_html(self, installment):
        snapshot = installment.document_snapshot or {}
        if snapshot.get('calculation_version') != 'bill_amount_v2':
            snapshot = self._build_document_snapshot(installment)
        html = self._build_bill_html(installment, snapshot)
        installment.document_snapshot = snapshot
        installment.document_html = html
        if installment.bill_number and not installment.bill_total:
            installment.bill_total = Decimal(str(snapshot.get('bill_total') or installment.amount or 0))
        installment.save(update_fields=['document_snapshot', 'document_html', 'bill_total'])
        return html

    def _build_document_image(self, installment):
        try:
            from PIL import Image, ImageDraw, ImageFont
        except ImportError as exc:
            raise RuntimeError('Pillow is required to generate bill documents.') from exc

        snapshot = installment.document_snapshot or self._build_document_snapshot(installment)
        branch = installment.enrollment.branch
        branch_header = self._branch_header_payload(branch)
        branch_address_lines = snapshot.get('branch_address_lines') or branch_header['branch_address_lines']
        branch_phone = snapshot.get('branch_phone') or branch_header['branch_phone']
        document_number = snapshot.get('document_number') or self._document_number(installment)
        document_title = 'OFFICIAL PAYMENT BILL' if snapshot.get('document_type') == 'bill' else 'PAYMENT RECEIPT'
        document_label = 'Bill No' if snapshot.get('document_type') == 'bill' else 'Receipt No'
        width, height = 1240, 1754
        margin = 82
        navy = '#1E3A5F'
        slate = '#334155'
        border = '#CBD5E1'
        light = '#F8FAFC'
        page = Image.new('RGB', (width, height), 'white')
        draw = ImageDraw.Draw(page)
        font = ImageFont.load_default()
        title_font = ImageFont.load_default()
        label_font = ImageFont.load_default()

        header_height = 190
        draw.rectangle((0, 0, width, header_height), fill=navy)
        logo_candidates = [
            Path(settings.BASE_DIR) / 'frontend' / 'public' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'brand-logo.png',
        ]
        for logo_path in logo_candidates:
            if logo_path.exists():
                logo = Image.open(logo_path).convert('RGBA')
                logo.thumbnail((112, 112))
                page.paste(logo, (margin, 34), logo)
                break

        def center_text(text, y, fill='white', selected_font=None):
            selected_font = selected_font or title_font
            bbox = draw.textbbox((0, 0), text, font=selected_font)
            draw.text(((width - (bbox[2] - bbox[0])) / 2, y), text, fill=fill, font=selected_font)

        center_text('INDRA INSTITUTE OF EDUCATION', 32)
        center_text('IT Training & Testing Services', 64)
        address_lines = []
        for line in branch_address_lines:
            address_lines.extend(wrap_text(draw, str(line), font, width - 360) or [str(line)])
        if branch_phone:
            address_lines.append(str(branch_phone))
        for line_index, line in enumerate(address_lines[:4]):
            center_text(str(line), 98 + (line_index * 22), selected_font=font)

        y = header_height
        draw.rectangle((0, y, width, y + 70), fill='white', outline=border)
        draw.text((margin, y + 24), document_title, fill='black', font=title_font)
        number_label = f'{document_label}: {document_number or ""}'
        number_bbox = draw.textbbox((0, 0), number_label, font=title_font)
        draw.text((width - margin - (number_bbox[2] - number_bbox[0]), y + 24), number_label, fill=navy, font=title_font)

        fields = [
            ('STUDENT NAME', snapshot.get('student_name') or ''),
            ('STUDENT ID', snapshot.get('student_number') or ''),
            ('COURSE', snapshot.get('course_name') or ''),
            ('BRANCH', snapshot.get('branch_name') or ''),
            ('PAYMENT MODE', snapshot.get('payment_mode') or ''),
            ('REFERENCE NO', snapshot.get('reference_number') or ''),
            ('PAYMENT DATE', snapshot.get('payment_date') or ''),
            ('INSTALLMENT', snapshot.get('installment_label') or ''),
            ('PAID AMOUNT', f"Rs {Decimal(str(snapshot.get('payment_amount') or 0)):,.2f}"),
            ('COURSE FEE', f"Rs {Decimal(str(snapshot.get('total_fees') or 0)):,.2f}"),
            ('TOTAL PAYMENT PAID', f"Rs {Decimal(str(snapshot.get('paid_amount') or 0)):,.2f}"),
            ('BALANCE', f"Rs {Decimal(str(snapshot.get('balance') or 0)):,.2f}"),
        ]

        y += 110
        col_width = (width - (margin * 2) - 44) / 2
        row_height = 80
        for index, (label, value) in enumerate(fields):
            col = index % 2
            row = index // 2
            x = margin + col * (col_width + 44)
            row_y = y + row * row_height
            draw.text((x, row_y), label, fill=slate, font=label_font)
            for line_index, line in enumerate(wrap_text(draw, str(value), font, int(col_width))[:3]):
                draw.text((x, row_y + 26 + (line_index * 22)), line, fill='black', font=font)

        table_y = y + 6 * row_height + 30
        draw.text((margin, table_y), 'PAYMENT SCHEDULE', fill='black', font=title_font)
        table_y += 34
        headers = ['INSTALLMENT', 'DUE DATE', 'AMOUNT', 'STATUS']
        col_widths = [420, 230, 230, 180]
        x = margin
        for header, col_width_item in zip(headers, col_widths):
            draw.rectangle((x, table_y, x + col_width_item, table_y + 42), fill=light, outline=border)
            draw.text((x + 12, table_y + 14), header, fill=slate, font=label_font)
            x += col_width_item
        y_row = table_y + 42
        for item in (snapshot.get('payment_schedule') or [])[:8]:
            values = [
                item.get('label') or '',
                self._receipt_date(item.get('due_date')) if item.get('due_date') else '',
                f"Rs {Decimal(str(item.get('amount') or 0)):,.2f}",
                item.get('status') or '',
            ]
            x = margin
            for value, col_width_item in zip(values, col_widths):
                draw.rectangle((x, y_row, x + col_width_item, y_row + 48), fill='white', outline=border)
                for line_index, line in enumerate(wrap_text(draw, str(value), font, col_width_item - 22)[:2]):
                    draw.text((x + 12, y_row + 12 + (line_index * 18)), line, fill='black', font=font)
                x += col_width_item
            y_row += 48

        info_y = y_row + 64
        draw.line((margin, info_y - 24, width - margin, info_y - 24), fill=border, width=2)
        draw.text((margin, info_y), 'GENERATED BY: Indra Institute of Education', fill=slate, font=font)
        draw.text((margin, info_y + 34), f"GENERATED ON: {self._receipt_date(snapshot.get('generated_at'))}", fill=slate, font=font)

        footer_y = height - 92
        draw.rectangle((0, footer_y, width, height), fill=navy)
        draw.text((margin, footer_y + 34), 'Fees once paid cannot be refunded.', fill='white', font=font)
        footer_text = 'This is a computer-generated document, no signature required.'
        footer_bbox = draw.textbbox((0, 0), footer_text, font=font)
        draw.text((width - margin - (footer_bbox[2] - footer_bbox[0]), footer_y + 34), footer_text, fill='white', font=font)

        return page

    def _build_document_pdf(self, installment):
        page = self._build_document_image(installment)
        output = io.BytesIO()
        page.save(output, format='PDF')
        output.seek(0)
        return output.read()

    def _build_document_png(self, installment):
        page = self._build_document_image(installment)
        output = io.BytesIO()
        page.save(output, format='PNG')
        output.seek(0)
        return output.read()

    @action(detail=True, methods=['get'], url_path='view-bill')
    def view_bill(self, request, pk=None):
        installment = self.get_object()
        document_number = installment.bill_number
        if not document_number:
            return Response({'detail': 'Bill has not been generated yet.'}, status=404)

        response = HttpResponse(self._immutable_document_html(installment), content_type='text/html; charset=utf-8')
        response['Content-Disposition'] = f'inline; filename="{document_number}.html"'
        return response

    @action(detail=True, methods=['post'], url_path='send-bill')
    def send_bill(self, request, pk=None):
        installment = self.get_object()
        if request.user.is_super_admin:
            return Response({'detail': 'Admins can generate and verify bills, but cannot send bills.'}, status=403)
        document_number = installment.bill_number
        if not document_number:
            return Response({'detail': 'Generate the bill before sending it.'}, status=400)

        enrollment = installment.enrollment
        image_filename = f'{document_number}.png'
        installment.document_snapshot = self._build_document_snapshot(installment)
        installment.document_html = self._build_bill_html(installment, installment.document_snapshot)
        installment.save(update_fields=['document_snapshot', 'document_html'])
        try:
            image_bytes = self._build_document_png(installment)
        except RuntimeError as exc:
            return Response({'detail': str(exc)}, status=503)
        whatsapp_message = self._bill_whatsapp_message(enrollment, installment)
        if self._bill_api_delivery_enabled() and WATIClient().is_configured:
            log = send_candidate_document(
                candidate_name=enrollment.name,
                phone=enrollment.phone,
                message_type=WhatsAppMessage.MsgType.MANUAL,
                caption=whatsapp_message,
                file_bytes=image_bytes,
                filename=image_filename,
                content_type='image/png',
                sent_by=request.user,
                related_model='payment_installment',
                related_id=installment.id,
            )
            detail = 'Bill sent successfully.' if log.status == WhatsAppMessage.MsgStatus.SENT else 'Bill send request failed.'
            return Response({
                'detail': detail,
                'phone': enrollment.phone,
                'document_number': document_number,
                'document_filename': image_filename,
                'whatsapp_message': whatsapp_message,
                'sent_at': log.sent_at or log.created_at,
                'sent_at_display': timezone.localtime(log.sent_at or log.created_at).strftime('%d-%b-%Y %I:%M %p'),
                'sent_by': request.user.full_name or request.user.username,
                **whatsapp_send_payload(log),
            })

        self._store_bill_image(request, installment, image_bytes)
        return Response({
            'detail': 'Bill image generated. Share it through WhatsApp.',
            'phone': enrollment.phone,
            'document_number': document_number,
            'document_filename': image_filename,
            'bill_image_data': base64.b64encode(image_bytes).decode('ascii'),
            'bill_image_content_type': 'image/png',
            'share_mode': 'browser_file_share',
            'whatsapp_message': whatsapp_message,
            'whatsapp_sent': False,
            'whatsapp_status': 'prepared',
            'whatsapp_error': '',
            'whatsapp_provider': 'browser_file_share',
        })

    @action(detail=True, methods=['post'], url_path='confirm-bill-sent')
    def confirm_bill_sent(self, request, pk=None):
        installment = self.get_object()
        if request.user.is_super_admin:
            return Response({'detail': 'Admins can generate and verify bills, but cannot send bills.'}, status=403)
        document_number = installment.bill_number
        if not document_number:
            return Response({'detail': 'Generate the bill before sending it.'}, status=400)

        enrollment = installment.enrollment
        message = self._bill_whatsapp_message(enrollment, installment)
        log = log_whatsapp_message(
            candidate_name=enrollment.name,
            phone=enrollment.phone,
            message_type=WhatsAppMessage.MsgType.MANUAL,
            message_body=message,
            result={
                'provider': 'browser_file_share',
                'document_number': document_number,
                'filename': f'{document_number}.png',
            },
            template_name=f'{document_number}.png',
            sent_by=request.user,
            related_model='payment_installment',
            related_id=installment.id,
            provider='browser_file_share',
        )
        return Response({
            'detail': 'Bill sent successfully.',
            'phone': enrollment.phone,
            'document_number': document_number,
            'document_filename': f'{document_number}.png',
            'whatsapp_message': message,
            'sent_at': log.sent_at or log.created_at,
            'sent_at_display': timezone.localtime(log.sent_at or log.created_at).strftime('%d-%b-%Y %I:%M %p'),
            'sent_by': request.user.full_name or request.user.username,
            **whatsapp_send_payload(log),
        })


class PublicBillView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, installment_id, document_number):
        installment = PaymentInstallment.objects.select_related(
            'payment',
            'enrollment__branch',
            'enrollment__course',
            'bill_generated_by',
        ).filter(pk=installment_id, bill_number=document_number).first()
        if not installment:
            return Response({'detail': 'Bill not found.'}, status=404)

        builder = PaymentInstallmentViewSet()
        try:
            pdf_bytes = builder._build_document_pdf(installment)
        except RuntimeError:
            response = HttpResponse(builder._immutable_document_html(installment), content_type='text/html; charset=utf-8')
            response['Content-Disposition'] = f'inline; filename="{document_number}.html"'
            return response

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{document_number}.pdf"'
        return response


class AdminReceiptViewSet(viewsets.ModelViewSet):
    """Admin-only standalone receipts for miscellaneous payments."""
    serializer_class = AdminReceiptSerializer
    permission_classes = [IsSuperAdmin]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['purpose', 'payment_mode', 'payment_date']
    search_fields = ['receipt_number', 'name', 'phone', 'purpose']
    ordering_fields = ['payment_date', 'generated_on', 'amount', 'name', 'purpose']
    ordering = ['-payment_date', '-created_at']
    DEFAULT_BRANCH_NAME = 'Gandhipuram'
    SYSTEM_BRANCH_ADDRESS = 'Indra Institute of Education, Gandhipuram, Coimbatore'
    SYSTEM_BRANCH_PHONE = '9000000001'

    def get_queryset(self):
        return AdminReceipt.objects.select_related('generated_by__branch').all()

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except Exception:
            logger.exception('Admin receipt list failed.')
            return Response({'detail': 'Failed to load receipts.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
        try:
            return super().retrieve(request, *args, **kwargs)
        except Exception:
            logger.exception('Admin receipt retrieve failed: receipt_id=%s.', kwargs.get('pk'))
            return Response({'detail': 'Failed to load receipt.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except Exception:
            logger.exception('Admin receipt create failed.')
            return Response({'detail': 'Failed to create receipt.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _generate_receipt_number(self):
        year = timezone.localdate().year
        prefix = f'IIE-REC-{year}-'
        count = AdminReceipt.objects.filter(receipt_number__startswith=prefix).count()
        return f'{prefix}{count + 1:04d}'

    def perform_create(self, serializer):
        receipt_number = self._generate_receipt_number()
        while AdminReceipt.objects.filter(receipt_number=receipt_number).exists():
            sequence = int(receipt_number.rsplit('-', 1)[1]) + 1
            receipt_number = f'IIE-REC-{timezone.localdate().year}-{sequence:04d}'
        serializer.save(
            receipt_number=receipt_number,
            generated_by=self.request.user,
            generated_on=timezone.now(),
        )

    def _receipt_date(self, value):
        if not value:
            return timezone.localdate().strftime('%d/%m/%Y')
        if isinstance(value, str):
            value = parse_date(value)
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y')
        return str(value)

    def _receipt_datetime(self, value):
        if not value:
            value = timezone.now()
        if isinstance(value, str):
            parsed = parse_date(value)
            if parsed:
                return parsed.strftime('%d/%m/%Y')
        if hasattr(value, 'utcoffset'):
            value = timezone.localtime(value)
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y %I:%M %p')
        return str(value)

    def _logo_src(self):
        logo_candidates = [
            Path(settings.BASE_DIR) / 'frontend' / 'public' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'brand-logo.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'brand-logo.png',
            Path(settings.BASE_DIR) / 'frontend' / 'dist' / 'receipt-logo.png',
        ]
        for logo_path in logo_candidates:
            if logo_path.exists():
                return f"data:image/png;base64,{base64.b64encode(logo_path.read_bytes()).decode('ascii')}"
        return ''

    def _branch_header_payload(self, branch):
        try:
            default_branch = Branch.objects.filter(name__iexact=self.DEFAULT_BRANCH_NAME).first()
        except Exception:
            logger.exception('Admin receipt branch lookup failed.')
            default_branch = None
        branch = default_branch or branch
        lines = []
        if branch:
            lines = [line.strip() for line in str(branch.address or '').splitlines() if line.strip()]
            city_line = ', '.join([
                value for value in [
                    str(branch.city or '').strip(),
                    str(branch.state or '').strip(),
                    str(branch.pincode or '').strip(),
                ]
                if value
            ])
            if city_line:
                lines.append(city_line)
        fallback_address = getattr(settings, 'ADMIN_RECEIPT_DEFAULT_BRANCH_ADDRESS', self.SYSTEM_BRANCH_ADDRESS)
        fallback_phone = getattr(settings, 'ADMIN_RECEIPT_DEFAULT_BRANCH_PHONE', self.SYSTEM_BRANCH_PHONE)
        return {
            'branch_name': branch.name if branch and branch.name else getattr(settings, 'ADMIN_RECEIPT_DEFAULT_BRANCH_NAME', 'Gandhipuram Branch'),
            'branch_address_lines': lines or [fallback_address],
            'branch_phone': branch.phone if branch and branch.phone else fallback_phone,
        }

    def _amount_in_words(self, amount):
        amount = Decimal(str(amount or 0)).quantize(Decimal('1'))
        number = int(amount)
        if number == 0:
            return 'Zero Rupees Only'
        ones = [
            '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
            'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
            'Seventeen', 'Eighteen', 'Nineteen',
        ]
        tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

        def under_hundred(value):
            if value < 20:
                return ones[value]
            return ' '.join(part for part in [tens[value // 10], ones[value % 10]] if part)

        def under_thousand(value):
            if value < 100:
                return under_hundred(value)
            return ' '.join(part for part in [ones[value // 100], 'Hundred', under_hundred(value % 100)] if part)

        parts = []
        for divisor, label in ((10000000, 'Crore'), (100000, 'Lakh'), (1000, 'Thousand')):
            if number >= divisor:
                parts.append(f'{under_thousand(number // divisor)} {label}')
                number %= divisor
        if number:
            parts.append(under_thousand(number))
        return f'{" ".join(parts)} Rupees Only'

    def _build_receipt_html(self, receipt):
        logo_src = self._logo_src()
        generated_by = receipt.generated_by.full_name if receipt.generated_by else 'Admin'
        branch_header = self._branch_header_payload(receipt.generated_by.branch if receipt.generated_by_id and receipt.generated_by else None)
        branch_address_html = ''.join(f'<p>{escape(str(line))}</p>' for line in branch_header['branch_address_lines'])
        branch_phone_html = f'<p>{escape(str(branch_header["branch_phone"]))}</p>' if branch_header.get('branch_phone') else ''
        amount = Decimal(str(receipt.amount or 0))
        notes_row = ''
        if receipt.notes:
            notes_row = f'<div class="field field-wide"><div class="label">NOTES</div><div class="value">{escape(receipt.notes)}</div></div>'
        return f"""<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <title>{escape(receipt.receipt_number or 'Receipt')}</title>
    <style>
      * {{ box-sizing: border-box; }}
      body {{ font-family: Libertine, "Linux Libertine", "Libertinus Serif", Georgia, "Times New Roman", serif; color: #111827; margin: 14px; background: #F8FAFC; }}
      .sheet {{ max-width: 780px; margin: 0 auto; border: 1px solid #CBD5E1; background: white; box-shadow: 0 10px 30px rgba(15, 23, 42, 0.08); }}
      .header {{ position: relative; display: flex; align-items: center; justify-content: center; min-height: 126px; padding: 12px 24px; background: #1E3A5F; color: white; }}
      .logo {{ position: absolute; left: 22px; top: 18px; display: flex; align-items: center; justify-content: flex-start; }}
      .logo img {{ width: auto; height: 48px; object-fit: contain; display: block; }}
      .brand {{ width: 100%; min-width: 0; padding: 0 86px; text-align: center; }}
      .brand h1 {{ margin: 0 0 4px; font-size: 20px; font-weight: 800; letter-spacing: 0.08em; text-transform: uppercase; color: white; line-height: 1.15; }}
      .brand .tagline {{ margin: 0 0 7px; font-size: 12.5px; font-weight: 600; color: #F8FAFC; line-height: 1.2; }}
      .brand .address {{ margin: 0; }}
      .brand .address p {{ margin: 1px 0; font-size: 10.8px; line-height: 1.24; color: #F8FAFC; }}
      .receipt-bar {{ display: flex; align-items: center; justify-content: space-between; gap: 16px; min-height: 38px; padding: 9px 18px; border-bottom: 1px solid #CBD5E1; background: #ffffff; }}
      .receipt-title {{ font-size: 14px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.14em; color: #111827; line-height: 1.1; }}
      .receipt-no {{ font-size: 12px; font-weight: 800; color: #1E3A5F; line-height: 1.1; }}
      .section {{ padding: 14px 18px; border-bottom: 1px solid #CBD5E1; background: #ffffff; }}
      .grid {{ display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); column-gap: 24px; row-gap: 10px; }}
      .field {{ display: grid; grid-template-columns: 145px minmax(0, 220px); gap: 10px; align-items: baseline; min-height: 22px; }}
      .field-wide {{ grid-column: 1 / -1; grid-template-columns: 145px minmax(0, 1fr); }}
      .label {{ font-size: 9.8px; text-transform: uppercase; letter-spacing: 0.07em; color: #334155; line-height: 1.2; }}
      .value {{ font-size: 12.2px; font-weight: 800; color: #111827; line-height: 1.35; text-align: left; }}
      .amount {{ color: #1E3A5F; }}
      .generated-info {{ padding: 18px 24px; background: white; color: #334155; font-size: 11.5px; line-height: 1.8; }}
      .generated-info p {{ margin: 0; }}
      .generated-info .info-label {{ font-weight: 700; }}
      .generated-info .info-value {{ font-weight: 400; color: #111827; }}
      .bottom {{ display: flex; align-items: center; justify-content: space-between; gap: 24px; min-height: 52px; margin-top: 8px; padding: 10px 24px; background: #1E3A5F; border-top: 1px solid #d6dce5; color: white; font-size: 13px; line-height: 1.35; }}
      @media print {{
        body {{ margin: 0; background: white; }}
        .sheet {{ border: 1px solid #CBD5E1; max-width: none; box-shadow: none; }}
        .header, .bottom {{ print-color-adjust: exact; -webkit-print-color-adjust: exact; }}
      }}
    </style>
  </head>
  <body>
    <div class="sheet">
      <div class="header">
        <div class="logo">
          {'<img src="' + logo_src + '" alt="Indra Institute of Education logo">' if logo_src else ''}
        </div>
        <div class="brand">
          <h1>INDRA INSTITUTE OF EDUCATION</h1>
          <div class="tagline">IT Training &amp; Testing Services</div>
          <div class="address">
            {branch_address_html}
            {branch_phone_html}
          </div>
        </div>
      </div>
      <div class="receipt-bar">
        <div class="receipt-title">PAYMENT RECEIPT</div>
        <div class="receipt-no">Receipt No: {escape(receipt.receipt_number)}</div>
      </div>
      <div class="section">
        <div class="grid">
          <div class="field"><div class="label">RECEIPT NO</div><div class="value">{escape(receipt.receipt_number)}</div></div>
          <div class="field"><div class="label">PAYMENT DATE</div><div class="value">{escape(self._receipt_date(receipt.payment_date))}</div></div>
          <div class="field"><div class="label">NAME</div><div class="value">{escape(receipt.name)}</div></div>
          <div class="field"><div class="label">PHONE NUMBER</div><div class="value">{escape(receipt.phone)}</div></div>
          <div class="field"><div class="label">PURPOSE</div><div class="value">{escape(receipt.purpose)}</div></div>
          <div class="field"><div class="label">PAYMENT MODE</div><div class="value">{escape(receipt.get_payment_mode_display())}</div></div>
          <div class="field"><div class="label">AMOUNT</div><div class="value amount">Rs {amount:,.2f}</div></div>
          {notes_row}
        </div>
      </div>
      <div class="generated-info">
        <p><span class="info-label">GENERATED BY:</span> <span class="info-value">{escape(generated_by)}</span></p>
        <p><span class="info-label">GENERATED ON:</span> <span class="info-value">{escape(self._receipt_date(receipt.generated_on))}</span></p>
      </div>
      <div class="bottom">
        <div>Fees once paid cannot be refunded.</div>
        <div>This is a computer-generated bill, no signature required.</div>
      </div>
    </div>
  </body>
</html>"""

    def _build_receipt_pdf(self, receipt):
        try:
            from PIL import Image, ImageDraw, ImageFont
        except ImportError as exc:
            raise RuntimeError('Pillow is required to generate receipt PDFs.') from exc

        width, height = 1240, 1754
        margin = 76
        navy = '#1E3A5F'
        ink = '#111827'
        slate = '#334155'
        muted = '#64748b'
        border = '#CBD5E1'
        soft = '#F8FAFC'
        page = Image.new('RGB', (width, height), 'white')
        draw = ImageDraw.Draw(page)

        def load_font(weight, size):
            font_candidates = [
                Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'fonts' / f'libertinus-serif-{weight}.woff2',
                Path(settings.BASE_DIR) / 'staticfiles' / 'assets' / f'libertinus-serif-{weight}.woff2',
                Path('/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf'),
                Path('/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf'),
                Path('C:/Windows/Fonts/georgia.ttf'),
                Path('C:/Windows/Fonts/times.ttf'),
            ]
            for font_path in font_candidates:
                try:
                    if font_path.exists():
                        return ImageFont.truetype(str(font_path), size=size)
                except Exception:
                    continue
            return ImageFont.load_default()

        title_font = load_font(700, 42)
        header_font = load_font(700, 30)
        section_font = load_font(700, 25)
        label_font = load_font(700, 17)
        value_font = load_font(600, 24)
        body_font = load_font(400, 21)
        small_font = load_font(400, 18)

        branch_header = self._branch_header_payload(None)
        header_height = 245
        draw.rectangle((0, 0, width, header_height), fill=navy)
        logo_candidates = [
            Path(settings.BASE_DIR) / 'frontend' / 'public' / 'iie-white.png',
            Path(settings.BASE_DIR) / 'frontend' / 'src' / 'assets' / 'brand-logo.png',
        ]
        for logo_path in logo_candidates:
            if logo_path.exists():
                logo = Image.open(logo_path).convert('RGBA')
                logo.thumbnail((126, 126))
                page.paste(logo, (margin, 44), logo)
                break

        def center_text(text, y, fill='white', selected_font=None):
            selected_font = selected_font or title_font
            bbox = draw.textbbox((0, 0), text, font=selected_font)
            draw.text(((width - (bbox[2] - bbox[0])) / 2, y), text, fill=fill, font=selected_font)

        center_text('INDRA INSTITUTE OF EDUCATION', 34, selected_font=title_font)
        center_text(branch_header['branch_name'], 88, selected_font=header_font)
        address_lines = []
        for line in branch_header['branch_address_lines']:
            address_lines.extend(wrap_text(draw, str(line), body_font, width - 400) or [str(line)])
        if branch_header.get('branch_phone'):
            address_lines.append(f'Phone: {branch_header["branch_phone"]}')
        for line_index, line in enumerate(address_lines[:5]):
            center_text(str(line), 128 + (line_index * 25), selected_font=body_font)

        y = header_height + 34
        draw.rounded_rectangle((margin, y, width - margin, y + 92), radius=12, fill=soft, outline=border, width=2)
        draw.text((margin + 26, y + 28), 'PAYMENT RECEIPT', fill=ink, font=header_font)
        receipt_label = f'Receipt Number: {receipt.receipt_number or "-"}'
        receipt_bbox = draw.textbbox((0, 0), receipt_label, font=value_font)
        draw.text((width - margin - 26 - (receipt_bbox[2] - receipt_bbox[0]), y + 32), receipt_label, fill=navy, font=value_font)

        y += 128
        amount = Decimal(str(receipt.amount or 0))
        amount_words = self._amount_in_words(amount)
        generated_by = receipt.generated_by.full_name if receipt.generated_by else 'Admin'
        reference_number = str(getattr(receipt, 'reference_number', '') or '').strip()

        def section_title(title, current_y):
            draw.text((margin, current_y), title, fill=navy, font=section_font)
            draw.line((margin, current_y + 36, width - margin, current_y + 36), fill=border, width=2)
            return current_y + 58

        def draw_field(label, value, x, current_y, box_width, box_height=86):
            draw.rounded_rectangle((x, current_y, x + box_width, current_y + box_height), radius=10, fill=soft, outline=border, width=1)
            draw.text((x + 18, current_y + 15), label.upper(), fill=muted, font=label_font)
            lines = wrap_text(draw, str(value or '-'), value_font, box_width - 36)
            for line_index, line in enumerate(lines[:2]):
                draw.text((x + 18, current_y + 43 + (line_index * 26)), line, fill=ink, font=value_font if line_index == 0 else body_font)

        col_gap = 24
        col_width = (width - (margin * 2) - col_gap) // 2

        y = section_title('Receipt Details', y)
        draw_field('Receipt Number', receipt.receipt_number, margin, y, col_width)
        draw_field('Payment Date', self._receipt_date(receipt.payment_date), margin + col_width + col_gap, y, col_width)
        y += 122

        y = section_title('Received From', y)
        draw_field('Student Name', receipt.name, margin, y, col_width)
        draw_field('Phone Number', receipt.phone, margin + col_width + col_gap, y, col_width)
        y += 122

        y = section_title('Payment Information', y)
        purpose_height = 102
        draw_field('Purpose', receipt.purpose, margin, y, col_width, purpose_height)
        draw_field('Payment Mode', receipt.get_payment_mode_display(), margin + col_width + col_gap, y, col_width, purpose_height)
        y += purpose_height + 16
        if reference_number:
            draw_field('Reference Number', reference_number, margin, y, width - (margin * 2), 82)
            y += 110

        y = section_title('Amount', y)
        amount_box_height = 132
        draw.rounded_rectangle((margin, y, width - margin, y + amount_box_height), radius=12, fill=soft, outline=border, width=1)
        draw.text((margin + 22, y + 18), 'AMOUNT', fill=muted, font=label_font)
        draw.text((margin + 22, y + 50), f'Rs. {amount:,.2f}', fill=navy, font=title_font)
        draw.text((margin + 460, y + 18), 'AMOUNT IN WORDS', fill=muted, font=label_font)
        for line_index, line in enumerate(wrap_text(draw, amount_words, value_font, width - margin - 482)[:3]):
            draw.text((margin + 460, y + 50 + (line_index * 30)), line, fill=ink, font=value_font)
        y += amount_box_height + 38

        y = section_title('Generated Details', y)
        draw_field('Generated By', generated_by, margin, y, col_width)
        draw_field('Generated Date & Time', self._receipt_datetime(receipt.generated_on), margin + col_width + col_gap, y, col_width)
        y += 124

        if receipt.notes:
            y = section_title('Notes', y)
            note_height = 96
            draw.rounded_rectangle((margin, y, width - margin, y + note_height), radius=10, fill=soft, outline=border, width=1)
            for line_index, line in enumerate(wrap_text(draw, receipt.notes, body_font, width - (margin * 2) - 36)[:3]):
                draw.text((margin + 18, y + 18 + (line_index * 25)), line, fill=ink, font=body_font)

        footer_y = height - 92
        draw.rectangle((0, footer_y, width, height), fill=navy)
        draw.text((margin, footer_y + 24), 'Indra Institute of Education', fill='white', font=value_font)
        draw.text((margin, footer_y + 55), 'Fees once paid cannot be refunded.', fill='white', font=small_font)
        footer_text = 'This is a computer-generated receipt, no signature required.'
        footer_bbox = draw.textbbox((0, 0), footer_text, font=small_font)
        draw.text((width - margin - (footer_bbox[2] - footer_bbox[0]), footer_y + 38), footer_text, fill='white', font=small_font)

        output = io.BytesIO()
        page.save(output, format='PDF', resolution=120.0)
        output.seek(0)
        return output.read()

    @action(detail=True, methods=['get'], url_path='view-receipt')
    def view_receipt(self, request, pk=None):
        try:
            receipt = self.get_object()
            response = HttpResponse(self._build_receipt_pdf(receipt), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{receipt.receipt_number or f"receipt-{receipt.id}"}.pdf"'
            return response
        except Exception:
            logger.exception('Admin receipt PDF view failed: receipt_id=%s.', pk)
            return Response({'detail': 'Failed to open receipt PDF.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='download-receipt')
    def download_receipt(self, request, pk=None):
        try:
            receipt = self.get_object()
            response = HttpResponse(self._build_receipt_pdf(receipt), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{receipt.receipt_number or f"receipt-{receipt.id}"}.pdf"'
            return response
        except Exception:
            logger.exception('Admin receipt PDF download failed: receipt_id=%s.', pk)
            return Response({'detail': 'Failed to download receipt PDF.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class DeleteCandidatesView(APIView):
    permission_classes = [IsSuperAdmin]

    record_models = {
        'lead': Lead,
        'walkin': WalkIn,
        'enrollment': Enrollment,
    }

    def base_queryset(self, model):
        if model is Lead:
            return Lead.objects.select_related('branch', 'course', 'created_by', 'deleted_by', 'walkin')
        if model is WalkIn:
            return WalkIn.objects.select_related('branch', 'course', 'created_by', 'deleted_by', 'lead', 'enrollment')
        return Enrollment.objects.select_related('branch', 'course', 'created_by', 'enrolled_by', 'deleted_by', 'lead', 'walkin', 'walkin__lead')

    def filtered_queryset(self, model, request, deleted):
        qs = self.base_queryset(model).filter(is_deleted=deleted)
        branch_id = request.query_params.get('branch')
        search = (request.query_params.get('search') or '').strip()
        if branch_id and branch_id != 'all':
            qs = qs.filter(branch_id=branch_id)
        if search:
            search_filter = Q(name__icontains=search) | Q(phone__icontains=search)
            if model is Lead:
                search_filter |= Q(lead_number__icontains=search)
            elif model is WalkIn:
                search_filter |= Q(candidate_number__icontains=search)
            else:
                search_filter |= Q(student_number__icontains=search)
            qs = qs.filter(search_filter)
        return qs

    def get(self, request):
        deleted = request.query_params.get('deleted') == 'true'
        rows = []
        for model in (Lead, WalkIn, Enrollment):
            rows.extend(candidate_payload(candidate) for candidate in self.filtered_queryset(model, request, deleted))
        rows.sort(
            key=lambda row: row['deleted_at'] or row['added_date'] or timezone.now(),
            reverse=True,
        )
        return Response(rows)

    def get_candidate(self, record_type, pk):
        model = self.record_models.get(record_type)
        if not model:
            raise DjangoValidationError('Invalid candidate type.')
        return self.base_queryset(model).get(pk=pk)

    def post(self, request):
        record_type = request.data.get('record_type')
        candidate_id = request.data.get('id')
        action_name = request.data.get('action')
        model = self.record_models.get(record_type)
        if not model:
            return Response({'detail': 'Invalid candidate type.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            candidate = self.get_candidate(record_type, candidate_id)
        except model.DoesNotExist:
            return Response({'detail': 'Candidate not found.'}, status=status.HTTP_404_NOT_FOUND)

        records = linked_candidate_records(candidate)
        if action_name == 'delete':
            deleted_at = timezone.now()
            for record in records:
                record.is_deleted = True
                record.deleted_at = deleted_at
                record.deleted_by = request.user
                record.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by', 'updated_at'])
            return Response({'detail': 'Candidate moved to deleted history.'})
        if action_name == 'restore':
            for record in records:
                record.is_deleted = False
                record.deleted_at = None
                record.deleted_by = None
                record.save(update_fields=['is_deleted', 'deleted_at', 'deleted_by', 'updated_at'])
            return Response({'detail': 'Candidate restored.'})
        return Response({'detail': 'Invalid action.'}, status=status.HTTP_400_BAD_REQUEST)


def next_birthday(dob, today):
    try:
        birthday = dob.replace(year=today.year)
    except ValueError:
        birthday = dob.replace(year=today.year, month=2, day=28)
    if birthday < today:
        try:
            birthday = dob.replace(year=today.year + 1)
        except ValueError:
            birthday = dob.replace(year=today.year + 1, month=2, day=28)
    return birthday


def month_window(year, month):
    start = timezone.datetime(year, month, 1).date()
    end = start.replace(day=monthrange(year, month)[1])
    return start, end


def previous_months(count=3, start_date=None):
    current = start_date or timezone.localdate()
    year = current.year
    month = current.month
    result = []
    for _ in range(count):
        result.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return result


def calculate_user_monthly_rating(user, year=None, month=None):
    return calculate_kpi_rating(user, year, month)


class DashboardSummaryView(APIView):
    """
    GET /api/dashboard/summary/
    Returns KPI tiles for the logged-in user (branch-scoped for staff).
    """
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        user    = request.user
        today   = timezone.localdate()
        year = today.year
        month = today.month

        # Branch filter
        lead_qs   = visible_candidate_queryset(Lead.objects.all())
        walkin_qs = visible_candidate_queryset(WalkIn.objects.all())
        enroll_qs = visible_candidate_queryset(Enrollment.objects.all())
        pay_qs    = visible_payment_queryset(Payment.objects.all())
        collection_qs = PaymentInstallment.objects.all()
        transfer_qs = BranchTransferRequest.objects.filter(status=BranchTransferRequest.Status.PENDING)
        target_qs = BranchTarget.objects.filter(year=year, month=month)
        selected_branch = None

        if not user.is_super_admin:
            lead_qs = lead_qs.filter(branch=user.branch)
            walkin_qs = walkin_qs.filter(branch=user.branch)
            enroll_qs = enroll_qs.filter(branch=user.branch)
            pay_qs = pay_qs.filter(enrollment__branch=user.branch)
            collection_qs = collection_qs.filter(enrollment__branch=user.branch)
            transfer_qs = transfer_qs.filter(Q(current_branch=user.branch) | Q(requested_branch=user.branch))
            target_qs = target_qs.filter(branch=user.branch)
            selected_branch = user.branch
        else:
            branch_id = request.query_params.get('branch')
            if branch_id:
                selected_branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
                if selected_branch:
                    lead_qs = lead_qs.filter(branch=selected_branch)
                    walkin_qs = walkin_qs.filter(branch=selected_branch)
                    enroll_qs = enroll_qs.filter(branch=selected_branch)
                    pay_qs = pay_qs.filter(enrollment__branch=selected_branch)
                    collection_qs = collection_qs.filter(enrollment__branch=selected_branch)
                    transfer_qs = transfer_qs.filter(Q(current_branch=selected_branch) | Q(requested_branch=selected_branch))
                    target_qs = target_qs.filter(branch=selected_branch)

        total_fee_amount = pay_qs.aggregate(total=Sum('total_fees'))['total'] or 0
        total_paid_amount = pay_qs.aggregate(total=Sum('paid_amount'))['total'] or 0
        value_enroll_qs = reporting_enrollment_queryset(enroll_qs)
        total_value_amount = value_enroll_qs.aggregate(
            total=Sum('net_payable_fee')
        )['total'] or 0
        dashboard_monthly_enroll_qs = value_enroll_qs.filter(enrollment_date__year=year, enrollment_date__month=month)
        value_this_month = dashboard_monthly_enroll_qs.aggregate(
            total=Sum('net_payable_fee')
        )['total'] or 0
        current_month_collected_amount = collection_qs.filter(
            payment_date__year=year,
            payment_date__month=month,
        ).aggregate(total=Sum('amount'))['total'] or 0
        target_count = target_qs.count()
        target_totals = target_qs.aggregate(
            lead_target=Sum('lead_target'),
            walkin_target=Sum('walkin_target'),
            enroll_target=Sum('enroll_target'),
            value_target=Sum('revenue_target'),
        )
        birthday_rows = []
        if not user.is_super_admin:
            birthday_start = today
            birthday_end = today + timedelta(days=6)
            birthday_rows = [
                {
                    'id': enrollment.id,
                    'name': enrollment.name,
                    'phone': enrollment.phone,
                    'birthday_date': next_birthday(enrollment.dob, today),
                    'days_left': (next_birthday(enrollment.dob, today) - today).days,
                    'course_name': enrollment.course.name if enrollment.course else '',
                    'branch_name': enrollment.branch.name if enrollment.branch else '',
                }
                for enrollment in official_enrollment_queryset(enroll_qs).exclude(dob__isnull=True).select_related('course', 'branch').order_by('name')
                if birthday_start <= next_birthday(enrollment.dob, today) <= birthday_end
            ]
            birthday_rows.sort(key=lambda row: (row['days_left'], row['name']))

        week_start = today - timedelta(days=(today.weekday() + 1) % 7)
        week_end = week_start + timedelta(days=6)
        active_pending_due_qs = pay_qs.filter(
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            paid_amount__lt=F('total_fees'),
            next_payment_date__isnull=False,
            next_payment_date__lte=today,
        )
        weekly_pending_qs = pay_qs.filter(
            status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL],
            paid_amount__lt=F('total_fees'),
            next_payment_date__isnull=False,
            next_payment_date__gte=week_start,
            next_payment_date__lte=week_end,
        )
        weekly_totals = weekly_pending_qs.aggregate(
            total=Sum('total_fees'),
            paid=Sum('paid_amount'),
        )
        weekly_pending_amount = (weekly_totals['total'] or 0) - (weekly_totals['paid'] or 0)
        leads_this_month = lead_qs.filter(created_at__year=year, created_at__month=month).count()
        walkins_this_month = walkin_qs.filter(visit_date__year=year, visit_date__month=month).count()
        counted_enroll_qs = value_enroll_qs
        enroll_this_month = dashboard_monthly_enroll_qs.count()
        conversion_rate = bounded_ratio(enroll_this_month, walkins_this_month)

        return Response({
            'total_leads':        lead_qs.count(),
            'leads_this_month':   leads_this_month,
            'leads_followup_today': active_lead_follow_up_queryset(
                lead_qs,
                'next_follow_up_date',
                today,
            ).count(),
            'total_walkins':      walkin_qs.count(),
            'walkins_this_month': walkins_this_month,
            'walkins_followup_today': active_walkin_follow_up_queryset(
                walkin_qs,
                'follow_up_date',
                today,
            ).count(),
            'total_enrollments':  counted_enroll_qs.count(),
            'enroll_this_month':  enroll_this_month,
            'conversion_rate':    conversion_rate,
            'total_revenue':      total_value_amount,
            'total_value':        total_value_amount,
            'value_this_month':   value_this_month,
            'revenue_this_month': value_this_month,
            'current_month_collected_amount': current_month_collected_amount,
            'monthly_collection': current_month_collected_amount,
            'pending_payments':   active_pending_due_qs.count(),
            'this_week_pending_payments': weekly_pending_amount,
            'pending_transfer_requests': transfer_qs.count(),
            'pending_amount':     total_fee_amount - total_paid_amount,
            'targets_set':        target_count > 0,
            'lead_target':        target_totals['lead_target'] if target_count else None,
            'walkin_target':      target_totals['walkin_target'] if target_count else None,
            'enroll_target':      target_totals['enroll_target'] if target_count else None,
            'value_target':       target_totals['value_target'] if target_count else None,
            'revenue_target':     target_totals['value_target'] if target_count else None,
            'today_birthdays':    birthday_rows,
            'selected_branch_id':  selected_branch.id if selected_branch else None,
            'selected_branch_name': selected_branch.name if selected_branch else 'All Branches',
            'performance_scope': 'branch' if user.is_super_admin else 'user',
        })


class DashboardBranchComparisonView(APIView):
    """GET /api/dashboard/branch-comparison/ — Super admin only."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        branches  = Branch.objects.filter(is_active=True)
        month_str = request.query_params.get('month', timezone.now().strftime('%Y-%m'))

        data = []
        for branch in branches:
            year, month = map(int, month_str.split('-'))
            start_date = date(year, month, 1)
            end_date = date(year, month, monthrange(year, month)[1])
            metrics = scoped_metrics(branch=branch.id, start_date=start_date, end_date=end_date)
            data.append({
                'branch_id':   branch.id,
                'branch_name': branch.name,
                'leads':       metrics['leads'],
                'walkins':     metrics['walkins'],
                'enrollments': metrics['enrollments'],
                'value':       metrics['revenue'],
            })
        return Response(data)


class DashboardTrendView(APIView):
    """GET /api/dashboard/trends/?days=30 — daily trend data."""
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        days   = int(request.query_params.get('days', 30))
        user   = request.user
        result = []

        for i in range(days - 1, -1, -1):
            day = (timezone.now() - timedelta(days=i)).date()
            lead_qs   = visible_candidate_queryset(Lead.objects.filter(created_at__date=day))
            walkin_qs = visible_candidate_queryset(WalkIn.objects.filter(visit_date=day))
            enroll_qs = enrollment_value_queryset(visible_candidate_queryset(Enrollment.objects.filter(enrollment_date=day)))

            if not user.is_super_admin:
                lead_qs   = lead_qs.filter(branch=user.branch)
                walkin_qs = walkin_qs.filter(branch=user.branch)
                enroll_qs = enroll_qs.filter(branch=user.branch)

            result.append({
                'date':        str(day),
                'leads':       lead_qs.count(),
                'walkins':     walkin_qs.count(),
                'enrollments': enroll_qs.count(),
            })
        return Response(result)


class DashboardHistoricalAnalyticsView(APIView):
    """
    GET /api/dashboard/historical-analytics/?branch=all
    Returns current-month historical analytics compared across 2023-2025.
    """
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        user = request.user
        current_month = timezone.localdate().month
        queryset = HistoricalAnalyticsEntry.objects.filter(
            year__in=[2023, 2024, 2025],
            month=current_month,
        )
        selected_branch = None

        if user.is_super_admin:
            branch_id = request.query_params.get('branch')
            if branch_id and branch_id != 'all':
                selected_branch = Branch.objects.filter(pk=branch_id, is_active=True).first()
                if not selected_branch:
                    return Response({'detail': 'Branch not found.'}, status=status.HTTP_404_NOT_FOUND)
                queryset = queryset.filter(branch=selected_branch)
        else:
            selected_branch = user.branch
            if not selected_branch:
                queryset = queryset.none()
            else:
                queryset = queryset.filter(branch=selected_branch)

        year_rows = {
            year: {
                'year': year,
                'label': str(year),
                'leads': 0,
                'walkins': 0,
                'enrollments': 0,
                'value_amount': Decimal('0'),
            }
            for year in (2023, 2024, 2025)
        }

        for row in queryset.values('year').annotate(
            leads=Sum('leads_count'),
            walkins=Sum('walkins_count'),
            enrollments=Sum('enrollments_count'),
            value_amount=Sum('value_amount'),
        ):
            year = row['year']
            year_rows[year]['leads'] = row['leads'] or 0
            year_rows[year]['walkins'] = row['walkins'] or 0
            year_rows[year]['enrollments'] = row['enrollments'] or 0
            year_rows[year]['value_amount'] = row['value_amount'] or Decimal('0')

        return Response({
            'month': current_month,
            'month_name': month_abbr[current_month],
            'branch_id': selected_branch.id if selected_branch else None,
            'branch_name': selected_branch.name if selected_branch else 'All Branches',
            'results': [year_rows[year] for year in (2023, 2024, 2025)],
        })


class DashboardMyRatingView(APIView):
    """GET /api/dashboard/my-rating/ - current user's monthly star rating."""
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        if request.user.role == User.Role.SUPER_ADMIN:
            today = timezone.localdate()
            return Response({
                'id': None,
                'user': request.user.id,
                'username': request.user.username,
                'full_name': request.user.full_name,
                'branch_name': request.user.branch.name if request.user.branch else None,
                'year': today.year,
                'month': today.month,
                'score': 0,
                'stars': 0,
                'star_display': '',
                'breakdown': {},
                'created_at': None,
                'updated_at': None,
                'applies': False,
            })

        today = timezone.localdate()
        rating = calculate_user_monthly_rating(request.user, today.year, today.month)
        return Response(UserMonthlyRatingSerializer(rating).data)


class UserRatingReportView(APIView):
    """GET /api/reports/user-ratings/ - KPI counselor ratings ranked by current month."""
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        months = previous_months(3)
        users = User.objects.filter(is_active=True).exclude(role=User.Role.SUPER_ADMIN).select_related('branch').order_by('username')
        rows = []
        for user in users:
            ratings = []
            for year, month in months:
                rating = calculate_user_monthly_rating(user, year, month)
                ratings.append(UserMonthlyRatingSerializer(rating).data)
            current_rating = ratings[0] if ratings else None
            previous_rating = ratings[1] if len(ratings) > 1 else None
            current_score = current_rating['score'] if current_rating else 0
            previous_score = previous_rating['score'] if previous_rating else 0
            score_trend = current_score - previous_score
            rows.append({
                'user_id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'counselor_name': user.full_name or user.username,
                'branch_name': user.branch.name if user.branch else None,
                'current_month_score': current_score,
                'score': current_score,
                'stars': current_rating['stars'] if current_rating else 1,
                'star_display': current_rating['star_display'] if current_rating else 'â­â˜†â˜†â˜†â˜†',
                'percentage': current_score,
                'previous_month_score': previous_score,
                'previous_month_stars': previous_rating['stars'] if previous_rating else 1,
                'trend': score_trend,
                'trend_direction': 'up' if score_trend > 0 else 'down' if score_trend < 0 else 'flat',
                'ratings': ratings,
            })
        rows.sort(key=lambda row: (-row['score'], row['counselor_name'].lower()))
        for index, row in enumerate(rows, start=1):
            row['rank'] = index
        return Response({
            'months': [{'year': year, 'month': month} for year, month in months],
            'results': rows,
        })


class ConversionFunnelReportView(APIView):
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        year = int(request.query_params.get('year') or timezone.localdate().year)
        month = int(request.query_params.get('month') or timezone.localdate().month)
        branch_id = request.query_params.get('branch')
        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
        lead_qs = visible_candidate_queryset(Lead.objects.filter(created_at__year=year, created_at__month=month))
        metrics = scoped_metrics(branch=branch_id if branch_id and branch_id != 'all' else None, start_date=start_date, end_date=end_date)
        if branch_id and branch_id != 'all':
            lead_qs = lead_qs.filter(branch_id=branch_id)
        total_leads = lead_qs.count()
        contacted = lead_qs.filter(status__in=[Lead.Status.CONTACTED, Lead.Status.INTERESTED, Lead.Status.FOLLOW_UP, Lead.Status.WALK_IN, Lead.Status.ENROLLED, Lead.Status.CONVERTED, Lead.Status.CONVERTED_TO_WALKIN]).count()
        walkins = metrics['walkins']
        enrollments = metrics['enrollments']

        return Response({
            'filters': {'year': year, 'month': month, 'branch': branch_id or 'all'},
            'funnel': [
                {'stage': 'Leads', 'count': total_leads, 'conversion_percent': 100 if total_leads else 0},
                {'stage': 'Contacted', 'count': contacted, 'conversion_percent': bounded_ratio(contacted, total_leads)},
                {'stage': 'Walk-ins', 'count': walkins, 'conversion_percent': bounded_ratio(walkins, total_leads)},
                {'stage': 'Enrollments', 'count': enrollments, 'conversion_percent': bounded_ratio(enrollments, walkins)},
            ],
        })


class BranchPerformanceComparisonReportView(APIView):
    permission_classes = [IsSuperAdmin]

    def get(self, request):
        year = int(request.query_params.get('year') or timezone.localdate().year)
        month = int(request.query_params.get('month') or timezone.localdate().month)
        today = timezone.localdate()
        rows = []
        for branch in Branch.objects.filter(is_active=True).order_by('name'):
            start_date = date(year, month, 1)
            end_date = date(year, month, monthrange(year, month)[1])
            metrics = scoped_metrics(branch=branch.id, start_date=start_date, end_date=end_date)
            leads = visible_candidate_queryset(Lead.objects.filter(branch=branch, created_at__year=year, created_at__month=month))
            walkins = visible_candidate_queryset(WalkIn.objects.filter(branch=branch, visit_date__year=year, visit_date__month=month))
            payments = visible_payment_queryset(Payment.objects.filter(enrollment__branch=branch))
            transfers_out = LeadTransferHistory.objects.filter(from_branch=branch, created_at__year=year, created_at__month=month).count()
            transfers_in = LeadTransferHistory.objects.filter(to_branch=branch, created_at__year=year, created_at__month=month).count()
            target = BranchTarget.objects.filter(branch=branch, year=year, month=month).first()
            missed_leads = visible_candidate_queryset(Lead.objects.filter(branch=branch, next_follow_up_date__lt=today)).exclude(status__in=[Lead.Status.ENROLLED, Lead.Status.CONVERTED, Lead.Status.CONVERTED_TO_WALKIN, Lead.Status.NOT_INTERESTED, Lead.Status.DROPPED, Lead.Status.LOST]).count()
            missed_walkins = visible_candidate_queryset(WalkIn.objects.filter(branch=branch, follow_up_date__lt=today)).exclude(status__in=[WalkIn.Status.CONVERTED, WalkIn.Status.NOT_INTERESTED]).count()
            completed_followups = FollowUp.objects.filter(created_at__year=year, created_at__month=month).filter(
                Q(record_type=FollowUp.RecordType.LEAD, record_id__in=leads.values('id')) |
                Q(record_type=FollowUp.RecordType.WALKIN, record_id__in=walkins.values('id'))
            ).count()
            due_followups = missed_leads + missed_walkins + completed_followups
            value = metrics['revenue']
            target_score = 0
            if target:
                achieved = sum([
                    metrics['leads'] >= target.lead_target,
                    metrics['walkins'] >= target.walkin_target,
                    metrics['enrollments'] >= target.enroll_target,
                    value >= target.revenue_target,
                ])
                target_score = round((achieved / 4) * 100, 2)
            pending_totals = payments.filter(status__in=[Payment.Status.UNPAID, Payment.Status.PARTIAL]).aggregate(
                total=Sum('total_fees'), paid=Sum('paid_amount')
            )
            rows.append({
                'branch_id': branch.id,
                'branch_name': branch.name,
                'leads': metrics['leads'],
                'transferred_leads': transfers_out,
                'received_leads': transfers_in,
                'walkins': metrics['walkins'],
                'enrollments': metrics['enrollments'],
                'value': value,
                'target_achievement': target_score,
                'follow_up_completion': round((completed_followups / due_followups) * 100, 2) if due_followups else 100,
                'payment_pending': (pending_totals['total'] or 0) - (pending_totals['paid'] or 0),
                'missed_followups': missed_leads + missed_walkins,
            })
        return Response(rows)


class ExportLeadsExcelView(APIView):
    """GET /api/reports/export/leads/ — Download leads as Excel."""
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'leads', 'leads_export')
        try:
            import openpyxl
        except ImportError:
            return Response(
                {'detail': 'Excel export is unavailable because openpyxl is not installed.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        qs = visible_candidate_queryset(Lead.objects.select_related('course','branch','assigned_to')).order_by('-created_at')
        if not request.user.is_super_admin:
            qs = qs.filter(branch=request.user.branch)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leads'
        headers = ['Lead No','Name','Phone','Location','Course',
                   'Status','Source','Source Description','Walk-in Date','Converted At','Branch','Assigned To','Created By','Created']
        ws.append(headers)

        for lead in qs:
            ws.append([
                lead.lead_number, lead.name, lead.phone, lead.location,
                lead.course.name if lead.course else '',
                automated_lead_status_display(lead), lead.get_source_display(),
                getattr(lead, 'source_description', ''),
                str(lead.walkin_date) if lead.walkin_date else '',
                lead.converted_at.strftime('%Y-%m-%d %H:%M') if lead.converted_at else '',
                lead.branch.name if lead.branch else '',
                lead.assigned_to.full_name if lead.assigned_to else '',
                lead.created_by.full_name if lead.created_by else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leads_export.xlsx"'
        return response


class ExportEnrollmentsExcelView(APIView):
    """GET /api/reports/export/enrollments/"""
    permission_classes = [IsStaffOrAdmin]

    def get(self, request):
        if not request.user.is_super_admin:
            return Response({'detail': 'Only admin can export data.'}, status=status.HTTP_403_FORBIDDEN)
        return admin_template_export_response(request, 'enrollments', 'enrollments_export')
        try:
            import openpyxl
        except ImportError:
            return Response(
                {'detail': 'Excel export is unavailable because openpyxl is not installed.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )

        qs = visible_candidate_queryset(Enrollment.objects.select_related(
            'course','branch','enrolled_by','payment'
        )).order_by('-enrollment_date')
        if not request.user.is_super_admin:
            qs = qs.filter(branch=request.user.branch)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Enrollments'
        headers = ['Student ID','Name','Phone','Email','Course','Branch',
                   'Actual Fees','Discount','Final Fees','Start Date',
                   'Enrollment Date','Status','Payment Status','Paid','Balance']
        ws.append(headers)

        for enroll in qs:
            pay = getattr(enroll, 'payment', None)
            ws.append([
                enroll.student_number, enroll.name, enroll.phone, enroll.email,
                enroll.course.name, enroll.branch.name if enroll.branch else '',
                float(enroll.actual_fees), float(enroll.discount_amount), float(enroll.final_fees),
                str(enroll.start_date) if enroll.start_date else '',
                str(enroll.enrollment_date), enroll.get_status_display(),
                pay.get_status_display() if pay else '',
                float(pay.paid_amount) if pay else 0,
                float(pay.balance) if pay else float(enroll.final_fees),
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="enrollments_export.xlsx"'
        return response
